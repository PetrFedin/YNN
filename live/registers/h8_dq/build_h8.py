#!/usr/bin/env python3
"""
H8: data-quality gate на COGS/марже.

Проблема: часть W3/H5 unit_cost >> цена продажи → ложная отрицательная маржа;
отдельные sales-строки с revenue≈0/1 портят агрегаты.

1) Пометить/исправить неадекватный COGS (prefer FILE, else drop W3 if unit>2.5×price)
2) Исключить garbage lines из marts
3) Пересобрать margin marts + anomalies

Не SoT.
"""
from __future__ import annotations

import csv
import json
import re
import shutil
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[3]
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/h8_dq_20260724"
MART = ROOT / "live/marts"
W4 = ROOT / "live/registers/w4_sales_settle"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def fnum(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def scrub_sales() -> tuple[list[dict], dict]:
    src = W4 / "sales_lines.csv"
    bak = W4 / "sales_lines_pre_h8.csv"
    if not bak.exists():
        shutil.copy2(src, bak)
    sales = list(csv.DictReader(open(src, encoding="utf-8")))

    # FILE cogs из H3 extended (первичный 1С)
    file_cogs = {}
    h3 = ROOT / "live/registers/h3_new_docs/sales_extended_2024_2026.csv"
    if h3.exists():
        for r in csv.DictReader(open(h3, encoding="utf-8")):
            if r.get("cogs_rub") not in (None, ""):
                file_cogs[r["sales_line_id"]] = float(r["cogs_rub"])

    # median price by canonical_sku from sane lines
    prices = defaultdict(list)
    for s in sales:
        q = fnum(s.get("qty")) or 0
        rev = fnum(s.get("revenue_rub")) or 0
        can = s.get("canonical_sku") or ""
        if can and q > 0 and rev >= 500:  # ignore junk
            prices[can].append(rev / q)

    med_price = {}
    for can, ps in prices.items():
        ps = sorted(ps)
        med_price[can] = ps[len(ps) // 2]

    fixed = 0
    junk = 0
    cleared_w3 = 0
    restored_file = 0
    for s in sales:
        flags = []
        q = fnum(s.get("qty")) or 0
        rev = fnum(s.get("revenue_rub")) or 0
        cogs = fnum(s.get("cogs_rub"))
        unit = fnum(s.get("w3_unit_cost"))
        src_c = s.get("cogs_source") or ""
        can = s.get("canonical_sku") or ""
        price = (rev / q) if q > 0 else None
        sid = s.get("sales_line_id") or ""

        # stubs / returns
        if rev is None or (rev == 0 and s.get("revenue_rub") in (None, "")):
            # пустая выручка = stub-строка дерева 1С без суммы
            if s.get("revenue_rub") in (None, ""):
                flags.append("EMPTY_REVENUE_STUB")
                junk += 1
        elif rev < 0:
            flags.append("RETURN_OR_CREDIT")
        elif rev < 100 and q <= 1 and abs(rev) < 100:
            # только явный мусор уровня 0–1 ₽, не stub
            if 0 < rev < 100:
                flags.append("JUNK_REVENUE")
                junk += 1
        if q <= 0 and rev and rev > 0:
            flags.append("ZERO_QTY")

        inflated = False
        ref = med_price.get(can) or price
        if unit and ref and ref > 0 and unit > ref * 2.5 and src_c in ("W3_H5", "H5_STOCK", "W3"):
            inflated = True
        if unit and price and price >= 500 and unit > price * 2.5 and src_c in ("W3_H5", "H5_STOCK", "W3"):
            inflated = True

        if inflated:
            flags.append("INFLATED_UNIT_COST")
            s["w3_unit_cost"] = ""
            s["w3_cost_version_id"] = ""
            cleared_w3 += 1
            # восстановить FILE cogs если есть и адекватен
            fc = file_cogs.get(sid)
            if fc is not None and rev and rev >= 500 and fc <= rev * 5:
                s["cogs_rub"] = round(fc, 2)
                s["cogs_source"] = "FILE"
                s["margin_rub"] = round(rev - fc, 2)
                restored_file += 1
            else:
                s["cogs_rub"] = ""
                s["cogs_source"] = ""
                s["margin_rub"] = ""
            fixed += 1

        # FILE cogs insane: cogs > 5× revenue on line (only positive sales)
        cogs = fnum(s.get("cogs_rub"))
        if cogs is not None and rev and rev >= 500 and cogs > rev * 5:
            flags.append("INFLATED_FILE_COGS")
            s["cogs_rub"] = ""
            s["cogs_source"] = ""
            s["margin_rub"] = ""
            fixed += 1

        # recompute margin if still have both
        cogs2 = fnum(s.get("cogs_rub"))
        rev2 = fnum(s.get("revenue_rub"))
        if cogs2 is not None and rev2 is not None and "EMPTY_REVENUE_STUB" not in flags and "JUNK_REVENUE" not in flags:
            s["margin_rub"] = round(rev2 - cogs2, 2)

        s["dq_flags"] = "|".join(flags)
        # stubs out of margin; returns KEEP (net view)
        s["dq_exclude_from_margin"] = "Y" if ("EMPTY_REVENUE_STUB" in flags or "JUNK_REVENUE" in flags or "ZERO_QTY" in flags) else "N"

    fields = list(sales[0].keys())
    for col in ("dq_flags", "dq_exclude_from_margin"):
        if col not in fields:
            fields.append(col)
    write_csv(src, sales, fields)
    write_csv(W4 / "sales_lines_1c.csv", [s for s in sales if s["channel"] in ("B2B", "IM")], fields)
    write_csv(OUT / "sales_dq_flags.csv", [s for s in sales if s.get("dq_flags")], fields)

    stats = {
        "sales_lines": len(sales),
        "junk_revenue_lines": junk,
        "cleared_inflated_w3": cleared_w3,
        "restored_file_cogs": restored_file,
        "fixed_total_ops": fixed,
        "flagged_lines": sum(1 for s in sales if s.get("dq_flags")),
        "excluded_from_margin": sum(1 for s in sales if s.get("dq_exclude_from_margin") == "Y"),
        "with_cogs_after": sum(1 for s in sales if s.get("cogs_rub") not in (None, "")),
        "with_cost_link_after": sum(1 for s in sales if s.get("w3_cost_version_id")),
    }
    return sales, stats


def rebuild_marts(sales: list[dict]) -> dict:
    by_cm = defaultdict(lambda: {"revenue": 0.0, "revenue_costed": 0.0, "cogs": 0.0, "qty": 0.0, "lines": 0, "lines_with_cogs": 0})
    by_sku = defaultdict(lambda: {"revenue": 0.0, "cogs": 0.0, "qty": 0.0, "lines": 0, "name": "", "channels": set()})

    for s in sales:
        if s.get("dq_exclude_from_margin") == "Y":
            continue
        ch = s.get("channel") or ""
        pm = s.get("period_month") or ""
        rev = fnum(s.get("revenue_rub")) or 0.0
        cogs = fnum(s.get("cogs_rub"))
        qty = fnum(s.get("qty")) or 0.0
        by_cm[(ch, pm)]["revenue"] += rev
        by_cm[(ch, pm)]["qty"] += qty
        by_cm[(ch, pm)]["lines"] += 1
        if cogs is not None:
            by_cm[(ch, pm)]["cogs"] += cogs
            by_cm[(ch, pm)]["revenue_costed"] += rev
            by_cm[(ch, pm)]["lines_with_cogs"] += 1

        can = s.get("canonical_sku") or ""
        if can:
            by_sku[can]["revenue"] += rev
            by_sku[can]["qty"] += qty
            by_sku[can]["lines"] += 1
            if cogs is not None:
                by_sku[can]["cogs"] += cogs
            if not by_sku[can]["name"]:
                by_sku[can]["name"] = (s.get("sku_name") or "")[:100]
            by_sku[can]["channels"].add(ch)

    channel_month = []
    for (ch, pm), v in sorted(by_cm.items()):
        margin = v["revenue_costed"] - v["cogs"] if v["lines_with_cogs"] else None
        pct = (margin / v["revenue_costed"] * 100) if margin is not None and v["revenue_costed"] else None
        cov = v["lines_with_cogs"] / v["lines"] if v["lines"] else 0
        channel_month.append(
            {
                "channel": ch,
                "period_month": pm,
                "lines": v["lines"],
                "lines_with_cogs": v["lines_with_cogs"],
                "cogs_coverage": round(cov, 3),
                "qty": round(v["qty"], 2),
                "revenue_rub": round(v["revenue"], 2),
                "revenue_costed_rub": round(v["revenue_costed"], 2),
                "cogs_rub": round(v["cogs"], 2) if v["lines_with_cogs"] else "",
                "margin_rub": round(margin, 2) if margin is not None else "",
                "margin_pct": round(pct, 1) if pct is not None else "",
                "status": "OK" if cov >= 0.85 else ("PARTIAL" if cov >= 0.5 else "WEAK"),
            }
        )

    by_ch = defaultdict(lambda: {"revenue": 0.0, "revenue_costed": 0.0, "cogs": 0.0, "lines": 0, "lines_with_cogs": 0})
    for r in channel_month:
        by_ch[r["channel"]]["revenue"] += r["revenue_rub"]
        by_ch[r["channel"]]["revenue_costed"] += r["revenue_costed_rub"]
        by_ch[r["channel"]]["lines"] += r["lines"]
        by_ch[r["channel"]]["lines_with_cogs"] += r["lines_with_cogs"]
        if r["cogs_rub"] != "":
            by_ch[r["channel"]]["cogs"] += float(r["cogs_rub"])

    channel_total = []
    for ch, v in sorted(by_ch.items()):
        margin = v["revenue_costed"] - v["cogs"]
        pct = margin / v["revenue_costed"] * 100 if v["revenue_costed"] else 0
        channel_total.append(
            {
                "channel": ch,
                "lines": v["lines"],
                "cogs_coverage": round(v["lines_with_cogs"] / v["lines"], 3) if v["lines"] else 0,
                "revenue_rub": round(v["revenue"], 2),
                "revenue_costed_rub": round(v["revenue_costed"], 2),
                "cogs_rub": round(v["cogs"], 2),
                "margin_rub": round(margin, 2),
                "margin_pct": round(pct, 1),
            }
        )

    sku_rows = []
    for can, v in by_sku.items():
        if v["revenue"] <= 0 or v["cogs"] <= 0:
            continue
        margin = v["revenue"] - v["cogs"]
        pct = margin / v["revenue"] * 100
        sku_rows.append(
            {
                "canonical_sku": can,
                "name": v["name"],
                "channels": ",".join(sorted(v["channels"])),
                "lines": v["lines"],
                "qty": round(v["qty"], 2),
                "revenue_rub": round(v["revenue"], 2),
                "cogs_rub": round(v["cogs"], 2),
                "margin_rub": round(margin, 2),
                "margin_pct": round(pct, 1),
            }
        )
    sku_rows.sort(key=lambda x: x["margin_rub"], reverse=True)
    top = sku_rows[:40]
    bottom = sorted(sku_rows, key=lambda x: x["margin_pct"])[:40]
    neg = [r for r in sku_rows if r["margin_pct"] < 0]

    write_csv(MART / "margin_channel_month.csv", channel_month, list(channel_month[0].keys()))
    write_csv(MART / "margin_channel_total.csv", channel_total, list(channel_total[0].keys()))
    write_csv(MART / "margin_sku_top40.csv", top, list(top[0].keys()) if top else ["canonical_sku"])
    write_csv(MART / "margin_sku_bottom40.csv", bottom, list(bottom[0].keys()) if bottom else ["canonical_sku"])
    write_csv(OUT / "margin_channel_total.csv", channel_total, list(channel_total[0].keys()))

    anomalies = []
    for r in channel_month:
        pct = fnum(r.get("margin_pct"))
        flags = []
        if r.get("status") == "WEAK":
            flags.append("WEAK_COGS_COVERAGE")
        if pct is not None and pct < 0:
            flags.append("NEGATIVE_MARGIN")
        if flags:
            anomalies.append(
                {
                    "scope": "CHANNEL_MONTH",
                    "channel": r["channel"],
                    "period_month": r["period_month"],
                    "canonical_sku": "",
                    "revenue_rub": r["revenue_rub"],
                    "margin_pct": r["margin_pct"],
                    "cogs_coverage": r["cogs_coverage"],
                    "flags": "|".join(flags),
                    "priority": "HIGH" if "NEGATIVE_MARGIN" in flags else "MED",
                }
            )
    for r in neg:
        anomalies.append(
            {
                "scope": "SKU",
                "channel": r["channels"],
                "period_month": "",
                "canonical_sku": r["canonical_sku"],
                "revenue_rub": r["revenue_rub"],
                "margin_pct": r["margin_pct"],
                "cogs_coverage": "",
                "flags": "NEGATIVE_MARGIN",
                "priority": "HIGH",
            }
        )
    write_csv(
        MART / "margin_anomalies.csv",
        anomalies,
        ["scope", "channel", "period_month", "canonical_sku", "revenue_rub", "margin_pct", "cogs_coverage", "flags", "priority"],
    )
    write_csv(
        MART / "margin_negative_skus.csv",
        [
            {
                "scope": "SKU",
                "channel": r["channels"],
                "period_month": "",
                "canonical_sku": r["canonical_sku"],
                "revenue_rub": r["revenue_rub"],
                "margin_pct": r["margin_pct"],
                "cogs_coverage": "",
                "flags": "NEGATIVE_MARGIN",
                "priority": "HIGH",
                "name": r["name"],
                "margin_rub": r["margin_rub"],
            }
            for r in neg
        ],
        ["scope", "channel", "period_month", "canonical_sku", "revenue_rub", "margin_pct", "cogs_coverage", "flags", "priority", "name", "margin_rub"],
    )

    overall_costed = sum(c["revenue_costed_rub"] for c in channel_total)
    overall_margin = sum(c["margin_rub"] for c in channel_total)
    return {
        "channel_totals": channel_total,
        "overall_revenue": round(sum(c["revenue_rub"] for c in channel_total), 2),
        "overall_revenue_costed": round(overall_costed, 2),
        "overall_margin": round(overall_margin, 2),
        "overall_margin_pct": round(overall_margin / overall_costed * 100, 1) if overall_costed else 0,
        "negative_skus_after": len(neg),
        "anomaly_rows": len(anomalies),
    }


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)

    before_neg = sum(1 for _ in csv.DictReader(open(MART / "margin_negative_skus.csv"))) if (MART / "margin_negative_skus.csv").exists() else None
    before_pct = None
    if (MART / "margin_channel_total.csv").exists():
        rows = list(csv.DictReader(open(MART / "margin_channel_total.csv")))
        costed = sum(float(r.get("revenue_costed_rub") or r.get("revenue_rub") or 0) for r in rows)
        # approximate from file if margin present
        marg = sum(float(r.get("margin_rub") or 0) for r in rows)
        before_pct = round(marg / costed * 100, 1) if costed else None

    sales, dq = scrub_sales()
    mart = rebuild_marts(sales)

    summary = {
        "generated_at": NOW,
        "wave": "H8",
        "dq": dq,
        "margin_before_pct": before_pct,
        "margin_before_neg_skus": before_neg,
        "margin_after": mart,
        "finding": (
            f"H8 DQ: cleared {dq['cleared_inflated_w3']} inflated W3 costs, "
            f"junk {dq['junk_revenue_lines']}; "
            f"neg SKUs {before_neg}→{mart['negative_skus_after']}; "
            f"margin {before_pct}%→{mart['overall_margin_pct']}% (costed)."
        ),
        "next": "RACI ACCEPT; finance review remaining neg SKUs (mostly FILE 1C cogs)",
        "not_sot": True,
    }
    json.dump(summary, open(OUT / "h8_summary.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    json.dump(summary, open(EV / "h8_summary.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    json.dump(summary, open(MART / "h8_summary.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)

    if (W4 / "w4_summary.json").exists():
        w4s = json.load(open(W4 / "w4_summary.json"))
        w4s["h8_dq"] = dq
        w4s["generated_at"] = NOW
        json.dump(w4s, open(W4 / "w4_summary.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)

    wb = Workbook()
    ws = wb.active
    ws.title = "00_Summary"
    for i, (k, v) in enumerate(summary.items(), 1):
        ws.cell(i, 1, k)
        ws.cell(i, 2, json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v)
    fill = PatternFill("solid", fgColor="1F4E79")

    def add(name, rows):
        w = wb.create_sheet(name)
        if not rows:
            return
        headers = list(rows[0].keys())
        for c, h in enumerate(headers, 1):
            cell = w.cell(1, c, h)
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
        for ri, row in enumerate(rows[:3000], 2):
            for ci, h in enumerate(headers, 1):
                w.cell(ri, ci, row.get(h, ""))

    add("01_Channel", mart["channel_totals"])
    flagged = [s for s in sales if s.get("dq_flags")]
    add("02_DQ_Flags", flagged[:2000])
    wb.save(EV / "YANINA_H8_DQ_EVIDENCE.xlsx")

    (OUT / "README.md").write_text(f"# H8 DQ\n\n{NOW}\n\n{summary['finding']}\n", encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
