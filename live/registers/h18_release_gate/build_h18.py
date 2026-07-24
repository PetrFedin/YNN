#!/usr/bin/env python3
"""
H18: Controls Release Gate (провизорные ворота месяца).

Зачем:
- После H15–H17 денежные multi-controls зелёные, но нет единой политики
  «можно ли считать месяц закрытым для управленческого обзора».
- Собираем помесячный gate из уже существующих controls_dashboard + Owner Actions.
- Это подготовка к фазе C3 (ворота релиза), без объявления SoT и без RACI autofill.

Правило (PROVISIONAL, не SoT):
- Gate-controls: IM_ACQ_COMBO, TSUM_NET_MODEL, BANK_DDS_CORE,
  TAX_CASH_BANK, PAYROLL_MULTI, OPEX_MULTI
- PASS статус: CLOSE / SOFT / SOFT_GAP
- Месяц RELEASED, если все gate-controls PASS; иначе BLOCKED + причины
- Owner action links — подсказка, что запросить/решить

Не трогает RACI. Не меняет регистры продаж/банка.
"""
from __future__ import annotations

import csv
import json
import shutil
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[3]
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/h18_release_gate_20260724"
MART = ROOT / "live/marts"
CC = ROOT / "live/YANINA_LIVE_CONTROL_CENTER_20260723.xlsx"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")

# Провизорная политика ворот (можно ужесточить после RACI ACCEPT)
GATE_CONTROLS = [
    "IM_ACQ_COMBO",
    "TSUM_NET_MODEL",
    "BANK_DDS_CORE",
    "TAX_CASH_BANK",
    "PAYROLL_MULTI",
    "OPEX_MULTI",
]
PASS_STATUSES = {"CLOSE", "SOFT", "SOFT_GAP"}

# Связка control → owner action (H17), чтобы blocked месяц указывал «что делать»
CONTROL_TO_ACTIONS = {
    "IM_ACQ_COMBO": ["A-DATA-IM-01"],
    "TSUM_NET_MODEL": ["A-TSUM-RATE-01"],
    "BANK_DDS_CORE": [],  # структурный Salon/DDS; не P0 H17
    "TAX_CASH_BANK": [],
    "PAYROLL_MULTI": ["A-DATA-ZP-01"],
    "OPEX_MULTI": [],
    "B2B_SETTLE_BANK": ["A-DATA-B2B-01"],
}


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def load_dashboard() -> list[dict]:
    path = MART / "controls_dashboard.csv"
    return list(csv.DictReader(open(path, encoding="utf-8")))


def load_owner_actions() -> dict[str, dict]:
    path = MART / "owner_actions.csv"
    if not path.exists():
        return {}
    return {r["action_id"]: r for r in csv.DictReader(open(path, encoding="utf-8"))}


def is_pass(status: str) -> bool:
    return (status or "").upper() in PASS_STATUSES


def build_gate(rows: list[dict], actions: dict[str, dict]) -> tuple[list[dict], list[dict], list[dict]]:
    """Помесячный gate + blocked digests + control fail list."""
    by_month: dict[str, dict[str, dict]] = defaultdict(dict)
    for r in rows:
        cid = r.get("control_id") or ""
        pm = r.get("period_month") or ""
        # только YYYY-MM; агрегаты ALL/TOTAL не участвуют в monthly gate
        if not cid or not pm or pm.upper() in {"ALL", "TOTAL"} or len(pm) != 7 or pm[4] != "-":
            continue
        by_month[pm][cid] = r

    months = sorted(by_month.keys())
    gate_rows: list[dict] = []
    blocked: list[dict] = []
    fails: list[dict] = []

    for pm in months:
        ctrl = by_month[pm]
        fail_ids: list[str] = []
        soft_ids: list[str] = []
        missing: list[str] = []
        details: list[str] = []
        linked_actions: list[str] = []

        for cid in GATE_CONTROLS:
            if cid not in ctrl:
                missing.append(cid)
                fail_ids.append(cid)
                details.append(f"{cid}=MISSING")
                linked_actions.extend(CONTROL_TO_ACTIONS.get(cid, []))
                fails.append(
                    {
                        "period_month": pm,
                        "control_id": cid,
                        "status": "MISSING",
                        "metric": "",
                        "detail": "not in dashboard for month",
                        "owner_actions": "|".join(CONTROL_TO_ACTIONS.get(cid, [])),
                    }
                )
                continue
            st = (ctrl[cid].get("status") or "").upper()
            if is_pass(st):
                if st in ("SOFT", "SOFT_GAP"):
                    soft_ids.append(cid)
                continue
            fail_ids.append(cid)
            details.append(f"{cid}={st}")
            linked_actions.extend(CONTROL_TO_ACTIONS.get(cid, []))
            fails.append(
                {
                    "period_month": pm,
                    "control_id": cid,
                    "status": st,
                    "metric": ctrl[cid].get("metric") or "",
                    "detail": ctrl[cid].get("detail") or "",
                    "owner_actions": "|".join(CONTROL_TO_ACTIONS.get(cid, [])),
                }
            )

        # уникальные action ids
        linked_u = []
        for a in linked_actions:
            if a and a not in linked_u:
                linked_u.append(a)
        action_titles = []
        for aid in linked_u:
            if aid in actions:
                action_titles.append(f"{aid}:{actions[aid].get('title','')}")

        released = len(fail_ids) == 0
        verdict = "RELEASED" if released else "BLOCKED"
        gate_rows.append(
            {
                "period_month": pm,
                "verdict": verdict,
                "gate_controls_n": len(GATE_CONTROLS),
                "pass_n": len(GATE_CONTROLS) - len(fail_ids),
                "fail_n": len(fail_ids),
                "soft_pass_n": len(soft_ids),
                "fail_controls": "|".join(fail_ids),
                "soft_controls": "|".join(soft_ids),
                "missing_controls": "|".join(missing),
                "owner_action_ids": "|".join(linked_u),
                "owner_action_hints": " || ".join(action_titles),
                "policy": "PROVISIONAL_H18",
                "so_t": "N",
                "note": "; ".join(details) if details else "all gate controls PASS",
            }
        )
        if not released:
            blocked.append(gate_rows[-1])

    return gate_rows, blocked, fails


def write_policy() -> dict:
    policy = {
        "policy_id": "PROVISIONAL_H18",
        "generated_at": NOW,
        "so_t": False,
        "description": (
            "Провизорные ворота месяца для управленческого обзора. "
            "Не Source of Truth. После RACI ACCEPT пороги могут ужесточиться."
        ),
        "gate_controls": GATE_CONTROLS,
        "pass_statuses": sorted(PASS_STATUSES),
        "fail_example_statuses": ["OPEN", "WIDE_GAP", "GAP", "BANK_ONLY", "RIGHT_ONLY", "PARTIAL", "MISSING"],
        "release_rule": "RELEASED iff every gate_control status in pass_statuses",
        "control_to_owner_actions": CONTROL_TO_ACTIONS,
        "out_of_scope": [
            "RACI ACCEPT",
            "SKU identity / margin SoT",
            "Auto-fill Owner Packet decisions",
        ],
    }
    (OUT / "release_policy.json").write_text(json.dumps(policy, ensure_ascii=False, indent=2), encoding="utf-8")
    (MART / "release_policy.json").write_text(json.dumps(policy, ensure_ascii=False, indent=2), encoding="utf-8")
    return policy


def write_markdown(gate_rows: list[dict], blocked: list[dict], summary: dict):
    lines = [
        "# Controls Release Gate (H18)",
        "",
        f"Updated: {NOW}",
        "",
        "Провизорные ворота месяца. **Не SoT.**",
        "",
        f"- Months: **{summary['months_n']}**",
        f"- RELEASED: **{summary['released_n']}**",
        f"- BLOCKED: **{summary['blocked_n']}** ({summary['blocked_pct']}%)",
        f"- Gate controls: `{', '.join(GATE_CONTROLS)}`",
        f"- PASS statuses: `{', '.join(sorted(PASS_STATUSES))}`",
        "",
        "## BLOCKED months",
        "",
    ]
    if not blocked:
        lines.append("_нет_")
    else:
        lines.append("| Month | Fail | Soft pass | Owner actions |")
        lines.append("|-------|------|-----------|---------------|")
        for r in blocked:
            lines.append(
                f"| {r['period_month']} | {r['fail_controls'] or '—'} | "
                f"{r['soft_controls'] or '—'} | {r['owner_action_ids'] or '—'} |"
            )
    lines.extend(
        [
            "",
            "## Как читать",
            "",
            "1. `RELEASED` — все gate-controls CLOSE/SOFT (управленчески «можно смотреть месяц»).",
            "2. `BLOCKED` — есть OPEN/WIDE_GAP/…; смотрите `release_gate_blocked.csv` и Owner Actions.",
            "3. После RACI ACCEPT политику можно ужесточить (например, запретить SOFT).",
            "4. B2B settle сейчас только как ALL PARTIAL (39/54) — не в gate; запрос `A-DATA-B2B-01`.",
            "",
            "Файлы: `live/marts/release_gate_month.csv`, `release_gate_blocked.csv`, `release_gate_fails.csv`.",
            "",
        ]
    )
    text = "\n".join(lines)
    (OUT / "RELEASE_GATE.md").write_text(text, encoding="utf-8")
    (ROOT / "live/RELEASE_GATE.md").write_text(text, encoding="utf-8")
    (EV / "RELEASE_GATE.md").write_text(text, encoding="utf-8")


def update_live_cc(summary: dict):
    if not CC.exists():
        return
    wb = load_workbook(CC)
    if "H18_Gate" in wb.sheetnames:
        del wb["H18_Gate"]
    ws = wb.create_sheet("H18_Gate", 0)
    ws["A1"] = "H18 Controls Release Gate"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A2"] = NOW
    ws["A4"] = summary["finding"]
    ws["A5"] = "RELEASED"
    ws["B5"] = f"{summary['released_n']}/{summary['months_n']}"
    ws["A6"] = "BLOCKED"
    ws["B6"] = f"{summary['blocked_n']}/{summary['months_n']}"
    ws["A7"] = "Blocked %"
    ws["B7"] = summary["blocked_pct"]
    ws["A8"] = "Top fail controls"
    ws["B8"] = summary["top_fail_controls"]
    ws["A9"] = "Policy"
    ws["B9"] = "PROVISIONAL_H18 (not SoT)"
    ws["A10"] = "Docs"
    ws["B10"] = "live/RELEASE_GATE.md"
    wb.save(CC)


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)
    MART.mkdir(parents=True, exist_ok=True)

    dash = load_dashboard()
    actions = load_owner_actions()
    policy = write_policy()
    gate_rows, blocked, fails = build_gate(dash, actions)

    fields = list(gate_rows[0].keys()) if gate_rows else ["period_month", "verdict"]
    write_csv(MART / "release_gate_month.csv", gate_rows, fields)
    write_csv(OUT / "release_gate_month.csv", gate_rows, fields)
    write_csv(MART / "release_gate_blocked.csv", blocked, fields)
    fail_fields = list(fails[0].keys()) if fails else ["period_month", "control_id", "status"]
    write_csv(MART / "release_gate_fails.csv", fails, fail_fields)
    write_csv(OUT / "release_gate_fails.csv", fails, fail_fields)

    # Информационный контур B2B (не gate: только ALL PARTIAL в dashboard)
    b2b = next((r for r in dash if r.get("control_id") == "B2B_SETTLE_BANK"), None)
    b2b_note = None
    if b2b:
        b2b_note = f"{b2b.get('status')} {b2b.get('metric')} ({b2b.get('detail')})"

    released_n = sum(1 for r in gate_rows if r["verdict"] == "RELEASED")
    blocked_n = len(blocked)
    months_n = len(gate_rows)
    fail_ctr = Counter(f["control_id"] for f in fails)
    top_fail = ", ".join(f"{k}:{v}" for k, v in fail_ctr.most_common(5)) or "—"

    summary = {
        "wave": "H18",
        "generated_at": NOW,
        "path_choice": "Controls release gate — operationalize residuals without new files/RACI",
        "finding": (
            f"H18: provisional release gate — RELEASED {released_n}/{months_n}, "
            f"BLOCKED {blocked_n}/{months_n}. Top fails: {top_fail}."
            + (f" B2B settle (info): {b2b_note}." if b2b_note else "")
        ),
        "months_n": months_n,
        "released_n": released_n,
        "blocked_n": blocked_n,
        "blocked_pct": round(100.0 * blocked_n / months_n, 1) if months_n else 0.0,
        "top_fail_controls": top_fail,
        "b2b_settle_info": b2b_note,
        "gate_controls": GATE_CONTROLS,
        "pass_statuses": sorted(PASS_STATUSES),
        "policy_id": policy["policy_id"],
        "not_sot": True,
    }
    write_markdown(gate_rows, blocked, summary)
    (OUT / "h18_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    (EV / "h18_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    for name in (
        "release_gate_month.csv",
        "release_gate_fails.csv",
        "release_policy.json",
        "h18_summary.json",
        "RELEASE_GATE.md",
    ):
        src = OUT / name
        if src.exists():
            shutil.copy2(src, EV / name)
    shutil.copy2(MART / "release_gate_blocked.csv", EV / "release_gate_blocked.csv")

    update_live_cc(summary)
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
