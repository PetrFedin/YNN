#!/usr/bin/env python3
"""H73 — Designers payroll (KPI/smetka) ↔ collections / person-cost.

Зачем:
  5× зп_конструкторы_0N.26.xlsx содержат (а) KPI по месяцам на Лист1,
  (б) сметку 01.25 с артикулами × конструктор × часы/ставка.
  Нужна сквозная связка с H62/H65 без fake Accept.

Правила:
  - smetka fingerprint-dedupe (все 5 файлов несут одну и ту же сметку)
  - KPI Лист1 — единственный monthly signal
  - do_not_auto_accept=YES · so_t=N
"""
from __future__ import annotations

import csv
import hashlib
import json
import re
import shutil
import warnings
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

warnings.filterwarnings("ignore")

ROOT = Path(__file__).resolve().parents[3]
REG = Path(__file__).resolve().parent
MARTS = ROOT / "live/marts"
MAPS = ROOT / "live/maps"
EV = ROOT / "live/evidence/h73_designers_collection_bridge_20260729"
WAVE_B = ROOT / "live/client_pack/execution_wave_b"
DOWNLOADS = Path("/Users/petr/Downloads/YANINA документы")


def write_csv(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)


def norm_art(raw) -> str:
    if raw is None or raw == "":
        return ""
    s = str(raw).strip().upper().replace(" ", "").replace("Ё", "Е")
    s = re.sub(r"^(ПО|К)", "", s)
    m = re.search(r"(\d{1,2}-\d{2,4}[A-ZА-Я]?)", s)
    if m:
        return m.group(1)
    m = re.search(r"(0-\d+[A-ZА-Я]?|Т-\d+[A-ZА-Я]?)", s)
    return m.group(1) if m else ""


def norm_name(s: str) -> str:
    s = re.sub(r"\s+", " ", (s or "").strip())
    return s.split()[0] if s else ""


def catalog() -> list[dict]:
    out = []
    with (ROOT / "live/registers/00_SOURCE_CATALOG_107.csv").open(encoding="utf-8") as f:
        for r in csv.DictReader(f):
            if r["category"] != "payroll_designers":
                continue
            p = Path(r["path"])
            if not p.exists():
                p = DOWNLOADS / r["file_name"]
            out.append({**r, "_path": p})
    return out


def month_from_filename(name: str) -> str:
    m = re.search(r"_(\d{2})\.(\d{2})", name)
    return f"20{m.group(2)}-{m.group(1)}" if m else ""


def parse_smetka(path: Path, meta: dict) -> tuple[list[dict], str | None]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    if "сметк 01.25" not in wb.sheetnames:
        wb.close()
        return [], None
    ws = wb["сметк 01.25"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    fp = hashlib.md5(repr(rows[:80]).encode()).hexdigest()[:16]
    header_i = None
    for i, row in enumerate(rows):
        vals = [str(c or "").lower() for c in (row or [])[:8]]
        if any("арт" in v for v in vals) and any("конструктор" in v for v in vals):
            header_i = i
            break
    if header_i is None:
        return [], fp
    out = []

    def fnum(x):
        try:
            return float(x) if x not in (None, "") else None
        except (TypeError, ValueError):
            return None

    for i, row in enumerate(rows[header_i + 1 :], start=header_i + 2):
        if not row:
            continue
        art = norm_art(row[3] if len(row) > 3 else None)
        if not art:
            continue
        hours = fnum(row[7] if len(row) > 7 else None)
        rate = fnum(row[8] if len(row) > 8 else None)
        cost = fnum(row[9] if len(row) > 9 else None)
        out.append(
            {
                "smetka_line_id": f"SM-{fp}-{i}",
                "source_file": meta["file_name"],
                "source_file_id": meta["source_file_id"],
                "sheet": "сметк 01.25",
                "content_fingerprint": fp,
                "product_name": str(row[1] or "").strip(),
                "fabric_note": str(row[2] or "").strip()[:80],
                "article_raw": str(row[3] or "").strip(),
                "article_norm": art,
                "constructor": str(row[4] or "").strip(),
                "constructor_surname": norm_name(str(row[4] or "")),
                "client_or_project": str(row[5] or "").strip(),
                "work_type": str(row[6] or "").strip(),
                "hours": hours if hours is not None else "",
                "rate_rub": rate if rate is not None else "",
                "cost_rub": cost if cost is not None else "",
                "as_of_hint": "2025-01_smetka_sheet",
                "so_t": "N",
            }
        )
    return out, fp


def parse_kpi(path: Path, meta: dict, period_month: str) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    if "Лист1" not in wb.sheetnames:
        wb.close()
        return []
    rows = list(wb["Лист1"].iter_rows(values_only=True))
    wb.close()
    for row in rows[:8]:
        if row and isinstance(row[0], datetime):
            period_month = f"{row[0].year:04d}-{row[0].month:02d}"
            break
    out: list[dict] = []
    i = 0
    while i < len(rows):
        row = rows[i]
        a0 = row[0] if row else None
        if (
            a0
            and isinstance(a0, str)
            and i + 1 < len(rows)
            and str(rows[i + 1][0] or "").startswith("Задач")
        ):
            full = a0.strip()
            surname = norm_name(full)
            kpi_score = pay_net = None
            cut_n = podkroy_n = lekala_n = delivered_n = None
            collection_flags: list[str] = []
            j = i + 2
            while j < len(rows):
                r2 = rows[j]
                if not r2:
                    j += 1
                    continue
                b0 = r2[0]
                if (
                    isinstance(b0, str)
                    and j + 1 < len(rows)
                    and str(rows[j + 1][0] or "").startswith("Задач")
                ):
                    break
                # «Результат» часто без ФИО в col0 — читаем по col2
                if "Результат" in str(r2[2] or ""):
                    try:
                        kpi_score = float(r2[5]) if r2[5] not in (None, "") else None
                    except (TypeError, ValueError):
                        pass
                if (
                    isinstance(b0, str)
                    and surname.lower() in b0.lower()
                    and str(r2[3] or "").lower() == "итог"
                ):
                    try:
                        pay_net = float(r2[5]) if r2[5] not in (None, "") else None
                    except (TypeError, ValueError):
                        pass
                    j += 1
                    break
                task = str(b0 or "")
                qty = r2[4] if len(r2) > 4 else None
                if task.startswith("1.") and qty not in (None, ""):
                    try:
                        cut_n = float(qty)
                    except (TypeError, ValueError):
                        pass
                if task.startswith("2.") and qty not in (None, ""):
                    try:
                        podkroy_n = float(qty)
                    except (TypeError, ValueError):
                        pass
                if task.startswith("3.") and qty not in (None, ""):
                    try:
                        lekala_n = float(qty)
                    except (TypeError, ValueError):
                        pass
                if "Сдача готовых изделий" in task and qty not in (None, ""):
                    try:
                        delivered_n = float(qty)
                    except (TypeError, ValueError):
                        pass
                mcol = re.search(r"Коллекц\w*\s*(\d{2})", task, re.I)
                if mcol:
                    collection_flags.append(mcol.group(1))
                j += 1
            out.append(
                {
                    "kpi_row_id": f"KPI-{period_month}-{surname}-{i}",
                    "period_month": period_month,
                    "constructor_full": full,
                    "constructor_surname": surname,
                    "kpi_score": round(kpi_score, 4) if kpi_score is not None else "",
                    "pay_net_rub": pay_net if pay_net is not None else "",
                    "cut_qty": cut_n if cut_n is not None else "",
                    "podkroy_qty": podkroy_n if podkroy_n is not None else "",
                    "lekala_qty": lekala_n if lekala_n is not None else "",
                    "delivered_qty": delivered_n if delivered_n is not None else "",
                    "collection_flags": "|".join(sorted(set(collection_flags))),
                    "source_file": meta["file_name"],
                    "source_file_id": meta["source_file_id"],
                    "so_t": "N",
                }
            )
            i = j
            continue
        i += 1
    return out


def main() -> dict:
    files = catalog()
    smetka_all: list[dict] = []
    seen: set[str] = set()
    kpi_all: list[dict] = []
    for meta in files:
        period = month_from_filename(meta["file_name"])
        sm, fp = parse_smetka(meta["_path"], meta)
        if fp and fp not in seen:
            seen.add(fp)
            smetka_all.extend(sm)
        kpi_all.extend(parse_kpi(meta["_path"], meta, period))

    col = {
        r["article_norm"]: r
        for r in csv.DictReader(open(MARTS / "collection_sku_stock_bridge.csv", encoding="utf-8"))
    }
    pc = {
        r["article_norm"]: r
        for r in csv.DictReader(open(MARTS / "person_cost_collection_bridge.csv", encoding="utf-8"))
    }
    high = {
        r["article_norm"]: r
        for r in csv.DictReader(open(MARTS / "high_gap_owner_worksheet.csv", encoding="utf-8"))
    }

    bridge = []
    for r in smetka_all:
        a = r["article_norm"]
        c = col.get(a)
        p = pc.get(a)
        h = high.get(a)
        if c and p:
            bv = "LINKS_COLLECTION_AND_PC"
        elif c:
            bv = "LINKS_COLLECTION_ONLY"
        elif p:
            bv = "LINKS_PERSON_COST_ONLY"
        else:
            bv = "ORPHAN_SMETKA"
        bridge.append(
            {
                **r,
                "in_collection_sales": "Y" if c else "N",
                "collection_sale_eur": c["collection_sale_eur"] if c else "",
                "collection_link_coverage": c["link_coverage"] if c else "",
                "in_person_cost": "Y" if p else "N",
                "person_sources": p["person_sources"] if p else "",
                "person_lines_n": p["person_lines_n"] if p else "",
                "was_high_gap": "Y" if h else "N",
                "person_cost_hit_on_high": h["person_cost_hit"] if h else "",
                "bridge_value": bv,
                "do_not_auto_accept": "YES",
            }
        )

    by_c: dict = defaultdict(
        lambda: {
            "smetka_n": 0,
            "hours": 0.0,
            "cost": 0.0,
            "arts": set(),
            "col_arts": set(),
            "kpi_months": 0,
            "pay_sum": 0.0,
            "kpi_scores": [],
        }
    )
    for r in bridge:
        s = r["constructor_surname"] or r["constructor"]
        b = by_c[s]
        b["smetka_n"] += 1
        if r["hours"] != "":
            b["hours"] += float(r["hours"])
        if r["cost_rub"] != "":
            b["cost"] += float(r["cost_rub"])
        b["arts"].add(r["article_norm"])
        if r["in_collection_sales"] == "Y":
            b["col_arts"].add(r["article_norm"])
    for r in kpi_all:
        s = r["constructor_surname"]
        b = by_c[s]
        b["kpi_months"] += 1
        if r["pay_net_rub"] != "":
            b["pay_sum"] += float(r["pay_net_rub"])
        if r["kpi_score"] != "":
            b["kpi_scores"].append(float(r["kpi_score"]))

    constr = []
    for s, b in sorted(by_c.items(), key=lambda x: -x[1]["pay_sum"]):
        constr.append(
            {
                "constructor_surname": s,
                "smetka_lines_n": b["smetka_n"],
                "smetka_hours": round(b["hours"], 2),
                "smetka_cost_rub": round(b["cost"], 2),
                "unique_articles_n": len(b["arts"]),
                "collection_articles_n": len(b["col_arts"]),
                "kpi_months_n": b["kpi_months"],
                "pay_net_sum_rub": round(b["pay_sum"], 2),
                "kpi_score_avg": round(sum(b["kpi_scores"]) / len(b["kpi_scores"]), 3)
                if b["kpi_scores"]
                else "",
                "so_t": "N",
                "note": "smetka deduped once; KPI from monthly Лист1",
            }
        )

    kpi_month = []
    for m in sorted({r["period_month"] for r in kpi_all}):
        sub = [r for r in kpi_all if r["period_month"] == m]
        scores = [float(r["kpi_score"]) for r in sub if r["kpi_score"] != ""]
        kpi_month.append(
            {
                "period_month": m,
                "constructors_n": len(sub),
                "pay_net_sum_rub": round(
                    sum(float(r["pay_net_rub"]) for r in sub if r["pay_net_rub"] != ""), 2
                ),
                "cut_qty_sum": round(
                    sum(float(r["cut_qty"]) for r in sub if r["cut_qty"] != ""), 2
                ),
                "delivered_qty_sum": round(
                    sum(float(r["delivered_qty"]) for r in sub if r["delivered_qty"] != ""), 2
                ),
                "kpi_avg": round(sum(scores) / len(scores), 3) if scores else "",
                "so_t": "N",
            }
        )

    meta = {
        "horizon": "H73",
        "date": str(date.today()),
        "title": "Designers payroll KPI/smetka ↔ collections/person-cost",
        "files_n": len(files),
        "smetka_lines_n": len(smetka_all),
        "smetka_unique_articles_n": len({r["article_norm"] for r in smetka_all}),
        "kpi_rows_n": len(kpi_all),
        "constructors_n": len(constr),
        "smetka_links_collection_n": sum(1 for r in bridge if r["in_collection_sales"] == "Y"),
        "smetka_links_person_cost_n": sum(1 for r in bridge if r["in_person_cost"] == "Y"),
        "smetka_links_both_n": sum(1 for r in bridge if r["bridge_value"] == "LINKS_COLLECTION_AND_PC"),
        "collection_sale_eur_touched": round(
            sum(float(r["collection_sale_eur"]) for r in bridge if r["collection_sale_eur"] != ""),
            2,
        ),
        "kpi_pay_total_rub": round(
            sum(float(r["pay_net_rub"]) for r in kpi_all if r["pay_net_rub"] != ""), 2
        ),
        "no_fake_accept": True,
        "so_t": False,
        "note": "Five monthly xlsx share identical smetka; KPI Лист1 differs 2026-01..05",
    }

    REG.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)
    (REG / "meta.json").write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
    write_csv(REG / "designer_smetka_lines.csv", smetka_all)
    write_csv(REG / "designer_smetka_collection_bridge.csv", bridge)
    write_csv(REG / "designer_kpi_monthly.csv", kpi_all)
    write_csv(REG / "designer_constructor_summary.csv", constr)
    write_csv(REG / "designer_kpi_month_totals.csv", kpi_month)

    for name in [
        "designer_smetka_lines.csv",
        "designer_smetka_collection_bridge.csv",
        "designer_kpi_monthly.csv",
        "designer_constructor_summary.csv",
        "designer_kpi_month_totals.csv",
        "meta.json",
    ]:
        src = REG / name
        if name == "meta.json":
            shutil.copy2(src, MARTS / "h73_meta.json")
            shutil.copy2(src, MAPS / "h73_meta.json")
            shutil.copy2(src, EV / "meta.json")
        else:
            shutil.copy2(src, MARTS / name)
            shutil.copy2(src, MAPS / name)
            shutil.copy2(src, EV / name)

    # 34–37: не пересекаться с H67 `30_fabric_cash_bridge_metrics.csv`
    shutil.copy2(REG / "designer_smetka_collection_bridge.csv", WAVE_B / "34_designer_smetka_collection_bridge.csv")
    shutil.copy2(REG / "designer_kpi_monthly.csv", WAVE_B / "35_designer_kpi_monthly.csv")
    shutil.copy2(REG / "designer_constructor_summary.csv", WAVE_B / "36_designer_constructor_summary.csv")
    shutil.copy2(REG / "designer_kpi_month_totals.csv", WAVE_B / "37_designer_kpi_month_totals.csv")

    print(json.dumps(meta, ensure_ascii=False, indent=2))
    return meta


if __name__ == "__main__":
    main()
