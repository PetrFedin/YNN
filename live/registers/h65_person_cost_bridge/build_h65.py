#!/usr/bin/env python3
"""H65: person-cost + embroidery → collection/SKU/stock article bridge.

Closes the H64 COLLECTION_STYLE hole (43-xx / 47-xx ≠ goods 0-xxxx):
Мокеева/Жукова/Меркушина/вышивка carry MD model numbers that match showroom codes.

Not SoT. No auto-Accept. Cost amounts are indicative workshop cards only.
"""
from __future__ import annotations

import csv
import json
import re
import shutil
import warnings
import zipfile
import xml.etree.ElementTree as ET
from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

warnings.filterwarnings("ignore")

ROOT = Path(__file__).resolve().parents[3]
if not (ROOT / "live/registers/00_SOURCE_CATALOG_107.csv").exists():
    ROOT = Path.cwd()
REG = Path(__file__).resolve().parent
MARTS = ROOT / "live/marts"
MAPS = ROOT / "live/maps"
EV = ROOT / "live/evidence/h65_person_cost_bridge_20260729"
WAVE_B = ROOT / "live/client_pack/execution_wave_b"
NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


def write_csv(path: Path, fields, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)


def norm_art(raw) -> str:
    if raw is None or raw == "":
        return ""
    s = str(raw).strip().upper().replace(" ", "").replace("Ё", "Е")
    s = re.sub(r"^(ПО|К)", "", s)
    m = re.search(r"(\d{1,2}-\d{2,4}[A-ZА-Я]?)", s)
    if m:
        return m.group(1)
    m = re.search(r"(0-\d+[A-ZА-Я]?|Т-\d+[A-ZА-Я]?|ИМ-\d+)", s)
    return m.group(1) if m else ""


def catalog_paths(cat: str) -> list[dict]:
    out = []
    for r in csv.DictReader((ROOT / "live/registers/00_SOURCE_CATALOG_107.csv").open()):
        if r["category"] != cat:
            continue
        p = Path(r["path"])
        if not p.exists():
            p = ROOT / "documents" / r["file_name"]
        if not p.exists():
            # Downloads fallback used by local symlink workspace
            alt = Path("/Users/petr/Downloads/YANINA документы") / r["file_name"]
            if alt.exists():
                p = alt
        out.append({**r, "_path": p})
    return out


def col_row(ref: str):
    m = re.match(r"([A-Z]+)(\d+)", ref)
    col = 0
    for c in m.group(1):
        col = col * 26 + ord(c) - 64
    return col - 1, int(m.group(2))


def read_xlsx_via_zip(path: Path, sheet_xml: str = "xl/worksheets/sheet1.xml") -> dict[int, dict[int, str]]:
    """Fallback for workbooks with corrupt merge cells (Мокеева 2026)."""
    z = zipfile.ZipFile(path)
    ss: list[str] = []
    if "xl/sharedStrings.xml" in z.namelist():
        root = ET.fromstring(z.read("xl/sharedStrings.xml"))
        for si in root.findall("m:si", NS):
            texts = [t.text or "" for t in si.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t")]
            ss.append("".join(texts))
    root = ET.fromstring(z.read(sheet_xml))
    rows: dict[int, dict[int, str]] = {}
    for c in root.findall(".//m:c", NS):
        ref = c.get("r")
        if not ref:
            continue
        col, row = col_row(ref)
        t = c.get("t")
        v = c.find("m:v", NS)
        if v is None:
            continue
        val = v.text
        if t == "s" and val is not None:
            val = ss[int(val)]
        rows.setdefault(row, {})[col] = val
    z.close()
    return rows


def safe_float(v):
    try:
        if v is None or v == "":
            return None
        return float(v)
    except Exception:
        return None


def parse_mokeeva(path: Path, meta: dict, year_tag: str) -> list[dict]:
    out = []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        for i, row in enumerate(rows[1:], start=2):
            art = norm_art(row[6] if len(row) > 6 else None)
            if not art:
                continue
            # cost columns vary; take rightmost numeric-looking total if present
            cost = None
            for idx in (23, 24, 25, 22, 19, 16):
                if len(row) > idx:
                    cost = safe_float(row[idx])
                    if cost and cost > 100:
                        break
            out.append(
                {
                    "pc_line_id": f"MOK{year_tag}-{i}",
                    "source_person": "mokeeva",
                    "source_year": year_tag,
                    "client": str(row[2] or "").strip() if len(row) > 2 else "",
                    "product_name": str(row[5] or "").strip() if len(row) > 5 else "",
                    "article_raw": str(row[6] or "").strip(),
                    "article_norm": art,
                    "fabric_note": str(row[7] or "").strip()[:120] if len(row) > 7 else "",
                    "cost_rub_indicative": round(cost, 2) if cost else "",
                    "source_file": meta["file_name"],
                    "source_file_id": meta["source_file_id"],
                    "so_t": "N",
                }
            )
        return out
    except Exception:
        # corrupt merges → zip reader (Мокеева 2026)
        rows = read_xlsx_via_zip(path)
        for i, cols in sorted(rows.items()):
            if i == 1:
                continue
            art = norm_art(cols.get(6))
            if not art:
                continue
            cost = None
            for idx in (25, 24, 23, 22, 19, 16):
                cost = safe_float(cols.get(idx))
                if cost and cost > 100:
                    break
            out.append(
                {
                    "pc_line_id": f"MOK{year_tag}-{i}",
                    "source_person": "mokeeva",
                    "source_year": year_tag,
                    "client": str(cols.get(2) or "").strip(),
                    "product_name": str(cols.get(5) or "").strip(),
                    "article_raw": str(cols.get(6) or "").strip(),
                    "article_norm": art,
                    "fabric_note": str(cols.get(7) or "").strip()[:120],
                    "cost_rub_indicative": round(cost, 2) if cost else "",
                    "source_file": meta["file_name"],
                    "source_file_id": meta["source_file_id"],
                    "so_t": "N",
                }
            )
        return out


def parse_zhukova(path: Path, meta: dict) -> list[dict]:
    out = []
    wb = load_workbook(path, read_only=True, data_only=True)
    for sheet in wb.sheetnames:
        if sheet not in {"2024", "2025", "2026"}:
            continue
        ws = wb[sheet]
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i <= 2:
                continue
            art = norm_art(row[1] if row else None)
            if not art:
                continue
            # materials / sewing cost often around col 7+
            cost = None
            for idx in range(7, min(len(row), 20)):
                c = safe_float(row[idx])
                if c and c > 500:
                    cost = c
                    break
            out.append(
                {
                    "pc_line_id": f"ZHU{sheet}-{i}",
                    "source_person": "zhukova",
                    "source_year": sheet,
                    "client": "",
                    "product_name": str(row[2] or "").strip() if len(row) > 2 else "",
                    "article_raw": str(row[1] or "").strip(),
                    "article_norm": art,
                    "fabric_note": str(row[3] or "").strip()[:80] if len(row) > 3 else "",
                    "cost_rub_indicative": round(cost, 2) if cost else "",
                    "source_file": meta["file_name"],
                    "source_file_id": meta["source_file_id"],
                    "so_t": "N",
                }
            )
    wb.close()
    return out


def parse_merkushina(path: Path, meta: dict) -> list[dict]:
    out = []
    wb = load_workbook(path, read_only=True, data_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        # find header with Артикул
        header_i = None
        for i, row in enumerate(rows[:15]):
            vals = [str(c or "").lower() for c in row[:8]]
            if any("артикул" in v for v in vals):
                header_i = i
                break
        if header_i is None:
            continue
        for i, row in enumerate(rows[header_i + 1 :], start=header_i + 2):
            art = norm_art(row[1] if row else None)
            if not art:
                continue
            cost = None
            for idx in range(7, min(len(row), 18)):
                c = safe_float(row[idx])
                if c and c > 500:
                    cost = c
                    break
            out.append(
                {
                    "pc_line_id": f"MER-{sheet[:8]}-{i}",
                    "source_person": "merkushina",
                    "source_year": sheet[:12],
                    "client": "",
                    "product_name": str(row[2] or "").strip() if len(row) > 2 else "",
                    "article_raw": str(row[1] or "").strip(),
                    "article_norm": art,
                    "fabric_note": str(row[4] or "").strip()[:80] if len(row) > 4 else "",
                    "cost_rub_indicative": round(cost, 2) if cost else "",
                    "source_file": meta["file_name"],
                    "source_file_id": meta["source_file_id"],
                    "so_t": "N",
                }
            )
    wb.close()
    return out


def parse_embroidery(path: Path, meta: dict) -> list[dict]:
    out = []
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            continue
        art = norm_art(row[1] if row else None)
        if not art:
            continue
        hours = safe_float(row[7] if len(row) > 7 else None)
        out.append(
            {
                "pc_line_id": f"EMB-{i}",
                "source_person": "embroidery",
                "source_year": "2024-2026",
                "client": str(row[3] or "").strip() if len(row) > 3 else "",
                "product_name": str(row[2] or "").strip() if len(row) > 2 else "",
                "article_raw": str(row[1] or "").strip(),
                "article_norm": art,
                "fabric_note": str(row[5] or "").strip()[:120] if len(row) > 5 else "",
                "cost_rub_indicative": "",  # hours-based; money not reliable here
                "hours": hours if hours is not None else "",
                "source_file": meta["file_name"],
                "source_file_id": meta["source_file_id"],
                "so_t": "N",
            }
        )
    wb.close()
    return out


def article_family(art: str) -> str:
    if art.startswith("0-") or art.startswith("ИМ-") or art.startswith("Т-"):
        return "GOODS_STYLE"
    if re.match(r"^\d{2}-", art):
        return "COLLECTION_STYLE"
    return "OTHER"


def main():
    for d in (REG, MARTS, MAPS, EV, WAVE_B):
        d.mkdir(parents=True, exist_ok=True)

    lines: list[dict] = []
    for meta in catalog_paths("cost_person"):
        name = meta["file_name"].lower()
        p = meta["_path"]
        if not p.exists():
            continue
        if "мокеева" in name and "2025" in name:
            lines.extend(parse_mokeeva(p, meta, "2025"))
        elif "мокеева" in name and "2026" in name:
            lines.extend(parse_mokeeva(p, meta, "2026"))
        elif "жукова" in name:
            lines.extend(parse_zhukova(p, meta))
        elif "меркушина" in name:
            lines.extend(parse_merkushina(p, meta))

    for meta in catalog_paths("cost_embroidery"):
        if meta["_path"].exists():
            lines.extend(parse_embroidery(meta["_path"], meta))

    # ensure hours field on all
    for r in lines:
        r.setdefault("hours", "")

    # aggregate by article
    by_art: dict[str, dict] = defaultdict(
        lambda: {
            "persons": set(),
            "lines_n": 0,
            "clients": set(),
            "cost_sum": 0.0,
            "cost_n": 0,
            "hours_sum": 0.0,
            "products": set(),
            "years": set(),
        }
    )
    for r in lines:
        a = r["article_norm"]
        b = by_art[a]
        b["persons"].add(r["source_person"])
        b["lines_n"] += 1
        if r["client"]:
            b["clients"].add(r["client"][:40])
        if r["cost_rub_indicative"] != "":
            b["cost_sum"] += float(r["cost_rub_indicative"])
            b["cost_n"] += 1
        if r["hours"] != "":
            b["hours_sum"] += float(r["hours"])
        if r["product_name"]:
            b["products"].add(r["product_name"][:40])
        b["years"].add(str(r["source_year"]))

    # join collection / sku / stock bridges
    col = {r["article_norm"]: r for r in csv.DictReader((MARTS / "collection_sku_stock_bridge.csv").open())}
    sku = {r["article_norm"] for r in csv.DictReader((MARTS / "sku_master_normalized.csv").open()) if r["article_norm"]}
    stock = {r["article_norm"] for r in csv.DictReader((MARTS / "stock_cost_articles.csv").open()) if r["article_norm"]}

    bridge = []
    for art, info in sorted(by_art.items(), key=lambda x: -x[1]["lines_n"]):
        c = col.get(art)
        bridge.append(
            {
                "article_norm": art,
                "article_family": article_family(art),
                "person_sources": "|".join(sorted(info["persons"])),
                "person_lines_n": info["lines_n"],
                "person_years": "|".join(sorted(info["years"])),
                "clients_sample": "|".join(sorted(info["clients"])[:3]),
                "products_sample": "|".join(sorted(info["products"])[:3]),
                "cost_rub_sum_indicative": round(info["cost_sum"], 2) if info["cost_n"] else "",
                "cost_lines_n": info["cost_n"],
                "hours_sum": round(info["hours_sum"], 2) if info["hours_sum"] else "",
                "in_collection_sales": "Y" if c else "N",
                "collection_sale_eur": c["collection_sale_eur"] if c else "",
                "collection_link_coverage": c["link_coverage"] if c else "",
                "collection_priority": c["priority"] if c else "",
                "in_sku_master": "Y" if art in sku else "N",
                "in_stock_cost": "Y" if art in stock else "N",
                "bridge_value": (
                    "CLOSES_COLLECTION_GAP"
                    if c and c["link_coverage"] == "NONE"
                    else ("SUPPORTS_COLLECTION" if c else ("GOODS_COST_ONLY" if art in sku or art in stock else "ORPHAN_COST"))
                ),
                "so_t": "N",
                "do_not_auto_accept": "YES",
            }
        )

    closes = [r for r in bridge if r["bridge_value"] == "CLOSES_COLLECTION_GAP"]
    closes_sale = sum(float(r["collection_sale_eur"]) for r in closes if r["collection_sale_eur"] != "")
    high_before = sum(1 for r in col.values() if r["priority"] == "HIGH")
    high_closed = sum(1 for r in closes if r["collection_priority"] == "HIGH")
    col_style = [r for r in bridge if r["article_family"] == "COLLECTION_STYLE" and r["in_collection_sales"] == "Y"]

    # owner worksheet: HIGH collection gaps still open after person-cost
    still_open = []
    pc_arts = set(by_art)
    for art, r in sorted(col.items(), key=lambda x: -float(x[1]["collection_sale_eur"])):
        if r["priority"] != "HIGH":
            continue
        still_open.append(
            {
                "article_norm": art,
                "collection_sale_eur": r["collection_sale_eur"],
                "collection_types": r["collection_types"],
                "person_cost_hit": "Y" if art in pc_arts else "N",
                "owner_action": (
                    "CONFIRM person-cost card as MD cost proxy (not Accept)"
                    if art in pc_arts
                    else "NEED alias OR MD line / workshop card"
                ),
                "so_t": "N",
            }
        )

    meta = {
        "horizon": "H65",
        "date": str(date.today()),
        "title": "Person-cost/embroidery → collection article bridge",
        "person_cost_lines_n": len(lines),
        "unique_articles_n": len(by_art),
        "bridge_n": len(bridge),
        "closes_collection_gap_n": len(closes),
        "closes_collection_gap_sale_eur": round(closes_sale, 2),
        "high_gaps_total": high_before,
        "high_gaps_with_person_cost": high_closed,
        "collection_style_linked_n": len(col_style),
        "by_person": {
            p: sum(1 for r in lines if r["source_person"] == p)
            for p in sorted({r["source_person"] for r in lines})
        },
        "no_fake_accept": True,
        "so_t": False,
    }

    (REG / "meta.json").write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
    write_csv(REG / "person_cost_lines.csv", list(lines[0].keys()), lines)
    write_csv(REG / "person_cost_collection_bridge.csv", list(bridge[0].keys()), bridge)
    write_csv(REG / "person_cost_closes_gaps.csv", list(closes[0].keys()) if closes else ["article_norm"], closes[:80])
    write_csv(
        REG / "high_gap_owner_worksheet.csv",
        list(still_open[0].keys()) if still_open else ["article_norm"],
        still_open,
    )

    for name in [
        "person_cost_lines.csv",
        "person_cost_collection_bridge.csv",
        "person_cost_closes_gaps.csv",
        "high_gap_owner_worksheet.csv",
        "meta.json",
    ]:
        src = REG / name
        if name == "meta.json":
            shutil.copy2(src, MARTS / "h65_meta.json")
            shutil.copy2(src, MAPS / "h65_meta.json")
            shutil.copy2(src, EV / "meta.json")
        else:
            shutil.copy2(src, MARTS / name)
            shutil.copy2(src, MAPS / name)
            shutil.copy2(src, EV / name)

    shutil.copy2(REG / "person_cost_collection_bridge.csv", WAVE_B / "19_person_cost_collection_bridge.csv")
    shutil.copy2(REG / "person_cost_closes_gaps.csv", WAVE_B / "20_person_cost_closes_gaps.csv")
    shutil.copy2(REG / "high_gap_owner_worksheet.csv", WAVE_B / "21_high_gap_owner_worksheet.csv")

    print(json.dumps(meta, ensure_ascii=False, indent=2))
    print("TOP closes:")
    for r in sorted(closes, key=lambda x: -float(x["collection_sale_eur"] or 0))[:8]:
        print(r["article_norm"], r["collection_sale_eur"], r["person_sources"], r["person_lines_n"])


if __name__ == "__main__":
    main()
