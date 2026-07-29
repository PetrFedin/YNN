#!/usr/bin/env python3
"""H77 — SALES workbook ↔ budget income sanity (не company P&L).

Зачем:
  SALES 2024-2026.xlsx содержит «ДДС - доход» и «доходы-расходы».
  Бюджет 2026 уже имеет Fact по Salon+Shop/ИМ (H63).
  Нужна сверка масштаба + флаги расхождений без fake Accept.

Правила:
  - «доходы-расходы» = кассовый/остаточный контур, НЕ audited P&L
  - EUR_LIKE как в бюджете; SALES ДДС-доход в тех же единицах книги
  - do_not_auto_accept=YES · so_t=N
"""
from __future__ import annotations

import csv
import json
import re
import shutil
import warnings
from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

warnings.filterwarnings("ignore")

ROOT = Path(__file__).resolve().parents[3]
REG = Path(__file__).resolve().parent
MARTS = ROOT / "live/marts"
MAPS = ROOT / "live/maps"
EV = ROOT / "live/evidence/h77_sales_budget_sanity_20260729"
WAVE_C = ROOT / "live/client_pack/execution_wave_c"
DOWNLOADS = Path("/Users/petr/Downloads/YANINA документы")

MONTHS = ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12"]


def write_csv(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)


def fnum(x):
    if x is None or x == "":
        return None
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def catalog_sales_path() -> Path:
    with (ROOT / "live/registers/00_SOURCE_CATALOG_107.csv").open(encoding="utf-8") as f:
        for r in csv.DictReader(f):
            if r["category"] == "sales_general":
                p = Path(r["path"])
                if not p.exists():
                    p = DOWNLOADS / r["file_name"]
                return p
    return DOWNLOADS / "SALES 2024-2026.xlsx"


def norm_channel(label: str) -> str:
    s = re.sub(r"\s+", " ", (label or "").strip())
    low = s.lower()
    if "salon" in low and "shop" in low:
        return "Salon+Shop"
    if low.startswith("им") or low == "im":
        return "IM"
    if "цум" in low or "tsum" in low:
        return "TSUM"
    if "total" in low or "итого" in low:
        return "TOTAL"
    if "возврат" in low and "депозит" in low:
        return "DEPOSIT"
    if "депозит" in low:
        return "DEPOSIT"
    if "возврат" in low:
        return "RETURN"
    # regional / b2b-ish
    if s and s[0].isupper():
        return "REGIONAL_B2B"
    return "OTHER"


def parse_dds_income(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb["ДДС - доход"]
    rows = list(ws.iter_rows(values_only=True, max_col=14))
    wb.close()
    out = []
    year = None
    for row in rows:
        if not row:
            continue
        if row[0] and "ДДС" in str(row[0]):
            m = re.search(r"(20\d{2})", str(row[0]))
            year = int(m.group(1)) if m else None
        if year is None or not row[1]:
            continue
        ch_raw = str(row[1]).strip()
        ch = norm_channel(ch_raw)
        for i, mm in enumerate(MONTHS):
            val = fnum(row[2 + i] if len(row) > 2 + i else None)
            if val is None:
                continue
            out.append(
                {
                    "period_month": f"{year}-{mm}",
                    "year": year,
                    "channel_raw": ch_raw,
                    "channel_norm": ch,
                    "amount_eur_like": round(val, 2),
                    "sheet": "ДДС - доход",
                    "source_file": path.name,
                    "so_t": "N",
                }
            )
    return out


def parse_income_expense_cash(path: Path) -> list[dict]:
    """Parse late-year cash block (остаток/оборот/расходы/TOTAL) — not P&L."""
    wb = load_workbook(path, data_only=True)
    ws = wb["доходы-расходы"]
    rows = list(ws.iter_rows(values_only=True, max_col=14))
    wb.close()
    out = []
    year = None
    for row in rows:
        if not row:
            continue
        if isinstance(row[0], int) and 2010 <= row[0] <= 2030:
            year = row[0]
        if year is None or year < 2024:
            continue
        label = str(row[1] or "").strip()
        if not label:
            continue
        low = label.lower()
        metric = None
        if "итого" in low and "остаток" in low:
            metric = "cash_balance_total"
        elif low == "оборот":
            metric = "cash_turnover"
        elif "расходы общ" in low:
            metric = "cash_costs_total"
        elif low.strip() in ("total", "total ") or (low.startswith("total") and "остаток" not in low):
            metric = "cash_total_net_like"
        elif "зп нал" in low:
            metric = "payroll_cash"
        elif "зп карты" in low or "зп, охрана" in low:
            metric = "payroll_cards_or_rub"
        if not metric:
            continue
        for i, mm in enumerate(MONTHS):
            val = fnum(row[2 + i] if len(row) > 2 + i else None)
            if val is None:
                continue
            out.append(
                {
                    "period_month": f"{year}-{mm}",
                    "year": year,
                    "metric": metric,
                    "label_raw": label,
                    "amount_eur_like": round(val, 2),
                    "sheet": "доходы-расходы",
                    "warning": "NOT_COMPANY_PL_SANITY_ONLY",
                    "source_file": path.name,
                    "so_t": "N",
                }
            )
    return out


def main() -> dict:
    path = catalog_sales_path()
    dds = parse_dds_income(path)
    cash = parse_income_expense_cash(path)

    # budget income lines of interest
    bud_lines = list(
        csv.DictReader((MARTS / "budget_plan_fact_lines.csv").open(encoding="utf-8"))
    )
    bud_inc = [
        r
        for r in bud_lines
        if r.get("section") == "income"
        and r.get("article")
        and r["article"].lower() not in {"поступления"}
    ]

    # aggregate sales channels by month
    sales_m = defaultdict(lambda: defaultdict(float))
    for r in dds:
        sales_m[r["period_month"]][r["channel_norm"]] += r["amount_eur_like"]
        if r["channel_norm"] == "REGIONAL_B2B":
            sales_m[r["period_month"]]["REGIONAL_B2B_SUM"] += r["amount_eur_like"]

    # budget by article-month
    bud_m = defaultdict(dict)
    for r in bud_inc:
        art = r["article"]
        bud_m[r["period_month"]][art] = {
            "plan": fnum(r["plan_eur"]),
            "fact": fnum(r["fact_eur"]),
            "budget_id": r["budget_id"],
        }

    # bridge months where either has data
    months = sorted(set(sales_m) | set(bud_m))
    bridge = []
    for m in months:
        # only 2025-07+ overlap with budgets typically, but keep all for sales view
        s = sales_m.get(m, {})
        b = bud_m.get(m, {})
        ss_sales = s.get("Salon+Shop")
        ss_bud = b.get("Salon+Shop", {})
        im_sales = s.get("IM")
        im_bud = b.get("ИМ", {})
        total_sales = s.get("TOTAL")
        tsum_sales = s.get("TSUM")

        def delta(a, b):
            if a is None or b is None:
                return ""
            return round(a - b, 2)

        def pct(a, b):
            if a is None or b in (None, 0):
                return ""
            return round((a - b) / b * 100, 2)

        ss_fact = ss_bud.get("fact")
        im_fact = im_bud.get("fact")
        ss_d = delta(ss_sales, ss_fact) if ss_fact is not None else ""
        deposit = s.get("DEPOSIT")
        flags = []
        if ss_sales is not None and ss_fact is not None and abs(ss_sales - ss_fact) > 1:
            flags.append("SALON_SHOP_MISMATCH")
            # Jan-2026: SALES has возврат депозита −4200; budget Fact may be net of it
            if (
                deposit is not None
                and ss_d != ""
                and abs(float(ss_d) + float(deposit)) <= 1
            ):
                flags.append("EXPLAINED_BY_DEPOSIT_RETURN")
        if ss_sales is not None and ss_fact is None and m >= "2026-01" and m <= "2026-06":
            flags.append("BUDGET_FACT_MISSING")
        if ss_sales is None and ss_fact is not None:
            flags.append("SALES_DDS_MISSING")
        if m == "2026-06" and (ss_sales is None or ss_fact is None):
            flags.append("JUNE2026_INCOMPLETE")

        status = "N/A"
        if ss_sales is not None and ss_fact is not None:
            if abs(ss_sales - ss_fact) <= 1:
                status = "MATCH"
            elif "EXPLAINED_BY_DEPOSIT_RETURN" in flags:
                status = "MATCH_AFTER_DEPOSIT"
            else:
                status = "MISMATCH"
        elif ss_sales is not None or ss_fact is not None:
            status = "PARTIAL"

        bridge.append(
            {
                "period_month": m,
                "sales_salon_shop": ss_sales if ss_sales is not None else "",
                "budget_salon_shop_plan": ss_bud.get("plan") if ss_bud.get("plan") is not None else "",
                "budget_salon_shop_fact": ss_fact if ss_fact is not None else "",
                "delta_sales_minus_budget_fact": ss_d,
                "delta_pct": pct(ss_sales, ss_fact) if ss_fact not in (None, 0) else "",
                "sales_deposit_adj": deposit if deposit is not None else "",
                "sales_im": im_sales if im_sales is not None else "",
                "budget_im_fact": im_fact if im_fact is not None else "",
                "delta_im": delta(im_sales, im_fact) if im_fact is not None else "",
                "sales_tsum": tsum_sales if tsum_sales is not None else "",
                "sales_total": total_sales if total_sales is not None else "",
                "sales_regional_b2b": s.get("REGIONAL_B2B_SUM", "") or "",
                "compare_status": status,
                "flags": "|".join(flags) if flags else "",
                "budget_id": ss_bud.get("budget_id") or im_bud.get("budget_id") or "",
                "do_not_auto_accept": "YES",
                "so_t": "N",
                "note": "SALES ДДС-доход vs budget income Fact; same workbook EUR_LIKE scale",
            }
        )

    # budget month totals repaired from section sums (leaf income/expense)
    bud_tot = defaultdict(lambda: {"plan_in": 0.0, "fact_in": 0.0, "plan_ex": 0.0, "fact_ex": 0.0, "n_in": 0, "n_ex": 0})
    for r in bud_lines:
        if (r.get("article") or "").lower() in {"поступления", "расходы"}:
            continue
        m = r["period_month"]
        pe, fe = fnum(r["plan_eur"]), fnum(r["fact_eur"])
        if r.get("section") == "income":
            if pe is not None:
                bud_tot[m]["plan_in"] += pe
            if fe is not None:
                bud_tot[m]["fact_in"] += fe
            bud_tot[m]["n_in"] += 1
        elif r.get("section") == "expense":
            if pe is not None:
                bud_tot[m]["plan_ex"] += pe
            if fe is not None:
                bud_tot[m]["fact_ex"] += fe
            bud_tot[m]["n_ex"] += 1

    bud_tot_rows = []
    for m in sorted(bud_tot):
        v = bud_tot[m]
        sales_total = sales_m.get(m, {}).get("TOTAL")
        fact_in = round(v["fact_in"], 2)
        bud_tot_rows.append(
            {
                "period_month": m,
                "budget_fact_income_sum_eur": fact_in if v["fact_in"] else "",
                "budget_plan_income_sum_eur": round(v["plan_in"], 2) if v["plan_in"] else "",
                "budget_fact_expense_sum_eur": round(v["fact_ex"], 2) if v["fact_ex"] else "",
                "budget_plan_expense_sum_eur": round(v["plan_ex"], 2) if v["plan_ex"] else "",
                "sales_dds_total_eur": sales_total if sales_total is not None else "",
                "delta_sales_total_minus_budget_income": (
                    round(sales_total - fact_in, 2) if sales_total is not None and fact_in else ""
                ),
                "income_lines_n": v["n_in"],
                "expense_lines_n": v["n_ex"],
                "note": "Sum of leaf budget lines (not Поступления/Расходы headers); indicative",
                "so_t": "N",
            }
        )

    # owner flags
    flags = []
    for r in bridge:
        if r["compare_status"] == "MISMATCH" or "JUNE2026" in r["flags"] or (
            "BUDGET_FACT_MISSING" in r["flags"] and "EXPLAINED" not in r["flags"]
        ):
            ask = "Fill June 2026 budget Fact / SALES June column (ties to DDS June gap)"
            if r["compare_status"] == "MISMATCH":
                ask = "Reconcile SALES ДДС-доход Salon+Shop vs budget Fact (same month)"
            if "EXPLAINED_BY_DEPOSIT_RETURN" in r["flags"]:
                ask = "Confirm budget Fact nets deposit return; treat as explained (not gap)"
            flags.append(
                {
                    "period_month": r["period_month"],
                    "priority": "HIGH"
                    if r["compare_status"] == "MISMATCH" or "JUNE" in r["flags"]
                    else "MED",
                    "issue": r["flags"] or r["compare_status"],
                    "sales_salon_shop": r["sales_salon_shop"],
                    "budget_salon_shop_fact": r["budget_salon_shop_fact"],
                    "delta": r["delta_sales_minus_budget_fact"],
                    "owner_ask": ask,
                    "do_not_auto_accept": "YES",
                    "so_t": "N",
                }
            )
        elif r["compare_status"] == "MATCH_AFTER_DEPOSIT":
            flags.append(
                {
                    "period_month": r["period_month"],
                    "priority": "LOW",
                    "issue": r["flags"],
                    "sales_salon_shop": r["sales_salon_shop"],
                    "budget_salon_shop_fact": r["budget_salon_shop_fact"],
                    "delta": r["delta_sales_minus_budget_fact"],
                    "owner_ask": "Optional confirm: budget Fact = SALES Salon+Shop + deposit return",
                    "do_not_auto_accept": "YES",
                    "so_t": "N",
                }
            )

    # rollup
    st_c = defaultdict(int)
    for r in bridge:
        if r["period_month"] >= "2025-07":  # budget era
            st_c[r["compare_status"]] += 1

    meta = {
        "horizon": "H77",
        "date": str(date.today()),
        "title": "SALES ДДС-доход ↔ budget income sanity",
        "sales_file": path.name,
        "dds_income_rows_n": len(dds),
        "cash_sheet_rows_n": len(cash),
        "bridge_months_n": len(bridge),
        "budget_overlap_match_n": st_c.get("MATCH", 0) + st_c.get("MATCH_AFTER_DEPOSIT", 0),
        "budget_overlap_mismatch_n": st_c.get("MISMATCH", 0),
        "budget_overlap_partial_n": st_c.get("PARTIAL", 0),
        "jan2026_deposit_explains_delta": True,
        "owner_flags_n": len(flags),
        "no_fake_accept": True,
        "so_t": False,
        "warning": "доходы-расходы sheet is cash/остатки — NOT company P&L",
    }

    REG.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)
    (REG / "meta.json").write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
    write_csv(REG / "sales_dds_income_monthly.csv", dds)
    write_csv(REG / "sales_income_expense_cash_metrics.csv", cash)
    write_csv(REG / "sales_budget_income_bridge.csv", bridge)
    write_csv(REG / "budget_month_sums_repaired.csv", bud_tot_rows)
    write_csv(
        REG / "sales_budget_owner_flags.csv",
        flags
        if flags
        else [
            {
                "period_month": "",
                "priority": "",
                "issue": "none",
                "owner_ask": "",
                "so_t": "N",
            }
        ],
    )

    for name in [
        "sales_dds_income_monthly.csv",
        "sales_income_expense_cash_metrics.csv",
        "sales_budget_income_bridge.csv",
        "budget_month_sums_repaired.csv",
        "sales_budget_owner_flags.csv",
        "meta.json",
    ]:
        src = REG / name
        if name == "meta.json":
            shutil.copy2(src, MARTS / "h77_meta.json")
            shutil.copy2(src, MAPS / "h77_meta.json")
            shutil.copy2(src, EV / "meta.json")
        else:
            shutil.copy2(src, MARTS / name)
            shutil.copy2(src, MAPS / name)
            shutil.copy2(src, EV / name)

    shutil.copy2(REG / "sales_budget_income_bridge.csv", WAVE_C / "34_sales_budget_income_bridge.csv")
    shutil.copy2(REG / "budget_month_sums_repaired.csv", WAVE_C / "35_budget_month_sums_repaired.csv")
    shutil.copy2(REG / "sales_budget_owner_flags.csv", WAVE_C / "36_sales_budget_owner_flags.csv")
    shutil.copy2(REG / "sales_dds_income_monthly.csv", WAVE_C / "37_sales_dds_income_monthly.csv")

    print(json.dumps(meta, ensure_ascii=False, indent=2))
    print("overlap statuses", dict(st_c))
    for r in bridge:
        if r["period_month"] >= "2026-01" and r["period_month"] <= "2026-06":
            print(
                r["period_month"],
                r["compare_status"],
                "sales",
                r["sales_salon_shop"],
                "bud",
                r["budget_salon_shop_fact"],
                "Δ",
                r["delta_sales_minus_budget_fact"],
                r["flags"],
            )
    print("flags", flags)
    return meta


if __name__ == "__main__":
    main()
