#!/usr/bin/env python3
"""
H25: Domain operating board после RACI ACCEPT.

Зачем:
- ACCEPT доменов Cash/Bank/Tax/Payroll есть → превратить BLOCKED/gaps
  в персональные задачи owners (не общий «надо данные»).
- Зафиксировать OPEN-домены как явный ask на ФИО.
- Обновить release gate строками owner_domain / owner_fio.

Не выдумывает ФИО. Не объявляет полный SoT.
"""
from __future__ import annotations

import csv
import json
import shutil
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[3]
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/h25_domain_ops_20260724"
MART = ROOT / "live/marts"
PACKET = ROOT / "live/YANINA_OWNER_PACKET_RACI_AND_REQUESTS.xlsx"
CC = ROOT / "live/YANINA_LIVE_CONTROL_CENTER_20260723.xlsx"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF")

# control → domain (accepted or open)
CONTROL_DOMAIN = {
    "IM_ACQ_COMBO": "BANK",
    "BANK_DDS_CORE": "BANK",
    "TAX_CASH_BANK": "TAX",
    "TSUM_NET_MODEL": "B2B",  # commercial/legal; owner ещё OPEN
    "PAYROLL_MULTI": "PAYROLL",
    "OPEX_MULTI": "BANK",
    "B2B_SETTLE_BANK": "B2B",
}

ASK_FOR_OPEN = {
    "PRODUCT": "Нужен Owner Product/SKU MDM — ACCEPT sku_alias_master candidates",
    "COST": "Нужен Owner Costing — cost versions / quarantine 0-3243 follow-up",
    "PRODUCTION": "Нужен Owner Production — PO/BOM",
    "B2B": "Нужен Owner B2B/Commercial — open settlements + TSUM rate",
    "DATA_STEWARD": "Нужен Data steward — ужесточение release gate / monthly retests",
}


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def load_owners() -> dict[str, dict]:
    rows = list(csv.DictReader(open(MART / "sot_owners.csv", encoding="utf-8")))
    return {r["domain"]: r for r in rows}


def build_gap_board(owners: dict[str, dict]) -> list[dict]:
    fails = list(csv.DictReader(open(MART / "release_gate_fails.csv", encoding="utf-8")))
    board = []
    for f in fails:
        cid = f.get("control_id") or ""
        domain = CONTROL_DOMAIN.get(cid, "DATA_STEWARD")
        own = owners.get(domain, {})
        ost = own.get("status") or "OPEN_NEEDS_OWNER"
        # ACCEPTED и ACCEPTED_STUB (H26) — задача назначена
        board.append(
            {
                "period_month": f.get("period_month"),
                "control_id": cid,
                "status": f.get("status"),
                "metric": f.get("metric"),
                "detail": f.get("detail"),
                "owner_domain": domain,
                "owner_fio": own.get("owner") or "",
                "owner_status": ost,
                "priority": "P1" if ost == "ACCEPTED" else ("P1_STUB" if ost == "ACCEPTED_STUB" else "P0_NEED_OWNER"),
                "is_stub_owner": "Y" if ost == "ACCEPTED_STUB" else "N",
                "action": _action_for(cid, f.get("period_month") or "", f.get("status") or ""),
                "linked_request": (f.get("owner_actions") or ""),
                "updated_at": NOW,
            }
        )
    # sort: real accepted → stubs → open
    rank = {"ACCEPTED": 0, "ACCEPTED_STUB": 1}
    board.sort(
        key=lambda r: (rank.get(r["owner_status"], 2), r["period_month"], r["control_id"])
    )
    return board


def _action_for(cid: str, month: str, status: str) -> str:
    if cid == "IM_ACQ_COMBO":
        return f"Прислать эквайринг-реестр (Tinkoff/TBank/VTB) за {month}"
    if cid == "BANK_DDS_CORE":
        if month == "2026-06":
            return "Дослать строки ДДС 2026 за июнь (в файле сейчас пусто) или подтвердить отсутствие"
        return f"Разобрать WIDE_GAP bank↔DDS за {month} (Salon out / карты / статьи)"
    if cid == "TAX_CASH_BANK":
        return f"Сверить налоговый платёж/начисление за {month} (status={status})"
    if cid == "TSUM_NET_MODEL":
        return f"Дать агентский % ЦУМ из договора; сверить модель за {month}"
    return f"Закрыть control {cid} за {month}"


def enrich_gate(owners: dict[str, dict]) -> list[dict]:
    rows = list(csv.DictReader(open(MART / "release_gate_month.csv", encoding="utf-8")))
    fails = list(csv.DictReader(open(MART / "release_gate_fails.csv", encoding="utf-8")))
    by_m = defaultdict(list)
    for f in fails:
        by_m[f["period_month"]].append(f)

    out = []
    for r in rows:
        pm = r["period_month"]
        fl = by_m.get(pm, [])
        domains = sorted({CONTROL_DOMAIN.get(f["control_id"], "?") for f in fl}) if fl else []
        fios = []
        for d in domains:
            o = owners.get(d, {})
            if o.get("owner"):
                fios.append(f"{d}:{o['owner']}")
            else:
                fios.append(f"{d}:OPEN")
        nr = dict(r)
        nr["fail_owner_domains"] = "|".join(domains)
        nr["fail_owner_fios"] = " | ".join(fios)
        has_owner = any(
            owners.get(d, {}).get("status") in ("ACCEPTED", "ACCEPTED_STUB") for d in domains
        )
        has_stub = any(owners.get(d, {}).get("status") == "ACCEPTED_STUB" for d in domains)
        if r["verdict"] == "BLOCKED" and has_owner:
            nr["ops_status"] = "OWNER_ASSIGNED_GAP_STUB" if has_stub and not any(
                owners.get(d, {}).get("status") == "ACCEPTED" for d in domains
            ) else "OWNER_ASSIGNED_GAP"
        else:
            nr["ops_status"] = r["verdict"]
        out.append(nr)
    return out


def build_owner_packs(board: list[dict], owners: dict[str, dict]) -> dict[str, list[dict]]:
    packs: dict[str, list[dict]] = defaultdict(list)
    for r in board:
        key = r["owner_fio"] or f"OPEN:{r['owner_domain']}"
        packs[key].append(r)
    return packs


def build_open_domain_asks(owners: dict[str, dict]) -> list[dict]:
    rows = []
    for domain, meta in owners.items():
        if meta.get("status") != "OPEN_NEEDS_OWNER":
            continue
        rows.append(
            {
                "domain": domain,
                "status": meta["status"],
                "raci_source": meta.get("source"),
                "ask": ASK_FOR_OPEN.get(domain, "Назначить Owner FIO в RACI"),
                "blocks": {
                    "PRODUCT": "sku_alias_master ACCEPT",
                    "COST": "cost identity / BOM",
                    "PRODUCTION": "PO/BOM",
                    "B2B": "TSUM rate + B2B open 15",
                    "DATA_STEWARD": "gate hardening + retests",
                }.get(domain, "domain ownership"),
                "suggested_candidate": {
                    "B2B": "коммерция / байер (уточнить ФИО)",
                    "PRODUCT": "номенклатура 1С (уточнить ФИО)",
                    "COST": "Меркушина/Жукова как costing sources — нужен Owner над ними",
                    "PRODUCTION": "производство (уточнить ФИО)",
                    "DATA_STEWARD": "Сливяк или отдельный data steward",
                }.get(domain, ""),
                "decision_FIO": "",
                "decision_ACCEPT": "",
            }
        )
    return rows


def write_xlsx(board, gate, asks, packs):
    wb = Workbook()
    ws = wb.active
    ws.title = "00_Gap_Board"
    ws["A1"] = "YANINA Domain Ops Board (H25)"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A2"] = NOW
    ws.append([])
    if board:
        ws.append(list(board[0].keys()))
        for c in range(1, len(board[0]) + 1):
            ws.cell(4, c).fill = HDR_FILL
            ws.cell(4, c).font = HDR_FONT
        for r in board:
            ws.append(list(r.values()))

    def add_sheet(name, rows):
        w = wb.create_sheet(name)
        if not rows:
            w["A1"] = "empty"
            return
        w.append(list(rows[0].keys()))
        for c in range(1, len(rows[0]) + 1):
            w.cell(1, c).fill = HDR_FILL
            w.cell(1, c).font = HDR_FONT
        for r in rows:
            w.append(list(r.values()))

    add_sheet("Release_Gate_Owned", gate)
    add_sheet("Open_Domain_Asks", asks)

    # per-owner sheets (safe names)
    for owner, rows in packs.items():
        safe = (
            owner.replace(" ", "_")
            .replace(":", "_")
            .replace("/", "_")[:28]
        )
        add_sheet(f"Pack_{safe}", rows)

    path = ROOT / "live/YANINA_DOMAIN_OPS_H25.xlsx"
    wb.save(path)
    wb.save(OUT / "YANINA_DOMAIN_OPS_H25.xlsx")
    wb.save(EV / "YANINA_DOMAIN_OPS_H25.xlsx")
    return path


def update_packet(asks: list[dict], board: list[dict]):
    if not PACKET.exists():
        return
    wb = load_workbook(PACKET)
    name = "DOMAIN_OPS_H25"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, 3)
    ws["A1"] = "Domain Ops (H25) — задачи после ACCEPT"
    ws["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws["A2"] = NOW
    ws.append([])
    ws.append(["section", "key", "detail"])
    ws.append(["SUMMARY", "gap_rows", str(len(board))])
    ws.append(
        [
            "SUMMARY",
            "assigned_to_accepted_owners",
            str(sum(1 for r in board if r["owner_status"] == "ACCEPTED")),
        ]
    )
    ws.append(
        [
            "SUMMARY",
            "blocked_on_open_owners",
            str(sum(1 for r in board if r["owner_status"] != "ACCEPTED")),
        ]
    )
    ws.append([])
    ws.append(["OPEN_DOMAIN", "ask", "suggested_candidate"])
    for a in asks:
        ws.append([a["domain"], a["ask"], a["suggested_candidate"]])
    if "README" in wb.sheetnames:
        wb["README"]["A16"] = (
            f"H25 {NOW}: Domain Ops board — live/YANINA_DOMAIN_OPS_H25.xlsx + лист DOMAIN_OPS_H25"
        )
    wb.save(PACKET)


def update_cc(summary: dict):
    if not CC.exists():
        return
    wb = load_workbook(CC)
    if "H25_DomainOps" in wb.sheetnames:
        del wb["H25_DomainOps"]
    ws = wb.create_sheet("H25_DomainOps", 0)
    ws["A1"] = "H25 Domain Ops"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A2"] = NOW
    ws["A4"] = summary["finding"]
    ws["A5"] = "Gap rows"
    ws["B5"] = summary["gap_rows"]
    ws["A6"] = "Assigned (accepted owners)"
    ws["B6"] = summary["assigned_n"]
    ws["A7"] = "Need open-domain FIO"
    ws["B7"] = summary["need_owner_n"]
    ws["A8"] = "Pack"
    ws["B8"] = "live/YANINA_DOMAIN_OPS_H25.xlsx"
    wb.save(CC)


def write_md(board, asks, summary):
    by_owner = Counter(r["owner_fio"] or f"OPEN:{r['owner_domain']}" for r in board)
    lines = [
        "# Domain Ops (H25)",
        "",
        f"Updated: {NOW}",
        "",
        "После RACI ACCEPT — дыры разложены по owners.",
        "",
        f"- Gap rows: **{summary['gap_rows']}**",
        f"- На accepted owners: **{summary['assigned_n']}**",
        f"- Упираются в OPEN domain: **{summary['need_owner_n']}**",
        "",
        "## По владельцам",
        "",
    ]
    for k, n in by_owner.most_common():
        lines.append(f"- **{k}**: {n} задач")
    lines.extend(["", "## OPEN domains — нужны ФИО", ""])
    for a in asks:
        lines.append(f"- **{a['domain']}**: {a['ask']} _(кандидат: {a['suggested_candidate']})_")
    lines.extend(
        [
            "",
            "## Файлы",
            "- `live/YANINA_DOMAIN_OPS_H25.xlsx`",
            "- `live/marts/domain_gap_board.csv`",
            "- `live/marts/open_domain_asks.csv`",
            "- `live/marts/release_gate_month.csv` (с owner columns)",
            "",
        ]
    )
    text = "\n".join(lines)
    (OUT / "DOMAIN_OPS.md").write_text(text, encoding="utf-8")
    (ROOT / "live/DOMAIN_OPS.md").write_text(text, encoding="utf-8")
    (EV / "DOMAIN_OPS.md").write_text(text, encoding="utf-8")


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)

    owners = load_owners()
    board = build_gap_board(owners)
    gate = enrich_gate(owners)
    asks = build_open_domain_asks(owners)
    packs = build_owner_packs(board, owners)

    write_csv(MART / "domain_gap_board.csv", board, list(board[0].keys()) if board else ["period_month"])
    write_csv(OUT / "domain_gap_board.csv", board, list(board[0].keys()) if board else ["period_month"])
    write_csv(MART / "open_domain_asks.csv", asks, list(asks[0].keys()) if asks else ["domain"])
    write_csv(MART / "release_gate_month.csv", gate, list(gate[0].keys()) if gate else ["period_month"])

    # per-owner csv packs
    pack_dir = MART / "owner_packs"
    pack_dir.mkdir(parents=True, exist_ok=True)
    for owner, rows in packs.items():
        safe = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in owner)[:40]
        write_csv(pack_dir / f"{safe}.csv", rows, list(rows[0].keys()))

    xlsx = write_xlsx(board, gate, asks, packs)
    update_packet(asks, board)

    assigned = sum(1 for r in board if r["owner_status"] in ("ACCEPTED", "ACCEPTED_STUB"))
    real = sum(1 for r in board if r["owner_status"] == "ACCEPTED")
    stub = sum(1 for r in board if r["owner_status"] == "ACCEPTED_STUB")
    need = sum(1 for r in board if r["owner_status"] not in ("ACCEPTED", "ACCEPTED_STUB"))
    summary = {
        "wave": "H25",
        "generated_at": NOW,
        "path_choice": "Domain ops board — assign BLOCKED gaps to accepted owners",
        "finding": (
            f"H25: {len(board)} gap rows → {assigned} assigned "
            f"(real={real}, stub={stub}), open={need}. "
            f"Open domain asks: {len(asks)}."
        ),
        "gap_rows": len(board),
        "assigned_n": assigned,
        "assigned_real_n": real,
        "assigned_stub_n": stub,
        "need_owner_n": need,
        "open_domain_asks": len(asks),
        "xlsx": str(xlsx.relative_to(ROOT)),
        "not_sot": True,
    }
    write_md(board, asks, summary)
    update_cc(summary)
    (OUT / "h25_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    (EV / "h25_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    for name in ("domain_gap_board.csv", "DOMAIN_OPS.md"):
        shutil.copy2(OUT / name, EV / name)
    shutil.copy2(MART / "open_domain_asks.csv", EV / "open_domain_asks.csv")

    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
