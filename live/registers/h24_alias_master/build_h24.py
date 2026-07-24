#!/usr/bin/env python3
"""
H24: SKU alias master (controlled candidates) после RACI ACCEPT.

Зачем: фаза C1 — зафиксировать alias registry из H17 candidates.
Product Owner ещё OPEN → статус CONTROLLED_CANDIDATE (не auto-apply в sales).
После назначения DOM-PRODUCT можно ACCEPT registry целиком.

Не меняет COGS/sales joins.
"""
from __future__ import annotations

import csv
import json
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[3]
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/h24_alias_master_20260724"
MART = ROOT / "live/marts"
CC = ROOT / "live/YANINA_LIVE_CONTROL_CENTER_20260723.xlsx"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)

    src = MART / "sku_alias_candidates.csv"
    if not src.exists():
        raise SystemExit("missing sku_alias_candidates.csv — run H17 first")
    cands = list(csv.DictReader(open(src, encoding="utf-8")))

    master = []
    for r in cands:
        master.append(
            {
                "canonical_sku": r.get("canonical_sku"),
                "alias_status": "CONTROLLED_CANDIDATE",
                "fixed_lines": r.get("fixed_lines"),
                "channels": r.get("channels"),
                "fix_types": r.get("fix_types"),
                "sale_name_sample": r.get("sale_name_sample"),
                "wrong_or_alt_cost_name_sample": r.get("wrong_or_alt_cost_name_sample"),
                "preferred_cost_version_ids": r.get("preferred_cost_version_ids"),
                "proposed_rule": r.get("proposed_rule"),
                "owner_domain": "DOM-PRODUCT",
                "owner_status": "OPEN_NEEDS_OWNER",
                "registry_accept": "PENDING_PRODUCT_OWNER",
                "applied_to_sales": "N",
                "so_t": "N",
                "note": "Promoted from H17 after RACI ACCEPT domains; do not auto-apply until Product Owner named",
                "updated_at": NOW,
            }
        )

    # also encode accepted finance exceptions as non-alias policies
    ex_path = MART / "margin_exceptions.csv"
    if ex_path.exists():
        for r in csv.DictReader(open(ex_path, encoding="utf-8")):
            if r.get("policy") == "WHOLESALE_OK_LOSS":
                master.append(
                    {
                        "canonical_sku": r.get("canonical_sku"),
                        "alias_status": "MARGIN_EXCEPTION_ACCEPTED",
                        "fixed_lines": r.get("lines"),
                        "channels": r.get("channels"),
                        "fix_types": r.get("policy"),
                        "sale_name_sample": "",
                        "wrong_or_alt_cost_name_sample": "",
                        "preferred_cost_version_ids": "",
                        "proposed_rule": "Keep FILE cost; flag WHOLESALE_OK_LOSS",
                        "owner_domain": "DOM-B2B/FINANCE",
                        "owner_status": "ACCEPTED_VIA_H23",
                        "registry_accept": "ACCEPT",
                        "applied_to_sales": "Y",
                        "so_t": "DOMAIN_OWNED",
                        "note": "Owner-accepted commercial loss exception",
                        "updated_at": NOW,
                    }
                )
            elif r.get("policy") == "COST_IDENTITY_QUARANTINE":
                master.append(
                    {
                        "canonical_sku": r.get("canonical_sku"),
                        "alias_status": "QUARANTINE_ACCEPTED",
                        "fixed_lines": r.get("lines"),
                        "channels": r.get("channels"),
                        "fix_types": r.get("policy"),
                        "sale_name_sample": "",
                        "wrong_or_alt_cost_name_sample": "",
                        "preferred_cost_version_ids": "",
                        "proposed_rule": "No COGS until sweatshirt cost version exists",
                        "owner_domain": "DOM-COST/PRODUCT",
                        "owner_status": "ACCEPTED_VIA_H23",
                        "registry_accept": "ACCEPT",
                        "applied_to_sales": "Y",
                        "so_t": "DOMAIN_OWNED",
                        "note": "Owner-accepted quarantine",
                        "updated_at": NOW,
                    }
                )

    fields = list(master[0].keys())
    write_csv(MART / "sku_alias_master.csv", master, fields)
    write_csv(OUT / "sku_alias_master.csv", master, fields)
    write_csv(ROOT / "live/registers/h24_alias_master/sku_alias_master.csv", master, fields)

    cand_n = sum(1 for r in master if r["alias_status"] == "CONTROLLED_CANDIDATE")
    acc_n = sum(1 for r in master if r["registry_accept"] == "ACCEPT")

    summary = {
        "wave": "H24",
        "generated_at": NOW,
        "finding": (
            f"H24: sku_alias_master created — {cand_n} controlled candidates (pending Product Owner), "
            f"{acc_n} accepted exception/quarantine rows from H23."
        ),
        "controlled_candidates": cand_n,
        "accepted_rows": acc_n,
        "applied_to_sales_joins": False,
        "not_sot": True,
        "path": "live/marts/sku_alias_master.csv",
    }
    (OUT / "h24_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    (EV / "h24_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    shutil.copy2(MART / "sku_alias_master.csv", EV / "sku_alias_master.csv")

    md = [
        "# SKU Alias Master (H24)",
        "",
        f"Updated: {NOW}",
        "",
        f"- Controlled candidates (pending Product Owner): **{cand_n}**",
        f"- Accepted exception/quarantine rows: **{acc_n}**",
        "- Sales joins: **not auto-applied** for candidates",
        "",
        "File: `live/marts/sku_alias_master.csv`",
        "",
        "Next: назначить DOM-PRODUCT Owner → ACCEPT candidates → optional apply rules.",
        "",
    ]
    (ROOT / "live/SKU_ALIAS_MASTER.md").write_text("\n".join(md), encoding="utf-8")
    (EV / "SKU_ALIAS_MASTER.md").write_text("\n".join(md), encoding="utf-8")

    if CC.exists():
        wb = load_workbook(CC)
        if "H24_Alias" in wb.sheetnames:
            del wb["H24_Alias"]
        ws = wb.create_sheet("H24_Alias", 0)
        ws["A1"] = "H24 SKU Alias Master"
        ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
        ws["A2"] = NOW
        ws["A4"] = summary["finding"]
        ws["A5"] = "Candidates"
        ws["B5"] = cand_n
        ws["A6"] = "Accepted rows"
        ws["B6"] = acc_n
        wb.save(CC)

    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
