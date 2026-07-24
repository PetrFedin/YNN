#!/usr/bin/env python3
"""
H29: Канал индивидуального пошива / МД отдельно от товарных goods marts.

Зачем (P0 channel mix из H27):
- Бриф 2025: 83% инд.пошив / 8% ИМ / 9% B2B(ЦУМ)
- W4 sales_lines = только товар B2B/IM/TSUM — услуги МД там нет
- Уже есть proxy: sales_dds_income_eur.csv канал Salon+Shop ≈83.9% в 2025

Делаем:
1) md_income_month_eur — Salon+Shop (+ классификация каналов SALES DDS)
2) channel_mix_income_eur — сверка с брифом по годам
3) md_opex_month — расходы ДДС direction_activity Модный дом
4) md_bridge_month — доход EUR@100 vs opex RUB (indicative)
5) Не мешаем с margin_channel_total (goods+COGS)

Не парсим МД.xlsx в этом шаге (можно H30) — сначала закрываем бриф из уже разобранного.
"""
from __future__ import annotations

import csv
import json
import shutil
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[3]
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/h29_md_channel_20260724"
MART = ROOT / "live/marts"
W1 = ROOT / "live/registers/w1_bank_cash"
W4 = ROOT / "live/registers/w4_sales_settle"
CC = ROOT / "live/YANINA_LIVE_CONTROL_CENTER_20260723.xlsx"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")

# Политика курса из брифа H27
FX_EUR_RUB = 100.0

BRIEF_2025 = {"MD_INDIVIDUAL": 83.0, "IM": 8.0, "TSUM_B2B": 9.0}


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def fnum(v):
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(" ", "").replace(",", "."))
    except (TypeError, ValueError):
        return None


def map_income_channel(raw: str) -> str:
    """Таксономия SALES DDS → управленческие каналы брифа."""
    c = (raw or "").strip()
    if c in ("TOTAL",):
        return "TOTAL"
    if c == "Salon+Shop":
        return "MD_INDIVIDUAL"
    if c in ("ИМ", "возврат ИМ"):
        return "IM"
    if c == "ЦУМ":
        return "TSUM_B2B"
    if c.startswith("возврат"):
        return "OTHER_RETURNS"
    # города / мелкий опт в SALES DDS — не путать с W4 B2B goods
    return "OTHER_REGIONAL"


def build_income() -> tuple[list[dict], list[dict], list[dict]]:
    raw_rows = list(csv.DictReader(open(W4 / "sales_dds_income_eur.csv", encoding="utf-8")))
    month_ch = []
    for r in raw_rows:
        if (r.get("channel") or "") == "TOTAL":
            continue
        mapped = map_income_channel(r.get("channel") or "")
        eur = fnum(r.get("amount_eur")) or 0.0
        month_ch.append(
            {
                "period_month": r.get("period_month"),
                "channel_raw": r.get("channel"),
                "channel": mapped,
                "amount_eur": round(eur, 2),
                "amount_rub_fx100": round(eur * FX_EUR_RUB, 2),
                "fx_policy": FX_EUR_RUB,
                "source": r.get("source") or "SALES_DDS_INCOME",
                "source_file_id": r.get("source_file_id"),
                "so_t": "N",
                "note": "Income from SALES DDS sheet; not W4 goods sales_lines",
            }
        )

    # rollup by month × mapped channel
    by = defaultdict(float)
    for r in month_ch:
        by[(r["period_month"], r["channel"])] += r["amount_eur"]
    rollup = []
    for (pm, ch), eur in sorted(by.items()):
        rollup.append(
            {
                "period_month": pm,
                "channel": ch,
                "amount_eur": round(eur, 2),
                "amount_rub_fx100": round(eur * FX_EUR_RUB, 2),
                "fx_policy": FX_EUR_RUB,
            }
        )

    # year mix vs brief
    year_ch = defaultdict(lambda: defaultdict(float))
    for r in rollup:
        y = (r["period_month"] or "")[:4]
        if not y.isdigit():
            continue
        year_ch[y][r["channel"]] += r["amount_eur"]

    mix = []
    for y in sorted(year_ch.keys()):
        tot = sum(year_ch[y].values()) or 1.0
        for ch, eur in sorted(year_ch[y].items(), key=lambda x: -x[1]):
            share = 100.0 * eur / tot
            brief = BRIEF_2025.get(ch) if y == "2025" else None
            delta = round(share - brief, 1) if brief is not None else ""
            mix.append(
                {
                    "year": y,
                    "channel": ch,
                    "amount_eur": round(eur, 2),
                    "share_pct": round(share, 1),
                    "brief_share_pct": brief if brief is not None else "",
                    "delta_vs_brief_pp": delta,
                    "fx_policy": FX_EUR_RUB,
                    "basis": "SALES_DDS_INCOME_EUR",
                }
            )
        # TOTAL
        mix.append(
            {
                "year": y,
                "channel": "TOTAL",
                "amount_eur": round(sum(year_ch[y].values()), 2),
                "share_pct": 100.0,
                "brief_share_pct": "",
                "delta_vs_brief_pp": "",
                "fx_policy": FX_EUR_RUB,
                "basis": "SALES_DDS_INCOME_EUR",
            }
        )
    return month_ch, rollup, mix


def build_md_opex() -> list[dict]:
    by = defaultdict(lambda: {"rub": 0.0, "eur": 0.0, "n": 0})
    for r in csv.DictReader(open(W1 / "cash_lines.csv", encoding="utf-8")):
        da = (r.get("direction_activity") or "").lower().replace("ё", "е")
        if "модн" not in da:
            continue
        if r.get("ledger") != "B":
            continue
        pm = r.get("period_month") or ""
        rub = fnum(r.get("amount_rub")) or 0.0
        eur = fnum(r.get("amount_eur")) or 0.0
        by[pm]["rub"] += rub
        by[pm]["eur"] += eur
        by[pm]["n"] += 1
    rows = []
    for pm in sorted(by.keys()):
        v = by[pm]
        rows.append(
            {
                "period_month": pm,
                "direction_activity": "Модный дом",
                "expense_rub": round(v["rub"], 2),
                "expense_eur": round(v["eur"], 2),
                "n_lines": v["n"],
                "ledger": "B",
                "note": "DDS outflows tagged Модный дом; not product COGS",
            }
        )
    return rows


def build_bridge(income_roll: list[dict], opex: list[dict]) -> list[dict]:
    inc = defaultdict(float)
    for r in income_roll:
        if r["channel"] == "MD_INDIVIDUAL":
            inc[r["period_month"]] += r["amount_rub_fx100"]
    op = {r["period_month"]: r["expense_rub"] for r in opex}
    months = sorted(set(inc) | set(op))
    rows = []
    for pm in months:
        revenue = inc.get(pm, 0.0)
        expense = op.get(pm, 0.0)
        rows.append(
            {
                "period_month": pm,
                "md_income_rub_fx100": round(revenue, 2),
                "md_opex_rub": round(expense, 2),
                "md_contrib_rub_indicative": round(revenue - expense, 2),
                "note": (
                    "Indicative only: income from SALES DDS EUR@100; "
                    "opex=DDS Модный дом (may include materials/payroll overlapping COGS elsewhere)"
                ),
                "so_t": "N",
            }
        )
    return rows


def update_channel_mix_brief(mix: list[dict]):
    """Refresh channel_mix_vs_brief with income-basis 2025."""
    rows = []
    m2025 = [r for r in mix if r["year"] == "2025" and r["channel"] != "TOTAL"]
    tot = next((r["amount_eur"] for r in mix if r["year"] == "2025" and r["channel"] == "TOTAL"), 0)
    for r in m2025:
        rows.append(
            {
                "scope": "brief_vs_sales_dds_2025",
                "channel": r["channel"],
                "brief_share_pct": r.get("brief_share_pct") or "",
                "our_share_pct": r["share_pct"],
                "our_amount_eur": r["amount_eur"],
                "delta_pp": r.get("delta_vs_brief_pp") or "",
                "note": "Income EUR basis (SALES DDS); closes H27 MF-CHANNEL-MIX for income view",
            }
        )
    # keep goods note
    rows.append(
        {
            "scope": "goods_costed_marts",
            "channel": "B2B+IM+TSUM",
            "brief_share_pct": "",
            "our_share_pct": "",
            "our_amount_eur": "",
            "delta_pp": "",
            "note": "Goods marts intentionally exclude MD services — see margin_channel_total",
        }
    )
    write_csv(MART / "channel_mix_vs_brief.csv", rows, list(rows[0].keys()))
    return tot


def update_model_flag():
    path = MART / "model_flags_h27.csv"
    if not path.exists():
        return
    rows = list(csv.DictReader(open(path, encoding="utf-8")))
    for r in rows:
        if r.get("flag_id") == "MF-CHANNEL-MIX-2025":
            r["status"] = "ADDRESSED_H29_INCOME_VIEW"
            r["action"] = (
                "Income view via SALES DDS: MD_INDIVIDUAL≈Salon+Shop. "
                "Goods marts remain separate. Optional: parse МД.xlsx line-level (H30)."
            )
    write_csv(path, rows, list(rows[0].keys()))


def write_md(summary: dict):
    lines = [
        "# H29 — Канал МД / индивидуальный пошив",
        "",
        f"Updated: {NOW}",
        "",
        "## Зачем",
        "Бриф: 83% выручки — услуги инд.пошива. В W4 goods этого канала нет.",
        "Берём уже разобранный SALES DDS (`Salon+Shop`) + opex «Модный дом».",
        "",
        "## 2025 income mix (EUR, SALES DDS)",
        "",
    ]
    for ch, share, brief in summary.get("mix_2025", []):
        lines.append(f"- **{ch}**: {share}% (бриф {brief if brief is not None else '—'}%)")
    lines.extend(
        [
            "",
            f"- FX policy: EUR×**{FX_EUR_RUB}**",
            f"- MD opex months: **{summary.get('opex_months')}**",
            "",
            "## Политика",
            "1. `MD_INDIVIDUAL` — услуги (income EUR), не путать с B2B/IM/TSUM goods.",
            "2. `margin_channel_total` — только товар + COGS.",
            "3. Bridge MD income@100 − DDS Модный дом — indicative, возможен overlap с COGS/payroll.",
            "4. Следующий шаг: line-level из `МД — копия.xlsx` (H30), если нужны заказы.",
            "",
            "Files:",
            "- `live/marts/md_income_month_eur.csv`",
            "- `live/marts/channel_mix_income_eur.csv`",
            "- `live/marts/md_opex_month.csv`",
            "- `live/marts/md_bridge_month.csv`",
            "",
        ]
    )
    text = "\n".join(lines)
    (OUT / "MD_CHANNEL.md").write_text(text, encoding="utf-8")
    (ROOT / "live/MD_CHANNEL.md").write_text(text, encoding="utf-8")
    (EV / "MD_CHANNEL.md").write_text(text, encoding="utf-8")


def update_cc(summary: dict):
    if not CC.exists():
        return
    wb = load_workbook(CC)
    if "H29_MD_Channel" in wb.sheetnames:
        del wb["H29_MD_Channel"]
    ws = wb.create_sheet("H29_MD_Channel", 0)
    ws["A1"] = "H29 MD / Individual sewing channel"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A2"] = NOW
    ws["A4"] = summary["finding"]
    ws["A5"] = "Doc"
    ws["B5"] = "live/MD_CHANNEL.md"
    wb.save(CC)


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)

    detail, rollup, mix = build_income()
    opex = build_md_opex()
    bridge = build_bridge(rollup, opex)
    update_channel_mix_brief(mix)
    update_model_flag()

    md_only = [r for r in detail if r["channel"] == "MD_INDIVIDUAL"]
    write_csv(MART / "md_income_detail_eur.csv", detail, list(detail[0].keys()))
    write_csv(MART / "md_income_month_eur.csv", md_only, list(md_only[0].keys()) if md_only else ["period_month"])
    # also rollup MD only monthly from rollup
    md_roll = [r for r in rollup if r["channel"] == "MD_INDIVIDUAL"]
    write_csv(MART / "md_income_month_rollup_eur.csv", md_roll, list(md_roll[0].keys()) if md_roll else ["period_month"])
    write_csv(MART / "channel_mix_income_eur.csv", mix, list(mix[0].keys()))
    write_csv(OUT / "channel_mix_income_eur.csv", mix, list(mix[0].keys()))
    write_csv(MART / "md_opex_month.csv", opex, list(opex[0].keys()) if opex else ["period_month"])
    write_csv(MART / "md_bridge_month.csv", bridge, list(bridge[0].keys()) if bridge else ["period_month"])

    mix_2025 = []
    for r in mix:
        if r["year"] == "2025" and r["channel"] in BRIEF_2025:
            mix_2025.append((r["channel"], r["share_pct"], r["brief_share_pct"]))

    summary = {
        "wave": "H29",
        "generated_at": NOW,
        "finding": (
            "H29: MD_INDIVIDUAL channel from SALES DDS Salon+Shop; "
            + ", ".join(f"{c}={s}% (brief {b}%)" for c, s, b in mix_2025)
            + f"; MD opex months={len(opex)}. Goods marts unchanged."
        ),
        "mix_2025": mix_2025,
        "opex_months": len(opex),
        "md_income_months": len(md_roll),
        "fx_policy": FX_EUR_RUB,
        "not_sot": True,
        "md_xlsx_parsed": False,
    }
    write_md(summary)
    update_cc(summary)
    (OUT / "h29_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    (EV / "h29_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    for name in (
        "channel_mix_income_eur.csv",
        "md_opex_month.csv",
        "md_bridge_month.csv",
        "MD_CHANNEL.md",
        "h29_summary.json",
    ):
        src = MART / name if (MART / name).exists() else OUT / name
        if src.exists():
            shutil.copy2(src, EV / name)
    shutil.copy2(MART / "md_income_month_eur.csv", EV / "md_income_month_eur.csv")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
