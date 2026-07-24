#!/usr/bin/env python3
"""
H10: channel cash recon + stronger B2B settle↔bank + SKU identity registry.

Зачем:
1) IM/TSUM выручка плохо стыкуется построчно с банком (эквайринг/агент) —
   нужен месячный recon по классам входящих платежей.
2) B2B settle↔bank: улучшить нормализацию имён + окно ±2 мес.
3) Зафиксировать dual-SKU identity из H9 как реестр (не SoT, но master-кандидат).

Не SoT. RACI не трогаем.
"""
from __future__ import annotations

import csv
import json
import re
import shutil
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[3]
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/h10_channel_cash_20260724"
MART = ROOT / "live/marts"
W1 = ROOT / "live/registers/w1_bank_cash"
W4 = ROOT / "live/registers/w4_sales_settle"
H9 = ROOT / "live/registers/h9_cost_identity"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def fnum(v):
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(" ", "").replace(",", "."))
    except (TypeError, ValueError):
        return None


def month_idx(pm: str) -> int | None:
    if not pm or "-" not in pm:
        return None
    y, m = pm.split("-")[:2]
    try:
        return int(y) * 12 + int(m)
    except ValueError:
        return None


def norm_name(s: str) -> str:
    s = (s or "").upper().replace("Ё", "Е")
    s = re.sub(r"[\"«»]", " ", s)
    s = re.sub(
        r"\b(ООО|АО|ЗАО|ПАО|ИП|ОБЩЕСТВО|ОГРАНИЧЕННОЙ|ОТВЕТСТВЕННОСТЬЮ|Г|МОСКВА|САНКТ|ПЕТЕРБУРГ)\b",
        " ",
        s,
    )
    s = re.sub(r"Р/С\s*\d+", " ", s)
    s = re.sub(r"[^A-ZА-Я0-9]+", " ", s)
    return " ".join(t for t in s.split() if len(t) >= 3)


def name_overlap(a: str, b: str) -> float:
    ta = set(norm_name(a).split())
    tb = set(norm_name(b).split())
    if not ta or not tb:
        return 0.0
    return len(ta & tb) / len(ta | tb)


def classify_in(p: dict) -> str:
    """Класс входящего платежа для channel cash recon."""
    c = (p.get("counterparty_raw") or "").upper().replace("Ё", "Е")
    pur = (p.get("purpose") or "").upper().replace("Ё", "Е")

    if any(x in c for x in ("ДЕКОР", "ЯНИНА ЮЛИЯ")) or "МЕЖДУ СВОИМИ" in pur or "МЕЖДУ СВОИМИ" in c:
        return "INTERNAL"
    if "ЗАЧИСЛЕНИЕ НА КАРТУ" in c and "СВОИМИ" in c:
        return "INTERNAL"

    # эквайринг IM
    if any(x in c for x in ("ТИНЬКОФФ", "TINKOFF", "ТБАНК")):
        return "ACQ_IM"
    if "ЯНДЕКС" in c or "YANDEX" in c:
        return "ACQ_IM"
    if "ЭКВАЙР" in pur or "ВОЗМЕЩЕНИЕ СР-В ПО ОПЕРАЦИЯМ ЭКВАЙРИНГА" in pur:
        return "ACQ_IM"
    if "РЕЕСТР ОПЕРАЦИЙ" in pur and any(x in c for x in ("ТИНЬКОФФ", "ТБАНК", "АЛЬФА")):
        return "ACQ_IM"

    # агент ЦУМ
    if "АГЕНТСК" in pur or "РЕАЛИЗОВАННЫЙ ТОВАР" in pur or "ДП00288281" in pur:
        return "TSUM_AGENT"
    if "ЦУМ" in c or "ТОРГОВЫЙ ДОМ ЦУМ" in c:
        return "TSUM_AGENT"

    if "МЕРКУРИ" in c:
        return "B2B_NAMED"

    # налог: не путать с формулировкой «БЕЗ НАЛОГА (НДС)» в обычных оплатах покупателей
    tax_cp = any(x in c for x in ("УФК", "КАЗНАЧЕЙ", "ФНС", "ОСФР", "СФР "))
    tax_pur = any(
        x in pur
        for x in (
            "УПЛАТА НАЛОГА",
            "ЕНС",
            "НАЛОГ НА ПРИБЫЛЬ",
            "УСН",
            "НДФЛ",
            "СТРАХОВЫМ ВЗНОСАМ",
        )
    )
    if tax_cp or tax_pur:
        return "TAX_REFUND_OR_OTHER"

    return "OTHER_IN"


def build_bank_in_classes() -> tuple[list[dict], dict]:
    pays = [
        p
        for p in csv.DictReader(open(W1 / "bank_payments.csv", encoding="utf-8"))
        if p.get("direction") == "in" and p.get("is_internal") != "Y"
    ]
    rows = []
    for p in pays:
        cls = classify_in(p)
        # is_internal already filtered; INTERNAL class = LE transfers still marked in file as not internal sometimes
        rows.append(
            {
                "bank_payment_id": p.get("bank_payment_id"),
                "period_month": p.get("period_month"),
                "payment_date": p.get("payment_date"),
                "amount": p.get("amount"),
                "counterparty_raw": p.get("counterparty_raw"),
                "purpose": (p.get("purpose") or "")[:200],
                "source_bank": p.get("source_bank"),
                "legal_entity_id": p.get("legal_entity_id"),
                "in_class": cls,
            }
        )
    stats = {
        "payments": len(rows),
        "by_class_count": dict(Counter(r["in_class"] for r in rows)),
        "by_class_amount": {
            k: round(sum(fnum(r["amount"]) or 0 for r in rows if r["in_class"] == k), 2)
            for k in sorted({r["in_class"] for r in rows})
        },
    }
    return rows, stats


def sales_by_channel_month() -> dict[tuple[str, str], float]:
    out = defaultdict(float)
    for s in csv.DictReader(open(W4 / "sales_lines.csv", encoding="utf-8")):
        if s.get("dq_exclude_from_margin") == "Y":
            continue
        rev = fnum(s.get("revenue_rub")) or 0
        # returns negative stay in channel totals for cash recon (net)
        ch = s.get("channel") or ""
        pm = s.get("period_month") or ""
        out[(ch, pm)] += rev
    return out


def status_gap(sales: float, cash: float) -> str:
    if sales == 0 and cash == 0:
        return "EMPTY"
    if sales == 0 or cash == 0:
        return "OPEN"
    gap = abs(cash - sales) / max(abs(sales), abs(cash))
    if gap <= 0.08:
        return "CLOSE"
    if gap <= 0.25:
        return "SOFT"
    return "OPEN"


def build_channel_recon(bank_rows: list[dict]) -> tuple[list[dict], list[dict], dict]:
    sales = sales_by_channel_month()
    cash = defaultdict(float)
    for r in bank_rows:
        cash[(r["in_class"], r.get("period_month") or "")] += fnum(r["amount"]) or 0

    # IM: ACQ_IM lag0 + lag1 (эквайринг часто с лагом)
    im_months = sorted({m for (ch, m) in sales if ch == "IM" and m})
    im_rows = []
    for m in im_months:
        s = sales.get(("IM", m), 0.0)
        c0 = cash.get(("ACQ_IM", m), 0.0)
        # lag1: cash next month vs this sales
        mi = month_idx(m)
        c1 = 0.0
        if mi:
            for (cls, pm), amt in cash.items():
                if cls == "ACQ_IM" and month_idx(pm) == mi + 1:
                    c1 += amt
        # best of lag0 / blend
        st0 = status_gap(s, c0)
        # also try cash month vs prior sales already covered by iterating
        im_rows.append(
            {
                "channel": "IM",
                "period_month": m,
                "sales_revenue_rub": round(s, 2),
                "bank_in_rub": round(c0, 2),
                "bank_in_next_month_rub": round(c1, 2),
                "gap_rub": round(c0 - s, 2),
                "gap_pct": round((c0 - s) / s * 100, 1) if s else "",
                "status": st0,
                "bank_class": "ACQ_IM",
                "note": "Tinkoff/TBank/VTB acquiring reimbursements",
            }
        )

    tsum_months = sorted({m for (ch, m) in sales if ch == "TSUM" and m})
    tsum_rows = []
    for m in tsum_months:
        s = sales.get(("TSUM", m), 0.0)
        c0 = cash.get(("TSUM_AGENT", m), 0.0)
        tsum_rows.append(
            {
                "channel": "TSUM",
                "period_month": m,
                "sales_revenue_rub": round(s, 2),
                "bank_in_rub": round(c0, 2),
                "bank_in_next_month_rub": "",
                "gap_rub": round(c0 - s, 2),
                "gap_pct": round((c0 - s) / s * 100, 1) if s else "",
                "status": status_gap(s, c0),
                "bank_class": "TSUM_AGENT",
                "note": "Агентские возмещения; часто net of commission / partial month",
            }
        )

    all_rows = im_rows + tsum_rows
    stats = {
        "im": dict(Counter(r["status"] for r in im_rows)),
        "tsum": dict(Counter(r["status"] for r in tsum_rows)),
        "im_sales_total": round(sum(r["sales_revenue_rub"] for r in im_rows), 2),
        "im_bank_total": round(sum(r["bank_in_rub"] for r in im_rows), 2),
        "tsum_sales_total": round(sum(r["sales_revenue_rub"] for r in tsum_rows), 2),
        "tsum_bank_total": round(sum(r["bank_in_rub"] for r in tsum_rows), 2),
        "coverage_im": round(
            sum(r["bank_in_rub"] for r in im_rows) / sum(r["sales_revenue_rub"] for r in im_rows), 3
        )
        if sum(r["sales_revenue_rub"] for r in im_rows)
        else None,
        "coverage_tsum": round(
            sum(r["bank_in_rub"] for r in tsum_rows) / sum(r["sales_revenue_rub"] for r in tsum_rows), 3
        )
        if sum(r["sales_revenue_rub"] for r in tsum_rows)
        else None,
    }
    # monthly class totals mart
    class_month = []
    for (cls, pm), amt in sorted(cash.items()):
        class_month.append({"in_class": cls, "period_month": pm, "amount_rub": round(amt, 2)})
    return all_rows, class_month, stats


def rematch_b2b_settlements(bank_rows: list[dict]) -> tuple[list[dict], dict]:
    settles = list(csv.DictReader(open(W4 / "settlements.csv", encoding="utf-8")))
    bak = W4 / "settlements_pre_h10.csv"
    if not bak.exists():
        shutil.copy2(W4 / "settlements.csv", bak)

    # candidate bank: обычные входящие (не эквайринг/агент/internal)
    bank = [
        r
        for r in bank_rows
        if r["in_class"] in ("OTHER_IN", "B2B_NAMED") and (fnum(r["amount"]) or 0) >= 1000
    ]
    # доп. индекс: оплаты по накладным часто содержат № документа в purpose
    def doc_nums(text: str) -> set[str]:
        u = (text or "").upper().replace("Ё", "Е")
        nums = set(re.findall(r"НАКЛАДН\w*\s*№?\s*(\d{1,6})", u))
        nums |= set(re.findall(r"№\s*(\d{1,6})\b", u))
        # отбросить дни/месяцы и годы
        out = set()
        for n in nums:
            if n in {"2023", "2024", "2025", "2026"}:
                continue
            if len(n) <= 2 and int(n) <= 31:
                continue
            out.add(n)
        return out
    # also need original payment ids from full bank for already-linked exclusion
    used = {s["bank_payment_id"] for s in settles if s.get("bank_payment_id")}

    by_month = defaultdict(list)
    for p in bank:
        if p["bank_payment_id"] in used:
            continue
        by_month[p.get("period_month") or ""].append(p)

    matches = []
    newly = 0
    for st in settles:
        if st.get("bank_payment_id"):
            continue
        if st.get("channel") != "B2B":
            continue
        rev = fnum(st.get("revenue_rub")) or 0
        if rev < 1000 or not st.get("period_month"):
            continue
        mi = month_idx(st["period_month"])
        buyer = st.get("buyer") or ""
        st_docs = doc_nums(st.get("document") or "")
        best = None
        for pm, plist in by_month.items():
            pmi = month_idx(pm)
            if mi is None or pmi is None or abs(pmi - mi) > 2:
                continue
            for p in plist:
                if p["bank_payment_id"] in used:
                    continue
                amt = fnum(p.get("amount")) or 0
                ratio = abs(amt - rev) / max(amt, rev)
                if ratio > 0.015:
                    continue
                ov = name_overlap(buyer, p.get("counterparty_raw") or "")
                # purpose may contain buyer tokens
                if ov < 0.3:
                    bt = set(norm_name(buyer).split())
                    pur = norm_name(p.get("purpose") or "")
                    if bt and any(t in pur for t in bt if len(t) >= 5):
                        ov = max(ov, 0.35)
                # номер накладной в назначении платежа
                p_docs = doc_nums(p.get("purpose") or "")
                doc_hit = bool(st_docs & p_docs)
                if doc_hit:
                    ov = max(ov, 0.45)
                if ov < 0.3:
                    continue
                conf = "HIGH" if (ov >= 0.5 and ratio <= 0.01) or (doc_hit and ov >= 0.45) else (
                    "MED" if ov >= 0.35 else "LOW"
                )
                if conf == "LOW":
                    continue
                cand = (conf, ov, ratio, abs(pmi - mi), p, doc_hit)
                order = {"HIGH": 0, "MED": 1}
                key = (order[cand[0]], -cand[1], cand[2], cand[3])
                if best is None or key < (
                    order[best[0]],
                    -best[1],
                    best[2],
                    best[3],
                ):
                    best = cand
        if not best:
            continue
        conf, ov, ratio, lag, p, doc_hit = best
        used.add(p["bank_payment_id"])
        st["bank_payment_id"] = p["bank_payment_id"]
        st["status"] = f"LINKED_H10_{conf}"
        newly += 1
        matches.append(
            {
                "settlement_id": st["settlement_id"],
                "document": st.get("document", ""),
                "buyer": buyer,
                "revenue_rub": st.get("revenue_rub", ""),
                "period_month": st.get("period_month", ""),
                "bank_payment_id": p["bank_payment_id"],
                "payment_date": p.get("payment_date", ""),
                "bank_amount": p.get("amount", ""),
                "bank_period_month": p.get("period_month", ""),
                "counterparty_raw": p.get("counterparty_raw", ""),
                "name_overlap": round(ov, 3),
                "amount_ratio": round(ratio, 4),
                "month_lag": lag,
                "doc_hit": "Y" if doc_hit else "N",
                "confidence": conf,
                "match_method": "h10_b2b_name_amount_window",
            }
        )

    write_csv(W4 / "settlements.csv", settles, list(settles[0].keys()))
    linked = sum(1 for s in settles if s.get("bank_payment_id"))
    linked_b2b = sum(1 for s in settles if s.get("channel") == "B2B" and s.get("bank_payment_id"))
    open_b2b = sum(1 for s in settles if s.get("channel") == "B2B" and not s.get("bank_payment_id"))
    stats = {
        "newly_linked": newly,
        "linked_total": linked,
        "linked_b2b": linked_b2b,
        "open_b2b": open_b2b,
        "by_confidence": dict(Counter(m["confidence"] for m in matches)),
        "linked_rev_b2b": round(
            sum(fnum(s.get("revenue_rub")) or 0 for s in settles if s.get("channel") == "B2B" and s.get("bank_payment_id")),
            2,
        ),
        "open_rev_b2b": round(
            sum(fnum(s.get("revenue_rub")) or 0 for s in settles if s.get("channel") == "B2B" and not s.get("bank_payment_id")),
            2,
        ),
    }
    return matches, stats


def build_sku_identity_registry() -> tuple[list[dict], dict]:
    """Реестр dual-identity из H9 collision fixes + finance packet."""
    rows = []
    coll_path = H9 / "cost_collisions_fixed.csv"
    if coll_path.exists():
        for r in csv.DictReader(open(coll_path, encoding="utf-8")):
            rows.append(
                {
                    "canonical_sku": r.get("canonical_sku"),
                    "sale_name": r.get("sale_name"),
                    "wrong_cost_name": r.get("bom_name") if r.get("fix_type") == "CATEGORY_MISMATCH" else r.get("bom_name"),
                    "preferred_cost_name": r.get("bom_name"),
                    "preferred_cost_version_id": r.get("cost_version_id"),
                    "fix_type": r.get("fix_type") or "UNIT_RATIO",
                    "channel": r.get("channel"),
                    "sales_line_id": r.get("sales_line_id"),
                    "status": "AUTO_FIXED_H9",
                    "note": "Один артикул связан с разными продуктами в источниках",
                }
            )
    # unique SKU summary
    by_sku = defaultdict(lambda: {"lines": 0, "sale_names": set(), "fix_types": set(), "channels": set()})
    for r in rows:
        s = by_sku[r["canonical_sku"]]
        s["lines"] += 1
        if r.get("sale_name"):
            s["sale_names"].add(r["sale_name"][:80])
        s["fix_types"].add(r.get("fix_type") or "")
        s["channels"].add(r.get("channel") or "")

    summary = []
    for sku, v in sorted(by_sku.items(), key=lambda x: -x[1]["lines"]):
        summary.append(
            {
                "canonical_sku": sku,
                "fixed_lines": v["lines"],
                "channels": ",".join(sorted(v["channels"])),
                "fix_types": "|".join(sorted(x for x in v["fix_types"] if x)),
                "sale_name_sample": next(iter(v["sale_names"]), ""),
                "registry_status": "ACTIVE_ALIAS_RISK",
                "action": "Не смешивать cost/stock без сверки имени",
            }
        )

    # add finance leftovers
    fin = MART / "finance_neg_sku_review.csv"
    if fin.exists():
        for r in csv.DictReader(open(fin, encoding="utf-8")):
            if r.get("canonical_sku") in by_sku:
                continue
            summary.append(
                {
                    "canonical_sku": r.get("canonical_sku"),
                    "fixed_lines": 0,
                    "channels": r.get("channels"),
                    "fix_types": r.get("review_reason"),
                    "sale_name_sample": r.get("name"),
                    "registry_status": "FINANCE_REVIEW",
                    "action": r.get("action"),
                }
            )

    stats = {"collision_line_rows": len(rows), "sku_registry_rows": len(summary)}
    return summary, stats


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)
    MART.mkdir(parents=True, exist_ok=True)

    bank_rows, bank_stats = build_bank_in_classes()
    write_csv(
        OUT / "bank_in_classified.csv",
        bank_rows,
        list(bank_rows[0].keys()) if bank_rows else ["bank_payment_id"],
    )
    write_csv(
        MART / "bank_in_classified.csv",
        bank_rows,
        list(bank_rows[0].keys()) if bank_rows else ["bank_payment_id"],
    )

    recon, class_month, recon_stats = build_channel_recon(bank_rows)
    write_csv(
        MART / "recon_channel_cash_month.csv",
        recon,
        list(recon[0].keys()) if recon else ["channel"],
    )
    write_csv(OUT / "recon_channel_cash_month.csv", recon, list(recon[0].keys()) if recon else ["channel"])
    write_csv(
        MART / "bank_in_by_class_month.csv",
        class_month,
        ["in_class", "period_month", "amount_rub"],
    )

    matches, settle_stats = rematch_b2b_settlements(bank_rows)
    write_csv(
        OUT / "settle_bank_h10_b2b.csv",
        matches,
        list(matches[0].keys()) if matches else ["settlement_id"],
    )
    write_csv(
        W4 / "soft_matches_settle_bank_h10.csv",
        matches,
        list(matches[0].keys()) if matches else ["settlement_id"],
    )
    write_csv(
        MART / "settle_bank_b2b_links.csv",
        matches,
        list(matches[0].keys()) if matches else ["settlement_id"],
    )

    registry, reg_stats = build_sku_identity_registry()
    write_csv(
        MART / "sku_dual_identity_registry.csv",
        registry,
        list(registry[0].keys()) if registry else ["canonical_sku"],
    )
    write_csv(OUT / "sku_dual_identity_registry.csv", registry, list(registry[0].keys()) if registry else ["canonical_sku"])

    summary = {
        "wave": "H10",
        "generated_at": NOW,
        "finding": (
            f"H10: bank IN classified; IM cash coverage {recon_stats.get('coverage_im')}, "
            f"TSUM agent coverage {recon_stats.get('coverage_tsum')}; "
            f"B2B settle +{settle_stats['newly_linked']} "
            f"(linked B2B {settle_stats['linked_b2b']}, open {settle_stats['open_b2b']}); "
            f"SKU identity registry {reg_stats['sku_registry_rows']}."
        ),
        "bank_in": bank_stats,
        "channel_recon": recon_stats,
        "settle_b2b": settle_stats,
        "sku_identity": reg_stats,
        "not_sot": True,
    }
    (OUT / "h10_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    (EV / "h10_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    for name in (
        "bank_in_classified.csv",
        "recon_channel_cash_month.csv",
        "settle_bank_h10_b2b.csv",
        "sku_dual_identity_registry.csv",
        "h10_summary.json",
    ):
        src = OUT / name
        if src.exists():
            shutil.copy2(src, EV / name)

    wb = Workbook()
    ws = wb.active
    ws.title = "H10_Summary"
    ws["A1"] = "H10 Channel Cash"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A2"] = NOW
    ws["A4"] = summary["finding"]
    ws["A6"] = "IM coverage"
    ws["B6"] = recon_stats.get("coverage_im")
    ws["A7"] = "TSUM coverage"
    ws["B7"] = recon_stats.get("coverage_tsum")
    ws["A8"] = "B2B newly linked"
    ws["B8"] = settle_stats["newly_linked"]
    ws["A9"] = "B2B linked / open"
    ws["B9"] = f"{settle_stats['linked_b2b']} / {settle_stats['open_b2b']}"
    ws["A10"] = "SKU registry"
    ws["B10"] = reg_stats["sku_registry_rows"]

    ws2 = wb.create_sheet("BankInClasses")
    ws2.append(["in_class", "amount"])
    for k, v in bank_stats["by_class_amount"].items():
        ws2.append([k, v])
    if matches:
        ws3 = wb.create_sheet("B2B_Links")
        ws3.append(list(matches[0].keys()))
        for r in matches:
            ws3.append(list(r.values()))
    wb.save(OUT / "H10_CHANNEL_CASH.xlsx")
    wb.save(EV / "H10_CHANNEL_CASH.xlsx")

    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
