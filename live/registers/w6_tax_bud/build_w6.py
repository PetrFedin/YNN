#!/usr/bin/env python3
"""
W6: REG-TAX + REG-BUD staging — закрытие налогового и бюджетного контура.

TAX:
- каталог обязательств из 93 файлов (metadata + PDF header);
- cash-налоги из W5 opex (статьи «налог*»);
- сверка tax_cash ↔ bank tax-like.

BUD:
- Бюджет 2025 H2 + Бюджет 2026 (план/факт EUR по статьям/месяцам);
- сверка факт расходов ↔ DDS EUR / W2 payroll EUR-ish.

Не SoT: суммы из налоговых PDF не полностью извлечены (цифры разнесены по клеткам формы).
"""
from __future__ import annotations

import csv
import hashlib
import json
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from pypdf import PdfReader

ROOT = Path(__file__).resolve().parents[3]
DOCS = ROOT / "documents"
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/w6_tax_bud_20260724"
W1 = ROOT / "live/registers/w1_bank_cash"
W2 = ROOT / "live/registers/w2_payroll"
W5 = ROOT / "live/registers/w5_sup_exp_mat"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")


def nfc(s: str) -> str:
    return unicodedata.normalize("NFC", s or "")


def sha16(*parts) -> str:
    raw = "|".join("" if p is None else str(p) for p in parts)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]


def to_float(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def resolve(name: str) -> Path | None:
    t = nfc(name)
    for p in DOCS.iterdir():
        if nfc(p.name) == t:
            return p
    return None


def load_catalog():
    return list(csv.DictReader(open(ROOT / "live/registers/00_SOURCE_CATALOG_93.csv", encoding="utf-8-sig")))


# ── TAX catalog from files + light PDF header ─────────────────────
TAX_TYPE_MAP = {
    "tax_6ndfl": "6NDFL",
    "tax_insurance": "RSV",
    "tax_nds": "NDS",
    "tax_usn": "USN",
    "tax_ens": "ENS",
    "tax_list": "REPORT_LIST",
}


def parse_tax_meta_from_name(name: str) -> dict:
    n = nfc(name)
    entity = "LE-OOO-DEKOR" if "ДЕКОР" in n.upper() or "декор" in n.lower() else (
        "LE-IP-YANINA" if ("Янина" in n or "ЯНИНА" in n) else ""
    )
    year = None
    m = re.search(r"за\s+(\d{4})\s*г", n, re.I)
    if m:
        year = m.group(1)
    else:
        m = re.search(r"(20\d{2})", n)
        year = m.group(1) if m else ""
    quarter = ""
    mq = re.search(r"(\d)\s*квартал", n, re.I)
    if mq:
        quarter = mq.group(1)
    period = f"{year}-Q{quarter}" if year and quarter else (year or "")
    return {"legal_entity_id": entity, "year": year or "", "quarter": quarter, "period": period}


def pdf_header_bits(path: Path) -> dict:
    try:
        r = PdfReader(str(path))
        text = "\n".join((r.pages[i].extract_text() or "") for i in range(min(2, len(r.pages))))
    except Exception as e:
        return {"pdf_ok": "N", "pdf_error": str(e)[:80], "inn": "", "pages": 0, "knd": ""}
    # spaced INN: 7 7 0 7 ...
    inn = ""
    m = re.search(r"ИНН\s*((?:\d\s*){10,12})", text)
    if m:
        inn = re.sub(r"\s+", "", m.group(1))
    knd = ""
    m = re.search(r"КНД\s*(\d{7})", text)
    if m:
        knd = m.group(1)
    year = ""
    m = re.search(r"(?:Календарный год|Отчетный год)\s*((?:\d\s*){4})", text)
    if m:
        year = re.sub(r"\s+", "", m.group(1))
    return {
        "pdf_ok": "Y",
        "pdf_error": "",
        "inn": inn,
        "pages": len(r.pages),
        "knd": knd,
        "pdf_year": year,
    }


def build_tax_obligations(catalog: list[dict]) -> list[dict]:
    out = []
    for row in catalog:
        cat = row.get("category") or ""
        if not cat.startswith("tax"):
            continue
        fname = row["file_name"]
        meta = parse_tax_meta_from_name(fname)
        tax_type = TAX_TYPE_MAP.get(cat, cat)
        path = resolve(fname)
        pdf = pdf_header_bits(path) if path and fname.lower().endswith(".pdf") else {"pdf_ok": "N", "inn": "", "pages": 0, "knd": "", "pdf_year": "", "pdf_error": "missing"}
        # prefer PDF year if present
        year = pdf.get("pdf_year") or meta["year"]
        period = meta["period"]
        if year and meta["quarter"]:
            period = f"{year}-Q{meta['quarter']}"
        elif year:
            period = year
        # entity by INN if known
        legal = meta["legal_entity_id"]
        if pdf.get("inn") == "770701688220":
            legal = "LE-IP-YANINA"
        elif pdf.get("inn") and pdf.get("inn") != "770701688220" and "ДЕКОР" in fname.upper():
            legal = "LE-OOO-DEKOR"
        oid = "OB-" + sha16(row["master_file_id"], tax_type, period, legal)
        out.append(
            {
                "obligation_id": oid,
                "tax_type": tax_type,
                "legal_entity_id": legal,
                "inn": pdf.get("inn") or "",
                "period": period,
                "year": year or "",
                "quarter": meta["quarter"],
                "knd": pdf.get("knd") or "",
                "amount": "",  # full form extract deferred
                "amount_status": "NOT_EXTRACTED_FROM_PDF",
                "ens_allocation_id": "",
                "bank_payment_id": "",
                "source_file_id": row["master_file_id"],
                "source_file_name": fname,
                "pdf_pages": pdf.get("pages") or "",
                "pdf_ok": pdf.get("pdf_ok") or "N",
                "status": "DECLARED_FILE",
            }
        )
    return out


def tax_cash_from_w5() -> list[dict]:
    """Налоговый cash из W5 opex: налоги/ЕНС/взносы — без коммерческого страхования."""
    p = W5 / "expense_opex_only.csv"
    if not p.exists():
        p = W5 / "expense_lines.csv"
    if not p.exists():
        return []
    # коммерческое страхование ≠ налоговые/соцвзносы
    skip_art = re.compile(r"страховк[аи]\s+(помещен|а/м|авто)|страхование\s+(а/м|авто|помещен)|дмс\b", re.I)
    keep_art = re.compile(
        r"налог|ндфл|усн|ндс|енс|единый налоговый|нсипз|нс\s*и\s*пз|страховые\s+взнос|страхов\.?\s*взнос",
        re.I,
    )
    out = []
    for r in csv.DictReader(open(p, encoding="utf-8")):
        if r.get("line_class") and r["line_class"] != "EXPENSE":
            continue
        art = r.get("article_name") or ""
        if skip_art.search(art):
            continue
        if not keep_art.search(art):
            # статья «страхование» без уточнения — пропускаем (не однозначный налог)
            if re.search(r"^страхование$", art.strip(), re.I):
                continue
            continue
        out.append(
            {
                "tax_cash_id": "TC-" + sha16(r["expense_line_id"]),
                "period_month": r["period_month"],
                "article_name": r["article_name"],
                "account_bucket": r.get("account_bucket", ""),
                "legal_entity_id": r.get("legal_entity_id", ""),
                "amount_rub": r["amount_rub"],
                "expense_line_id": r["expense_line_id"],
                "source_file_id": r.get("source_file_id", ""),
            }
        )
    return out


def bank_tax_like() -> list[dict]:
    """
    Платежи в бюджет/ФНС/УФК (не коммерческие «в т.ч. НДС» и не контрагент «ИП Казначей»).
    Сигнал: Казначейство России / УФК / ФНС / ОСФР + ЕНС / налоговый платёж.
    """
    p = W1 / "bank_payments.csv"
    if not p.exists():
        return []
    # контрагент бюджета (не ФИО «Казначей …»)
    cp_ok = re.compile(
        r"казначейство\s+россии|уфк\b|фнс\s+россии|\(фнс|осфр|фонд\s+пенсион|социальн",
        re.I,
    )
    purpose_ok = re.compile(
        r"единый\s+налоговый|енс\b|налог|ндфл|усн|ндс|страхов|взнос|пени|штраф",
        re.I,
    )
    out = []
    for r in csv.DictReader(open(p, encoding="utf-8")):
        if r.get("direction") != "out" or r.get("is_internal") == "Y":
            continue
        purpose = r.get("purpose") or ""
        cp = r.get("counterparty_raw") or ""
        if not cp_ok.search(cp):
            continue
        # УФК без налогового назначения (аренда ДГИ и т.п.) отсекаем, если нет tax purpose
        if not purpose_ok.search(purpose) and not re.search(r"фнс|осфр|пенсион|социальн", cp, re.I):
            continue
        out.append(r)
    return out


def recon_tax(tax_cash, bank_tax):
    cash_m = defaultdict(float)
    for r in tax_cash:
        cash_m[r["period_month"]] += float(r["amount_rub"] or 0)
    bank_m = defaultdict(float)
    for r in bank_tax:
        bank_m[r["period_month"]] += float(r["amount"] or 0)
    rows = []
    for m in sorted(set(cash_m) | set(bank_m)):
        c, b = cash_m.get(m, 0), bank_m.get(m, 0)
        status = "N/A"
        delta = ""
        if c and b:
            delta = round(c - b, 2)
            ratio = abs(c - b) / max(c, b)
            if ratio <= 0.05:
                status = "CLOSE"
            elif ratio <= 0.15:
                status = "SOFT_GAP"
            elif ratio <= 0.30:
                status = "WIDE_GAP"
            else:
                status = "GAP"
        elif c and not b:
            status = "CASH_ONLY"
        elif b and not c:
            status = "BANK_ONLY"
        rows.append(
            {
                "period_month": m,
                "tax_cash_rub": round(c, 2),
                "bank_tax_like_rub": round(b, 2),
                "delta": delta,
                "status": status,
                "note": "tax_cash=W5 opex tax articles; bank=Казначейство/УФК/ФНС/ОСФР only",
            }
        )
    return rows


# ── BUDGET ────────────────────────────────────────────────────────
def parse_budget_wide(path: Path, source_file_id: str, version: str) -> list[dict]:
    """Парсит wide plan/fact по месяцам (оба бюджета)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 3:
        return []

    # find month header row (dates) and plan/fact row
    month_row = None
    label_row = None
    for i, r in enumerate(rows[:5]):
        dates = [c for c in (r or []) if isinstance(c, (datetime, date))]
        if len(dates) >= 2:
            month_row = i
        joined = " ".join(str(c) for c in (r or []) if c is not None).lower()
        if "план" in joined and "факт" in joined:
            label_row = i

    if month_row is None or label_row is None:
        return []

    # columns: for each month date at col j, plan=j fact=j+1 OR find pairs from label row
    months = []  # list of (plan_col, fact_col, YYYY-MM)
    mr = rows[month_row]
    lr = rows[label_row]
    for j, c in enumerate(mr):
        if isinstance(c, datetime):
            ym = c.strftime("%Y-%m")
        elif isinstance(c, date):
            ym = c.strftime("%Y-%m")
        else:
            continue
        # expect plan at j, fact at j+1 if labels say so
        plan_col, fact_col = j, j + 1
        if j < len(lr) and str(lr[j] or "").strip().lower() == "факт":
            # rare
            plan_col, fact_col = j - 1, j
        months.append((plan_col, fact_col, ym))

    out = []
    start = max(month_row, label_row) + 1
    section = ""
    for rnum, r in enumerate(rows[start:], start=start + 1):
        if not r:
            continue
        # article name usually col1 (col0 is row number)
        name = None
        if len(r) > 1 and r[1] is not None and not isinstance(r[1], (int, float)):
            name = str(r[1]).strip()
        elif r[0] is not None and not isinstance(r[0], (int, float)):
            name = str(r[0]).strip()
        if not name:
            continue
        if name.upper().startswith("БЮДЖЕТ"):
            continue
        low = name.lower()
        if low in ("поступления", "расходы"):
            section = name
        for plan_col, fact_col, ym in months:
            plan = to_float(r[plan_col]) if len(r) > plan_col else None
            fact = to_float(r[fact_col]) if len(r) > fact_col else None
            if plan is None and fact is None:
                continue
            if (plan or 0) == 0 and (fact or 0) == 0:
                continue
            out.append(
                {
                    "budget_line_id": "BL-" + sha16(source_file_id, version, ym, name, plan, fact),
                    "budget_version": version,
                    "period_month": ym,
                    "section": section,
                    "article_name": name,
                    "plan_eur": round(plan, 2) if plan is not None else "",
                    "fact_eur": round(fact, 2) if fact is not None else "",
                    "var_eur": round((fact or 0) - (plan or 0), 2) if plan is not None and fact is not None else "",
                    "currency": "EUR",
                    "source_file_id": source_file_id,
                    "source_row_id": f"r{rnum}",
                }
            )
    return out


def recon_budget(budget_lines):
    """Факт расходов бюджета vs DDS EUR; ЗП факт vs W2 card/cash approx."""
    # DDS EUR from cash_lines amount_eur ledger B
    dds_eur = defaultdict(float)
    if (W1 / "cash_lines.csv").exists():
        for c in csv.DictReader(open(W1 / "cash_lines.csv", encoding="utf-8")):
            if c.get("ledger") != "B":
                continue
            dds_eur[c["period_month"]] += float(c["amount_eur"] or 0)

    # budget expenses fact (article == Расходы top or sum of expense section children)
    # Use top-level 'Расходы' row if present
    bud_exp = {}
    bud_payroll = defaultdict(float)
    bud_income = {}
    for b in budget_lines:
        art = b["article_name"]
        m = b["period_month"]
        fact = b["fact_eur"]
        if fact == "":
            continue
        fact = float(fact)
        if art == "Расходы":
            bud_exp[m] = fact
        elif art == "Поступления":
            bud_income[m] = fact
        elif "заработн" in art.lower() or art.lower().startswith("зарплат"):
            bud_payroll[m] += fact

    # W2 payroll cash+card in RUB / 100 as rough EUR if rate~100 (project convention in DDS)
    w2 = defaultdict(float)
    if (W2 / "recon_zp_dds_bank.csv").exists():
        for r in csv.DictReader(open(W2 / "recon_zp_dds_bank.csv", encoding="utf-8")):
            # prefer dist totals
            cash = float(r.get("dist_cash_total") or 0) or 0
            card = float(r.get("dist_card_total") or 0) or float(r.get("zp_card_sheet_sum") or 0) or 0
            if cash or card:
                w2[r["period_month"]] = (cash + card) / 100.0
            elif r.get("zp_ctrl_rub"):
                w2[r["period_month"]] = float(r["zp_ctrl_rub"]) / 100.0

    months = sorted(set(bud_exp) | set(dds_eur) | set(bud_payroll) | set(w2) | set(bud_income))
    rows = []
    for m in months:
        be = bud_exp.get(m)
        de = dds_eur.get(m, 0)
        status = "N/A"
        delta = ""
        if be is not None and de:
            delta = round(be - de, 2)
            ratio = abs(be - de) / max(abs(be), abs(de))
            if ratio <= 0.10:
                status = "CLOSE"
            elif ratio <= 0.20:
                status = "SOFT_GAP"
            else:
                status = "GAP"
        elif be is not None and not de:
            status = "BUD_ONLY"
        elif de and be is None:
            status = "DDS_ONLY"

        bp = bud_payroll.get(m, 0)
        wp = w2.get(m, 0)
        st_p = "N/A"
        d_p = ""
        if bp and wp:
            d_p = round(bp - wp, 2)
            ratio = abs(bp - wp) / max(bp, wp)
            if ratio <= 0.10:
                st_p = "CLOSE"
            elif ratio <= 0.20:
                st_p = "SOFT_GAP"
            else:
                st_p = "GAP"

        rows.append(
            {
                "period_month": m,
                "budget_income_fact_eur": round(bud_income[m], 2) if m in bud_income else "",
                "budget_expense_fact_eur": round(be, 2) if be is not None else "",
                "dds_b_eur": round(de, 2),
                "delta_bud_exp_vs_dds": delta,
                "status_bud_exp_vs_dds": status,
                "budget_payroll_fact_eur": round(bp, 2) if bp else "",
                "w2_payroll_eur_approx": round(wp, 2) if wp else "",
                "delta_payroll": d_p,
                "status_payroll": st_p,
                "note": "budget EUR; W2/100 ≈ EUR at rate 100 used in DDS files",
            }
        )
    return rows


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)
    catalog = load_catalog()

    obligations = build_tax_obligations(catalog)
    tax_cash = tax_cash_from_w5()
    bank_tax = bank_tax_like()
    tax_recon = recon_tax(tax_cash, bank_tax)

    budget = []
    p = resolve("Бюджет 2026.xlsx")
    if p:
        budget.extend(parse_budget_wide(p, "FILE-012", "BUD-2026"))
    p = resolve("Бюджет 2-е полугодие 2025(1).xlsx")
    if p:
        budget.extend(parse_budget_wide(p, "FILE-011", "BUD-2025H2"))
    bud_recon = recon_budget(budget)

    # DEKOR INN capture if any
    dekor_inn = sorted({o["inn"] for o in obligations if o["legal_entity_id"] == "LE-OOO-DEKOR" and o["inn"]})

    write_csv(
        OUT / "tax_obligations.csv",
        obligations,
        [
            "obligation_id", "tax_type", "legal_entity_id", "inn", "period", "year", "quarter",
            "knd", "amount", "amount_status", "ens_allocation_id", "bank_payment_id",
            "source_file_id", "source_file_name", "pdf_pages", "pdf_ok", "status",
        ],
    )
    write_csv(
        OUT / "tax_cash_lines.csv",
        tax_cash,
        ["tax_cash_id", "period_month", "article_name", "account_bucket", "legal_entity_id", "amount_rub", "expense_line_id", "source_file_id"],
    )
    write_csv(
        OUT / "bank_tax_like.csv",
        [
            {
                "bank_payment_id": r["bank_payment_id"],
                "period_month": r["period_month"],
                "payment_date": r["payment_date"],
                "amount": r["amount"],
                "counterparty_raw": r["counterparty_raw"],
                "purpose": (r.get("purpose") or "")[:120],
                "source_file_id": r.get("source_file_id", ""),
            }
            for r in bank_tax
        ],
        ["bank_payment_id", "period_month", "payment_date", "amount", "counterparty_raw", "purpose", "source_file_id"],
    )
    write_csv(OUT / "recon_tax_cash_bank.csv", tax_recon, list(tax_recon[0].keys()) if tax_recon else ["period_month"])
    write_csv(
        OUT / "budget_lines.csv",
        budget,
        ["budget_line_id", "budget_version", "period_month", "section", "article_name", "plan_eur", "fact_eur", "var_eur", "currency", "source_file_id", "source_row_id"],
    )
    write_csv(OUT / "recon_budget.csv", bud_recon, list(bud_recon[0].keys()) if bud_recon else ["period_month"])

    tax_status = Counter(r["status"] for r in tax_recon if r["tax_cash_rub"] or r["bank_tax_like_rub"])
    bud_status = Counter(r["status_bud_exp_vs_dds"] for r in bud_recon if r["budget_expense_fact_eur"] or r["dds_b_eur"])
    pay_status = Counter(r["status_payroll"] for r in bud_recon if r["status_payroll"] != "N/A")

    by_type = Counter(o["tax_type"] for o in obligations)
    summary = {
        "generated_at": NOW,
        "wave": "W6",
        "tax_obligations": len(obligations),
        "tax_by_type": dict(by_type),
        "tax_pdf_ok": sum(1 for o in obligations if o["pdf_ok"] == "Y"),
        "dekor_inn_candidates": dekor_inn,
        "tax_cash_lines": len(tax_cash),
        "bank_tax_like_payments": len(bank_tax),
        "recon_tax_status": dict(tax_status),
        "tax_close_soft": [r["period_month"] for r in tax_recon if r["status"] in ("CLOSE", "SOFT_GAP")],
        "budget_lines": len(budget),
        "budget_versions": sorted({b["budget_version"] for b in budget}),
        "recon_budget_exp_vs_dds": dict(bud_status),
        "recon_budget_payroll": dict(pay_status),
        "budget_exp_close_soft": [r["period_month"] for r in bud_recon if r["status_bud_exp_vs_dds"] in ("CLOSE", "SOFT_GAP")],
        "finding": (
            f"W6: {len(obligations)} tax obligations ({dict(by_type)}); "
            f"tax_cash↔bank: {dict(tax_status)}; "
            f"budget {len(budget)} lines; bud_exp↔DDS: {dict(bud_status)}; "
            f"bud_payroll↔W2: {dict(pay_status)}."
        ),
        "architecture_complete_waves": ["W1", "W2", "W3", "W4", "W5", "W6"],
        "next": "RACI/SoT gate; optional deep PDF amount extract; SUP.inn→bank match",
    }
    json.dump(summary, open(OUT / "w6_summary.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    json.dump(summary, open(EV / "w6_summary.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)

    wb = Workbook()
    ws = wb.active
    ws.title = "00_Summary"
    for i, (k, v) in enumerate(summary.items(), 1):
        ws.cell(i, 1, k)
        ws.cell(i, 2, json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v)
    fill = PatternFill("solid", fgColor="1F4E79")

    def add(name, rows):
        w = wb.create_sheet(name)
        if not rows:
            return
        headers = list(rows[0].keys())
        for c, h in enumerate(headers, 1):
            cell = w.cell(1, c, h)
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
        for ri, row in enumerate(rows[:4000], 2):
            for ci, h in enumerate(headers, 1):
                w.cell(ri, ci, row.get(h, ""))

    add("01_Tax_Obligations", obligations)
    add("02_Tax_Recon", tax_recon)
    add("03_Budget_Recon", bud_recon)
    add("04_Budget_Lines", budget[:2000])
    wb.save(EV / "YANINA_W6_TAX_BUD_EVIDENCE.xlsx")

    (OUT / "README.md").write_text(
        f"""# W6 TAX / BUD

Generated: {NOW}

- `tax_obligations.csv` — {len(obligations)}
- `tax_cash_lines.csv` — {len(tax_cash)}
- `budget_lines.csv` — {len(budget)}
- tax_cash↔bank CLOSE/SOFT: {summary['tax_close_soft']}
- bud_exp↔DDS CLOSE/SOFT: {summary['budget_exp_close_soft']}

Evidence: `../../evidence/w6_tax_bud_20260724/`

Waves W1–W6 complete (Controlled Staging).
""",
        encoding="utf-8",
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
