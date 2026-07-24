#!/usr/bin/env python3
"""
W2: REG-EMP + REG-PAYROLL staging + сверка ZP ↔ DDS ↔ BANK.

Зачем:
- закрыть H2P grain на CLOSE-месяцах денежного скелета;
- provisional employee_id / payroll_line_id / payroll_batch_id;
- контроль: итог ЗП vs статья «оплата труда» в ДДС; карты vs bank out.

Не SoT: без RACI и без стабильного кадрового ID из 1С.
"""
from __future__ import annotations

import csv
import hashlib
import json
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[3]
DOCS = ROOT / "documents"
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/w2_payroll_20260724"
W1 = ROOT / "live/registers/w1_bank_cash"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")

MONTH_RU = {
    "январ": "01",
    "феврал": "02",
    "март": "03",
    "апрел": "04",
    "ма": "05",  # май / май
    "июн": "06",
    "июл": "07",
    "август": "08",
    "сентябр": "09",
    "октябр": "10",
    "ноябр": "11",
    "декабр": "12",
}


def nfc(s: str) -> str:
    return unicodedata.normalize("NFC", s or "")


def sha16(*parts) -> str:
    raw = "|".join("" if p is None else str(p) for p in parts)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]


def to_float(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def parse_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def norm_fio(s: str) -> str:
    s = nfc(str(s or "")).replace("\xa0", " ").strip()
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"\(.*?\)", "", s).strip()
    return s.lower()


def fio_key(s: str) -> str:
    """Ключ сопоставления: фамилия + инициал имени."""
    parts = norm_fio(s).replace(".", " ").split()
    if not parts:
        return ""
    fam = parts[0]
    name_i = parts[1][0] if len(parts) > 1 and parts[1] else ""
    return f"{fam}|{name_i}"


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def load_catalog():
    return {
        r["file_name"]: r
        for r in csv.DictReader(open(ROOT / "live/registers/00_SOURCE_CATALOG_93.csv", encoding="utf-8-sig"))
    }


def resolve_doc(name: str) -> Path | None:
    target = nfc(name)
    for p in DOCS.iterdir():
        if nfc(p.name) == target:
            return p
    # soft: startswith
    for p in DOCS.iterdir():
        if nfc(p.name).startswith(target[:12]):
            return p
    return None


# ── EMP from Штатка ───────────────────────────────────────────────
def build_employees() -> list[dict]:
    path = resolve_doc("Штатка ИП.xlsx")
    if not path:
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    employees = []
    section = "UNKNOWN"
    for r in rows:
        if not r:
            continue
        a0 = str(r[0]).strip() if r[0] is not None else ""
        a1 = str(r[1]).strip() if len(r) > 1 and r[1] is not None else ""
        a2 = str(r[2]).strip() if len(r) > 2 and r[2] is not None else ""
        a3 = str(r[3]).strip() if len(r) > 3 and r[3] is not None else ""
        # section headers often: '', 'АУП и окладники', '32 чел'
        if a1 and not a0 and ("чел" in a2.lower() or a1.isupper() or "оклад" in a1.lower() or "мастер" in a1.lower() or "вышив" in a1.lower() or "конструк" in a1.lower() or "цех" in a1.lower() or "ауп" in a1.lower()):
            section = a1
            continue
        if not a1 or not re.match(r"^\d+$", a0):
            # sometimes number in col0
            if a0.isdigit() and a1:
                pass
            else:
                continue
        fio = a1
        role = a2
        note = a3
        emp_id = "EMP-" + sha16("IP", fio_key(fio), role)
        employees.append(
            {
                "employee_id": emp_id,
                "fio": fio,
                "fio_key": fio_key(fio),
                "role": role,
                "section": section,
                "employer_id": "LE-IP-YANINA",
                "formal_status": "unregistered" if "не оформл" in note.lower() else ("ok" if note == "" else note),
                "notes": note,
                "source_file_id": "FILE-093",
                "valid_from": "2026-01-01",
            }
        )
    return employees


# ── Monthly ZP month from filename ────────────────────────────────
def month_from_zp_name(name: str) -> str | None:
    n = nfc(name).lower()
    m = re.search(r"(\d{4})", n)
    year = m.group(1) if m else None
    mon = None
    for k, v in MONTH_RU.items():
        if k in n:
            mon = v
            break
    if year and mon:
        return f"{year}-{mon}"
    return None


def find_header_row(rows, must_have_substrings):
    for i, r in enumerate(rows[:30]):
        joined = " | ".join(str(c).lower() for c in r if c is not None)
        if all(s in joined for s in must_have_substrings):
            return i, [str(c).strip() if c is not None else "" for c in r]
    return None, None


def col_idx(header, *cands):
    low = [h.lower() for h in header]
    for c in cands:
        for i, h in enumerate(low):
            if c in h:
                return i
    return None


def parse_person_sheet(ws, sheet_name: str, source_file_id: str, accrual_month: str, group: str) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    # detect month from first cells
    for r in rows[:5]:
        for c in r[:3]:
            d = parse_date(c)
            if d:
                accrual_month = d.strftime("%Y-%m")
                break

    # точный заголовок (не путать «Конструктора» title с колонкой «Конструктор»)
    hdr_i, header = None, None
    if sheet_name == "окладники":
        hdr_i, header = find_header_row(rows, ["направление", "оклад"])
    elif sheet_name == "вышивальщицы":
        hdr_i, header = find_header_row(rows, ["фио"])
    elif "конструктор" in sheet_name:
        for i, r in enumerate(rows[:20]):
            if r and str(r[0] or "").strip().lower() == "конструктор":
                hdr_i = i
                header = [str(c).strip() if c is not None else "" for c in r]
                break
    elif "мастер" in sheet_name:
        for i, r in enumerate(rows[:20]):
            if r and str(r[0] or "").strip().lower() in ("мастера", "мастер"):
                # нужен ряд с «итого к выплате» / «карта»
                joined = " ".join(str(c).lower() for c in r if c is not None)
                if "оклад" in joined and ("итого" in joined or "карта" in joined):
                    hdr_i = i
                    header = [str(c).strip() if c is not None else "" for c in r]
                    break

    if hdr_i is None:
        return []

    # map columns by header text
    i_fio = 0
    # for окладники fio is col0 without header name
    if sheet_name == "окладники":
        i_dir = col_idx(header, "направ")
        i_oklad = col_idx(header, "оклад по табелю") or col_idx(header, "оклад")
        i_total = col_idx(header, "итого")
        # prefer last meaningful
        i_card = col_idx(header, "карта")
        i_hand = col_idx(header, "итог")
        # find "Итого на руки по факту" specifically
        for i, h in enumerate(header):
            hl = h.lower()
            if "по факту" in hl:
                i_fact = i
                break
        else:
            i_fact = col_idx(header, "на руки")
        i_prem = col_idx(header, "премия юл")
        start = hdr_i + 1
    elif sheet_name == "вышивальщицы":
        # header may be 2 rows
        i_fio = col_idx(header, "фио") or 0
        i_oklad = None
        i_card = col_idx(header, "карта")
        i_fact = None
        for i, h in enumerate(header):
            if "по факту" in h.lower() or h.lower() == "на руки по факту":
                i_fact = i
            if h.lower() == "итог" or "итог" in h.lower() and "выслуг" in h.lower():
                pass
        i_hand = col_idx(header, "итог")
        # "Итого на руки"
        for i, h in enumerate(header):
            if "на руки" in h.lower() and "факт" not in h.lower():
                i_hand = i
            if "по факту" in h.lower():
                i_fact = i
        i_prem = col_idx(header, "премия")
        i_dir = None
        i_total = col_idx(header, "итого, с выслугой") or col_idx(header, "итого")
        start = hdr_i + 1
        # skip subheader row if empty fio
        if start < len(rows) and (rows[start][i_fio] in (None, "")):
            start += 1
    elif "конструктор" in sheet_name:
        i_fio = col_idx(header, "конструктор") or 0
        i_oklad = col_idx(header, "оклад по табелю") or col_idx(header, "оклад")
        i_total = col_idx(header, "итого к выплате") or col_idx(header, "итого")
        i_card = col_idx(header, "карта")
        i_fact = None
        i_hand = None
        for i, h in enumerate(header):
            if "по факту" in h.lower():
                i_fact = i
            elif "на руки" in h.lower():
                i_hand = i
        i_prem = col_idx(header, "премия")
        i_dir = None
        start = hdr_i + 1
    else:  # мастера
        i_fio = col_idx(header, "мастер") or 0
        i_oklad = col_idx(header, "оклад")
        i_total = col_idx(header, "итого к выплате") or col_idx(header, "итого")
        i_card = col_idx(header, "карта")
        i_fact = None
        i_hand = None
        for i, h in enumerate(header):
            hl = h.lower()
            if "по факту" in hl:
                i_fact = i
            elif "на руки" in hl:
                i_hand = i
        # masters sheet often has more cols after header row 3
        # look at next header-ish
        i_prem = col_idx(header, "премия")
        i_dir = None
        start = hdr_i + 1

    batch_id = "PB-" + sha16(source_file_id, accrual_month, group)
    out = []
    for rnum, r in enumerate(rows[start:], start=start + 1):
        if not r:
            continue
        fio_raw = r[i_fio] if i_fio is not None and len(r) > i_fio else None
        if fio_raw is None:
            continue
        fio = str(fio_raw).strip()
        if not fio or fio.lower().startswith("итого") or fio.lower() in ("х", "x"):
            continue
        if fio.replace(".", "", 1).isdigit():
            continue

        def g(idx):
            if idx is None or len(r) <= idx:
                return None
            return to_float(r[idx])

        gross = g(i_total) or g(i_oklad)
        card = g(i_card) if "i_card" in dir() else g(i_card)
        card = g(i_card)
        net_fact = g(i_fact) if i_fact is not None else None
        net_hand = g(i_hand) if i_hand is not None else None
        net = net_fact if net_fact is not None else net_hand
        prem = g(i_prem)
        direction = ""
        if i_dir is not None and len(r) > i_dir and r[i_dir]:
            direction = str(r[i_dir]).strip()

        # skip empty money rows
        if all(v is None or v == 0 for v in (gross, card, net, prem)):
            continue

        line_id = "PL-" + sha16(source_file_id, sheet_name, rnum, accrual_month, fio_key(fio))
        emp_id = "EMP-" + sha16("IP", fio_key(fio), "")
        cash_net = None
        if net is not None and card is not None:
            cash_net = max(net, 0)  # "на руки" already net of card in many sheets
        # In окладники: Итого на руки ≈ cash; Карта separate; fact ≈ cash rounded
        out.append(
            {
                "payroll_line_id": line_id,
                "payroll_batch_id": batch_id,
                "employee_id": emp_id,
                "fio": fio,
                "fio_key": fio_key(fio),
                "employer_id": "LE-IP-YANINA",
                "accrual_month": accrual_month,
                "payment_month": accrual_month,  # assumption W2; lag later
                "group": group,
                "direction_code": direction,
                "gross_accrual": round(gross, 2) if gross is not None else "",
                "card_amount": round(card, 2) if card is not None else "",
                "cash_amount": round(net, 2) if net is not None else "",
                "premium_owner": round(prem, 2) if prem is not None else "",
                "source_file_id": source_file_id,
                "source_sheet": sheet_name,
                "source_row_id": f"r{rnum}",
                "match_status": "UNMATCHED",
                "bank_payment_id": "",
                "cash_line_id": "",
            }
        )
    return out


def parse_cards_sheet(ws, source_file_id: str, accrual_month: str) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    hdr_i, header = find_header_row(rows, ["фио"])
    if hdr_i is None:
        return []
    i_fio = col_idx(header, "фио") or 1
    i_total = None
    for i, h in enumerate(header):
        if h.lower() == "итого" or h.lower() == "итог":
            i_total = i
    i_avans = col_idx(header, "аванс")
    i_zp = col_idx(header, "зарплата на карты")
    out = []
    batch_id = "PB-" + sha16(source_file_id, accrual_month, "CARDS")
    for rnum, r in enumerate(rows[hdr_i + 1 :], start=hdr_i + 2):
        if not r or len(r) <= i_fio or r[i_fio] is None:
            continue
        fio = str(r[i_fio]).strip()
        if not fio or fio.lower().startswith("итого"):
            continue
        total = to_float(r[i_total]) if i_total is not None and len(r) > i_total else None
        av = to_float(r[i_avans]) if i_avans is not None and len(r) > i_avans else None
        zp = to_float(r[i_zp]) if i_zp is not None and len(r) > i_zp else None
        if total is None and (av is not None or zp is not None):
            total = (av or 0) + (zp or 0)
        if total is None:
            continue
        out.append(
            {
                "payroll_line_id": "PL-" + sha16(source_file_id, "карты", rnum, accrual_month, fio_key(fio)),
                "payroll_batch_id": batch_id,
                "employee_id": "EMP-" + sha16("IP", fio_key(fio), ""),
                "fio": fio,
                "fio_key": fio_key(fio),
                "employer_id": "LE-IP-YANINA",
                "accrual_month": accrual_month,
                "payment_month": accrual_month,
                "group": "CARDS",
                "direction_code": str(r[2]).strip() if len(r) > 2 and r[2] else "",
                "gross_accrual": "",
                "card_amount": round(total, 2),
                "cash_amount": "",
                "premium_owner": "",
                "source_file_id": source_file_id,
                "source_sheet": "карты",
                "source_row_id": f"r{rnum}",
                "match_status": "UNMATCHED",
                "bank_payment_id": "",
                "cash_line_id": "",
            }
        )
    return out


def parse_distribution_totals(ws, accrual_month: str, source_file_id: str) -> dict:
    """Грубые итоги нал/карты из листа Распределение — для контроля."""
    rows = list(ws.iter_rows(values_only=True))
    cash_total = None
    card_total = None
    # heuristics: look for 'Итого' near cash block and sum of card block
    for r in rows:
        vals = [c for c in r if c is not None]
        joined = " ".join(str(c).lower() for c in vals)
        if "итого" in joined and cash_total is None:
            # last numeric often grand total
            nums = [to_float(c) for c in r if to_float(c) is not None]
            if nums:
                cash_total = max(nums)  # often the grand total cell
        if "карты" in joined and "за" in joined:
            pass
    # better: from april 2026 layout row with Итого and 6011700
    for r in rows:
        labels = [str(c).strip().lower() if c is not None else "" for c in r]
        if any(x == "итого" for x in labels):
            nums = [to_float(c) for c in r if to_float(c) is not None]
            if nums:
                # take rightmost
                cash_total = nums[-1]
                break
    # card totals row after "Карты"
    in_cards = False
    card_sums = []
    for r in rows:
        labels = " ".join(str(c).lower() for c in r if c is not None)
        if "карты за" in labels or labels.strip().startswith("карты"):
            in_cards = True
            continue
        if in_cards:
            if any(str(c).strip().lower() == "итого" for c in r if c is not None):
                nums = [to_float(c) for c in r if to_float(c) is not None]
                if nums:
                    card_total = nums[-1]
                break
            # accumulate group rows
            if any(str(c).strip().upper() in ("ОКЛАДНИКИ", "КОНСТРУКТОРЫ", "МАСТЕРА", "ВЫШИВКА") for c in r if c is not None):
                nums = [to_float(c) for c in r if to_float(c) is not None]
                if nums:
                    card_sums.append(nums[-1])
    if card_total is None and card_sums:
        card_total = sum(card_sums)
    return {
        "accrual_month": accrual_month,
        "source_file_id": source_file_id,
        "dist_cash_total": cash_total,
        "dist_card_total": card_total,
    }


SHEET_GROUP = {
    "окладники": "SALARY",
    "вышивальщицы": "EMBROIDERY",
    "ведомость конструкторы": "DESIGNERS",
    "ведомость мастера": "MASTERS",
}


def build_payroll(catalog: dict) -> tuple[list[dict], list[dict], list[dict]]:
    lines = []
    cards = []
    batches_meta = []
    gaps = []
    for p in sorted(DOCS.iterdir(), key=lambda x: nfc(x.name)):
        name = nfc(p.name)
        if not name.startswith("ЗП за") or not name.endswith(".xlsx"):
            continue
        month = month_from_zp_name(name)
        if not month:
            gaps.append({"gap_id": "GAP-W2-MONTH", "file": name, "issue": "cannot parse month"})
            continue
        meta = catalog.get(name) or catalog.get(p.name) or {}
        # catalog may have NFD май
        if not meta:
            for k, v in catalog.items():
                if nfc(k) == name:
                    meta = v
                    break
        fid = meta.get("master_file_id") or ("FILE-ZP-" + month)
        wb = load_workbook(p, read_only=True, data_only=True)
        for sheet, group in SHEET_GROUP.items():
            if sheet not in wb.sheetnames:
                continue
            lines.extend(parse_person_sheet(wb[sheet], sheet, fid, month, group))
        if "карты" in wb.sheetnames:
            cards.extend(parse_cards_sheet(wb["карты"], fid, month))
        if "Распределение" in wb.sheetnames:
            batches_meta.append(parse_distribution_totals(wb["Распределение"], month, fid))
        wb.close()
    return lines, cards, batches_meta


# ── Reconcile vs DDS / BANK ───────────────────────────────────────
def dds_payroll_by_month() -> dict[str, float]:
    cash = list(csv.DictReader(open(W1 / "cash_lines.csv", encoding="utf-8")))
    out = defaultdict(float)
    for c in cash:
        if c.get("ledger") != "B":
            continue
        art = (c.get("article_name") or "").lower()
        if "оплата труда" not in art:
            continue
        # all cash types for labor (нал+б/нал)
        out[c["period_month"]] += float(c["amount_rub"] or 0)
    return out


def dds_payroll_bn_by_month() -> dict[str, float]:
    cash = list(csv.DictReader(open(W1 / "cash_lines.csv", encoding="utf-8")))
    out = defaultdict(float)
    for c in cash:
        if c.get("ledger") != "B":
            continue
        if "оплата труда" not in (c.get("article_name") or "").lower():
            continue
        if "б/нал" not in (c.get("cash_type") or "").lower():
            continue
        out[c["period_month"]] += float(c["amount_rub"] or 0)
    return out


def bank_card_like_by_month() -> dict[str, float]:
    """Исходящие платежи, похожие на ЗП на карты (эвристика)."""
    pays = list(csv.DictReader(open(W1 / "bank_payments.csv", encoding="utf-8")))
    out = defaultdict(float)
    rx = re.compile(r"заработ|зарплат|\bзп\b|реестр|зачислен|фот|оплат[аы] труда", re.I)
    for p in pays:
        if p.get("direction") != "out":
            continue
        if p.get("is_internal") == "Y":
            continue
        purpose = p.get("purpose") or ""
        if rx.search(purpose):
            out[p["period_month"]] += float(p["amount"] or 0)
    return out


def reconcile(lines, cards, dist_meta):
    dds_all = dds_payroll_by_month()
    dds_bn = dds_payroll_bn_by_month()
    bank_zp = bank_card_like_by_month()

    by_month_cash = defaultdict(float)
    by_month_card_from_people = defaultdict(float)
    by_month_gross = defaultdict(float)
    n_lines = Counter()
    for L in lines:
        m = L["accrual_month"]
        n_lines[m] += 1
        if L["cash_amount"] != "":
            by_month_cash[m] += float(L["cash_amount"])
        if L["card_amount"] != "":
            by_month_card_from_people[m] += float(L["card_amount"])
        if L["gross_accrual"] != "":
            by_month_gross[m] += float(L["gross_accrual"])

    by_month_cards_sheet = defaultdict(float)
    for c in cards:
        by_month_cards_sheet[c["accrual_month"]] += float(c["card_amount"] or 0)

    dist = {d["accrual_month"]: d for d in dist_meta}
    months = sorted(set(n_lines) | set(dds_all) | set(by_month_cards_sheet))
    rows = []
    for m in months:
        zp_cash = by_month_cash[m]
        zp_card_people = by_month_card_from_people[m]
        zp_card_sheet = by_month_cards_sheet[m]
        zp_total_approx = zp_cash + zp_card_people
        # prefer cards sheet as bank-facing card total
        card_for_bank = zp_card_sheet or zp_card_people
        d_all = dds_all.get(m, 0.0)
        d_bn = dds_bn.get(m, 0.0)
        b_zp = bank_zp.get(m, 0.0)
        dist_cash = (dist.get(m) or {}).get("dist_cash_total")
        dist_card = (dist.get(m) or {}).get("dist_card_total")

        # Primary control: Распределение (нал+карты) vs DDS «оплата труда»
        # Fallback: sum line cash+card (может задвоить/недобрать из‑за структуры листов)
        if dist_cash not in (None, "") or dist_card not in (None, ""):
            zp_ctrl = float(dist_cash or 0) + float(dist_card or 0)
            zp_ctrl_src = "distribution"
        else:
            zp_ctrl = zp_total_approx
            zp_ctrl_src = "lines_cash_plus_card"

        delta = zp_ctrl - d_all if (zp_ctrl or d_all) else None
        status = "N/A"
        if zp_ctrl and d_all:
            tol = max(5000.0, 0.03 * max(zp_ctrl, d_all))
            if abs(delta) <= tol:
                status = "CLOSE"
            elif abs(delta) / max(zp_ctrl, d_all) <= 0.10:
                status = "SOFT_GAP"
            elif abs(delta) / max(zp_ctrl, d_all) <= 0.20:
                status = "WIDE_GAP"
            else:
                status = "GAP"
        elif zp_ctrl and not d_all:
            status = "ZP_ONLY"
        elif d_all and not zp_ctrl:
            status = "DDS_ONLY"

        # card sheet (preferred) vs bank zp-like
        card_bank_status = "N/A"
        card_delta = ""
        if card_for_bank and b_zp:
            card_delta = round(card_for_bank - b_zp, 2)
            if abs(card_for_bank - b_zp) <= max(5000, 0.05 * max(card_for_bank, b_zp)):
                card_bank_status = "CLOSE"
            elif abs(card_for_bank - b_zp) / max(card_for_bank, b_zp) <= 0.15:
                card_bank_status = "SOFT"
            else:
                card_bank_status = "GAP"
        elif card_for_bank and not b_zp:
            card_bank_status = "NO_BANK_HINT"

        rows.append(
            {
                "period_month": m,
                "n_payroll_lines": n_lines[m],
                "zp_ctrl_rub": round(zp_ctrl, 2),
                "zp_ctrl_source": zp_ctrl_src,
                "zp_cash_sum": round(zp_cash, 2),
                "zp_card_people_sum": round(zp_card_people, 2),
                "zp_card_sheet_sum": round(zp_card_sheet, 2),
                "zp_total_cash_plus_card": round(zp_total_approx, 2),
                "zp_gross_sum": round(by_month_gross[m], 2),
                "dist_cash_total": dist_cash if dist_cash is not None else "",
                "dist_card_total": dist_card if dist_card is not None else "",
                "dds_payroll_all_rub": round(d_all, 2),
                "dds_payroll_bn_rub": round(d_bn, 2),
                "delta_zp_vs_dds": round(delta, 2) if delta is not None else "",
                "status_zp_vs_dds": status,
                "bank_zp_like_out": round(b_zp, 2),
                "delta_card_vs_bank_zp": card_delta,
                "status_card_vs_bank": card_bank_status,
            }
        )
    return rows


def soft_match_cards_to_bank(cards: list[dict], month: str, limit: int = 40) -> list[dict]:
    pays = [
        p
        for p in csv.DictReader(open(W1 / "bank_payments.csv", encoding="utf-8"))
        if p.get("period_month") == month and p.get("direction") == "out" and p.get("is_internal") != "Y"
    ]
    by_amt = defaultdict(list)
    for p in pays:
        by_amt[round(float(p["amount"] or 0), 2)].append(p)
    used = set()
    matches = []
    for c in cards:
        if c["accrual_month"] != month:
            continue
        amt = round(float(c["card_amount"] or 0), 2)
        cands = [p for p in by_amt.get(amt, []) if p["bank_payment_id"] not in used]
        if not cands:
            continue
        p = cands[0]
        used.add(p["bank_payment_id"])
        matches.append(
            {
                "period_month": month,
                "payroll_line_id": c["payroll_line_id"],
                "fio": c["fio"],
                "card_amount": amt,
                "bank_payment_id": p["bank_payment_id"],
                "payment_date": p["payment_date"],
                "counterparty_raw": p["counterparty_raw"],
                "purpose": (p.get("purpose") or "")[:80],
                "match_method": "exact_amount_same_month",
                "confidence": "LOW",
            }
        )
        if len(matches) >= limit:
            break
    return matches


def enrich_employees_from_payroll(emps: list[dict], lines: list[dict], cards: list[dict]) -> list[dict]:
    by_key = {e["fio_key"]: e for e in emps}
    for src in list(lines) + list(cards):
        k = src["fio_key"]
        if not k:
            continue
        if k not in by_key:
            emp = {
                "employee_id": src["employee_id"],
                "fio": src["fio"],
                "fio_key": k,
                "role": "",
                "section": "FROM_PAYROLL_ONLY",
                "employer_id": "LE-IP-YANINA",
                "formal_status": "unknown",
                "notes": "seen in payroll, not in штатка snapshot",
                "source_file_id": src["source_file_id"],
                "valid_from": src["accrual_month"] + "-01",
            }
            emps.append(emp)
            by_key[k] = emp
        # relink payroll employee_id to staff id when key matches
        src["employee_id"] = by_key[k]["employee_id"]
    return emps


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)
    catalog = load_catalog()

    emps = build_employees()
    lines, cards, dist_meta = build_payroll(catalog)
    emps = enrich_employees_from_payroll(emps, lines, cards)

    recon = reconcile(lines, cards, dist_meta)
    # pilot: prefer CLOSE zp month among W1 close months, else 2026-04
    w1_close = set()
    try:
        for r in csv.DictReader(open(W1 / "recon_bank_vs_dds_month.csv", encoding="utf-8")):
            if r.get("status") in ("CLOSE", "SOFT_GAP"):
                w1_close.add(r["period_month"])
    except FileNotFoundError:
        pass

    pilot_candidates = [r for r in recon if r["status_zp_vs_dds"] in ("CLOSE", "SOFT_GAP") and r["n_payroll_lines"]]
    if not pilot_candidates:
        pilot_candidates = [r for r in recon if r["n_payroll_lines"]]
    # prefer 2026-04
    pilot = "2026-04"
    if any(r["period_month"] == "2026-04" for r in recon if r["n_payroll_lines"]):
        pilot = "2026-04"
    elif pilot_candidates:
        pilot = sorted(pilot_candidates, key=lambda x: x["period_month"], reverse=True)[0]["period_month"]

    matches = soft_match_cards_to_bank(cards, pilot, limit=60)

    emp_fields = [
        "employee_id", "fio", "fio_key", "role", "section", "employer_id",
        "formal_status", "notes", "source_file_id", "valid_from",
    ]
    pay_fields = [
        "payroll_line_id", "payroll_batch_id", "employee_id", "fio", "fio_key",
        "employer_id", "accrual_month", "payment_month", "group", "direction_code",
        "gross_accrual", "card_amount", "cash_amount", "premium_owner",
        "source_file_id", "source_sheet", "source_row_id", "match_status",
        "bank_payment_id", "cash_line_id",
    ]

    write_csv(OUT / "employees.csv", emps, emp_fields)
    write_csv(OUT / "payroll_lines.csv", lines, pay_fields)
    write_csv(OUT / "payroll_cards.csv", cards, pay_fields)
    write_csv(OUT / "payroll_distribution_meta.csv", dist_meta, ["accrual_month", "source_file_id", "dist_cash_total", "dist_card_total"])
    write_csv(OUT / "recon_zp_dds_bank.csv", recon, list(recon[0].keys()) if recon else ["period_month"])
    write_csv(OUT / "soft_matches_cards_bank.csv", matches, list(matches[0].keys()) if matches else ["period_month"])

    status_c = Counter(r["status_zp_vs_dds"] for r in recon if r["n_payroll_lines"])
    summary = {
        "generated_at": NOW,
        "wave": "W2",
        "employees": len(emps),
        "employees_from_staff": sum(1 for e in emps if e["section"] != "FROM_PAYROLL_ONLY"),
        "payroll_lines": len(lines),
        "payroll_card_rows": len(cards),
        "months_with_zp_files": sorted({L["accrual_month"] for L in lines}),
        "recon_status_counts_zp_vs_dds": dict(status_c),
        "close_soft_months": [r["period_month"] for r in recon if r["status_zp_vs_dds"] in ("CLOSE", "SOFT_GAP") and r["n_payroll_lines"]],
        "pilot_month": pilot,
        "pilot_card_bank_matches": len(matches),
        "pilot_row": next((r for r in recon if r["period_month"] == pilot), None),
        "finding": (
            f"W2: {len(emps)} employees, {len(lines)} payroll lines, {len(cards)} card rows. "
            f"ZP↔DDS: {dict(status_c)}. Pilot {pilot}: card↔bank exact matches {len(matches)} (LOW)."
        ),
        "next": "W3 SKU/COST or harden payroll_batch_id payment_month lag vs bank dates",
    }
    json.dump(summary, open(OUT / "w2_summary.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    json.dump(summary, open(EV / "w2_summary.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)

    # evidence xlsx
    wb = Workbook()
    ws = wb.active
    ws.title = "00_Summary"
    for i, (k, v) in enumerate(summary.items(), 1):
        ws.cell(i, 1, k)
        ws.cell(i, 2, json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v)
    fill = PatternFill("solid", fgColor="1F4E79")

    def add(name, rows):
        w = wb.create_sheet(name)
        if not rows:
            return
        headers = list(rows[0].keys())
        for c, h in enumerate(headers, 1):
            cell = w.cell(1, c, h)
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
        for ri, row in enumerate(rows, 2):
            for ci, h in enumerate(headers, 1):
                w.cell(ri, ci, row.get(h, ""))

    add("01_Recon", recon)
    add("02_Soft_Matches", matches)
    add("03_Dist_Meta", dist_meta)
    wb.save(EV / "YANINA_W2_PAYROLL_EVIDENCE.xlsx")

    (OUT / "README.md").write_text(
        f"""# W2 EMP / PAYROLL

Generated: {NOW}

- `employees.csv` — {len(emps)} (штатка + seen-in-payroll)
- `payroll_lines.csv` — {len(lines)} (окладники/вышивка/конструкторы/мастера)
- `payroll_cards.csv` — {len(cards)} (лист «карты»)
- `recon_zp_dds_bank.csv` — помесячно ZP↔DDS↔bank
- Pilot: **{pilot}**

Evidence: `../../evidence/w2_payroll_20260724/`

Controlled Staging — не SoT.
""",
        encoding="utf-8",
    )

    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
