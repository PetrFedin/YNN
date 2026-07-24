#!/usr/bin/env python3
"""
H26: Временные заглушки Owners на OPEN-доменах.

Запрос пользователя: если ФИО не дано — поставить временные stubs и продолжать.

Правила:
- Явные имена TEMP_STUB_* / «ВРЕМЯНКА …» — не выдавать за реальных людей
- decision = ACCEPT_STUB (не путать с финальным ACCEPT ФИО)
- status = ACCEPTED_STUB
- so_t по stubs = N; domain_owned_stub = Y
- Пересобрать sot_owners, domain gap board, alias master owner hooks
"""
from __future__ import annotations

import csv
import json
import shutil
import subprocess
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[3]
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/h26_temp_stubs_20260724"
MART = ROOT / "live/marts"
PACKET = ROOT / "live/YANINA_OWNER_PACKET_RACI_AND_REQUESTS.xlsx"
CC = ROOT / "live/YANINA_LIVE_CONTROL_CENTER_20260723.xlsx"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")
TODAY = datetime.now().strftime("%Y%m%d")
EVIDENCE = "chat: временные заглушки пока нет ФИО (H26)"

# item_id / role → stub FIO
STUBS = {
    ("SRC-CTRL-01", "Approver"): "ВРЕМЯНКА Payroll Approver (stub)",
    ("SRC-CTRL-02", "Owner"): "ВРЕМЯНКА Costing Owner (stub)",
    ("SRC-CTRL-02", "Approver"): "ВРЕМЯНКА Costing Approver (stub)",
    ("SRC-CTRL-03", "Owner"): "ВРЕМЯНКА B2B Owner (stub)",
    ("SRC-CTRL-03", "Approver"): "ВРЕМЯНКА B2B Approver (stub)",
    ("DOM-PRODUCT", "Owner"): "ВРЕМЯНКА Product Owner (stub)",
    ("DOM-COST", "Owner"): "ВРЕМЯНКА Cost Owner (stub)",
    ("DOM-PROD", "Owner"): "ВРЕМЯНКА Production Owner (stub)",
    ("DOM-B2B", "Owner"): "ВРЕМЯНКА B2B Owner (stub)",
    ("DOM-DATA", "Owner"): "ВРЕМЯНКА Data Steward (stub)",
}

DOMAIN_STUBS = {
    "PRODUCT": "ВРЕМЯНКА Product Owner (stub)",
    "COST": "ВРЕМЯНКА Cost Owner (stub)",
    "PRODUCTION": "ВРЕМЯНКА Production Owner (stub)",
    "B2B": "ВРЕМЯНКА B2B Owner (stub)",
    "DATA_STEWARD": "ВРЕМЯНКА Data Steward (stub)",
}

HDR_FILL = PatternFill("solid", fgColor="833C0C")
HDR_FONT = Font(bold=True, color="FFFFFF")


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def apply_raci_stubs(wb) -> list[dict]:
    ws = wb["RACI"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    applied = []
    for r in range(2, (ws.max_row or 1) + 1):
        item_id = ws.cell(r, idx["item_id"] + 1).value
        role = ws.cell(r, idx["role"] + 1).value
        key = (item_id, role)
        if key not in STUBS:
            continue
        fio = ws.cell(r, idx["named_person_FIO"] + 1).value
        # только если пусто или уже stub
        if fio and "stub" not in str(fio).lower() and "времен" not in str(fio).lower():
            continue
        stub = STUBS[key]
        ws.cell(r, idx["named_person_FIO"] + 1).value = stub
        ws.cell(r, idx["decision_ACCEPT_REJECT"] + 1).value = "ACCEPT_STUB"
        ws.cell(r, idx["approval_date_YYYYMMDD"] + 1).value = TODAY
        ws.cell(r, idx["evidence_link_or_file"] + 1).value = EVIDENCE
        ws.cell(r, idx["status"] + 1).value = "ACCEPTED_STUB"
        if "notes" in idx:
            ws.cell(r, idx["notes"] + 1).value = (
                "TEMPORARY STUB — заменить на реальное ФИО; не SoT person"
            )
        applied.append({"item_id": item_id, "role": role, "stub": stub})
    return applied


def update_sot_owners() -> list[dict]:
    path = MART / "sot_owners.csv"
    rows = list(csv.DictReader(open(path, encoding="utf-8")))
    out = []
    for r in rows:
        domain = r["domain"]
        if domain in DOMAIN_STUBS and (not r.get("owner") or r.get("status") == "OPEN_NEEDS_OWNER"):
            r = dict(r)
            r["owner"] = DOMAIN_STUBS[domain]
            r["status"] = "ACCEPTED_STUB"
            r["source"] = (r.get("source") or "") + "|H26_STUB"
            r["stub"] = "Y"
            r["replace_with_real_fio"] = "Y"
        else:
            r = dict(r)
            r.setdefault("stub", "N" if r.get("status") == "ACCEPTED" else r.get("stub", ""))
            r.setdefault("replace_with_real_fio", "N" if r.get("status") == "ACCEPTED" else "")
        out.append(r)
    fields = list(out[0].keys())
    write_csv(path, out, fields)
    write_csv(OUT / "sot_owners.csv", out, fields)
    return out


def update_alias_master():
    path = MART / "sku_alias_master.csv"
    if not path.exists():
        return 0
    rows = list(csv.DictReader(open(path, encoding="utf-8")))
    n = 0
    for r in rows:
        if r.get("alias_status") == "CONTROLLED_CANDIDATE":
            r["owner_status"] = "ACCEPTED_STUB"
            r["registry_accept"] = "ACCEPT_STUB"
            r["note"] = (
                "H26: Product Owner = temporary stub; candidates controlled under stub; "
                "still NOT auto-applied to sales joins"
            )
            r["updated_at"] = NOW
            n += 1
    write_csv(path, rows, list(rows[0].keys()))
    write_csv(OUT / "sku_alias_master.csv", rows, list(rows[0].keys()))
    return n


def update_open_asks():
    path = MART / "open_domain_asks.csv"
    if not path.exists():
        return
    rows = list(csv.DictReader(open(path, encoding="utf-8")))
    for r in rows:
        domain = r.get("domain")
        if domain in DOMAIN_STUBS:
            r["decision_FIO"] = DOMAIN_STUBS[domain]
            r["decision_ACCEPT"] = "ACCEPT_STUB"
            r["status_now"] = "ACCEPTED_STUB_PENDING_REAL_FIO"
    # ensure status_now in fields
    fields = list(rows[0].keys())
    if "status_now" not in fields:
        fields.append("status_now")
    write_csv(path, rows, fields)


def update_sot_policy():
    path = ROOT / "live/SOT_POLICY.md"
    extra = f"""

---

## H26 — временные заглушки ({NOW})

На OPEN-доменах поставлены **TEMPORARY STUBS** (не реальные ФИО):

| Domain | Stub |
|--------|------|
| PRODUCT | ВРЕМЯНКА Product Owner (stub) |
| COST | ВРЕМЯНКА Cost Owner (stub) |
| PRODUCTION | ВРЕМЯНКА Production Owner (stub) |
| B2B | ВРЕМЯНКА B2B Owner (stub) |
| DATA_STEWARD | ВРЕМЯНКА Data Steward (stub) |

`decision = ACCEPT_STUB`. Заменить на реальные ФИО → обычный `ACCEPT`.
Пока stubs: domain ops может назначать задачи, **полный SoT person-level не заявлен**.
"""
    if path.exists():
        text = path.read_text(encoding="utf-8")
        if "H26 — временные заглушки" not in text:
            path.write_text(text.rstrip() + "\n" + extra, encoding="utf-8")
    (EV / "SOT_POLICY.md").write_text(path.read_text(encoding="utf-8") if path.exists() else extra, encoding="utf-8")


def rebuild_domain_ops():
    """Перезапуск H25 после stubs, чтобы B2B gaps получили stub owner."""
    script = ROOT / "live/registers/h25_domain_ops/build_h25.py"
    subprocess.check_call([sys.executable, str(script)], cwd=str(ROOT))


def update_cc(summary: dict):
    if not CC.exists():
        return
    wb = load_workbook(CC)
    if "H26_Stubs" in wb.sheetnames:
        del wb["H26_Stubs"]
    ws = wb.create_sheet("H26_Stubs", 0)
    ws["A1"] = "H26 Temporary Owner Stubs"
    ws["A1"].font = Font(bold=True, size=14, color="833C0C")
    ws["A2"] = NOW
    ws["A4"] = summary["finding"]
    ws["A5"] = "Stubs applied"
    ws["B5"] = summary["raci_stubs_n"]
    ws["A6"] = "Replace with real FIO"
    ws["B6"] = "YES — required"
    wb.save(CC)


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)
    shutil.copy2(PACKET, EV / "YANINA_OWNER_PACKET_RACI_AND_REQUESTS.pre_h26.xlsx")

    wb = load_workbook(PACKET)
    applied = apply_raci_stubs(wb)
    if "README" in wb.sheetnames:
        wb["README"]["A17"] = (
            f"H26 {NOW}: TEMP owner stubs on OPEN domains (ACCEPT_STUB). "
            "Replace with real FIO ASAP."
        )
        wb["README"]["A17"].font = Font(bold=True, color="833C0C")

    # stub log sheet
    if "H26_STUBS" in wb.sheetnames:
        del wb["H26_STUBS"]
    ws = wb.create_sheet("H26_STUBS", 2)
    ws["A1"] = "Temporary owner stubs (H26)"
    ws["A1"].font = Font(bold=True, size=13, color="833C0C")
    ws["A2"] = f"{NOW} | {EVIDENCE}"
    ws.append([])
    ws.append(["item_id", "role", "stub_fio", "decision", "replace"])
    for c in range(1, 6):
        ws.cell(4, c).fill = HDR_FILL
        ws.cell(4, c).font = HDR_FONT
    for a in applied:
        ws.append([a["item_id"], a["role"], a["stub"], "ACCEPT_STUB", "Y"])
    wb.save(PACKET)
    shutil.copy2(PACKET, EV / "YANINA_OWNER_PACKET_RACI_AND_REQUESTS.xlsx")

    owners = update_sot_owners()
    alias_n = update_alias_master()
    update_open_asks()
    update_sot_policy()
    rebuild_domain_ops()

    # freeze status
    freeze = {
        "wave": "H26",
        "generated_at": NOW,
        "status": "DOMAIN_OWNED_WITH_STUBS",
        "raci_accept": True,
        "stubs": True,
        "so_t": False,
        "note": "Temporary stubs in place; replace with real FIO",
    }
    (MART / "staging_freeze.json").write_text(json.dumps(freeze, ensure_ascii=False, indent=2), encoding="utf-8")

    # recount board
    board = list(csv.DictReader(open(MART / "domain_gap_board.csv", encoding="utf-8")))
    assigned = sum(1 for r in board if r.get("owner_fio"))
    stub_assigned = sum(1 for r in board if "времен" in (r.get("owner_fio") or "").lower() or "stub" in (r.get("owner_fio") or "").lower())

    summary = {
        "wave": "H26",
        "generated_at": NOW,
        "finding": (
            f"H26: {len(applied)} RACI stubs (ACCEPT_STUB); alias candidates under Product stub; "
            f"domain board rebuilt — {assigned}/{len(board)} gaps have owner fio "
            f"({stub_assigned} on stubs). Replace stubs with real FIO."
        ),
        "raci_stubs_n": len(applied),
        "alias_candidates_under_stub": alias_n,
        "gap_rows": len(board),
        "gaps_with_owner": assigned,
        "gaps_on_stubs": stub_assigned,
        "stub_domains": DOMAIN_STUBS,
        "not_sot": True,
        "replace_stubs": True,
    }
    update_cc(summary)
    (OUT / "h26_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    (EV / "h26_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    write_csv(OUT / "raci_stubs_applied.csv", applied, list(applied[0].keys()) if applied else ["item_id"])
    shutil.copy2(OUT / "raci_stubs_applied.csv", EV / "raci_stubs_applied.csv")
    shutil.copy2(MART / "sot_owners.csv", EV / "sot_owners.csv")

    md = [
        "# Temporary Owner Stubs (H26)",
        "",
        f"Updated: {NOW}",
        "",
        "Заглушки **не люди**. Заменить на реальные ФИО.",
        "",
        "| item | role | stub |",
        "|------|------|------|",
    ]
    for a in applied:
        md.append(f"| {a['item_id']} | {a['role']} | {a['stub']} |")
    md.extend(["", f"Gaps with owner after rebuild: {assigned}/{len(board)}", ""])
    (OUT / "TEMP_STUBS.md").write_text("\n".join(md), encoding="utf-8")
    (ROOT / "live/TEMP_STUBS.md").write_text("\n".join(md), encoding="utf-8")
    (EV / "TEMP_STUBS.md").write_text("\n".join(md), encoding="utf-8")

    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
