#!/usr/bin/env python3
"""
H11: TSUM net-rate model + B2B multi-payment settle links.

Зачем:
1) Агентские ЦУМ — это net после комиссии; построчный lag0 даёт coverage ~0.47.
   Парсим период из purpose + калибруем net-rate → ожидаемый cash vs факт.
2) B2B open — крупные ИП платят несколькими платежами; нужен multi-match.

Не SoT. RACI не трогаем.
"""
from __future__ import annotations

import csv
import json
import re
import shutil
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[3]
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/h11_tsum_b2b_20260724"
MART = ROOT / "live/marts"
W4 = ROOT / "live/registers/w4_sales_settle"
H10 = ROOT / "live/registers/h10_channel_cash"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")

PERIOD_RE = re.compile(
    r"ПЕРИОД\s+С\s+(\d{2}\.\d{2}\.\d{4})\s+ПО\s+(\d{2}\.\d{2}\.\d{4})",
    re.I,
)


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def fnum(v):
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(" ", "").replace(",", "."))
    except (TypeError, ValueError):
        return None


def month_idx(pm: str) -> int | None:
    if not pm or "-" not in pm:
        return None
    y, m = pm.split("-")[:2]
    try:
        return int(y) * 12 + int(m)
    except ValueError:
        return None


def norm_name(s: str) -> str:
    s = (s or "").upper().replace("Ё", "Е")
    s = re.sub(r"[\"«»]", " ", s)
    s = re.sub(
        r"\b(ООО|АО|ЗАО|ПАО|ИП|ОБЩЕСТВО|ОГРАНИЧЕННОЙ|ОТВЕТСТВЕННОСТЬЮ|Г|МОСКВА|САНКТ|ПЕТЕРБУРГ)\b",
        " ",
        s,
    )
    s = re.sub(r"Р/С\s*\d+", " ", s)
    s = re.sub(r"[^A-ZА-Я0-9]+", " ", s)
    return " ".join(t for t in s.split() if len(t) >= 3)


def name_overlap(a: str, b: str) -> float:
    ta = set(norm_name(a).split())
    tb = set(norm_name(b).split())
    if not ta or not tb:
        return 0.0
    return len(ta & tb) / len(ta | tb)


def surname(buyer: str) -> str:
    toks = norm_name(buyer).split()
    return toks[0] if toks else ""


def status_gap(a: float, b: float) -> str:
    if a == 0 and b == 0:
        return "EMPTY"
    if a == 0 or b == 0:
        return "OPEN"
    gap = abs(b - a) / max(abs(a), abs(b))
    if gap <= 0.08:
        return "CLOSE"
    if gap <= 0.25:
        return "SOFT"
    return "OPEN"


def allocate_tsum_agent(bank_rows: list[dict]) -> tuple[list[dict], dict[str, float], dict]:
    """Разделяем damage vs sales-agent; аллоцируем по периоду из purpose."""
    detail = []
    alloc = defaultdict(float)
    damage_total = 0.0
    agent_total = 0.0
    parsed = 0

    for p in bank_rows:
        if p.get("in_class") != "TSUM_AGENT":
            continue
        pur = p.get("purpose") or ""
        amt = fnum(p.get("amount")) or 0.0
        if "УЩЕРБ" in pur.upper() or "ПРЕТЕНЗИ" in pur.upper():
            damage_total += amt
            detail.append(
                {
                    **{k: p.get(k) for k in ("bank_payment_id", "period_month", "payment_date", "amount", "purpose")},
                    "sub_class": "TSUM_DAMAGE",
                    "alloc_months": "",
                }
            )
            continue

        agent_total += amt
        m = PERIOD_RE.search(pur)
        if not m:
            pm = p.get("period_month") or ""
            alloc[pm] += amt
            detail.append(
                {
                    **{k: p.get(k) for k in ("bank_payment_id", "period_month", "payment_date", "amount", "purpose")},
                    "sub_class": "TSUM_AGENT_SALES",
                    "alloc_months": pm,
                }
            )
            continue

        parsed += 1
        d1 = datetime.strptime(m.group(1), "%d.%m.%Y")
        d2 = datetime.strptime(m.group(2), "%d.%m.%Y")
        days: dict[str, int] = defaultdict(int)
        d = d1
        while d <= d2:
            days[d.strftime("%Y-%m")] += 1
            d += timedelta(days=1)
        total_days = sum(days.values()) or 1
        parts = []
        for pm, n in sorted(days.items()):
            share = amt * n / total_days
            alloc[pm] += share
            parts.append(f"{pm}:{round(share, 2)}")
        detail.append(
            {
                **{k: p.get(k) for k in ("bank_payment_id", "period_month", "payment_date", "amount", "purpose")},
                "sub_class": "TSUM_AGENT_SALES",
                "alloc_months": "|".join(parts),
            }
        )

    stats = {
        "agent_sales_total": round(agent_total, 2),
        "damage_total": round(damage_total, 2),
        "parsed_period_rows": parsed,
        "alloc_months": len([k for k, v in alloc.items() if v > 0]),
    }
    return detail, dict(alloc), stats


def build_tsum_net_model(alloc: dict[str, float]) -> tuple[list[dict], dict]:
    sales = defaultdict(float)
    for s in csv.DictReader(open(W4 / "sales_lines.csv", encoding="utf-8")):
        if s.get("dq_exclude_from_margin") == "Y":
            continue
        if s.get("channel") != "TSUM":
            continue
        sales[s.get("period_month") or ""] += fnum(s.get("revenue_rub")) or 0.0

    # калибровка net-rate по месяцам где оба >0 и ratio в [0.25, 0.85] (без выбросов)
    rates = []
    for m, s in sales.items():
        a = alloc.get(m, 0.0)
        if s > 0 and a > 0:
            r = a / s
            if 0.25 <= r <= 0.85:
                rates.append(r)
    rates_sorted = sorted(rates)
    net_rate = rates_sorted[len(rates_sorted) // 2] if rates_sorted else 0.47

    rows = []
    for m in sorted(set(sales) | set(alloc)):
        if not m:
            continue
        s = sales.get(m, 0.0)
        a = alloc.get(m, 0.0)
        expected = s * net_rate
        # статус против модели (не против gross sales)
        st_model = status_gap(expected, a) if s else ("OPEN" if a else "EMPTY")
        st_gross = status_gap(s, a) if s else ("OPEN" if a else "EMPTY")
        rows.append(
            {
                "channel": "TSUM",
                "period_month": m,
                "sales_revenue_rub": round(s, 2),
                "agent_allocated_rub": round(a, 2),
                "net_rate_used": round(net_rate, 4),
                "expected_net_cash_rub": round(expected, 2),
                "gap_vs_model_rub": round(a - expected, 2),
                "gap_vs_model_pct": round((a - expected) / expected * 100, 1) if expected else "",
                "status_vs_model": st_model,
                "status_vs_gross": st_gross,
                "implied_net_rate": round(a / s, 4) if s else "",
                "note": "agent cash allocated by purpose period; model=median net-rate",
            }
        )

    stats = {
        "net_rate_median": round(net_rate, 4),
        "calibration_months": len(rates),
        "vs_model": dict(Counter(r["status_vs_model"] for r in rows if r["sales_revenue_rub"] > 0)),
        "vs_gross": dict(Counter(r["status_vs_gross"] for r in rows if r["sales_revenue_rub"] > 0)),
        "sales_total": round(sum(sales.values()), 2),
        "alloc_total": round(sum(alloc.values()), 2),
        "close_soft_model": sum(
            1 for r in rows if r["sales_revenue_rub"] > 0 and r["status_vs_model"] in ("CLOSE", "SOFT")
        ),
        "months_with_sales": sum(1 for r in rows if r["sales_revenue_rub"] > 0),
    }
    return rows, stats


def rematch_b2b_multipay(bank_rows: list[dict]) -> tuple[list[dict], dict]:
    settles = list(csv.DictReader(open(W4 / "settlements.csv", encoding="utf-8")))
    bak = W4 / "settlements_pre_h11.csv"
    if not bak.exists():
        shutil.copy2(W4 / "settlements.csv", bak)

    # ensure extra columns
    extra_cols = ["bank_payment_ids", "linked_amount_rub", "link_coverage"]
    for s in settles:
        for c in extra_cols:
            s.setdefault(c, s.get(c, ""))

    used = set()
    for s in settles:
        if s.get("bank_payment_id"):
            used.add(s["bank_payment_id"])
        for x in (s.get("bank_payment_ids") or "").split("|"):
            if x:
                used.add(x)

    bank = [
        r
        for r in bank_rows
        if r.get("in_class") in ("OTHER_IN", "B2B_NAMED")
        and (fnum(r.get("amount")) or 0) >= 1000
        and r.get("bank_payment_id") not in used
    ]

    matches = []
    newly_full = 0
    newly_partial = 0

    for st in settles:
        if st.get("bank_payment_id") or st.get("bank_payment_ids"):
            continue
        if st.get("channel") != "B2B":
            continue
        rev = fnum(st.get("revenue_rub")) or 0
        if rev < 1000:
            continue
        buyer = st.get("buyer") or ""
        sur = surname(buyer)
        if len(sur) < 4:
            continue
        mi = month_idx(st.get("period_month") or "")

        cands = []
        for p in bank:
            if p["bank_payment_id"] in used:
                continue
            cn = norm_name(p.get("counterparty_raw") or "")
            pn = norm_name(p.get("purpose") or "")
            if sur not in cn and sur not in pn:
                continue
            ov = name_overlap(buyer, p.get("counterparty_raw") or "")
            if ov < 0.25 and sur not in cn:
                continue
            amt = fnum(p.get("amount")) or 0
            pmi = month_idx(p.get("period_month") or "")
            lag = abs(pmi - mi) if mi is not None and pmi is not None else 99
            if lag > 4:
                continue
            cands.append({"p": p, "amt": amt, "ov": ov, "lag": lag, "ratio": abs(amt - rev) / max(amt, rev)})

        if not cands:
            continue

        # 1) exact single
        exact = [c for c in cands if c["ratio"] <= 0.015 and c["ov"] >= 0.25]
        exact.sort(key=lambda x: (x["ratio"], -x["ov"], x["lag"]))
        chosen = []
        kind = None
        if exact:
            chosen = [exact[0]]
            kind = "FULL_EXACT"
        else:
            # 2) greedy multi within ±3% of rev
            pool = [c for c in cands if c["ov"] >= 0.3]
            pool.sort(key=lambda x: (-x["amt"], x["lag"]))
            acc = 0.0
            for c in pool:
                if acc >= rev * 0.98:
                    break
                if acc + c["amt"] <= rev * 1.03:
                    # avoid using same payment twice
                    if c["p"]["bank_payment_id"] in used:
                        continue
                    if any(x["p"]["bank_payment_id"] == c["p"]["bank_payment_id"] for x in chosen):
                        continue
                    chosen.append(c)
                    acc += c["amt"]
            if chosen and abs(acc - rev) / rev <= 0.05:
                kind = "FULL_PARTS"
            elif chosen and acc / rev >= 0.4:
                kind = "PARTIAL"
            else:
                chosen = []
                kind = None

        if not chosen or not kind:
            continue

        ids = [c["p"]["bank_payment_id"] for c in chosen]
        for i in ids:
            used.add(i)
        linked_amt = round(sum(c["amt"] for c in chosen), 2)
        cov = round(linked_amt / rev, 3)

        st["bank_payment_ids"] = "|".join(ids)
        st["linked_amount_rub"] = linked_amt
        st["link_coverage"] = cov
        # primary id = largest payment
        primary = max(chosen, key=lambda x: x["amt"])
        st["bank_payment_id"] = primary["p"]["bank_payment_id"]
        if kind.startswith("FULL"):
            st["status"] = f"LINKED_H11_{kind}"
            newly_full += 1
        else:
            st["status"] = "LINKED_H11_PARTIAL"
            newly_partial += 1

        matches.append(
            {
                "settlement_id": st["settlement_id"],
                "document": st.get("document", ""),
                "buyer": buyer,
                "revenue_rub": rev,
                "period_month": st.get("period_month", ""),
                "bank_payment_ids": "|".join(ids),
                "n_payments": len(ids),
                "linked_amount_rub": linked_amt,
                "link_coverage": cov,
                "match_kind": kind,
                "name_overlap_max": round(max(c["ov"] for c in chosen), 3),
                "counterparty_sample": (primary["p"].get("counterparty_raw") or "")[:80],
            }
        )

    fields = list(settles[0].keys())
    for c in extra_cols:
        if c not in fields:
            fields.append(c)
    write_csv(W4 / "settlements.csv", settles, fields)

    linked_b2b = sum(1 for s in settles if s.get("channel") == "B2B" and s.get("bank_payment_id"))
    open_b2b = sum(1 for s in settles if s.get("channel") == "B2B" and not s.get("bank_payment_id"))
    stats = {
        "newly_full": newly_full,
        "newly_partial": newly_partial,
        "newly_total": newly_full + newly_partial,
        "linked_b2b": linked_b2b,
        "open_b2b": open_b2b,
        "by_kind": dict(Counter(m["match_kind"] for m in matches)),
        "linked_rev_b2b": round(
            sum(fnum(s.get("revenue_rub")) or 0 for s in settles if s.get("channel") == "B2B" and s.get("bank_payment_id")),
            2,
        ),
        "open_rev_b2b": round(
            sum(fnum(s.get("revenue_rub")) or 0 for s in settles if s.get("channel") == "B2B" and not s.get("bank_payment_id")),
            2,
        ),
        "new_linked_rev": round(sum(m["revenue_rub"] for m in matches), 2),
    }
    return matches, stats


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)
    MART.mkdir(parents=True, exist_ok=True)

    bank_rows = list(csv.DictReader(open(H10 / "bank_in_classified.csv", encoding="utf-8")))

    detail, alloc, alloc_stats = allocate_tsum_agent(bank_rows)
    write_csv(
        OUT / "tsum_agent_allocations.csv",
        detail,
        list(detail[0].keys()) if detail else ["bank_payment_id"],
    )

    tsum_rows, tsum_stats = build_tsum_net_model(alloc)
    write_csv(
        MART / "recon_tsum_net_model.csv",
        tsum_rows,
        list(tsum_rows[0].keys()) if tsum_rows else ["period_month"],
    )
    write_csv(OUT / "recon_tsum_net_model.csv", tsum_rows, list(tsum_rows[0].keys()) if tsum_rows else ["period_month"])

    matches, settle_stats = rematch_b2b_multipay(bank_rows)
    write_csv(
        MART / "settle_bank_b2b_multipay.csv",
        matches,
        list(matches[0].keys()) if matches else ["settlement_id"],
    )
    write_csv(OUT / "settle_bank_b2b_multipay.csv", matches, list(matches[0].keys()) if matches else ["settlement_id"])
    write_csv(
        W4 / "soft_matches_settle_bank_h11.csv",
        matches,
        list(matches[0].keys()) if matches else ["settlement_id"],
    )

    summary = {
        "wave": "H11",
        "generated_at": NOW,
        "finding": (
            f"H11: TSUM net-rate {tsum_stats['net_rate_median']} → "
            f"CLOSE/SOFT vs model {tsum_stats['close_soft_model']}/{tsum_stats['months_with_sales']}; "
            f"B2B multi-pay +{settle_stats['newly_total']} "
            f"(full {settle_stats['newly_full']}, partial {settle_stats['newly_partial']}) → "
            f"linked {settle_stats['linked_b2b']}/open {settle_stats['open_b2b']}."
        ),
        "tsum_alloc": alloc_stats,
        "tsum_model": tsum_stats,
        "settle_b2b": settle_stats,
        "not_sot": True,
    }
    (OUT / "h11_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    (EV / "h11_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    for name in (
        "tsum_agent_allocations.csv",
        "recon_tsum_net_model.csv",
        "settle_bank_b2b_multipay.csv",
        "h11_summary.json",
    ):
        src = OUT / name
        if src.exists():
            shutil.copy2(src, EV / name)
    shutil.copy2(MART / "recon_tsum_net_model.csv", EV / "recon_tsum_net_model.csv")

    wb = Workbook()
    ws = wb.active
    ws.title = "H11_Summary"
    ws["A1"] = "H11 TSUM net + B2B multi"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A2"] = NOW
    ws["A4"] = summary["finding"]
    ws["A6"] = "TSUM net-rate"
    ws["B6"] = tsum_stats["net_rate_median"]
    ws["A7"] = "CLOSE/SOFT vs model"
    ws["B7"] = f"{tsum_stats['close_soft_model']}/{tsum_stats['months_with_sales']}"
    ws["A8"] = "B2B newly"
    ws["B8"] = settle_stats["newly_total"]
    ws["A9"] = "B2B linked/open"
    ws["B9"] = f"{settle_stats['linked_b2b']}/{settle_stats['open_b2b']}"
    if matches:
        ws2 = wb.create_sheet("B2B_Multi")
        ws2.append(list(matches[0].keys()))
        for r in matches:
            ws2.append(list(r.values()))
    ws3 = wb.create_sheet("TSUM_Model")
    if tsum_rows:
        ws3.append(list(tsum_rows[0].keys()))
        for r in tsum_rows:
            ws3.append(list(r.values()))
    wb.save(OUT / "H11_TSUM_B2B.xlsx")
    wb.save(EV / "H11_TSUM_B2B.xlsx")

    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
