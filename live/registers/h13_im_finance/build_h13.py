#!/usr/bin/env python3
"""
H13: IM combo recon (IP↔ALL-LE) + quarantine SKU 0-3243 + finance unit economics.

Зачем:
1) Часть IM OPEN закрывается, если учесть эквайринг Декора (тот же канал продаж).
2) 0-3243 sales=свитшот, cost=худи/юбка — COGS в маржу не пускаем до ручной сверки.
3) B2B neg SKU — unit economics для Owner Packet (не SoT).

Не SoT. RACI не трогаем.
"""
from __future__ import annotations

import csv
import json
import shutil
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[3]
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/h13_im_finance_20260724"
MART = ROOT / "live/marts"
W4 = ROOT / "live/registers/w4_sales_settle"
H10 = ROOT / "live/registers/h10_channel_cash"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")

QUARANTINE_SKUS = {
    "0-3243": "Sale=свитшот Be a poem; W3/H5 cost=худи/юбка — identity collision",
}


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def fnum(v):
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(" ", "").replace(",", "."))
    except (TypeError, ValueError):
        return None


def month_idx(pm: str) -> int | None:
    if not pm or "-" not in pm:
        return None
    y, m = pm.split("-")[:2]
    try:
        return int(y) * 12 + int(m)
    except ValueError:
        return None


def status_gap(a: float, b: float) -> str:
    if a == 0 and b == 0:
        return "EMPTY"
    if a == 0 or b == 0:
        return "OPEN"
    gap = abs(b - a) / max(abs(a), abs(b))
    if gap <= 0.08:
        return "CLOSE"
    if gap <= 0.25:
        return "SOFT"
    return "OPEN"


def cash_at(src: dict[str, float], m: str, lag: int) -> float:
    mi = month_idx(m)
    if mi is None:
        return 0.0
    target = mi + lag
    return sum(amt for pm, amt in src.items() if pm and month_idx(pm) == target)


def build_im_combo() -> tuple[list[dict], dict]:
    acq_ip: dict[str, float] = defaultdict(float)
    acq_all: dict[str, float] = defaultdict(float)
    acq_dek: dict[str, float] = defaultdict(float)
    for r in csv.DictReader(open(H10 / "bank_in_classified.csv", encoding="utf-8")):
        if r.get("in_class") != "ACQ_IM":
            continue
        amt = fnum(r.get("amount")) or 0.0
        pm = r.get("period_month") or ""
        acq_all[pm] += amt
        if r.get("legal_entity_id") == "LE-IP-YANINA":
            acq_ip[pm] += amt
        elif r.get("legal_entity_id") == "LE-OOO-DEKOR":
            acq_dek[pm] += amt

    sales: dict[str, float] = defaultdict(float)
    for s in csv.DictReader(open(W4 / "sales_lines.csv", encoding="utf-8")):
        if s.get("dq_exclude_from_margin") == "Y":
            continue
        if s.get("channel") != "IM":
            continue
        sales[s.get("period_month") or ""] += fnum(s.get("revenue_rub")) or 0.0

    rows = []
    for m in sorted(pm for pm in sales if pm):
        s = sales[m]
        ip_opts = [(f"IP_lag{lag:+d}", cash_at(acq_ip, m, lag)) for lag in (-1, 0, 1)]
        all_opts = [(f"ALL_lag{lag:+d}", cash_at(acq_all, m, lag)) for lag in (-1, 0, 1)]
        ip_best = min(ip_opts, key=lambda x: abs(x[1] - s) if x[1] else 1e18)
        all_best = min(all_opts, key=lambda x: abs(x[1] - s) if x[1] else 1e18)
        # combo: лучший по abs gap
        combo = ip_best if abs(ip_best[1] - s) <= abs(all_best[1] - s) else all_best
        st = status_gap(s, combo[1])
        rows.append(
            {
                "channel": "IM",
                "period_month": m,
                "sales_revenue_rub": round(s, 2),
                "bank_ip_best_rub": round(ip_best[1], 2),
                "bank_ip_best_lag": ip_best[0],
                "status_ip": status_gap(s, ip_best[1]),
                "bank_all_best_rub": round(all_best[1], 2),
                "bank_all_best_lag": all_best[0],
                "status_all": status_gap(s, all_best[1]),
                "bank_combo_rub": round(combo[1], 2),
                "combo_source": combo[0],
                "dek_lag0_rub": round(acq_dek.get(m, 0.0), 2),
                "gap_rub": round(combo[1] - s, 2),
                "gap_pct": round((combo[1] - s) / s * 100, 1) if s else "",
                "status": st,
                "note": "combo=best(IP±1, ALL-LE±1); ALL includes DEKOR acquiring",
            }
        )

    stats = {
        "months": len(rows),
        "status_combo": dict(Counter(r["status"] for r in rows)),
        "status_ip": dict(Counter(r["status_ip"] for r in rows)),
        "status_all": dict(Counter(r["status_all"] for r in rows)),
        "close_soft_combo": sum(1 for r in rows if r["status"] in ("CLOSE", "SOFT")),
        "close_soft_ip": sum(1 for r in rows if r["status_ip"] in ("CLOSE", "SOFT")),
        "combo_pref": dict(Counter(r["combo_source"].split("_")[0] for r in rows)),
        "open_months": [r["period_month"] for r in rows if r["status"] == "OPEN"],
    }
    return rows, stats


def quarantine_collision_cogs() -> tuple[list[dict], dict]:
    src = W4 / "sales_lines.csv"
    bak = W4 / "sales_lines_pre_h13.csv"
    if not bak.exists():
        shutil.copy2(src, bak)

    sales = list(csv.DictReader(open(src, encoding="utf-8")))
    touched = []
    n = 0
    for s in sales:
        can = s.get("canonical_sku") or ""
        if can not in QUARANTINE_SKUS:
            continue
        if s.get("dq_exclude_from_margin") == "Y":
            continue
        cogs = fnum(s.get("cogs_rub"))
        if cogs is None:
            continue
        # снять COGS из маржи, сохранить old для аудита
        touched.append(
            {
                "sales_line_id": s.get("sales_line_id"),
                "canonical_sku": can,
                "channel": s.get("channel"),
                "sku_name": (s.get("sku_name") or "")[:100],
                "revenue_rub": s.get("revenue_rub"),
                "old_cogs_rub": s.get("cogs_rub"),
                "old_cogs_source": s.get("cogs_source"),
                "old_margin_rub": s.get("margin_rub"),
                "reason": QUARANTINE_SKUS[can],
            }
        )
        s["cogs_rub"] = ""
        s["cogs_source"] = ""
        s["margin_rub"] = ""
        s["w3_unit_cost"] = ""
        s["w3_cost_version_id"] = ""
        flags = [x for x in (s.get("dq_flags") or "").split("|") if x]
        if "COST_IDENTITY_QUARANTINE" not in flags:
            flags.append("COST_IDENTITY_QUARANTINE")
        s["dq_flags"] = "|".join(flags)
        # не exclude entirely — revenue остаётся в coverage, но без cogs
        n += 1

    write_csv(src, sales, list(sales[0].keys()))
    return touched, {"quarantined_lines": n, "skus": list(QUARANTINE_SKUS)}


def rebuild_margin_marts(sales: list[dict]) -> dict:
    by_ch = defaultdict(lambda: {"revenue": 0.0, "revenue_costed": 0.0, "cogs": 0.0, "lines": 0, "lines_with_cogs": 0})
    by_sku = defaultdict(
        lambda: {"revenue": 0.0, "cogs": 0.0, "costed_rev": 0.0, "qty": 0.0, "lines": 0, "name": "", "channels": set()}
    )
    by_cm = defaultdict(
        lambda: {"revenue": 0.0, "revenue_costed": 0.0, "cogs": 0.0, "qty": 0.0, "lines": 0, "lines_with_cogs": 0}
    )

    for s in sales:
        if s.get("dq_exclude_from_margin") == "Y":
            continue
        ch = s.get("channel") or ""
        pm = s.get("period_month") or ""
        rev = fnum(s.get("revenue_rub")) or 0.0
        cogs = fnum(s.get("cogs_rub"))
        qty = fnum(s.get("qty")) or 0.0
        by_ch[ch]["revenue"] += rev
        by_ch[ch]["lines"] += 1
        by_cm[(ch, pm)]["revenue"] += rev
        by_cm[(ch, pm)]["qty"] += qty
        by_cm[(ch, pm)]["lines"] += 1
        if cogs is not None:
            by_ch[ch]["cogs"] += cogs
            by_ch[ch]["revenue_costed"] += rev
            by_ch[ch]["lines_with_cogs"] += 1
            by_cm[(ch, pm)]["cogs"] += cogs
            by_cm[(ch, pm)]["revenue_costed"] += rev
            by_cm[(ch, pm)]["lines_with_cogs"] += 1
        can = s.get("canonical_sku") or ""
        if can:
            by_sku[can]["revenue"] += rev
            by_sku[can]["qty"] += qty
            by_sku[can]["lines"] += 1
            if cogs is not None:
                by_sku[can]["cogs"] += cogs
                by_sku[can]["costed_rev"] += rev
            if not by_sku[can]["name"]:
                by_sku[can]["name"] = (s.get("sku_name") or "")[:100]
            by_sku[can]["channels"].add(ch)

    channel_total = []
    for ch, v in sorted(by_ch.items()):
        margin = v["revenue_costed"] - v["cogs"] if v["lines_with_cogs"] else None
        pct = (margin / v["revenue_costed"] * 100) if margin is not None and v["revenue_costed"] else None
        cov = v["lines_with_cogs"] / v["lines"] if v["lines"] else 0
        channel_total.append(
            {
                "channel": ch,
                "lines": v["lines"],
                "cogs_coverage": round(cov, 3),
                "revenue_rub": round(v["revenue"], 2),
                "revenue_costed_rub": round(v["revenue_costed"], 2),
                "cogs_rub": round(v["cogs"], 2),
                "margin_rub": round(margin, 2) if margin is not None else "",
                "margin_pct": round(pct, 1) if pct is not None else "",
            }
        )

    channel_month = []
    for (ch, pm), v in sorted(by_cm.items()):
        margin = v["revenue_costed"] - v["cogs"] if v["lines_with_cogs"] else None
        pct = (margin / v["revenue_costed"] * 100) if margin is not None and v["revenue_costed"] else None
        cov = v["lines_with_cogs"] / v["lines"] if v["lines"] else 0
        channel_month.append(
            {
                "channel": ch,
                "period_month": pm,
                "lines": v["lines"],
                "lines_with_cogs": v["lines_with_cogs"],
                "cogs_coverage": round(cov, 3),
                "qty": round(v["qty"], 2),
                "revenue_rub": round(v["revenue"], 2),
                "revenue_costed_rub": round(v["revenue_costed"], 2),
                "cogs_rub": round(v["cogs"], 2) if v["lines_with_cogs"] else "",
                "margin_rub": round(margin, 2) if margin is not None else "",
                "margin_pct": round(pct, 1) if pct is not None else "",
                "status": "OK" if cov >= 0.85 else ("PARTIAL" if cov >= 0.5 else "WEAK"),
            }
        )

    sku_rows = []
    for can, v in by_sku.items():
        if not v["costed_rev"]:
            continue
        margin = v["costed_rev"] - v["cogs"]
        pct = margin / v["costed_rev"] * 100 if v["costed_rev"] else None
        sku_rows.append(
            {
                "canonical_sku": can,
                "name": v["name"],
                "channels": ",".join(sorted(v["channels"])),
                "lines": v["lines"],
                "qty": round(v["qty"], 2),
                "revenue_rub": round(v["revenue"], 2),
                "cogs_rub": round(v["cogs"], 2),
                "margin_rub": round(margin, 2),
                "margin_pct": round(pct, 1) if pct is not None else "",
            }
        )
    sku_rows.sort(key=lambda x: x["margin_rub"], reverse=True)
    neg = [r for r in sku_rows if r["margin_rub"] < 0]

    write_csv(
        MART / "margin_channel_total.csv",
        channel_total,
        ["channel", "lines", "cogs_coverage", "revenue_rub", "revenue_costed_rub", "cogs_rub", "margin_rub", "margin_pct"],
    )
    write_csv(
        MART / "margin_channel_month.csv",
        channel_month,
        [
            "channel",
            "period_month",
            "lines",
            "lines_with_cogs",
            "cogs_coverage",
            "qty",
            "revenue_rub",
            "revenue_costed_rub",
            "cogs_rub",
            "margin_rub",
            "margin_pct",
            "status",
        ],
    )
    write_csv(
        MART / "margin_sku_top40.csv",
        sku_rows[:40],
        ["canonical_sku", "name", "channels", "lines", "qty", "revenue_rub", "cogs_rub", "margin_rub", "margin_pct"],
    )
    write_csv(
        MART / "margin_sku_bottom40.csv",
        list(reversed(sku_rows[-40:])),
        ["canonical_sku", "name", "channels", "lines", "qty", "revenue_rub", "cogs_rub", "margin_rub", "margin_pct"],
    )
    write_csv(
        MART / "margin_negative_skus.csv",
        [
            {
                "scope": "SKU",
                "channel": r["channels"],
                "period_month": "",
                "canonical_sku": r["canonical_sku"],
                "revenue_rub": r["revenue_rub"],
                "margin_pct": r["margin_pct"],
                "cogs_coverage": "",
                "flags": "NEGATIVE_MARGIN",
                "priority": "HIGH",
                "name": r["name"],
                "margin_rub": r["margin_rub"],
            }
            for r in neg
        ],
        [
            "scope",
            "channel",
            "period_month",
            "canonical_sku",
            "revenue_rub",
            "margin_pct",
            "cogs_coverage",
            "flags",
            "priority",
            "name",
            "margin_rub",
        ],
    )

    tot_rev = sum(r["revenue_costed_rub"] for r in channel_total)
    tot_cogs = sum(r["cogs_rub"] for r in channel_total)
    tot_m = tot_rev - tot_cogs
    return {
        "negative_skus": len(neg),
        "overall_revenue_costed": round(tot_rev, 2),
        "overall_margin": round(tot_m, 2),
        "overall_margin_pct": round(tot_m / tot_rev * 100, 1) if tot_rev else None,
    }


def finance_unit_economics(sales: list[dict]) -> list[dict]:
    """Unit economics для оставшихся neg / wholesale B2B."""
    by = defaultdict(
        lambda: {
            "revenue": 0.0,
            "cogs": 0.0,
            "qty": 0.0,
            "lines": 0,
            "name": "",
            "channels": set(),
            "sources": set(),
            "flags": set(),
        }
    )
    for s in sales:
        if s.get("dq_exclude_from_margin") == "Y":
            continue
        can = s.get("canonical_sku") or ""
        if not can:
            continue
        rev = fnum(s.get("revenue_rub"))
        cogs = fnum(s.get("cogs_rub"))
        qty = fnum(s.get("qty")) or 0
        if rev is None:
            continue
        by[can]["revenue"] += rev
        by[can]["qty"] += qty
        by[can]["lines"] += 1
        by[can]["channels"].add(s.get("channel") or "")
        if cogs is not None:
            by[can]["cogs"] += cogs
            by[can]["sources"].add(s.get("cogs_source") or "")
        if s.get("dq_flags"):
            by[can]["flags"].update(x for x in s["dq_flags"].split("|") if x)
        if not by[can]["name"]:
            by[can]["name"] = (s.get("sku_name") or "")[:120]

    rows = []
    for can, v in by.items():
        if v["cogs"] <= 0 and "COST_IDENTITY_QUARANTINE" not in v["flags"]:
            continue
        margin = v["revenue"] - v["cogs"] if v["cogs"] else None
        if margin is not None and margin >= 0 and "COST_IDENTITY_QUARANTINE" not in v["flags"]:
            continue
        unit_price = v["revenue"] / v["qty"] if v["qty"] else None
        unit_cost = v["cogs"] / v["qty"] if v["qty"] and v["cogs"] else None
        reason = "WHOLESALE_BELOW_STOCK_COST"
        if "COST_IDENTITY_QUARANTINE" in v["flags"]:
            reason = "COST_IDENTITY_QUARANTINE"
            margin = None
        rows.append(
            {
                "canonical_sku": can,
                "name": v["name"],
                "channels": ",".join(sorted(v["channels"])),
                "lines": v["lines"],
                "qty": round(v["qty"], 2),
                "revenue_rub": round(v["revenue"], 2),
                "cogs_rub": round(v["cogs"], 2) if v["cogs"] else "",
                "margin_rub": round(margin, 2) if margin is not None else "",
                "margin_pct": round(margin / v["revenue"] * 100, 1) if margin is not None and v["revenue"] else "",
                "unit_price_rub": round(unit_price, 2) if unit_price else "",
                "unit_cost_rub": round(unit_cost, 2) if unit_cost else "",
                "unit_gap_rub": round(unit_price - unit_cost, 2) if unit_price and unit_cost else "",
                "cogs_sources": "|".join(sorted(x for x in v["sources"] if x)),
                "review_reason": reason,
                "action": "FINANCE_CONFIRM_PRICE_OR_COST"
                if reason.startswith("WHOLESALE")
                else "RESOLVE_SKU_IDENTITY_THEN_RELINK_COST",
            }
        )
    rows.sort(key=lambda x: (0 if x["margin_rub"] == "" else float(x["margin_rub"])))
    return rows


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)
    MART.mkdir(parents=True, exist_ok=True)

    im_rows, im_stats = build_im_combo()
    write_csv(MART / "recon_im_combo.csv", im_rows, list(im_rows[0].keys()) if im_rows else ["period_month"])
    write_csv(OUT / "recon_im_combo.csv", im_rows, list(im_rows[0].keys()) if im_rows else ["period_month"])

    # refresh channel cash mart IM rows from combo; keep TSUM
    old = list(csv.DictReader(open(MART / "recon_channel_cash_month.csv", encoding="utf-8")))
    tsum = [r for r in old if r.get("channel") == "TSUM"]
    merged = []
    for r in im_rows:
        merged.append(
            {
                "channel": "IM",
                "period_month": r["period_month"],
                "sales_revenue_rub": r["sales_revenue_rub"],
                "bank_in_rub": r["bank_combo_rub"],
                "bank_in_all_le_rub": r["bank_all_best_rub"],
                "bank_in_next_month_rub": "",
                "best_bank_in_rub": r["bank_combo_rub"],
                "best_gap_rub": r["gap_rub"],
                "best_status": r["status"],
                "best_lag_months": r["combo_source"],
                "gap_rub": r["gap_rub"],
                "gap_pct": r["gap_pct"],
                "status": r["status"],
                "bank_class": "ACQ_IM_COMBO",
                "note": r["note"],
            }
        )
    for r in tsum:
        merged.append(r)
    fields = [
        "channel",
        "period_month",
        "sales_revenue_rub",
        "bank_in_rub",
        "bank_in_all_le_rub",
        "bank_in_next_month_rub",
        "best_bank_in_rub",
        "best_gap_rub",
        "best_status",
        "best_lag_months",
        "gap_rub",
        "gap_pct",
        "status",
        "bank_class",
        "note",
    ]
    write_csv(MART / "recon_channel_cash_month.csv", merged, fields)

    touched, qstats = quarantine_collision_cogs()
    write_csv(
        OUT / "cogs_quarantine.csv",
        touched,
        list(touched[0].keys()) if touched else ["sales_line_id"],
    )
    write_csv(MART / "cogs_quarantine.csv", touched, list(touched[0].keys()) if touched else ["sales_line_id"])

    sales = list(csv.DictReader(open(W4 / "sales_lines.csv", encoding="utf-8")))
    margin_stats = rebuild_margin_marts(sales)
    fin = finance_unit_economics(sales)
    write_csv(MART / "finance_neg_sku_review.csv", fin, list(fin[0].keys()) if fin else ["canonical_sku"])
    write_csv(OUT / "finance_unit_economics.csv", fin, list(fin[0].keys()) if fin else ["canonical_sku"])
    write_csv(
        MART / "margin_anomalies.csv",
        [
            {
                "scope": "SKU",
                "channel": r["channels"],
                "period_month": "",
                "canonical_sku": r["canonical_sku"],
                "revenue_rub": r["revenue_rub"],
                "margin_pct": r["margin_pct"],
                "cogs_coverage": "",
                "flags": r["review_reason"],
                "priority": "HIGH",
            }
            for r in fin
        ],
        ["scope", "channel", "period_month", "canonical_sku", "revenue_rub", "margin_pct", "cogs_coverage", "flags", "priority"],
    )

    summary = {
        "wave": "H13",
        "generated_at": NOW,
        "finding": (
            f"H13: IM combo CLOSE/SOFT {im_stats['close_soft_combo']}/{im_stats['months']} "
            f"(IP-only was {im_stats['close_soft_ip']}); "
            f"quarantine {qstats['quarantined_lines']} COGS lines ({', '.join(qstats['skus'])}); "
            f"neg SKUs now {margin_stats['negative_skus']}; margin {margin_stats['overall_margin_pct']}%."
        ),
        "im_combo": im_stats,
        "quarantine": qstats,
        "margin": margin_stats,
        "finance_rows": len(fin),
        "not_sot": True,
    }
    (OUT / "h13_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    (EV / "h13_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    for name in (
        "recon_im_combo.csv",
        "cogs_quarantine.csv",
        "finance_unit_economics.csv",
        "h13_summary.json",
    ):
        src = OUT / name
        if src.exists():
            shutil.copy2(src, EV / name)

    wb = Workbook()
    ws = wb.active
    ws.title = "H13_Summary"
    ws["A1"] = "H13 IM combo + finance"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A2"] = NOW
    ws["A4"] = summary["finding"]
    ws["A6"] = "IM CLOSE/SOFT combo"
    ws["B6"] = f"{im_stats['close_soft_combo']}/{im_stats['months']}"
    ws["A7"] = "Quarantine lines"
    ws["B7"] = qstats["quarantined_lines"]
    ws["A8"] = "Neg SKUs"
    ws["B8"] = margin_stats["negative_skus"]
    ws["A9"] = "Margin %"
    ws["B9"] = margin_stats["overall_margin_pct"]
    if fin:
        ws2 = wb.create_sheet("FinanceUnits")
        ws2.append(list(fin[0].keys()))
        for r in fin:
            ws2.append(list(r.values()))
    wb.save(OUT / "H13_IM_FINANCE.xlsx")
    wb.save(EV / "H13_IM_FINANCE.xlsx")

    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
