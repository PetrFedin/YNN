#!/usr/bin/env python3
"""
H22: Staging Freeze + integrity certificate.

Зачем (последний автономный шаг ≥9/10 без RACI/файлов):
- Зафиксировать контрольные суммы ключевых артефактов.
- Прогнать инварианты (не «новые сверки», а проверка что контур цел).
- Явный STOP: дальше только Owner ACCEPT / новые данные.

Не меняет регистры продаж/банка. Не трогает RACI.
"""
from __future__ import annotations

import csv
import hashlib
import json
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[3]
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/h22_staging_freeze_20260724"
MART = ROOT / "live/marts"
CC = ROOT / "live/YANINA_LIVE_CONTROL_CENTER_20260723.xlsx"
PACKET = ROOT / "live/YANINA_OWNER_PACKET_RACI_AND_REQUESTS.xlsx"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")

# Ключевые файлы freeze
ARTIFACTS = [
    "STATUS.md",
    "live/RESULTS_AND_NEXT_PLAN.md",
    "live/OWNER_ACTIONS.md",
    "live/RELEASE_GATE.md",
    "live/FINANCE_RECOMMENDATIONS.md",
    "live/H21_APPLY.md",
    "live/YANINA_OWNER_PACKET_RACI_AND_REQUESTS.xlsx",
    "live/YANINA_LIVE_CONTROL_CENTER_20260723.xlsx",
    "live/marts/controls_summary.csv",
    "live/marts/release_gate_month.csv",
    "live/marts/margin_channel_total.csv",
    "live/marts/margin_channel_total_clean.csv",
    "live/marts/margin_exceptions.csv",
    "live/marts/finance_recommendations.csv",
    "live/marts/owner_actions.csv",
    "live/registers/w4_sales_settle/sales_lines.csv",
    "live/registers/w1_bank_cash/bank_payments.csv",
]


def sha256(path: Path) -> str | None:
    if not path.exists():
        return None
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def fnum(v):
    try:
        return float(str(v).replace(" ", "").replace(",", "."))
    except (TypeError, ValueError):
        return None


def run_invariants() -> list[dict]:
    """Инварианты целостности — PASS/FAIL."""
    checks = []

    def add(cid, ok, detail):
        checks.append({"check_id": cid, "status": "PASS" if ok else "FAIL", "detail": detail})

    # 1 RACI empty (expected until owner)
    filled = 0
    if PACKET.exists():
        wb = load_workbook(PACKET, read_only=True, data_only=True)
        ws = wb["RACI"]
        hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        di = hdr.index("decision_ACCEPT_REJECT")
        filled = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if r[di])
        wb.close()
    add("INV-RACI-EMPTY", filled == 0, f"decision filled={filled} (expected 0 until owner)")

    # 2 sales flags for exception SKUs
    sales = list(csv.DictReader(open(ROOT / "live/registers/w4_sales_settle/sales_lines.csv", encoding="utf-8")))
    ex_skus = {"0-2497", "0-2496", "0-2493A", "0-3243"}
    flagged = [s for s in sales if s.get("canonical_sku") in ex_skus and s.get("margin_exception") == "Y"]
    add("INV-H21-FLAGS", len(flagged) >= 8, f"flagged exception lines={len(flagged)}")

    # 3 quarantine 0-3243 has no cogs
    q = [s for s in sales if s.get("canonical_sku") == "0-3243" and s.get("dq_exclude_from_margin") != "Y"]
    q_ok = all(not s.get("cogs_rub") for s in q) if q else False
    add("INV-3243-NO-COGS", q_ok, f"0-3243 open lines={len(q)}, all cogs blank={q_ok}")

    # 4 release gate files
    gate = ROOT / "live/marts/release_gate_month.csv"
    if gate.exists():
        rows = list(csv.DictReader(open(gate, encoding="utf-8")))
        rel = sum(1 for r in rows if r.get("verdict") == "RELEASED")
        blk = sum(1 for r in rows if r.get("verdict") == "BLOCKED")
        add("INV-GATE-COUNTS", rel + blk == len(rows) and len(rows) == 30, f"RELEASED={rel} BLOCKED={blk} n={len(rows)}")
    else:
        add("INV-GATE-COUNTS", False, "missing release_gate_month.csv")

    # 5 controls summary multi greens
    ctrl = {r["control_id"]: r for r in csv.DictReader(open(MART / "controls_summary.csv", encoding="utf-8"))}
    for cid in ("PAYROLL_MULTI", "OPEX_MULTI"):
        pct = fnum(ctrl.get(cid, {}).get("close_soft_pct"))
        add(f"INV-{cid}", pct == 100.0, f"close_soft_pct={pct}")

    # 6 clean margin >= reported (wholesale losses removed)
    rep = list(csv.DictReader(open(MART / "margin_channel_total.csv", encoding="utf-8")))
    cln = list(csv.DictReader(open(MART / "margin_channel_total_clean.csv", encoding="utf-8")))
    rep_m = next(fnum(r["margin_pct"]) for r in rep if r["channel"] == "TOTAL")
    cln_m = next(fnum(r["margin_pct"]) for r in cln if r["channel"] == "TOTAL")
    add("INV-CLEAN-GE-REPORTED", cln_m is not None and rep_m is not None and cln_m >= rep_m - 0.05, f"reported={rep_m} clean={cln_m}")

    # 7 catalog 107
    cat = ROOT / "live/registers/00_SOURCE_CATALOG_107.csv"
    n = sum(1 for _ in open(cat, encoding="utf-8")) - 1 if cat.exists() else 0
    add("INV-CATALOG-107", n == 107, f"catalog_rows={n}")

    # 8 bank payments present
    bp = ROOT / "live/registers/w1_bank_cash/bank_payments.csv"
    bn = sum(1 for _ in open(bp, encoding="utf-8")) - 1 if bp.exists() else 0
    add("INV-BANK-PAYMENTS", bn >= 4900, f"bank_payments={bn}")

    return checks


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)

    manifest = []
    for rel in ARTIFACTS:
        p = ROOT / rel
        digest = sha256(p)
        manifest.append(
            {
                "path": rel,
                "exists": p.exists(),
                "size_bytes": p.stat().st_size if p.exists() else 0,
                "sha256": digest or "",
            }
        )

    checks = run_invariants()
    passed = sum(1 for c in checks if c["status"] == "PASS")
    failed = sum(1 for c in checks if c["status"] == "FAIL")

    freeze = {
        "wave": "H22",
        "generated_at": NOW,
        "freeze_id": f"STAGING_FREEZE_{NOW.replace(' ', '_').replace(':', '')}",
        "status": "FROZEN_AWAITING_OWNER" if failed == 0 else "FROZEN_WITH_INVARIANT_FAILS",
        "so_t": False,
        "raci_accept": False,
        "invariants_pass": passed,
        "invariants_fail": failed,
        "stop_policy": (
            "No further autonomous hardenings (H23+) without Owner Packet decision "
            "or new source files. Repeating 'делай дальше' without input will be refused."
        ),
        "owner_next": [
            "RACI decision_ACCEPT_REJECT",
            "Confirm/reject RECOMMENDATIONS_H20 / H21 provisional flags",
            "Optional: DATA_REQUESTS_NOW files",
        ],
        "finding": (
            f"H22 staging freeze: invariants {passed}/{passed+failed} PASS. "
            f"Status={'FROZEN_AWAITING_OWNER' if failed==0 else 'FROZEN_WITH_INVARIANT_FAILS'}. "
            "Autonomous path closed."
        ),
    }

    # write outputs
    fields_m = list(manifest[0].keys())
    with open(OUT / "freeze_manifest.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields_m)
        w.writeheader()
        w.writerows(manifest)
    with open(MART / "freeze_manifest.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields_m)
        w.writeheader()
        w.writerows(manifest)
    with open(OUT / "freeze_invariants.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["check_id", "status", "detail"])
        w.writeheader()
        w.writerows(checks)
    with open(MART / "freeze_invariants.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["check_id", "status", "detail"])
        w.writeheader()
        w.writerows(checks)

    (OUT / "h22_summary.json").write_text(json.dumps({**freeze, "manifest_n": len(manifest)}, ensure_ascii=False, indent=2), encoding="utf-8")
    (EV / "h22_summary.json").write_text(json.dumps({**freeze, "manifest_n": len(manifest)}, ensure_ascii=False, indent=2), encoding="utf-8")
    (MART / "staging_freeze.json").write_text(json.dumps(freeze, ensure_ascii=False, indent=2), encoding="utf-8")

    md = [
        "# STAGING FREEZE (H22)",
        "",
        f"Updated: {NOW}",
        "",
        f"**Status:** `{freeze['status']}`",
        f"**Invariants:** {passed} PASS / {failed} FAIL",
        "",
        "## STOP",
        "",
        freeze["stop_policy"],
        "",
        "## Owner next",
        "",
    ]
    for x in freeze["owner_next"]:
        md.append(f"1. {x}" if False else f"- {x}")
    md.extend(["", "## Invariants", "", "| Check | Status | Detail |", "|-------|--------|--------|"])
    for c in checks:
        md.append(f"| `{c['check_id']}` | {c['status']} | {c['detail']} |")
    md.extend(
        [
            "",
            "## Manifest",
            f"{len(manifest)} artifacts hashed → `live/marts/freeze_manifest.csv`",
            "",
            "Evidence: `live/evidence/h22_staging_freeze_20260724/`",
            "",
        ]
    )
    text = "\n".join(md)
    (OUT / "STAGING_FREEZE.md").write_text(text, encoding="utf-8")
    (ROOT / "live/STAGING_FREEZE.md").write_text(text, encoding="utf-8")
    (ROOT / "STAGING_FREEZE.md").write_text(text, encoding="utf-8")
    (EV / "STAGING_FREEZE.md").write_text(text, encoding="utf-8")
    shutil.copy2(OUT / "freeze_manifest.csv", EV / "freeze_manifest.csv")
    shutil.copy2(OUT / "freeze_invariants.csv", EV / "freeze_invariants.csv")

    # Live CC
    if CC.exists():
        wb = load_workbook(CC)
        if "H22_Freeze" in wb.sheetnames:
            del wb["H22_Freeze"]
        ws = wb.create_sheet("H22_Freeze", 0)
        ws["A1"] = "H22 Staging Freeze"
        ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
        ws["A2"] = NOW
        ws["A4"] = freeze["finding"]
        ws["A5"] = "Status"
        ws["B5"] = freeze["status"]
        ws["A6"] = "Invariants"
        ws["B6"] = f"{passed}/{passed+failed}"
        ws["A7"] = "Stop"
        ws["B7"] = "No H23+ without owner input"
        ws["A8"] = "Doc"
        ws["B8"] = "live/STAGING_FREEZE.md"
        wb.save(CC)

    print(json.dumps({**freeze, "manifest_n": len(manifest)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
