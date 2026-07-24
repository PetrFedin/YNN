#!/usr/bin/env python3
"""
H19: Sync Owner Packet with H17 actions + H18 release gate.

Зачем:
- Owner Packet всё ещё описывает состояние ~H3/W1 и не ведёт к текущим BLOCKED/P0–P3.
- Синхронизация повышает шанс реального RACI ACCEPT без автозаполнения решений.
- Один файл для владельцев: RACI + актуальные actions + blocked months + data requests.

НЕ заполняет decision_ACCEPT_REJECT. Не SoT.
"""
from __future__ import annotations

import csv
import json
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[3]
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/h19_owner_packet_sync_20260724"
MART = ROOT / "live/marts"
PACKET = ROOT / "live/YANINA_OWNER_PACKET_RACI_AND_REQUESTS.xlsx"
CC = ROOT / "live/YANINA_LIVE_CONTROL_CENTER_20260723.xlsx"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF")
WARN = Font(bold=True, color="C00000")


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def style_header(ws, ncols: int, row: int = 1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def autosize(ws, max_width: int = 48):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = 12
        for cell in col[:40]:
            if cell.value is not None:
                width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def clear_sheet(wb, name: str):
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


def load_rows(path: Path) -> list[dict]:
    if not path.exists():
        return []
    return list(csv.DictReader(open(path, encoding="utf-8")))


def update_readme(ws):
    # preserve structure: overwrite key lines
    ws["A1"] = "YANINA — ПАКЕТ ДЛЯ ВЛАДЕЛЬЦЕВ (обязательный разблокирующий шаг)"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A3"] = (
        "Зачем: без decision_ACCEPT_REJECT по RACI нельзя Controlled Staging → Source of Truth."
    )
    ws["A4"] = (
        "Что заполнить СЕЙЧАС: лист RACI (колонка decision_ACCEPT_REJECT) — это P0. "
        "Затем H17_ACTIONS (finance/SKU) и DATA_REQUESTS_NOW (файлы на BLOCKED)."
    )
    ws["A5"] = "Куда вернуть: сохранить файл → live/evidence/ или прислать в чат."
    ws["A7"] = (
        f"H19 {NOW}: пакет синхронизирован с H17 Owner Actions + H18 Release Gate "
        "(18/30 RELEASED, 12/30 BLOCKED)."
    )
    ws["A7"].font = Font(bold=True)
    ws["A8"] = (
        "RACI draft: Сливяк Галина = Bank/Tax/Payroll Owner; "
        "Мамушкина Елена = Cash Owner. decision_ACCEPT_REJECT пусто — заполняете вы."
    )
    ws["A9"] = "AI НЕ подставляет ACCEPT/REJECT и НЕ выдумывает ФИО."
    ws["A9"].font = WARN
    ws["A10"] = (
        "Уже сделано staging: W1–W6 + H1–H18 (маржа ~53%, payroll/opex multi 30/30, "
        "release gate, owner actions)."
    )
    ws["A11"] = "Смотреть рядом: live/OWNER_ACTIONS.md · live/RELEASE_GATE.md · STATUS.md"
    ws["A12"] = (
        "Новые листы H19: H17_ACTIONS, H18_BLOCKED, DATA_REQUESTS_NOW, DECISION_LOG."
    )


def update_fill_checklist(ws):
    ws.delete_rows(1, ws.max_row)
    ws["A1"] = "Чеклист разблокировки SoT — актуальный (H19)"
    ws["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws["A2"] = f"Обновлено: {NOW} | AI не подставляет ACCEPT"
    headers = ["step", "what", "where", "why", "status_now", "action_for_you"]
    ws.append([])
    ws.append(headers)
    style_header(ws, len(headers), 4)
    rows = [
        (
            "1",
            "ACCEPT/REJECT по RACI draft (Мамушкина/Сливяк + Approver)",
            "RACI → decision_ACCEPT_REJECT",
            "Единственный SoT-гейт (ST24-G01 / A-RACI-01)",
            "ПУСТО",
            "Проставить ACCEPT или REJECT+правка ФИО",
        ),
        (
            "2",
            "Finance exceptions: 0-2493A / 0-2496 / 0-2497",
            "H17_ACTIONS / DECISION_LOG",
            "B2B цена ~10K vs cost 12–13.5K",
            "TODO P1",
            "OK commercial loss / ошибка cost / прайс",
        ),
        (
            "3",
            "SKU identity 0-3243 (свитшот vs худи)",
            "H17_ACTIONS",
            "COGS quarantine на IM",
            "TODO P1",
            "Дать правильный cost version или alias",
        ),
        (
            "4",
            "Данные на BLOCKED months (IM/TSUM/DDS)",
            "H18_BLOCKED + DATA_REQUESTS_NOW",
            "12/30 месяцев BLOCKED в release gate",
            "OPEN",
            "Прислать эквайринг / % ЦУМ / ДДС июнь 2026",
        ),
        (
            "5",
            "B2B open 15 (~2.51M)",
            "DATA_REQUESTS_NOW",
            "Settlements без платежа в текущих выписках",
            "OPEN",
            "Платежи / взаимозачёты / подтвердить отсутствие",
        ),
        (
            "6",
            "Ведомости ЗП на месяцы без lines (2024, 2025-07+)",
            "DATA_REQUESTS_NOW",
            "Person-level payroll SoT",
            "DDS↔bank уже CLOSE",
            "Дослать ведомости или REJECT с причиной",
        ),
        (
            "7",
            "Контакты email/telegram Owners",
            "RACI",
            "Эскалации",
            "ПУСТО",
            "email_or_contact",
        ),
    ]
    for r in rows:
        ws.append(list(r))
    ws.append([])
    ws.append(["Что уже снято staging-ом (не заменяет RACI):"])
    ws.append(
        [
            "W1–W6 + H1–H18: bank 4933, sales 2826, margin ~53%, "
            "PAYROLL/OPEX multi 30/30, release gate 18/30, owner action pack"
        ]
    )
    autosize(ws)


def write_table_sheet(wb, name: str, title: str, rows: list[dict], fields: list[str]):
    ws = clear_sheet(wb, name)
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws["A2"] = f"Обновлено H19 {NOW} | Не SoT | decision колонки заполняете вы"
    ws.append([])
    ws.append(fields)
    style_header(ws, len(fields), 4)
    for r in rows:
        ws.append([r.get(f, "") for f in fields])
    autosize(ws)
    return ws


def build_h17_actions_rows() -> list[dict]:
    src = load_rows(MART / "owner_actions.csv")
    out = []
    for r in src:
        out.append(
            {
                "action_id": r.get("action_id"),
                "priority": r.get("priority"),
                "owner_hint": r.get("owner_hint"),
                "category": r.get("category"),
                "title": r.get("title"),
                "detail": r.get("detail"),
                "blocks": r.get("blocks"),
                "evidence_path": r.get("evidence_path"),
                "status_now": r.get("status") or "TODO",
                "decision_ACCEPT_REJECT_OR_DONE": "",  # owner fills
                "decision_note": "",
                "decision_date_YYYYMMDD": "",
            }
        )
    return out


def build_h18_blocked_rows() -> list[dict]:
    src = load_rows(MART / "release_gate_blocked.csv")
    out = []
    for r in src:
        out.append(
            {
                "period_month": r.get("period_month"),
                "verdict": r.get("verdict"),
                "fail_controls": r.get("fail_controls"),
                "soft_controls": r.get("soft_controls"),
                "owner_action_ids": r.get("owner_action_ids"),
                "note": r.get("note"),
                "data_needed": "",
                "owner_decision": "",
            }
        )
    return out


def build_data_requests_now() -> list[dict]:
    """Конкретные запросы из текущего gate/actions (поверх старых ST24)."""
    return [
        {
            "request_id": "DR-NOW-01",
            "priority": "P0",
            "linked_action": "A-RACI-01",
            "request": "Проставить decision_ACCEPT_REJECT в листе RACI",
            "required_fields": "decision_ACCEPT_REJECT, approval_date_YYYYMMDD, approver_FIO",
            "owner_role": "Юлия / Owners",
            "named_owner_hint": "Мамушкина / Сливяк",
            "blocks": "SoT",
            "status": "OPEN",
            "file_received": "",
            "notes": "Без этого всё остальное staging",
        },
        {
            "request_id": "DR-NOW-02",
            "priority": "P1",
            "linked_action": "A-FIN-0-2497|A-FIN-0-2496|A-FIN-0-2493A",
            "request": "Подтвердить B2B wholesale below cost (3 SKU)",
            "required_fields": "sku, decision(OK_LOSS|FIX_COST|FIX_PRICE), evidence",
            "owner_role": "Финансы / коммерция",
            "named_owner_hint": "",
            "blocks": "Margin SoT B2B",
            "status": "OPEN",
            "file_received": "",
            "notes": "evidence: live/marts/finance_b2b_loss_evidence.csv",
        },
        {
            "request_id": "DR-NOW-03",
            "priority": "P1",
            "linked_action": "A-FIN-0-3243",
            "request": "Разрешить identity 0-3243 (свитшот Be a poem)",
            "required_fields": "canonical_sku, correct_cost_version_id OR alias_map",
            "owner_role": "Производство / 1С",
            "named_owner_hint": "",
            "blocks": "IM COGS quarantine",
            "status": "OPEN",
            "file_received": "",
            "notes": "соседний 0-3244 ещё хуже по unit cost",
        },
        {
            "request_id": "DR-NOW-04",
            "priority": "P2",
            "linked_action": "A-DATA-IM-01",
            "request": "Эквайринг-реестры на IM OPEN: 2024-08, 2025-01/08/10, 2026-03/04",
            "required_fields": "period_month, acquirer, amount, legal_entity, settlement_date",
            "owner_role": "Сливяк / банк",
            "named_owner_hint": "Сливяк Галина",
            "blocks": "IM gate + release months",
            "status": "OPEN",
            "file_received": "",
            "notes": "цель IM CLOSE/SOFT 80%→90%+",
        },
        {
            "request_id": "DR-NOW-05",
            "priority": "P2",
            "linked_action": "A-DATA-B2B-01",
            "request": "Платежи/взаимозачёты на B2B open 15 (~2.51M ₽)",
            "required_fields": "settlement_id, buyer, amount, bank_payment_ref OR offset_act",
            "owner_role": "Сливяк / B2B",
            "named_owner_hint": "",
            "blocks": "B2B settle coverage",
            "status": "OPEN",
            "file_received": "",
            "notes": "live/marts/data_request_b2b_open.csv",
        },
        {
            "request_id": "DR-NOW-06",
            "priority": "P2",
            "linked_action": "A-DATA-ZP-01",
            "request": "Ведомости ЗП на месяцы без payroll_lines (2024 + 2025-07..12 + 2026-06)",
            "required_fields": "period_month, employee, accrual, card/cash split",
            "owner_role": "Сливяк / кадры",
            "named_owner_hint": "Сливяк Галина",
            "blocks": "Payroll person-level",
            "status": "OPEN",
            "file_received": "",
            "notes": "DDS↔bank уже CLOSE; нужны lines",
        },
        {
            "request_id": "DR-NOW-07",
            "priority": "P2",
            "linked_action": "A-TSUM-RATE-01",
            "request": "Агентский % ЦУМ из договора (+ ДДС строки за 2026-06)",
            "required_fields": "rate_pct, period_from, period_to; DDS June rows if available",
            "owner_role": "Юридический / финансы",
            "named_owner_hint": "",
            "blocks": "TSUM model SoT; BANK_DDS 2026-06 BANK_ONLY",
            "status": "OPEN",
            "file_received": "",
            "notes": "в текущем ДДС 2026 нет июня — это дыра источника, не парсера",
        },
        {
            "request_id": "DR-NOW-08",
            "priority": "P3",
            "linked_action": "A-FIN-WATCH-01",
            "request": "Ревью ~10 SKU unit≫BOM (без автофикса)",
            "required_fields": "sku, keep_FILE_cost|switch_BOM|incomplete_BOM",
            "owner_role": "Производство / финансы",
            "named_owner_hint": "",
            "blocks": "качество маржи TSUM/IM",
            "status": "OPEN",
            "file_received": "",
            "notes": "live/marts/cost_identity_review_priority.csv",
        },
    ]


def build_decision_log(actions: list[dict]) -> list[dict]:
    rows = []
    for r in actions:
        rows.append(
            {
                "decision_id": r["action_id"],
                "priority": r["priority"],
                "topic": r["title"],
                "owner_hint": r["owner_hint"],
                "decision_ACCEPT_REJECT_OR_DONE": "",
                "chosen_option": "",
                "comment": "",
                "decided_by_FIO": "",
                "date_YYYYMMDD": "",
            }
        )
    return rows


def update_gap_mitigation(ws):
    """Обновить staging_status под текущую реальность, не трогая gap_id."""
    # rewrite sheet with refreshed statuses
    updates = {
        "ST24-G01": ("не закрыт — ждёт decision_ACCEPT_REJECT", "ФИО + ACCEPT/REJECT в RACI"),
        "ST24-G02": ("частично: Jan/Feb 2026 lines есть; quarantine Excel quality", "архив до #REF по желанию"),
        "ST24-G03": ("частично: W2+H15 multi 30/30; lines не на все месяцы", "ведомости 2024 / 2025-07+"),
        "ST24-G04": ("частично W2", "полный accrual ledger"),
        "ST24-G05": ("сильно закрыт: Alfa+VTB+Sber Salon+card ~4933", "подтвердить полноту / дослать кассу"),
        "ST24-G06": ("сильно закрыт W6 tax cash↔bank ~97%", "договорной % не нужен; ENS allocation опционально"),
        "ST24-G07": ("сильно закрыт W3/H9; alias candidates 16 (H17)", "утвердить alias registry после RACI"),
        "ST24-G08": ("частично W3/H5/H9; quarantine 0-3243", "finance decisions + cost versions"),
        "ST24-G09": ("не закрыт", "PO / batch size"),
        "ST24-G10": ("частично W5", "BOM + выдачи"),
        "ST24-G11": ("частично: 39 linked / 15 open ~2.51M", "платежи/взаимозачёты на open 15"),
        "ST24-G12": ("частично: B2B факт до 2026-05 в файле; Jun пуст", "июнь–дек 2026 по мере появления"),
        "ST24-G13": ("частично: TSUM net-rate 0.4668 эвристика", "агентский % из договора"),
        "ST24-G14": ("частично: H18 release gate готов", "назначить data steward после RACI"),
    }
    # read existing
    headers = [c.value for c in ws[1]]
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        d = dict(zip(headers, row))
        gid = d.get("gap_id")
        if gid in updates:
            d["staging_status"], d["still_needs_from_owner"] = updates[gid]
        data.append(d)
    ws.delete_rows(1, ws.max_row)
    ws.append(headers)
    style_header(ws, len(headers), 1)
    for d in data:
        ws.append([d.get(h) for h in headers])
    autosize(ws)


def append_legacy_note_to_data_requests(ws):
    """Пометить старый лист как legacy, не удаляя историю."""
    ws.insert_rows(1, 2)
    ws["A1"] = (
        f"LEGACY ST24 requests (H3 era). Актуальные запросы — лист DATA_REQUESTS_NOW. "
        f"Обновлено H19 {NOW}."
    )
    ws["A1"].font = Font(italic=True, color="833C0C")


def update_live_cc(summary: dict):
    if not CC.exists():
        return
    wb = load_workbook(CC)
    if "H19_OwnerSync" in wb.sheetnames:
        del wb["H19_OwnerSync"]
    ws = wb.create_sheet("H19_OwnerSync", 0)
    ws["A1"] = "H19 Owner Packet Sync"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A2"] = NOW
    ws["A4"] = summary["finding"]
    ws["A5"] = "Packet"
    ws["B5"] = "live/YANINA_OWNER_PACKET_RACI_AND_REQUESTS.xlsx"
    ws["A6"] = "New sheets"
    ws["B6"] = "H17_ACTIONS | H18_BLOCKED | DATA_REQUESTS_NOW | DECISION_LOG"
    ws["A7"] = "RACI ACCEPT filled?"
    ws["B7"] = "NO (intentionally)"
    wb.save(CC)


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)

    if not PACKET.exists():
        raise SystemExit(f"missing packet: {PACKET}")

    # backup before mutate
    bak = EV / "YANINA_OWNER_PACKET_RACI_AND_REQUESTS.pre_h19.xlsx"
    shutil.copy2(PACKET, bak)

    wb = load_workbook(PACKET)
    update_readme(wb["README"])
    update_fill_checklist(wb["FILL_CHECKLIST"])
    update_gap_mitigation(wb["GAP_MITIGATION"])
    append_legacy_note_to_data_requests(wb["DATA_REQUESTS"])

    actions = build_h17_actions_rows()
    blocked = build_h18_blocked_rows()
    dr_now = build_data_requests_now()
    dlog = build_decision_log(actions)

    write_table_sheet(
        wb,
        "H17_ACTIONS",
        "H17 Owner Actions — проставьте decision (пусто = TODO)",
        actions,
        list(actions[0].keys()) if actions else ["action_id"],
    )
    write_table_sheet(
        wb,
        "H18_BLOCKED",
        "H18 Release Gate — BLOCKED months (что мешает закрыть месяц)",
        blocked,
        list(blocked[0].keys()) if blocked else ["period_month"],
    )
    write_table_sheet(
        wb,
        "DATA_REQUESTS_NOW",
        "Актуальные data requests (H17/H18) — приоритетнее legacy ST24",
        dr_now,
        list(dr_now[0].keys()),
    )
    write_table_sheet(
        wb,
        "DECISION_LOG",
        "Журнал решений владельцев (заполняете вы)",
        dlog,
        list(dlog[0].keys()),
    )

    # move new sheets near front after README
    order = [
        "README",
        "FILL_CHECKLIST",
        "RACI",
        "H17_ACTIONS",
        "H18_BLOCKED",
        "DATA_REQUESTS_NOW",
        "DECISION_LOG",
        "CANDIDATES_FROM_DOCS",
        "GAP_MITIGATION",
        "DATA_REQUESTS",
    ]
    for i, name in enumerate(order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

    wb.save(PACKET)
    shutil.copy2(PACKET, EV / "YANINA_OWNER_PACKET_RACI_AND_REQUESTS.xlsx")
    shutil.copy2(PACKET, OUT / "YANINA_OWNER_PACKET_RACI_AND_REQUESTS.xlsx")

    # also export CSVs for marts
    write_csv(MART / "data_requests_now.csv", dr_now, list(dr_now[0].keys()))
    write_csv(OUT / "data_requests_now.csv", dr_now, list(dr_now[0].keys()))

    # verify RACI decisions still empty
    wb2 = load_workbook(PACKET, read_only=True, data_only=True)
    raci = wb2["RACI"]
    hdr = [c.value for c in next(raci.iter_rows(min_row=1, max_row=1))]
    di = hdr.index("decision_ACCEPT_REJECT")
    filled = 0
    for row in raci.iter_rows(min_row=2, values_only=True):
        if row[di]:
            filled += 1
    wb2.close()

    summary = {
        "wave": "H19",
        "generated_at": NOW,
        "path_choice": "Sync Owner Packet to H17/H18 — maximize chance of real RACI ACCEPT",
        "finding": (
            f"H19: Owner Packet synced — sheets H17_ACTIONS({len(actions)}), "
            f"H18_BLOCKED({len(blocked)}), DATA_REQUESTS_NOW({len(dr_now)}), DECISION_LOG. "
            f"RACI decision cells still empty ({filled} filled)."
        ),
        "actions_n": len(actions),
        "blocked_n": len(blocked),
        "data_requests_now_n": len(dr_now),
        "raci_decisions_filled": filled,
        "packet": str(PACKET.relative_to(ROOT)),
        "backup": str(bak.relative_to(ROOT)),
        "not_sot": True,
        "did_not_autofill_accept": True,
    }
    (OUT / "h19_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    (EV / "h19_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    shutil.copy2(MART / "data_requests_now.csv", EV / "data_requests_now.csv")

    update_live_cc(summary)

    # short handoff md
    md = [
        "# Owner Packet Sync (H19)",
        "",
        f"Updated: {NOW}",
        "",
        "Файл: `live/YANINA_OWNER_PACKET_RACI_AND_REQUESTS.xlsx`",
        "",
        "1. **RACI** → колонка `decision_ACCEPT_REJECT` (P0)",
        "2. **H17_ACTIONS** → finance/SKU decisions",
        "3. **H18_BLOCKED** → какие месяцы нельзя закрывать",
        "4. **DATA_REQUESTS_NOW** → какие файлы прислать",
        "5. **DECISION_LOG** → журнал ваших решений",
        "",
        "AI не проставил ACCEPT. Backup pre-H19 лежит в evidence.",
        "",
    ]
    (OUT / "OWNER_PACKET_SYNC.md").write_text("\n".join(md), encoding="utf-8")
    (ROOT / "live/OWNER_PACKET_SYNC.md").write_text("\n".join(md), encoding="utf-8")
    (EV / "OWNER_PACKET_SYNC.md").write_text("\n".join(md), encoding="utf-8")

    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
