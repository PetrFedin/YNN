#!/usr/bin/env python3
"""
H6: витрины контроля + полный складской срез + усиление O2C.

1) Margin mart: channel×month и top/bottom SKU (из W4 после H5)
2) Остатки по всем складам FILE-098 (8 локаций + итого)
3) Пересверка settlements ↔ bank IN (все банки, exact+name)

Не SoT / не RACI ACCEPT.
"""
from __future__ import annotations

import csv
import hashlib
import json
import re
import shutil
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[3]
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/h6_marts_20260724"
DOCS = ROOT / "documents"
W1 = ROOT / "live/registers/w1_bank_cash"
W3 = ROOT / "live/registers/w3_sku_cost"
W4 = ROOT / "live/registers/w4_sales_settle"
W5 = ROOT / "live/registers/w5_sup_exp_mat"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")

CY2LAT = str.maketrans({
    "А": "A", "В": "B", "Е": "E", "К": "K", "М": "M", "Н": "H",
    "О": "O", "Р": "P", "С": "C", "Т": "T", "Х": "X",
    "Ё": "E", "ё": "E",
})


def nfc(s: str) -> str:
    return unicodedata.normalize("NFC", s or "")


def sha16(*parts) -> str:
    raw = "|".join("" if p is None else str(p) for p in parts)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def resolve(name: str) -> Path | None:
    t = nfc(name)
    for p in DOCS.iterdir():
        if nfc(p.name) == t:
            return p
    return None


def fnum(v) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def normalize_article(raw) -> str:
    if raw is None:
        return ""
    s = nfc(str(raw)).strip().upper().replace("Ё", "Е").translate(CY2LAT)
    s = s.replace("Т-", "T-").replace(" ", "").replace("\xa0", "")
    return s


def canonical_sku(raw) -> str:
    s = normalize_article(raw)
    return re.sub(r"/\d+[A-Z]?$", "", s)


def norm_name(s: str) -> str:
    s = nfc(s).lower()
    s = re.sub(r"[\"«»'`]", "", s)
    s = re.sub(r"\b(ооо|ип|зао|оао|ао|пао)\b", " ", s)
    s = re.sub(r"[^a-zа-я0-9]+", " ", s, flags=re.I)
    return re.sub(r"\s+", " ", s).strip()


# ── 1) Margin marts ───────────────────────────────────────────────
def build_margin_marts() -> tuple[list[dict], list[dict], list[dict], dict]:
    sales = list(csv.DictReader(open(W4 / "sales_lines.csv", encoding="utf-8")))
    by_cm = defaultdict(lambda: {"revenue": 0.0, "revenue_costed": 0.0, "cogs": 0.0, "qty": 0.0, "lines": 0, "lines_with_cogs": 0})
    by_sku = defaultdict(lambda: {"revenue": 0.0, "cogs": 0.0, "qty": 0.0, "lines": 0, "name": "", "channel_set": set()})

    for s in sales:
        ch = s.get("channel") or ""
        pm = s.get("period_month") or ""
        rev = fnum(s.get("revenue_rub")) or 0.0
        cogs = fnum(s.get("cogs_rub"))
        qty = fnum(s.get("qty")) or 0.0
        key = (ch, pm)
        by_cm[key]["revenue"] += rev
        by_cm[key]["qty"] += qty
        by_cm[key]["lines"] += 1
        if cogs is not None:
            by_cm[key]["cogs"] += cogs
            by_cm[key]["revenue_costed"] += rev
            by_cm[key]["lines_with_cogs"] += 1

        can = s.get("canonical_sku") or canonical_sku(s.get("article_raw"))
        if can:
            by_sku[can]["revenue"] += rev
            by_sku[can]["qty"] += qty
            by_sku[can]["lines"] += 1
            if cogs is not None:
                by_sku[can]["cogs"] += cogs
            if not by_sku[can]["name"]:
                by_sku[can]["name"] = (s.get("sku_name") or "")[:100]
            by_sku[can]["channel_set"].add(ch)

    channel_month = []
    for (ch, pm), v in sorted(by_cm.items()):
        # маржа только на покрытых COGS строках (без завышения %)
        margin = v["revenue_costed"] - v["cogs"] if v["lines_with_cogs"] else None
        pct = (margin / v["revenue_costed"] * 100) if margin is not None and v["revenue_costed"] else None
        cov = v["lines_with_cogs"] / v["lines"] if v["lines"] else 0
        channel_month.append(
            {
                "channel": ch,
                "period_month": pm,
                "lines": v["lines"],
                "lines_with_cogs": v["lines_with_cogs"],
                "cogs_coverage": round(cov, 3),
                "qty": round(v["qty"], 2),
                "revenue_rub": round(v["revenue"], 2),
                "revenue_costed_rub": round(v["revenue_costed"], 2),
                "cogs_rub": round(v["cogs"], 2) if v["lines_with_cogs"] else "",
                "margin_rub": round(margin, 2) if margin is not None else "",
                "margin_pct": round(pct, 1) if pct is not None else "",
                "status": "OK" if cov >= 0.85 else ("PARTIAL" if cov >= 0.5 else "WEAK"),
            }
        )

    # totals by channel
    by_ch = defaultdict(lambda: {"revenue": 0.0, "revenue_costed": 0.0, "cogs": 0.0, "lines": 0, "lines_with_cogs": 0})
    for r in channel_month:
        by_ch[r["channel"]]["revenue"] += r["revenue_rub"]
        by_ch[r["channel"]]["lines"] += r["lines"]
        by_ch[r["channel"]]["lines_with_cogs"] += r["lines_with_cogs"]
        by_ch[r["channel"]]["revenue_costed"] += r["revenue_costed_rub"]
        if r["cogs_rub"] != "":
            by_ch[r["channel"]]["cogs"] += float(r["cogs_rub"])

    channel_total = []
    for ch, v in sorted(by_ch.items()):
        margin = v["revenue_costed"] - v["cogs"]
        pct = margin / v["revenue_costed"] * 100 if v["revenue_costed"] else 0
        channel_total.append(
            {
                "channel": ch,
                "lines": v["lines"],
                "cogs_coverage": round(v["lines_with_cogs"] / v["lines"], 3) if v["lines"] else 0,
                "revenue_rub": round(v["revenue"], 2),
                "revenue_costed_rub": round(v["revenue_costed"], 2),
                "cogs_rub": round(v["cogs"], 2),
                "margin_rub": round(margin, 2),
                "margin_pct": round(pct, 1),
            }
        )

    sku_rows = []
    for can, v in by_sku.items():
        if v["revenue"] <= 0 or v["cogs"] <= 0:
            continue
        margin = v["revenue"] - v["cogs"]
        pct = margin / v["revenue"] * 100
        sku_rows.append(
            {
                "canonical_sku": can,
                "name": v["name"],
                "channels": ",".join(sorted(v["channel_set"])),
                "lines": v["lines"],
                "qty": round(v["qty"], 2),
                "revenue_rub": round(v["revenue"], 2),
                "cogs_rub": round(v["cogs"], 2),
                "margin_rub": round(margin, 2),
                "margin_pct": round(pct, 1),
            }
        )
    sku_rows.sort(key=lambda x: x["margin_rub"], reverse=True)
    top = sku_rows[:40]
    bottom = sorted(sku_rows, key=lambda x: x["margin_pct"])[:40]

    stats = {
        "channel_month_rows": len(channel_month),
        "channel_totals": channel_total,
        "sku_with_margin": len(sku_rows),
        "overall_revenue": round(sum(c["revenue_rub"] for c in channel_total), 2),
        "overall_revenue_costed": round(sum(c["revenue_costed_rub"] for c in channel_total), 2),
        "overall_margin": round(sum(c["margin_rub"] for c in channel_total), 2),
        "overall_margin_pct": round(
            sum(c["margin_rub"] for c in channel_total) / sum(c["revenue_costed_rub"] for c in channel_total) * 100, 1
        )
        if sum(c["revenue_costed_rub"] for c in channel_total)
        else 0,
    }
    return channel_month, channel_total, top + [{"canonical_sku": "---BOTTOM---"}] + bottom, stats


# ── 2) Full warehouse stock ───────────────────────────────────────
def parse_all_warehouses() -> tuple[list[dict], list[dict], dict]:
    path = resolve("Движение товара по складам.xlsx")
    if not path:
        return [], [], {"error": "missing"}
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header = next(ws.iter_rows(min_row=1, max_row=1, max_col=50, values_only=True))
    blocks = []
    for c, name in enumerate(header):
        if not name or c < 2:
            continue
        nm = str(name)
        if nm.strip().lower() == "итого":
            continue  # skip total block to avoid double count
        blocks.append({"name": nm[:80], "start": c, "end_col": c + 3})

    rows = []
    for i, r in enumerate(ws.iter_rows(min_row=4, max_col=45, values_only=True), start=4):
        if not r or not r[0]:
            continue
        a0 = str(r[0]).strip()
        if "," not in a0:
            continue
        art, _, name = a0.partition(",")
        art, name = art.strip(), name.strip()
        if not art or any(k in a0.lower() for k in ("списан", "оприход", "перемещ", "накладн")):
            continue
        for b in blocks:
            end = r[b["end_col"]] if len(r) > b["end_col"] else None
            begin = r[b["start"]] if len(r) > b["start"] else None
            inn = r[b["start"] + 1] if len(r) > b["start"] + 1 else None
            out = r[b["start"] + 2] if len(r) > b["start"] + 2 else None
            if not any(isinstance(x, (int, float)) and x for x in (begin, inn, out, end)):
                continue
            rows.append(
                {
                    "stock_wh_id": "SW-" + sha16(art, b["name"]),
                    "article_raw": art,
                    "canonical_sku": canonical_sku(art),
                    "name": name[:120],
                    "warehouse": b["name"],
                    "qty_begin": round(float(begin), 3) if isinstance(begin, (int, float)) else "",
                    "qty_in": round(float(inn), 3) if isinstance(inn, (int, float)) else "",
                    "qty_out": round(float(out), 3) if isinstance(out, (int, float)) else "",
                    "qty_end": round(float(end), 3) if isinstance(end, (int, float)) else "",
                    "source_file_id": "FILE-098",
                    "source_row_id": f"r{i}",
                }
            )
    wb.close()

    # warehouse summary
    by_wh = defaultdict(lambda: {"qty_end": 0.0, "skus": set(), "rows": 0})
    for r in rows:
        q = fnum(r["qty_end"]) or 0
        if q == 0 and not fnum(r["qty_begin"]) and not fnum(r["qty_in"]):
            continue
        by_wh[r["warehouse"]]["rows"] += 1
        by_wh[r["warehouse"]]["qty_end"] += fnum(r["qty_end"]) or 0
        by_wh[r["warehouse"]]["skus"].add(r["canonical_sku"])

    summary = [
        {
            "warehouse": wh,
            "rows": v["rows"],
            "skus_with_activity": len(v["skus"]),
            "qty_end_sum": round(v["qty_end"], 2),
        }
        for wh, v in sorted(by_wh.items(), key=lambda x: -x[1]["qty_end"])
    ]
    stats = {
        "stock_rows": len(rows),
        "warehouses": len(summary),
        "qty_end_total": round(sum(s["qty_end_sum"] for s in summary), 2),
        "unique_skus": len({r["canonical_sku"] for r in rows}),
    }
    return rows, summary, stats


# ── 3) Stronger settle↔bank ───────────────────────────────────────
def rematch_settlements() -> tuple[list[dict], dict]:
    settles = list(csv.DictReader(open(W4 / "settlements.csv", encoding="utf-8")))
    bak = W4 / "settlements_pre_h6.csv"
    if not bak.exists():
        shutil.copy2(W4 / "settlements.csv", bak)

    bank = [
        p
        for p in csv.DictReader(open(W1 / "bank_payments.csv", encoding="utf-8"))
        if p.get("direction") == "in" and p.get("is_internal") != "Y"
    ]
    by_month = defaultdict(list)
    for p in bank:
        by_month[p["period_month"]].append(p)

    used = {s["bank_payment_id"] for s in settles if s.get("bank_payment_id")}
    matches = []
    newly = 0

    def name_ov(a, b) -> float:
        ta = {t for t in norm_name(a).split() if len(t) >= 4}
        tb = {t for t in norm_name(b).split() if len(t) >= 4}
        if not ta or not tb:
            return 0.0
        return len(ta & tb) / max(len(ta), len(tb))

    for st in settles:
        if st.get("bank_payment_id"):
            continue
        rev = fnum(st.get("revenue_rub")) or 0
        if rev <= 0:
            continue
        best = None
        for p in by_month.get(st.get("period_month") or "", []):
            if p["bank_payment_id"] in used:
                continue
            amt = fnum(p.get("amount")) or 0
            ratio = abs(amt - rev) / max(amt, rev, 1)
            if ratio > 0.02:
                continue
            ov = name_ov(st.get("buyer") or "", p.get("counterparty_raw") or "")
            # exact amount alone = LOW; with name = HIGH/MED
            if ov >= 0.5:
                conf = "HIGH" if ov >= 0.65 and ratio <= 0.01 else "MED"
            elif ratio <= 0.001:
                conf = "LOW"
            else:
                continue
            cand = (conf, ov, ratio, p)
            order = {"HIGH": 0, "MED": 1, "LOW": 2}
            if best is None or (order[cand[0]], -cand[1], cand[2]) < (order[best[0]], -best[1], best[2]):
                best = cand
        if not best:
            continue
        conf, ov, ratio, p = best
        used.add(p["bank_payment_id"])
        st["bank_payment_id"] = p["bank_payment_id"]
        st["status"] = f"LINKED_H6_{conf}"
        newly += 1
        matches.append(
            {
                "settlement_id": st["settlement_id"],
                "document": st.get("document", ""),
                "buyer": st.get("buyer", ""),
                "revenue_rub": st.get("revenue_rub", ""),
                "period_month": st.get("period_month", ""),
                "bank_payment_id": p["bank_payment_id"],
                "payment_date": p["payment_date"],
                "bank_amount": p["amount"],
                "counterparty_raw": p.get("counterparty_raw", ""),
                "source_bank": p.get("source_bank", ""),
                "name_overlap": round(ov, 3),
                "confidence": conf,
                "match_method": "amount_name_h6",
            }
        )

    fields = list(settles[0].keys())
    write_csv(W4 / "settlements.csv", settles, fields)
    linked = sum(1 for s in settles if s.get("bank_payment_id"))
    by_conf = Counter(m["confidence"] for m in matches)
    by_status = Counter(s.get("status") for s in settles if s.get("bank_payment_id"))
    stats = {
        "settlements": len(settles),
        "linked_total": linked,
        "newly_linked_h6": newly,
        "new_by_confidence": dict(by_conf),
        "linked_status": dict(by_status),
    }
    return matches, stats


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)

    cm, ct, sku_tb, margin_stats = build_margin_marts()
    wh_rows, wh_sum, wh_stats = parse_all_warehouses()
    matches, settle_stats = rematch_settlements()

    write_csv(
        OUT / "margin_channel_month.csv",
        cm,
        list(cm[0].keys()) if cm else ["channel"],
    )
    write_csv(OUT / "margin_channel_total.csv", ct, list(ct[0].keys()) if ct else ["channel"])
    write_csv(OUT / "margin_sku_top_bottom.csv", [r for r in sku_tb if r.get("canonical_sku") != "---BOTTOM---"][:80], list(sku_tb[0].keys()))
    # also split top/bottom cleanly
    sku_only = [r for r in sku_tb if r.get("canonical_sku") and r["canonical_sku"] != "---BOTTOM---"]
    # reconstruct top/bottom from file logic - rebuild
    sales_sku = []
    # simpler: rewrite from cm builder outputs stored - call internal again via saved
    # Actually sku_tb has top40 + separator + bottom40. Filter separator.
    top = []
    bottom = []
    mode = "top"
    for r in sku_tb:
        if r.get("canonical_sku") == "---BOTTOM---":
            mode = "bottom"
            continue
        (top if mode == "top" else bottom).append(r)
    write_csv(OUT / "margin_sku_top40.csv", top, list(top[0].keys()) if top else ["canonical_sku"])
    write_csv(OUT / "margin_sku_bottom40.csv", bottom, list(bottom[0].keys()) if bottom else ["canonical_sku"])

    # publish marts next to architecture-ish live folder
    mart = ROOT / "live/marts"
    mart.mkdir(exist_ok=True)
    write_csv(mart / "margin_channel_month.csv", cm, list(cm[0].keys()) if cm else ["channel"])
    write_csv(mart / "margin_channel_total.csv", ct, list(ct[0].keys()) if ct else ["channel"])
    write_csv(mart / "margin_sku_top40.csv", top, list(top[0].keys()) if top else ["canonical_sku"])
    write_csv(mart / "margin_sku_bottom40.csv", bottom, list(bottom[0].keys()) if bottom else ["canonical_sku"])

    write_csv(OUT / "stock_by_warehouse_full.csv", wh_rows, list(wh_rows[0].keys()) if wh_rows else ["stock_wh_id"])
    write_csv(W5 / "stock_by_warehouse.csv", wh_rows, list(wh_rows[0].keys()) if wh_rows else ["stock_wh_id"])
    write_csv(OUT / "stock_warehouse_summary.csv", wh_sum, list(wh_sum[0].keys()) if wh_sum else ["warehouse"])
    write_csv(W5 / "stock_warehouse_summary.csv", wh_sum, list(wh_sum[0].keys()) if wh_sum else ["warehouse"])

    write_csv(
        OUT / "settle_bank_h6.csv",
        matches,
        list(matches[0].keys()) if matches else ["settlement_id"],
    )
    write_csv(
        W4 / "soft_matches_settle_bank_h6.csv",
        matches,
        list(matches[0].keys()) if matches else ["settlement_id"],
    )

    summary = {
        "generated_at": NOW,
        "wave": "H6",
        "margin": margin_stats,
        "stock_warehouses": wh_stats,
        "stock_warehouse_summary": wh_sum,
        "settle_bank": settle_stats,
        "finding": (
            f"H6: margin {margin_stats['overall_margin_pct']}% on costed "
            f"{margin_stats.get('overall_revenue_costed', margin_stats['overall_revenue']):.0f} RUB "
            f"(B2B {next(c['margin_pct'] for c in margin_stats['channel_totals'] if c['channel']=='B2B')}% / "
            f"IM {next(c['margin_pct'] for c in margin_stats['channel_totals'] if c['channel']=='IM')}% / "
            f"TSUM {next(c['margin_pct'] for c in margin_stats['channel_totals'] if c['channel']=='TSUM')}%); "
            f"warehouses {wh_stats.get('warehouses')} qty_end={wh_stats.get('qty_end_total')}; "
            f"settle linked {settle_stats['linked_total']} (+{settle_stats['newly_linked_h6']} H6)."
        ),
        "next": "RACI ACCEPT; optional BOM / fabric issue→FG",
        "not_sot": True,
    }
    json.dump(summary, open(OUT / "h6_summary.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    json.dump(summary, open(EV / "h6_summary.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    json.dump(summary, open(mart / "h6_summary.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)

    # patch w4 summary
    if (W4 / "w4_summary.json").exists():
        w4s = json.load(open(W4 / "w4_summary.json"))
        w4s["h6_settle"] = settle_stats
        w4s["generated_at"] = NOW
        json.dump(w4s, open(W4 / "w4_summary.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)

    wb = Workbook()
    ws = wb.active
    ws.title = "00_Summary"
    for i, (k, v) in enumerate(summary.items(), 1):
        ws.cell(i, 1, k)
        ws.cell(i, 2, json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v)
    fill = PatternFill("solid", fgColor="1F4E79")

    def add(name, rows):
        w = wb.create_sheet(name)
        if not rows:
            return
        headers = list(rows[0].keys())
        for c, h in enumerate(headers, 1):
            cell = w.cell(1, c, h)
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
        for ri, row in enumerate(rows[:4000], 2):
            for ci, h in enumerate(headers, 1):
                v = row.get(h, "")
                if isinstance(v, (dict, list)):
                    v = json.dumps(v, ensure_ascii=False)
                w.cell(ri, ci, v)

    add("01_Margin_Channel", ct)
    add("02_Margin_Month", cm)
    add("03_SKU_Top", top)
    add("04_SKU_Bottom", bottom)
    add("05_WH_Summary", wh_sum)
    add("06_Settle_H6", matches)
    wb.save(EV / "YANINA_H6_MARTS_EVIDENCE.xlsx")

    (OUT / "README.md").write_text(
        f"""# H6 Marts & Controls

{NOW}

{summary['finding']}

Marts: `../../marts/`

Evidence: `../../evidence/h6_marts_20260724/`
""",
        encoding="utf-8",
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
