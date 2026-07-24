#!/usr/bin/env python3
"""
H28: TSUM margin — product COGS vs commission (dual view).

Зачем (P0 из H27): комиссия ЦУМ входит в Excel/FILE себестоимость (УНФ не разделяет).
Cash-сверка уже использует net-rate ≈0.4668. Нельзя ещё раз вычитать комиссию в P&L.

Метод:
1) Для TSUM строк с cogs_source=FILE: product unit = median W3 unit_cost по SKU
   (если есть); commission_proxy = FILE cogs − product_cogs.
2) Если W3 нет — product_cogs неизвестен; оставляем reported only.
3) Marts: tsum_margin_dual, margin_channel с TSUM_REPORTED / TSUM_PRODUCT.
4) Политика: operating/cash не double-count; flag MF-TSUM → ADDRESSED_H28.

Не переписывает sales_lines cogs_rub (reported SoT Excel). Только аналитические колонки/marts.
"""
from __future__ import annotations

import csv
import json
import shutil
import statistics
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[3]
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/h28_tsum_cogs_split_20260724"
MART = ROOT / "live/marts"
W3 = ROOT / "live/registers/w3_sku_cost"
W4 = ROOT / "live/registers/w4_sales_settle"
CC = ROOT / "live/YANINA_LIVE_CONTROL_CENTER_20260723.xlsx"
PACKET = ROOT / "live/YANINA_OWNER_PACKET_RACI_AND_REQUESTS.xlsx"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")

# Политика net-rate из H11 (cash model)
NET_RATE = 0.4668
AGENCY_TAKE = round(1.0 - NET_RATE, 4)


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def fnum(v):
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(" ", "").replace(",", "."))
    except (TypeError, ValueError):
        return None


def w3_median_units() -> dict[str, float]:
    by: dict[str, list[float]] = defaultdict(list)
    for r in csv.DictReader(open(W3 / "cost_versions.csv", encoding="utf-8")):
        sku = r.get("canonical_sku") or ""
        uc = fnum(r.get("unit_cost_rub"))
        if not sku or uc is None or uc <= 0:
            continue
        by[sku].append(uc)
    return {sku: statistics.median(vals) for sku, vals in by.items() if vals}


def build_dual_fixed(w3u: dict[str, float]) -> tuple[list[dict], dict]:
    lines = []
    stats = defaultdict(int)
    tot = defaultdict(float)
    for r in csv.DictReader(open(W4 / "sales_lines.csv", encoding="utf-8")):
        if r.get("channel") != "TSUM" or r.get("dq_exclude_from_margin") == "Y":
            continue
        rev = fnum(r.get("revenue_rub")) or 0.0
        qty = fnum(r.get("qty")) or 0.0
        cogs = fnum(r.get("cogs_rub"))
        src = r.get("cogs_source") or ""
        sku = r.get("canonical_sku") or ""
        unit_price = rev / qty if qty else None
        agency_comm_at_net = rev * AGENCY_TAKE

        product_cogs = None
        commission_proxy = None
        method = "NO_SPLIT"
        if src == "FILE" and sku in w3u and qty:
            product_cogs = round(w3u[sku] * qty, 2)
            if cogs is not None:
                commission_proxy = round(cogs - product_cogs, 2)
            method = "FILE_MINUS_W3_MEDIAN"
            stats["file_w3_split"] += 1
        elif src.startswith("W3"):
            product_cogs = cogs
            commission_proxy = 0.0
            method = "W3_AS_PRODUCT"
            stats["w3_as_product"] += 1
        else:
            stats["no_split"] += 1

        reported_margin = round(rev - cogs, 2) if cogs is not None else None
        product_margin = round(rev - product_cogs, 2) if product_cogs is not None else None
        embed_share = round(commission_proxy / rev, 4) if commission_proxy is not None and rev else None

        tot["revenue"] += rev
        if cogs is not None:
            tot["cogs_reported"] += cogs
        if product_cogs is not None:
            tot["product_cogs"] += product_cogs
            tot["revenue_split"] += rev
        if commission_proxy is not None:
            tot["commission_proxy"] += commission_proxy
        tot["agency_comm_model"] += agency_comm_at_net

        lines.append(
            {
                "sales_line_id": r.get("sales_line_id"),
                "period_month": r.get("period_month"),
                "canonical_sku": sku,
                "qty": qty,
                "revenue_rub": rev,
                "unit_price_rub": round(unit_price, 2) if unit_price is not None else "",
                "cogs_reported_rub": cogs if cogs is not None else "",
                "cogs_source": src,
                "product_cogs_rub": product_cogs if product_cogs is not None else "",
                "commission_proxy_rub": commission_proxy if commission_proxy is not None else "",
                "agency_comm_at_net_rate_rub": round(agency_comm_at_net, 2),
                "reported_margin_rub": reported_margin if reported_margin is not None else "",
                "product_margin_rub": product_margin if product_margin is not None else "",
                "commission_proxy_share_of_rev": embed_share if embed_share is not None else "",
                "split_method": method,
                "net_rate_policy": NET_RATE,
                "agency_take_policy": AGENCY_TAKE,
            }
        )
    return lines, {"counts": dict(stats), "totals": dict(tot)}


def month_rollups(lines: list[dict]) -> list[dict]:
    by = defaultdict(lambda: defaultdict(float))
    n = defaultdict(lambda: defaultdict(int))
    for r in lines:
        m = r["period_month"] or ""
        by[m]["revenue"] += fnum(r["revenue_rub"]) or 0
        if r.get("cogs_reported_rub") != "":
            by[m]["cogs_reported"] += fnum(r["cogs_reported_rub"]) or 0
            n[m]["reported"] += 1
        if r.get("product_cogs_rub") != "":
            by[m]["product_cogs"] += fnum(r["product_cogs_rub"]) or 0
            by[m]["commission_proxy"] += fnum(r["commission_proxy_rub"]) or 0
            by[m]["revenue_split"] += fnum(r["revenue_rub"]) or 0
            n[m]["split"] += 1
        by[m]["agency_comm_model"] += fnum(r["agency_comm_at_net_rate_rub"]) or 0
    rows = []
    for m in sorted(by.keys()):
        v = by[m]
        rep_m = v["revenue"] - v["cogs_reported"]
        prod_m = v["revenue_split"] - v["product_cogs"] if v["revenue_split"] else None
        rows.append(
            {
                "channel": "TSUM",
                "period_month": m,
                "revenue_rub": round(v["revenue"], 2),
                "cogs_reported_rub": round(v["cogs_reported"], 2),
                "reported_margin_rub": round(rep_m, 2),
                "reported_margin_pct": round(100 * rep_m / v["revenue"], 1) if v["revenue"] else "",
                "product_cogs_rub": round(v["product_cogs"], 2),
                "commission_proxy_rub": round(v["commission_proxy"], 2),
                "product_margin_rub": round(prod_m, 2) if prod_m is not None else "",
                "product_margin_pct": round(100 * prod_m / v["revenue_split"], 1) if v["revenue_split"] else "",
                "agency_comm_at_net_rate_rub": round(v["agency_comm_model"], 2),
                "split_lines": n[m]["split"],
                "reported_lines": n[m]["reported"],
                "policy": "H28_DUAL_VIEW",
            }
        )
    return rows


def channel_views(lines: list[dict], other_channels: list[dict]) -> list[dict]:
    """margin_channel_views: keep other channels; split TSUM into two rows."""
    rev = sum(fnum(r["revenue_rub"]) or 0 for r in lines)
    cogs_r = sum(fnum(r["cogs_reported_rub"]) or 0 for r in lines if r.get("cogs_reported_rub") != "")
    cogs_p = sum(fnum(r["product_cogs_rub"]) or 0 for r in lines if r.get("product_cogs_rub") != "")
    rev_p = sum(fnum(r["revenue_rub"]) or 0 for r in lines if r.get("product_cogs_rub") != "")
    comm = sum(fnum(r["commission_proxy_rub"]) or 0 for r in lines if r.get("commission_proxy_rub") != "")
    rows = list(other_channels)
    # replace TSUM if present
    rows = [r for r in rows if r.get("channel") not in ("TSUM", "TSUM_REPORTED", "TSUM_PRODUCT", "TOTAL")]
    rows.append(
        {
            "channel": "TSUM_REPORTED",
            "lines": len(lines),
            "costed_lines": sum(1 for r in lines if r.get("cogs_reported_rub") != ""),
            "revenue_rub": round(rev, 2),
            "cogs_rub": round(cogs_r, 2),
            "margin_rub": round(rev - cogs_r, 2),
            "margin_pct": round(100 * (rev - cogs_r) / rev, 1) if rev else "",
            "view": "Excel COGS (может включать комиссию ЦУМ)",
        }
    )
    rows.append(
        {
            "channel": "TSUM_PRODUCT",
            "lines": sum(1 for r in lines if r.get("product_cogs_rub") != ""),
            "costed_lines": sum(1 for r in lines if r.get("product_cogs_rub") != ""),
            "revenue_rub": round(rev_p, 2),
            "cogs_rub": round(cogs_p, 2),
            "margin_rub": round(rev_p - cogs_p, 2),
            "margin_pct": round(100 * (rev_p - cogs_p) / rev_p, 1) if rev_p else "",
            "view": f"Product COGS≈W3; commission_proxy={round(comm,2)} (FILE−W3)",
        }
    )
    # totals for reported view (B2B+IM+TSUM_REPORTED)
    # rebuild from file
    return rows


def load_other_channels() -> list[dict]:
    rows = []
    for r in csv.DictReader(open(MART / "margin_channel_total.csv", encoding="utf-8")):
        if r.get("channel") in ("TSUM", "TOTAL"):
            continue
        rows.append(
            {
                "channel": r["channel"],
                "lines": r.get("lines"),
                "costed_lines": r.get("costed_lines"),
                "revenue_rub": r.get("revenue_rub"),
                "cogs_rub": r.get("cogs_rub"),
                "margin_rub": r.get("margin_rub"),
                "margin_pct": r.get("margin_pct"),
                "view": "unchanged",
            }
        )
    return rows


def update_model_flag():
    path = MART / "model_flags_h27.csv"
    if not path.exists():
        return
    rows = list(csv.DictReader(open(path, encoding="utf-8")))
    for r in rows:
        if r.get("flag_id") == "MF-TSUM-COGS-COMMISSION":
            r["status"] = "ADDRESSED_H28_DUAL_VIEW"
            r["action"] = (
                "Dual view: TSUM_REPORTED vs TSUM_PRODUCT (FILE−W3). "
                "Do not subtract agency take again in P&L/cash bridge. "
                "Confirm with Merkushina how % is baked into ЦУМ cost cards."
            )
    write_csv(path, rows, list(rows[0].keys()))


def write_md(summary: dict, month_rows: list[dict]):
    lines = [
        "# H28 — TSUM product COGS vs commission",
        "",
        f"Updated: {NOW}",
        "",
        "## Зачем",
        "Комиссия ЦУМ сидит в FILE/Excel себестоимости. Cash уже net-rate. Dual view без double-count.",
        "",
        f"- Net-rate policy: **{NET_RATE}** (agency take **{AGENCY_TAKE}**)",
        f"- Split method: FILE cogs − median W3 unit × qty",
        f"- Lines split: **{summary['split_lines']}** / {summary['tsum_lines']}",
        "",
        "## Итоги",
        "",
        f"| View | Revenue | COGS | Margin | Margin % |",
        f"|------|---------|------|--------|----------|",
        f"| Reported (FILE) | {summary['rev']:,.0f} | {summary['cogs_reported']:,.0f} | {summary['margin_reported']:,.0f} | {summary['margin_reported_pct']}% |",
        f"| Product (W3) | {summary['rev_split']:,.0f} | {summary['product_cogs']:,.0f} | {summary['margin_product']:,.0f} | {summary['margin_product_pct']}% |",
        f"| Commission proxy (FILE−W3) | — | {summary['commission_proxy']:,.0f} | — | {summary['commission_proxy_share_pct']}% of rev |",
        f"| Agency take at net-rate | — | {summary['agency_comm_model']:,.0f} | — | {AGENCY_TAKE*100:.2f}% of rev |",
        "",
        "## Политика",
        "1. `TSUM_REPORTED` — как в Excel (для сверки с их файлами).",
        "2. `TSUM_PRODUCT` — товарная маржа (W3), комиссия вынесена в proxy.",
        "3. В operating bridge / cash **не** вычитать agency take поверх reported COGS.",
        "4. Уточнить у Меркушиной формулу комиссии в карточках ЦУМ.",
        "",
        "Files: `live/marts/tsum_margin_dual_lines.csv`, `tsum_margin_dual_month.csv`, `margin_channel_views_h28.csv`",
        "",
    ]
    (OUT / "TSUM_COGS_SPLIT.md").write_text("\n".join(lines), encoding="utf-8")
    (ROOT / "live/TSUM_COGS_SPLIT.md").write_text("\n".join(lines), encoding="utf-8")
    (EV / "TSUM_COGS_SPLIT.md").write_text("\n".join(lines), encoding="utf-8")


def update_cc(summary: dict):
    if not CC.exists():
        return
    wb = load_workbook(CC)
    if "H28_TSUM_Split" in wb.sheetnames:
        del wb["H28_TSUM_Split"]
    ws = wb.create_sheet("H28_TSUM_Split", 0)
    ws["A1"] = "H28 TSUM COGS split"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A2"] = NOW
    ws["A4"] = summary["finding"]
    ws["A5"] = "Reported GM%"
    ws["B5"] = summary["margin_reported_pct"]
    ws["A6"] = "Product GM%"
    ws["B6"] = summary["margin_product_pct"]
    ws["A7"] = "Comm proxy / Agency model"
    ws["B7"] = f"{summary['commission_proxy']:,.0f} / {summary['agency_comm_model']:,.0f}"
    ws["A8"] = "Doc"
    ws["B8"] = "live/TSUM_COGS_SPLIT.md"
    wb.save(CC)
    if PACKET.exists():
        wb2 = load_workbook(PACKET)
        if "README" in wb2.sheetnames:
            wb2["README"]["A19"] = (
                f"H28 {NOW}: TSUM dual margin — reported {summary['margin_reported_pct']}% vs "
                f"product {summary['margin_product_pct']}%. See TSUM_COGS_SPLIT.md"
            )
        wb2.save(PACKET)


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)

    w3u = w3_median_units()
    lines, meta = build_dual_fixed(w3u)
    months = month_rollups(lines)
    other = load_other_channels()
    views = channel_views(lines, other)

    # add totals for both views roughly
    def add_total(view_channel_prefix=None):
        pass

    write_csv(MART / "tsum_margin_dual_lines.csv", lines, list(lines[0].keys()))
    write_csv(OUT / "tsum_margin_dual_lines.csv", lines, list(lines[0].keys()))
    write_csv(MART / "tsum_margin_dual_month.csv", months, list(months[0].keys()) if months else ["period_month"])
    write_csv(MART / "margin_channel_views_h28.csv", views, list(views[0].keys()))
    write_csv(OUT / "margin_channel_views_h28.csv", views, list(views[0].keys()))

    tot = meta["totals"]
    rev = tot["revenue"]
    cogs_r = tot["cogs_reported"]
    prod = tot.get("product_cogs", 0)
    rev_s = tot.get("revenue_split", 0)
    comm = tot.get("commission_proxy", 0)
    agency = tot.get("agency_comm_model", 0)

    summary = {
        "wave": "H28",
        "generated_at": NOW,
        "finding": (
            f"H28: TSUM dual view — reported margin {round(100*(rev-cogs_r)/rev,1) if rev else None}% "
            f"vs product(W3) {round(100*(rev_s-prod)/rev_s,1) if rev_s else None}%; "
            f"commission proxy {round(comm):,.0f} ₽ (~{round(100*comm/rev_s,1) if rev_s else None}% rev) "
            f"vs agency@net-rate {round(agency):,.0f} ₽. No double-count policy set."
        ),
        "tsum_lines": len(lines),
        "split_lines": meta["counts"].get("file_w3_split", 0) + meta["counts"].get("w3_as_product", 0),
        "counts": meta["counts"],
        "net_rate": NET_RATE,
        "agency_take": AGENCY_TAKE,
        "rev": round(rev, 2),
        "cogs_reported": round(cogs_r, 2),
        "margin_reported": round(rev - cogs_r, 2),
        "margin_reported_pct": round(100 * (rev - cogs_r) / rev, 1) if rev else None,
        "rev_split": round(rev_s, 2),
        "product_cogs": round(prod, 2),
        "commission_proxy": round(comm, 2),
        "commission_proxy_share_pct": round(100 * comm / rev_s, 1) if rev_s else None,
        "margin_product": round(rev_s - prod, 2),
        "margin_product_pct": round(100 * (rev_s - prod) / rev_s, 1) if rev_s else None,
        "agency_comm_model": round(agency, 2),
        "not_sot": True,
        "sales_lines_cogs_unchanged": True,
    }
    write_md(summary, months)
    update_model_flag()
    update_cc(summary)
    (OUT / "h28_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    (EV / "h28_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    for name in ("tsum_margin_dual_lines.csv", "tsum_margin_dual_month.csv", "margin_channel_views_h28.csv", "TSUM_COGS_SPLIT.md"):
        src = OUT / name if (OUT / name).exists() else MART / name
        if src.exists():
            shutil.copy2(src, EV / name)
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
