#!/usr/bin/env python3
"""
W5: REG-SUP + REG-EXP + REG-MAT staging (P2P / материалы).

Источники:
- Список контрагентов по закупке → SUP
- Расходы 2024–2026 (помесячные листы, разрезы по счетам) → EXP
- Остатки ткани 1С на 31.05.2026 → MAT snapshot

Сверка: EXP итог месяца ↔ bank operating out / DDS.
Не SoT: нет PO/invoice ids; EXP — агрегат по статье×счёт, не платёжные строки.
"""
from __future__ import annotations

import csv
import hashlib
import json
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[3]
DOCS = ROOT / "documents"
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/w5_sup_exp_mat_20260724"
W1 = ROOT / "live/registers/w1_bank_cash"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")

INN_RE = re.compile(r"ИНН\s*[:\s]*([0-9]{10,12})", re.I)
SHEET_MONTH = re.compile(r"^(\d{2})-(\d{2})$")  # MM-YY


def nfc(s: str) -> str:
    return unicodedata.normalize("NFC", s or "")


def sha16(*parts) -> str:
    raw = "|".join("" if p is None else str(p) for p in parts)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]


def to_float(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def resolve(name: str) -> Path | None:
    t = nfc(name)
    for p in DOCS.iterdir():
        if nfc(p.name) == t:
            return p
    return None


# Классификация строк «Расходы*.xlsx» (файл смешанный: доходы+расходы+обороты)
INCOME_MARKERS = (
    "доход",
    "частные клиенты",
    "интернет-магазин",
    "цум (комиссионер)",
    "салон",
)
TRANSFER_MARKERS = (
    "внутренние обороты",
    "обороты между",
    "на карту ип",
    "ип (обороты",
)
OWNER_MARKERS = (
    "снято на хознужды",
    "дивиденд",
)
TOTAL_MARKERS = ("итого",)


def classify_expense_article(article: str) -> str:
    a = (article or "").strip().lower()
    if not a:
        return "EMPTY"
    if any(x in a for x in TOTAL_MARKERS):
        return "TOTAL"
    if any(x in a for x in TRANSFER_MARKERS):
        return "TRANSFER"
    if any(x in a for x in OWNER_MARKERS):
        return "OWNER_DRAW"
    if any(x in a for x in INCOME_MARKERS):
        return "INCOME"
    if a.startswith("расходы за") or a.startswith("доходы за"):
        return "HEADER"
    return "EXPENSE"


def map_account(col_name: str) -> dict:
    n = (col_name or "").strip().lower()
    if "декор" in n:
        return {"account_bucket": "DEKOR", "legal_entity_id": "LE-OOO-DEKOR", "bank_hint": "VTB_DEKOR?"}
    if "салон" in n:
        return {"account_bucket": "SALON", "legal_entity_id": "LE-IP-YANINA", "bank_hint": "SALON"}
    if "втб" in n:
        return {"account_bucket": "IP_VTB", "legal_entity_id": "LE-IP-YANINA", "bank_hint": "VTB"}
    if "альфа" in n:
        return {"account_bucket": "IP_ALFA", "legal_entity_id": "LE-IP-YANINA", "bank_hint": "ALFA"}
    if "райф" in n:
        return {"account_bucket": "IP_RAIF", "legal_entity_id": "LE-IP-YANINA", "bank_hint": "RAIF"}
    if "карта" in n:
        return {"account_bucket": "IP_CARD", "legal_entity_id": "LE-IP-YANINA", "bank_hint": "CARD"}
    if "итого" in n:
        return {"account_bucket": "TOTAL", "legal_entity_id": "", "bank_hint": ""}
    return {"account_bucket": "OTHER", "legal_entity_id": "LE-IP-YANINA", "bank_hint": col_name}


# ── SUP ───────────────────────────────────────────────────────────
def build_suppliers() -> list[dict]:
    path = resolve("Список контрагентов по закупке.xlsx")
    if not path:
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    out = []
    seen = set()
    for rnum, r in enumerate(rows[1:], start=2):
        if not r or not r[0]:
            continue
        name = str(r[0]).strip()
        contact = str(r[1]).strip() if len(r) > 1 and r[1] else ""
        info = str(r[2]).strip() if len(r) > 2 and r[2] else ""
        m = INN_RE.search(info)
        inn = m.group(1) if m else ""
        cid = "CP-" + sha16("SUP", inn or name)
        if cid in seen:
            continue
        seen.add(cid)
        out.append(
            {
                "counterparty_id": cid,
                "name": name,
                "inn": inn,
                "contact": contact,
                "info_raw": info[:300],
                "has_inn": "Y" if inn else "N",
                "source_file_id": "FILE-085",
                "source_row_id": f"r{rnum}",
            }
        )
    return out


# ── EXP ───────────────────────────────────────────────────────────
def parse_expenses_file(path: Path, source_file_id: str) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    out = []
    for sheet in wb.sheetnames:
        m = SHEET_MONTH.match(sheet.strip())
        if not m:
            continue
        mm, yy = m.group(1), m.group(2)
        year = 2000 + int(yy)
        period_month = f"{year}-{mm}"
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        # find header with Наименование
        hdr_i = None
        header = None
        for i, r in enumerate(rows[:8]):
            if not r:
                continue
            joined = " ".join(str(c) for c in r if c is not None)
            if "Наименование" in joined:
                hdr_i = i
                header = [str(c).strip() if c is not None else "" for c in r]
                break
        if hdr_i is None:
            continue
        # account columns: all except name and Итого (we'll still store total separately)
        name_i = 0
        for i, h in enumerate(header):
            if "наименование" in h.lower():
                name_i = i
                break
        acct_cols = []
        total_i = None
        for i, h in enumerate(header):
            if i == name_i or not h:
                continue
            if "итого" in h.lower():
                total_i = i
                continue
            acct_cols.append((i, h, map_account(h)))

        for rnum, r in enumerate(rows[hdr_i + 1 :], start=hdr_i + 2):
            if not r or r[name_i] is None:
                continue
            article = str(r[name_i]).strip()
            if not article:
                continue
            line_class = classify_expense_article(article)
            if line_class in ("HEADER", "EMPTY"):
                continue
            vals = []
            for i, h, meta in acct_cols:
                v = to_float(r[i]) if len(r) > i else None
                if v is None or v == 0:
                    continue
                vals.append((h, meta, v))
            total = to_float(r[total_i]) if total_i is not None and len(r) > total_i else None
            if not vals and (total is None or total == 0):
                continue
            article_id = sha16("ART", article.lower())
            # TOTAL: одна строка по итоговому столбцу (контроль vs DDS)
            if line_class == "TOTAL" and total is not None:
                out.append(
                    {
                        "expense_line_id": "EX-" + sha16(source_file_id, sheet, rnum, "TOTAL", article, total),
                        "period_month": period_month,
                        "article_id": article_id,
                        "article_name": article,
                        "line_class": line_class,
                        "account_bucket": "TOTAL",
                        "account_col": "Итого",
                        "legal_entity_id": "",
                        "bank_hint": "",
                        "amount_rub": round(total, 2),
                        "row_total_rub": round(total, 2),
                        "counterparty_id": "",
                        "bank_payment_id": "",
                        "source_file_id": source_file_id,
                        "source_sheet": sheet,
                        "source_row_id": f"r{rnum}",
                        "match_status": "CONTROL_TOTAL",
                    }
                )
                continue
            for h, meta, v in vals:
                out.append(
                    {
                        "expense_line_id": "EX-" + sha16(source_file_id, sheet, rnum, h, article, v),
                        "period_month": period_month,
                        "article_id": article_id,
                        "article_name": article,
                        "line_class": line_class,
                        "account_bucket": meta["account_bucket"],
                        "account_col": h,
                        "legal_entity_id": meta["legal_entity_id"],
                        "bank_hint": meta["bank_hint"],
                        "amount_rub": round(v, 2),
                        "row_total_rub": round(total, 2) if total is not None else "",
                        "counterparty_id": "",
                        "bank_payment_id": "",
                        "source_file_id": source_file_id,
                        "source_sheet": sheet,
                        "source_row_id": f"r{rnum}",
                        "match_status": "UNMATCHED",
                    }
                )
    wb.close()
    return out


def build_expenses() -> list[dict]:
    files = [
        ("Расходы 2024.xlsx", "FILE-060"),
        ("Расходы 2025.xlsx", "FILE-061"),
        ("Расходы 2026.xlsx", "FILE-062"),
    ]
    out = []
    for name, fid in files:
        p = resolve(name)
        if p:
            out.extend(parse_expenses_file(p, fid))
    return out


# ── MAT ───────────────────────────────────────────────────────────
def build_materials() -> tuple[list[dict], dict]:
    path = resolve("Остатки ткани из 1С на 31.05.2026.xlsx")
    if not path:
        return [], {}
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    out = []
    warehouse_total = {}
    skip_names = {"номенклатура"}
    for rnum, r in enumerate(rows[1:], start=2):
        if not r or r[0] is None:
            continue
        name = str(r[0]).strip()
        if name.lower() in skip_names:
            continue
        qty = to_float(r[1]) if len(r) > 1 else None
        amount = to_float(r[2]) if len(r) > 2 else None
        unit_cost = to_float(r[3]) if len(r) > 3 else None
        # warehouse aggregator
        if "склад" in name.lower() and (unit_cost is None or unit_cost == 0):
            warehouse_total[name] = {
                "qty": qty,
                "amount": amount,
            }
            continue
        if amount is None and qty is None:
            continue
        # extract article-like code at end if present
        art_m = re.search(r"(\d{1,2}-\d{3,4}[A-Za-zА-Яа-я]?)\s*$", name)
        article = art_m.group(1) if art_m else ""
        mat_id = "MAT-" + sha16("FABRIC", name)
        out.append(
            {
                "material_movement_id": mat_id,  # snapshot row as stock position
                "material_id": mat_id,
                "name": name,
                "article_hint": article,
                "qty": qty if qty is not None else "",
                "amount_rub": round(amount, 2) if amount is not None else "",
                "unit_cost_rub": round(unit_cost, 2) if unit_cost is not None else "",
                "snapshot_date": "2026-05-31",
                "warehouse": "Основной склад",
                "movement_type": "STOCK_SNAPSHOT",
                "source_file_id": "FILE-053",
                "source_row_id": f"r{rnum}",
            }
        )
    meta = {
        "n_skus": len(out),
        "qty_sum": round(sum(float(x["qty"] or 0) for x in out), 2),
        "amount_sum_leaves": round(sum(float(x["amount_rub"] or 0) for x in out), 2),
        "warehouse_totals": {
            k: {"qty": v["qty"], "amount": v["amount"]} for k, v in warehouse_total.items()
        },
    }
    return out, meta


# ── Recon EXP vs bank / DDS ───────────────────────────────────────
def recon_exp(expenses: list[dict]) -> list[dict]:
    exp_m = defaultdict(float)
    exp_by_bucket = defaultdict(lambda: defaultdict(float))
    total_row = {}
    for e in expenses:
        m = e["period_month"]
        cls = e.get("line_class") or "EXPENSE"
        if cls == "TOTAL":
            total_row[m] = float(e["amount_rub"] or 0)
            continue
        if cls != "EXPENSE":
            continue
        exp_m[m] += float(e["amount_rub"] or 0)
        exp_by_bucket[m][e["account_bucket"]] += float(e["amount_rub"] or 0)

    bank_out = defaultdict(float)
    if (W1 / "recon_bank_vs_dds_month.csv").exists():
        for r in csv.DictReader(open(W1 / "recon_bank_vs_dds_month.csv", encoding="utf-8")):
            key = "bank_out_operating_rub" if "bank_out_operating_rub" in r else "bank_out_rub"
            bank_out[r["period_month"]] = float(r.get(key) or 0)
    elif (W1 / "bank_payments.csv").exists():
        for p in csv.DictReader(open(W1 / "bank_payments.csv", encoding="utf-8")):
            if p.get("direction") == "out" and p.get("is_internal") != "Y":
                bank_out[p["period_month"]] += float(p["amount"] or 0)

    dds = defaultdict(float)
    dds_bn = defaultdict(float)
    if (W1 / "cash_lines.csv").exists():
        for c in csv.DictReader(open(W1 / "cash_lines.csv", encoding="utf-8")):
            if c.get("ledger") != "B":
                continue
            dds[c["period_month"]] += float(c["amount_rub"] or 0)
            if "б/нал" in (c.get("cash_type") or "").lower():
                dds_bn[c["period_month"]] += float(c["amount_rub"] or 0)

    def status_of(a, b):
        if a and b:
            ratio = abs(a - b) / max(a, b)
            if ratio <= 0.05:
                return "CLOSE", round(a - b, 2)
            if ratio <= 0.15:
                return "SOFT_GAP", round(a - b, 2)
            if ratio <= 0.30:
                return "WIDE_GAP", round(a - b, 2)
            return "GAP", round(a - b, 2)
        if a and not b:
            return "LEFT_ONLY", ""
        if b and not a:
            return "RIGHT_ONLY", ""
        return "N/A", ""

    months = sorted(set(exp_m) | set(bank_out) | set(dds) | set(total_row))
    rows = []
    for m in months:
        e = exp_m.get(m, 0)
        b = bank_out.get(m, 0)
        d = dds.get(m, 0)
        dbn = dds_bn.get(m, 0)
        tot = total_row.get(m, 0)
        exp_hub = exp_by_bucket[m].get("IP_VTB", 0) + exp_by_bucket[m].get("IP_ALFA", 0)
        st_bank, d_bank = status_of(e, b)
        st_hub, d_hub = status_of(exp_hub, b)
        st_tot, d_tot = status_of(tot, d)
        st_tot_bn, d_tot_bn = status_of(tot, dbn)
        rows.append(
            {
                "period_month": m,
                "exp_opex_rub": round(e, 2),
                "exp_hub_vtb_alfa_rub": round(exp_hub, 2),
                "exp_total_row_rub": round(tot, 2) if tot else "",
                "bank_out_operating_rub": round(b, 2),
                "dds_b_all_rub": round(d, 2),
                "dds_b_bn_rub": round(dbn, 2),
                "delta_opex_vs_bank": d_bank,
                "status_opex_vs_bank": st_bank,
                "delta_hub_vs_bank": d_hub,
                "status_hub_vs_bank": st_hub,
                "delta_total_row_vs_dds": d_tot,
                "status_total_row_vs_dds": st_tot,
                "delta_total_row_vs_dds_bn": d_tot_bn,
                "status_total_row_vs_dds_bn": st_tot_bn,
                "exp_IP_VTB": round(exp_by_bucket[m].get("IP_VTB", 0), 2),
                "exp_IP_ALFA": round(exp_by_bucket[m].get("IP_ALFA", 0), 2),
                "exp_DEKOR": round(exp_by_bucket[m].get("DEKOR", 0), 2),
                "note": "opex=line_class EXPENSE; TOTAL=ИТОГО vs DDS",
            }
        )
    return rows


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)

    suppliers = build_suppliers()
    expenses = build_expenses()
    materials, mat_meta = build_materials()
    recon = recon_exp(expenses)

    no_inn = sum(1 for s in suppliers if s["has_inn"] == "N")
    class_c = Counter(e.get("line_class", "EXPENSE") for e in expenses)
    opex = [e for e in expenses if e.get("line_class") == "EXPENSE"]

    write_csv(
        OUT / "suppliers.csv",
        suppliers,
        ["counterparty_id", "name", "inn", "contact", "info_raw", "has_inn", "source_file_id", "source_row_id"],
    )
    exp_fields = [
        "expense_line_id", "period_month", "article_id", "article_name", "line_class",
        "account_bucket", "account_col", "legal_entity_id", "bank_hint", "amount_rub",
        "row_total_rub", "counterparty_id", "bank_payment_id", "source_file_id",
        "source_sheet", "source_row_id", "match_status",
    ]
    write_csv(OUT / "expense_lines.csv", expenses, exp_fields)
    write_csv(OUT / "expense_opex_only.csv", opex, exp_fields)
    mat_fields = [
        "material_movement_id", "material_id", "name", "article_hint", "qty", "amount_rub",
        "unit_cost_rub", "snapshot_date", "warehouse", "movement_type", "source_file_id", "source_row_id",
    ]
    write_csv(OUT / "material_stock.csv", materials, mat_fields)
    write_csv(OUT / "recon_exp_bank_dds.csv", recon, list(recon[0].keys()) if recon else ["period_month"])

    status_opex = Counter(r["status_opex_vs_bank"] for r in recon if r["exp_opex_rub"])
    status_hub = Counter(r["status_hub_vs_bank"] for r in recon if r["exp_hub_vtb_alfa_rub"])
    status_tot = Counter(r["status_total_row_vs_dds"] for r in recon if r["exp_total_row_rub"])
    status_tot_bn = Counter(r["status_total_row_vs_dds_bn"] for r in recon if r["exp_total_row_rub"])
    close_tot = [r["period_month"] for r in recon if r["status_total_row_vs_dds"] in ("CLOSE", "SOFT_GAP")]
    close_hub = [r["period_month"] for r in recon if r["status_hub_vs_bank"] in ("CLOSE", "SOFT_GAP")]

    art = Counter()
    for e in opex:
        art[e["article_name"]] += float(e["amount_rub"] or 0)

    wh = mat_meta.get("warehouse_totals") or {}
    preferred_amt = preferred_qty = None
    for k, v in wh.items():
        preferred_amt = v.get("amount")
        preferred_qty = v.get("qty")
        break

    summary = {
        "generated_at": NOW,
        "wave": "W5",
        "suppliers": len(suppliers),
        "suppliers_with_inn": len(suppliers) - no_inn,
        "suppliers_without_inn": no_inn,
        "expense_lines_all_classes": len(expenses),
        "expense_line_classes": dict(class_c),
        "expense_opex_lines": len(opex),
        "expense_opex_total_rub": round(sum(float(e["amount_rub"] or 0) for e in opex), 2),
        "expense_months": len({e["period_month"] for e in expenses}),
        "material_skus": mat_meta.get("n_skus"),
        "material_warehouse_qty": preferred_qty,
        "material_warehouse_amount_rub": preferred_amt,
        "material_amount_leaves_rub": mat_meta.get("amount_sum_leaves"),
        "material_snapshot_date": "2026-05-31",
        "recon_opex_vs_bank": dict(status_opex),
        "recon_hub_vs_bank": dict(status_hub),
        "recon_total_row_vs_dds": dict(status_tot),
        "recon_total_row_vs_dds_bn": dict(status_tot_bn),
        "close_soft_total_vs_dds": close_tot,
        "close_soft_hub_vs_bank": close_hub,
        "top_opex_articles": art.most_common(8),
        "finding": (
            f"W5: {len(suppliers)} SUP ({len(suppliers)-no_inn} INN); "
            f"{len(opex)} opex lines; "
            f"fabric {mat_meta.get('n_skus')} SKU / warehouse {preferred_amt} RUB @2026-05-31. "
            f"ИТОГО↔DDS Б/Нал: {dict(status_tot_bn)}; "
            f"opex hub(VTB+Alfa)↔bank: {dict(status_hub)} CLOSE/SOFT={close_hub}."
        ),
        "next": "W6 TAX/BUD or map SUP.inn → bank.counterparty_inn",
    }
    json.dump(summary, open(OUT / "w5_summary.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    json.dump(summary, open(EV / "w5_summary.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    json.dump(mat_meta, open(OUT / "material_meta.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)

    wb = Workbook()
    ws = wb.active
    ws.title = "00_Summary"
    for i, (k, v) in enumerate(summary.items(), 1):
        ws.cell(i, 1, k)
        ws.cell(i, 2, json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v)
    fill = PatternFill("solid", fgColor="1F4E79")

    def add(name, rows):
        w = wb.create_sheet(name)
        if not rows:
            return
        headers = list(rows[0].keys())
        for c, h in enumerate(headers, 1):
            cell = w.cell(1, c, h)
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
        for ri, row in enumerate(rows[:5000], 2):
            for ci, h in enumerate(headers, 1):
                w.cell(ri, ci, row.get(h, ""))

    add("01_Recon", recon)
    add("02_Suppliers", suppliers)
    add("03_Top_Opex", [{"article": a, "amount_rub": round(v, 2)} for a, v in art.most_common(30)])
    wb.save(EV / "YANINA_W5_SUP_EXP_MAT_EVIDENCE.xlsx")

    (OUT / "README.md").write_text(
        f"""# W5 SUP / EXP / MAT

Generated: {NOW}

- `suppliers.csv` — {len(suppliers)}
- `expense_lines.csv` — all classes; `expense_opex_only.csv` — {len(opex)}
- `material_stock.csv` — {mat_meta.get('n_skus')} SKU; warehouse amt {preferred_amt}
- ИТОГО↔DDS CLOSE/SOFT: {close_tot}

Evidence: `../../evidence/w5_sup_exp_mat_20260724/`
""",
        encoding="utf-8",
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
