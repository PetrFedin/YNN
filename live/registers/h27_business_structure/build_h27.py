#!/usr/bin/env python3
"""
H27: Структура бизнеса + замена TEMP stubs на реальные ФИО из письма.

Источник: письмо/бриф от команды (юрструктура, каналы, RACI ролей, закупки, 1С).

Критичные правки модели (флаги, не автофикс маржи):
1) Комиссия ЦУМ входит в себестоимость (ограничение УНФ) — риск double-count с net-rate.
2) Каналы 2025: индивидуальный пошив 83% / ИМ 8% / B2B(ЦУМ) 9% — сверить с нашими B2B/IM/TSUM.
3) ООО Декор с 2026 без деятельности → ликвидация; сотрудники в ИП.
4) Фурнитура не в 1С — только ДДС; в COGS рыночная/фактическая.
5) Движение МД/инд. заказов не ведётся — остатков нет.

Не объявляет audited SoT. Обновляет RACI ACCEPT с реальными ФИО.
"""
from __future__ import annotations

import csv
import json
import shutil
import subprocess
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[3]
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/h27_business_structure_20260724"
MART = ROOT / "live/marts"
PACKET = ROOT / "live/YANINA_OWNER_PACKET_RACI_AND_REQUESTS.xlsx"
CC = ROOT / "live/YANINA_LIVE_CONTROL_CENTER_20260723.xlsx"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")
TODAY = datetime.now().strftime("%Y%m%d")
EVIDENCE = "owner brief 2026-07-24: структура бизнеса / роли / 1С / закупки"

HDR = PatternFill("solid", fgColor="1F4E79")
HDR_F = Font(bold=True, color="FFFFFF")

# RACI replacements: (item_id, role) → (fio, notes)
RACI_REAL = {
    ("SRC-CTRL-01", "Approver"): (
        "Мамушкина Елена",
        "Исполнительный директор; выдача наличных по ФОТ",
    ),
    ("SRC-CTRL-02", "Owner"): (
        "Мокеева Анна",
        "Производственный директор; себестоимость вместе с Богдашкиной/Меркушиной/Жуковой",
    ),
    ("SRC-CTRL-02", "Approver"): (
        "Коптева Марина",
        "Операционный директор",
    ),
    ("SRC-CTRL-03", "Owner"): (
        "Коптева Марина",
        "B2B менеджера нет — контролирует операционный директор",
    ),
    ("SRC-CTRL-03", "Approver"): (
        "Мамушкина Елена",
        "Исполнительный директор",
    ),
    ("DOM-PRODUCT", "Owner"): (
        "Коновалова Анна",
        "Товаровед; склад/материалы (с Дендериной замещают кладовщика)",
    ),
    ("DOM-COST", "Owner"): (
        "Мокеева Анна",
        "Себестоимость: Мокеева + Богдашкина + Меркушина + Жукова (primary = пр.директор)",
    ),
    ("DOM-PROD", "Owner"): (
        "Мокеева Анна",
        "Производственный директор (МД: мастера/конструктора); технолог Шалагинова",
    ),
    ("DOM-B2B", "Owner"): (
        "Коптева Марина",
        "Контроль B2B/ЦУМ; отдельного менеджера нет",
    ),
    ("DOM-DATA", "Owner"): (
        "Сливяк Галина",
        "Главный бухгалтер: 1С, выгрузки, свод CF/P&L",
    ),
}

DOMAIN_OWNERS = {
    "CASH": ("Мамушкина Елена", "ACCEPTED", "N"),
    "BANK": ("Сливяк Галина", "ACCEPTED", "N"),
    "TAX": ("Сливяк Галина", "ACCEPTED", "N"),
    "TAX_APPROVER": ("ЯНИНА ЮЛИЯ ФЕДОРОВНА", "ACCEPTED", "N"),
    "PAYROLL": ("Сливяк Галина", "ACCEPTED", "N"),  # свод; цеха — ниже в org chart
    "PRODUCT": ("Коновалова Анна", "ACCEPTED", "N"),
    "COST": ("Мокеева Анна", "ACCEPTED", "N"),
    "PRODUCTION": ("Мокеева Анна", "ACCEPTED", "N"),
    "B2B": ("Коптева Марина", "ACCEPTED", "N"),
    "DATA_STEWARD": ("Сливяк Галина", "ACCEPTED", "N"),
}


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def write_business_structure() -> Path:
    text = f"""# YANINA — Структура бизнеса (H27)

Updated: {NOW}  
Источник: бриф команды / письмо Петру. **Контролируемый intel**, не audited SoT.

---

## 1. Юридическая структура

| Юрлицо | Статус | Налог | Примечание |
|--------|--------|-------|------------|
| **ООО «Декор»** | С **2026** хоз. деятельность не ведёт, подготовка к ликвидации | УСН 6% | Оборот 2025 (налог) **20,6 млн**; сотрудники переведены в ИП с 2026 |
| **ООО «Салон Юлия Янина»** | Аренда нежилых помещений ИП (на балансе Салона) | УСН 6% | Сотрудников нет |
| **ИП Янина Ю.Ф.** | Пошив женской одежды — основной операционный контур | УСН 6%; с 2025 **НДС 5%** | Сотрудники **46**; пониженный тариф взносов МП обрабатывающих; оборот 2025 (налог) **125 млн** |

---

## 2. Каналы продаж ИП (доля выручки 2025)

| Канал | Доля 2025 | Как читать в наших регистрах |
|-------|-----------|------------------------------|
| Индивидуальный пошив (услуги) | **83%** | МД / couture / инд. заказы — движение остатков **не ведётся**, продажи по факту |
| Интернет-магазин (сайт, розница) | **8%** | Наш канал **IM** |
| B2B (агентский договор с **ЦУМ**) | **9%** | В брифе B2B = агентский ЦУМ; в staging у нас отдельно B2B wholesale + TSUM — **сверить/уточнить таксономию** |

---

## 3. Производственные подразделения ИП

1. **Собственное производство** — кутюр и индивидуальные заказы  
2. **Проектный отдел** (своё + аутсорс) — Модный дом и частично ИМ  
3. **Проектный отдел** (аутсорс) — B2B и тираж для ИМ  

Подрядчики: мини-производства, ателье, самозанятые. Тираж мал для фабрик.

---

## 4. Финансовая отчётность ИП

- CF в Excel с **2023**; упрощённый **P&L с 2026**; **баланса нет**
- CF: подотчётники + исполнит. директор → бухгалтер сводит + безнал → собственнику
- Упрощённый бухучёт в **1С**; управленческий учёт в **Excel**
- Курсовые разницы начали отражать в 2026; отчёты часто в EUR по курсу **100**

### Денежный периметр
- Салон + корп. карта — отражаются в «расходы»; можно дослать выписки  
- Личные счета Юлии для снятия — проходят в **ДДС**  
- Контрагенты в основном по предоплате; долгосрочная ДЗ/КЗ — инд. заказы (таблица МД от Елены)  
- Подотчёт: ежемесячный ДДС + остаток в общем остатке средств  

---

## 5. 1С и себестоимость (критично для модели)

| Факт | Следствие для staging |
|------|------------------------|
| В 1С продажи площадок и ИМ | OK для sales lines |
| **Производство / формирование себестоимости в 1С не ведётся** | COGS из Excel cost tables (как у нас W3/H5) — ожидаемо |
| Приход «от себя» по накладной по Excel-себестоимости | Направление (МД/ИМ/B2B) в таблицах cost |
| **Учёт фурнитуры нет** (ассортимент) | Только ДДС; в COGS — рыночная или факт на изделие |
| В 1С ткани/мех/кожа | Реестр поставщиков тканей из 1С |
| Движение МД и инд. заказов **нигде не ведётся** | Остатков МД не будет |
| **Комиссия ЦУМ входит в себестоимость** (УНФ не умеет отдельно) | Риск **двойного учёта** с нашей TSUM net-rate сверкой — см. model flags |

---

## 6. Закупки (Дендерина)

- Москва: нал 1–15 м, 2–3 раза/нед; безнал рулоны ~30 м, 2–3 раза/мес; фурнитура/расходники  
- Клеевые: мин 100 м, ~2 раза/год  
- Принты: Мск / НН / СПб, от 25 м, ~2 раза/год; СДЭК; оплата → безнал  
- Импорт Maritex / Pino: 2 раза/год (янв–фев, сен–окт), от 50 м; оплата налом собственником во Франции; логистика Nerjus, оплата доставки налом в Мск  
- Китай: 2 раза/год через **Мамушкину**; крупный опт, поставщик «неизвестен» в брифе  
- Рекламации: акт + фото → замена/перепечатка/компенсация (редко)

---

## 7. Оргструктура / RACI (из брифа)

| Домен | Люди |
|-------|------|
| Наличные | **Мамушкина Елена** — исполнит. директор |
| Безнал / 1С / налоги / выгрузки | **Сливяк Галина** — гл. бухгалтер |
| ФОТ вышивка | Богдашкина Евгения |
| ФОТ мастера/конструктора (МД) | Мокеева Анна |
| ФОТ проектный отдел | Жукова Анна |
| ФОТ окладники / свод / ведомости | Сливяк Галина |
| Выдача наличных ЗП | Мамушкина Елена |
| Закупки | Дендерина Ирина (+ Меркушина частично, Богдашкина — фурнитура вышивки) |
| Склад | Вакансия; зам. Дендерина + **Коновалова Анна** (товаровед) |
| Производство МД | Богдашкина, Мокеева, Шалагинова (технолог) |
| Площадки/ИМ | Меркушина Татьяна |
| МД/ИМ проект | Жукова Анна |
| Себестоимость | Богдашкина, Мокеева, Меркушина, Жукова |
| B2B | Менеджера нет → **Коптева Марина** (опер. директор) |
| B2C | Лимачева Инна |

---

## 8. Model flags (открыть в controls)

См. `live/marts/model_flags_h27.csv` и Owner Packet лист `MODEL_FLAGS_H27`.
"""
    path = ROOT / "live/BUSINESS_STRUCTURE.md"
    path.write_text(text, encoding="utf-8")
    (OUT / "BUSINESS_STRUCTURE.md").write_text(text, encoding="utf-8")
    (EV / "BUSINESS_STRUCTURE.md").write_text(text, encoding="utf-8")
    (ROOT / "BUSINESS_STRUCTURE.md").write_text(text, encoding="utf-8")
    return path


def model_flags() -> list[dict]:
    return [
        {
            "flag_id": "MF-TSUM-COGS-COMMISSION",
            "severity": "P0",
            "domain": "B2B/TSUM/COST",
            "title": "Комиссия ЦУМ включена в себестоимость (УНФ)",
            "impact": "Маржа TSUM занижена как «товарная»; net-rate сверка может double-count комиссию",
            "action": "Разделить в модели: product_cogs vs tsum_commission; не вычитать комиссию дважды",
            "status": "OPEN_MODEL",
            "owner": "Коптева Марина / Мокеева Анна / Сливяк",
        },
        {
            "flag_id": "MF-CHANNEL-MIX-2025",
            "severity": "P0",
            "domain": "SALES",
            "title": "Каналы 2025: 83% инд.пошив / 8% ИМ / 9% B2B(ЦУМ)",
            "impact": "Наш split B2B vs TSUM vs IM может не совпадать с налоговой/управленческой таксономией",
            "action": "Сверить margin_channel_total с брифом; выделить MD/инд.пошив отдельно от wholesale B2B",
            "status": "OPEN_MODEL",
            "owner": "Сливяк / Коптева / Мамушкина",
        },
        {
            "flag_id": "MF-DECOR-WINDDOWN-2026",
            "severity": "P1",
            "domain": "LEGAL",
            "title": "ООО Декор с 2026 без деятельности → ликвидация",
            "impact": "Платежи Декор после 2025 — не operating; сотрудники в ИП",
            "action": "Пометить LE-DEKOR 2026+ как wind-down в bank/tax marts",
            "status": "OPEN_MODEL",
            "owner": "Сливяк Галина",
        },
        {
            "flag_id": "MF-NO-FURNITURE-1C",
            "severity": "P1",
            "domain": "MATERIALS",
            "title": "Фурнитура не в 1С — только ДДС",
            "impact": "Materials mart неполон; COGS фурнитуры эвристический",
            "action": "Не ждать stock фурнитуры; оставить DDS + cost-table inclusion",
            "status": "ACCEPTED_CONSTRAINT",
            "owner": "Дендерина / Мокеева",
        },
        {
            "flag_id": "MF-NO-MD-STOCK",
            "severity": "P1",
            "domain": "PRODUCT",
            "title": "Нет движения/остатков МД и инд. заказов",
            "impact": "Нельзя построить stock SoT по кутюру",
            "action": "Sales-on-fact only; не блокировать SoT отсутствием остатков МД",
            "status": "ACCEPTED_CONSTRAINT",
            "owner": "Мокеева / Мамушкина",
        },
        {
            "flag_id": "MF-FX-RATE-100",
            "severity": "P2",
            "domain": "FINANCE",
            "title": "Отчёты в EUR по курсу 100; FX diff только с 2026",
            "impact": "Сверки RUB↔EUR чувствительны к курсу",
            "action": "Явно хранить fx_policy=100 в DDS marts; не смешивать с рыночным курсом без флага",
            "status": "OPEN_MODEL",
            "owner": "Сливяк Галина",
        },
        {
            "flag_id": "MF-VAT5-IP-2025",
            "severity": "P1",
            "domain": "TAX",
            "title": "ИП с 2025 плательщик НДС 5%",
            "impact": "Выручка/налоги в сверках должны учитывать НДС 5%",
            "action": "Проверить tax cash vs декларации с учётом НДС 5%",
            "status": "OPEN_MODEL",
            "owner": "Сливяк Галина",
        },
    ]


def org_chart_rows() -> list[dict]:
    return [
        {"area": "Cash", "role": "Исполнительный директор", "fio": "Мамушкина Елена", "scope": "Наличные, выдача ЗП, CF доходы"},
        {"area": "Bank/Tax/1C", "role": "Главный бухгалтер", "fio": "Сливяк Галина", "scope": "Безнал, налоги, 1С, свод CF/P&L, окладники"},
        {"area": "Owner", "role": "Собственник / Tax Approver", "fio": "Янина Юлия Фёдоровна", "scope": "ИП, утверждение"},
        {"area": "Ops", "role": "Операционный директор", "fio": "Коптева Марина", "scope": "B2B контроль, ops"},
        {"area": "B2C", "role": "Менеджер клиенты", "fio": "Лимачева Инна", "scope": "B2C"},
        {"area": "Production", "role": "Производственный директор", "fio": "Мокеева Анна", "scope": "Мастера, конструктора МД, себестоимость"},
        {"area": "Embroidery", "role": "Нач. вышивального цеха", "fio": "Богдашкина Евгения", "scope": "Вышивка, ФОТ, фурнитура вышивки, cost"},
        {"area": "Project IM/площадки", "role": "Рук. проектного отдела", "fio": "Меркушина Татьяна", "scope": "Площадки/ИМ, аутсорс закупки, cost"},
        {"area": "Project MD/IM", "role": "Рук. проектного отдела", "fio": "Жукова Анна", "scope": "МД/ИМ проект, ФОТ отдела, cost"},
        {"area": "Tech", "role": "Технолог", "fio": "Шалагинова Татьяна", "scope": "МД технология"},
        {"area": "Procurement", "role": "Снабженец", "fio": "Дендерина Ирина", "scope": "Закупки тканей, склад (зам.)"},
        {"area": "Warehouse", "role": "Товаровед", "fio": "Коновалова Анна", "scope": "Склад/материалы (зам. кладовщика)"},
        {"area": "Warehouse", "role": "Кладовщик", "fio": "ВАКАНСИЯ", "scope": "Склад"},
    ]


def apply_raci(wb) -> list[dict]:
    ws = wb["RACI"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    changed = []
    for r in range(2, (ws.max_row or 1) + 1):
        item_id = ws.cell(r, idx["item_id"] + 1).value
        role = ws.cell(r, idx["role"] + 1).value
        key = (item_id, role)
        if key not in RACI_REAL:
            continue
        fio, note = RACI_REAL[key]
        old = ws.cell(r, idx["named_person_FIO"] + 1).value
        ws.cell(r, idx["named_person_FIO"] + 1).value = fio
        ws.cell(r, idx["decision_ACCEPT_REJECT"] + 1).value = "ACCEPT"
        ws.cell(r, idx["approval_date_YYYYMMDD"] + 1).value = TODAY
        ws.cell(r, idx["evidence_link_or_file"] + 1).value = EVIDENCE
        ws.cell(r, idx["status"] + 1).value = "ACCEPTED"
        if "notes" in idx:
            ws.cell(r, idx["notes"] + 1).value = f"H27 replaced stub→real. {note}"
        changed.append({"item_id": item_id, "role": role, "old": old, "new": fio})
    return changed


def update_sot_owners():
    rows = []
    for domain, (fio, status, stub) in DOMAIN_OWNERS.items():
        rows.append(
            {
                "domain": domain,
                "owner": fio,
                "status": status,
                "source": "H27_BUSINESS_BRIEF",
                "stub": stub,
                "replace_with_real_fio": "N",
            }
        )
    write_csv(MART / "sot_owners.csv", rows, list(rows[0].keys()))
    write_csv(OUT / "sot_owners.csv", rows, list(rows[0].keys()))
    return rows


def update_alias_master_owners():
    path = MART / "sku_alias_master.csv"
    if not path.exists():
        return
    rows = list(csv.DictReader(open(path, encoding="utf-8")))
    for r in rows:
        if r.get("alias_status") == "CONTROLLED_CANDIDATE":
            r["owner_domain"] = "DOM-PRODUCT"
            r["owner_status"] = "ACCEPTED"
            r["registry_accept"] = "PENDING_PRODUCT_REVIEW"  # real owner named, still review candidates
            r["note"] = "H27: Product Owner=Коновалова Анна; candidates await review, not auto-applied"
            r["updated_at"] = NOW
        elif "QUARANTINE" in (r.get("alias_status") or "") or "MARGIN_EXCEPTION" in (r.get("alias_status") or ""):
            r["owner_status"] = "ACCEPTED"
            r["updated_at"] = NOW
    write_csv(path, rows, list(rows[0].keys()))


def write_packet_sheets(wb, flags, org, changed):
    # MODEL_FLAGS
    name = "MODEL_FLAGS_H27"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, 3)
    ws["A1"] = "Model flags from business brief (H27)"
    ws["A1"].font = Font(bold=True, size=13, color="C00000")
    ws["A2"] = NOW
    ws.append([])
    ws.append(list(flags[0].keys()))
    for c in range(1, len(flags[0]) + 1):
        ws.cell(4, c).fill = HDR
        ws.cell(4, c).font = HDR_F
    for f in flags:
        ws.append(list(f.values()))

    name = "ORG_CHART_H27"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, 4)
    ws["A1"] = "Org chart from brief"
    ws.append([])
    ws.append(list(org[0].keys()))
    for c in range(1, len(org[0]) + 1):
        ws.cell(3, c).fill = HDR
        ws.cell(3, c).font = HDR_F
    for r in org:
        ws.append(list(r.values()))

    if "README" in wb.sheetnames:
        wb["README"]["A18"] = (
            f"H27 {NOW}: stubs→real FIO from business brief; "
            "see BUSINESS_STRUCTURE.md + MODEL_FLAGS_H27 (TSUM commission in COGS!)"
        )
        wb["README"]["A18"].font = Font(bold=True, color="C00000")


def update_sot_policy():
    path = ROOT / "live/SOT_POLICY.md"
    block = f"""

---

## H27 — структура бизнеса и реальные owners ({NOW})

Stubs **сняты**. Owners из брифа:

| Domain | Owner |
|--------|-------|
| CASH | Мамушкина Елена |
| BANK / TAX / PAYROLL / DATA | Сливяк Галина |
| TAX Approver | Янина Ю.Ф. |
| PRODUCT | Коновалова Анна |
| COST / PRODUCTION | Мокеева Анна |
| B2B | Коптева Марина |

Полный текст: `live/BUSINESS_STRUCTURE.md`  
Model flags: `live/marts/model_flags_h27.csv` (особенно **комиссия ЦУМ в COGS**).
"""
    if path.exists():
        t = path.read_text(encoding="utf-8")
        # strip old H26 stub section emphasis by appending H27
        if "H27 — структура" not in t:
            path.write_text(t.rstrip() + "\n" + block, encoding="utf-8")
    shutil.copy2(path, EV / "SOT_POLICY.md")


def update_cc(summary):
    if not CC.exists():
        return
    wb = load_workbook(CC)
    if "H27_BizStruct" in wb.sheetnames:
        del wb["H27_BizStruct"]
    ws = wb.create_sheet("H27_BizStruct", 0)
    ws["A1"] = "H27 Business Structure"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A2"] = NOW
    ws["A4"] = summary["finding"]
    ws["A5"] = "Doc"
    ws["B5"] = "live/BUSINESS_STRUCTURE.md"
    ws["A6"] = "P0 flag"
    ws["B6"] = "TSUM commission inside COGS"
    ws["A7"] = "Channel mix 2025"
    ws["B7"] = "83% individual / 8% IM / 9% B2B(TSUM)"
    wb.save(CC)


def rebuild_ops():
    subprocess.check_call([sys.executable, str(ROOT / "live/registers/h25_domain_ops/build_h25.py")], cwd=str(ROOT))


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)
    shutil.copy2(PACKET, EV / "YANINA_OWNER_PACKET_RACI_AND_REQUESTS.pre_h27.xlsx")

    write_business_structure()
    flags = model_flags()
    org = org_chart_rows()
    write_csv(MART / "model_flags_h27.csv", flags, list(flags[0].keys()))
    write_csv(OUT / "model_flags_h27.csv", flags, list(flags[0].keys()))
    write_csv(MART / "org_chart_h27.csv", org, list(org[0].keys()))

    wb = load_workbook(PACKET)
    changed = apply_raci(wb)
    write_packet_sheets(wb, flags, org, changed)
    wb.save(PACKET)
    shutil.copy2(PACKET, EV / "YANINA_OWNER_PACKET_RACI_AND_REQUESTS.xlsx")

    owners = update_sot_owners()
    update_alias_master_owners()
    update_sot_policy()
    rebuild_ops()

    freeze = {
        "wave": "H27",
        "generated_at": NOW,
        "status": "DOMAIN_OWNED_FROM_BRIEF",
        "stubs_cleared": True,
        "so_t": False,
        "raci_accept": True,
        "model_flags_open": sum(1 for f in flags if f["status"] == "OPEN_MODEL"),
    }
    (MART / "staging_freeze.json").write_text(json.dumps(freeze, ensure_ascii=False, indent=2), encoding="utf-8")

    summary = {
        "wave": "H27",
        "generated_at": NOW,
        "finding": (
            f"H27: business structure ingested; {len(changed)} RACI stubs→real FIO; "
            f"{len(flags)} model flags (P0: TSUM commission in COGS; channel mix 83/8/9; Decor wind-down). "
            "See live/BUSINESS_STRUCTURE.md"
        ),
        "raci_replaced_n": len(changed),
        "raci_replaced": changed,
        "owners": owners,
        "model_flags_n": len(flags),
        "p0_flags": [f["flag_id"] for f in flags if f["severity"] == "P0"],
        "not_sot": True,
    }
    update_cc(summary)
    (OUT / "h27_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    (EV / "h27_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    shutil.copy2(MART / "model_flags_h27.csv", EV / "model_flags_h27.csv")
    shutil.copy2(MART / "org_chart_h27.csv", EV / "org_chart_h27.csv")
    shutil.copy2(MART / "sot_owners.csv", EV / "sot_owners.csv")

    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
