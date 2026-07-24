#!/usr/bin/env python3
"""
H17: Owner Action Pack + SKU alias candidates.

Путь: денежные controls на текущих данных закрыты (H15/H16).
Дальше без RACI/новых файлов нужен пакет решений и запросов,
чтобы команда закрыла SoT-гейт и точечные дыры.

1) OWNER_ACTIONS.csv / xlsx — RACI, finance exceptions, data requests
2) sku_alias_candidates.csv — из H9 collisions (master-кандидат, не SoT)
3) finance_b2b_loss_evidence.csv — строки продаж по 3 SKU
4) Обновить RESULTS/STATUS

Не SoT. Не заполняет RACI ACCEPT за пользователя.
"""
from __future__ import annotations

import csv
import json
import shutil
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[3]
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/h17_owner_pack_20260724"
MART = ROOT / "live/marts"
W4 = ROOT / "live/registers/w4_sales_settle"
H9 = ROOT / "live/registers/h9_cost_identity"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def fnum(v):
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(" ", "").replace(",", "."))
    except (TypeError, ValueError):
        return None


def build_owner_actions() -> list[dict]:
    actions = []

    def add(**kw):
        actions.append(
            {
                "action_id": kw.get("action_id"),
                "priority": kw.get("priority"),
                "owner_hint": kw.get("owner_hint"),
                "category": kw.get("category"),
                "title": kw.get("title"),
                "detail": kw.get("detail"),
                "blocks": kw.get("blocks"),
                "evidence_path": kw.get("evidence_path", ""),
                "status": "TODO",
                "decision_needed": kw.get("decision_needed", "Y"),
            }
        )

    add(
        action_id="A-RACI-01",
        priority="P0",
        owner_hint="Юлия / Сливяк / Мамушкина",
        category="RACI",
        title="ACCEPT или REJECT черновика RACI",
        detail=(
            "В Owner Packet: Мамушкина Елена = Cash Owner; Сливяк Галина = Bank/Tax/Payroll Owner. "
            "Заполнить decision_ACCEPT_REJECT. Без этого регистры остаются staging, не SoT."
        ),
        blocks="SoT / эталонные политики данных",
        evidence_path="live/YANINA_OWNER_PACKET_RACI_AND_REQUESTS.xlsx",
    )

    # finance SKUs
    fin_path = MART / "finance_neg_sku_review.csv"
    if fin_path.exists():
        for r in csv.DictReader(open(fin_path, encoding="utf-8")):
            sku = r.get("canonical_sku") or ""
            if r.get("review_reason") == "WHOLESALE_BELOW_STOCK_COST":
                add(
                    action_id=f"A-FIN-{sku}",
                    priority="P1",
                    owner_hint="Финансы / коммерция",
                    category="FINANCE_EXCEPTION",
                    title=f"Подтвердить B2B убыток {sku}",
                    detail=(
                        f"Цена/ед {r.get('unit_price_rub')} ₽ vs cost/ед {r.get('unit_cost_rub')} ₽ "
                        f"(gap {r.get('unit_gap_rub')} ₽). Выручка {r.get('revenue_rub')}, "
                        f"маржа {r.get('margin_rub')}. Варианты: OK commercial loss / ошибка cost / прайс."
                    ),
                    blocks="Чистота margin SoT по B2B",
                    evidence_path="live/marts/finance_b2b_loss_evidence.csv",
                )
            elif "QUARANTINE" in (r.get("review_reason") or ""):
                add(
                    action_id=f"A-FIN-{sku}",
                    priority="P1",
                    owner_hint="Производство / 1С номенклатура",
                    category="SKU_IDENTITY",
                    title=f"Разрешить identity {sku}",
                    detail=(
                        "Продажа = свитшот «Be a poem»; cost masters = худи/юбка; "
                        "соседний 0-3244 свитшот с unit≈43160 (ещё хуже). Нужен правильный cost version "
                        "или alias map."
                    ),
                    blocks="IM COGS на одной строке",
                    evidence_path="live/registers/h13_im_finance/cogs_quarantine.csv",
                )

    # IM OPEN months
    im_path = MART / "recon_im_combo.csv"
    if im_path.exists():
        opens = [r for r in csv.DictReader(open(im_path, encoding="utf-8")) if r.get("status") == "OPEN"]
        months = ",".join(r["period_month"] for r in opens)
        add(
            action_id="A-DATA-IM-01",
            priority="P2",
            owner_hint="Сливяк / банк",
            category="DATA_REQUEST",
            title="Эквайринг-реестры на IM OPEN-месяцы",
            detail=(
                f"Месяцы: {months}. Нужны полные возмещения Tinkoff/TBank/VTB (и Декор, если отдельно). "
                "Цель: IM CLOSE/SOFT с 80% → 90%+."
            ),
            blocks="IM cash coverage",
            evidence_path="live/marts/recon_im_combo.csv",
            decision_needed="N",
        )

    # B2B open
    settles = list(csv.DictReader(open(W4 / "settlements.csv", encoding="utf-8")))
    open_b2b = [s for s in settles if s.get("channel") == "B2B" and not s.get("bank_payment_id")]
    open_rev = sum(fnum(s.get("revenue_rub")) or 0 for s in open_b2b)
    buyers = sorted({(s.get("buyer") or "")[:40] for s in open_b2b})
    add(
        action_id="A-DATA-B2B-01",
        priority="P2",
        owner_hint="Сливяк / продажи B2B",
        category="DATA_REQUEST",
        title=f"Платежи/взаимозачёты на B2B open ({len(open_b2b)} шт, ~{open_rev:,.0f} ₽)",
        detail=(
            "В текущих выписках нет достаточных свободных платежей тех же контрагентов. "
            f"Покупатели: {'; '.join(buyers[:8])}{'…' if len(buyers) > 8 else ''}."
        ),
        blocks="B2B settle coverage",
        evidence_path="live/marts/data_request_b2b_open.csv",
        decision_needed="N",
    )

    # ZP vedomosti
    pay_path = MART / "recon_payroll_multi.csv"
    if pay_path.exists():
        no_lines = [r["period_month"] for r in csv.DictReader(open(pay_path, encoding="utf-8")) if int(r.get("lines_n") or 0) == 0]
        add(
            action_id="A-DATA-ZP-01",
            priority="P2",
            owner_hint="Сливяк / кадры",
            category="DATA_REQUEST",
            title="Ведомости ЗП на месяцы без payroll_lines",
            detail=(
                f"Месяцы без линий: {', '.join(no_lines)}. "
                "DDS↔bank уже CLOSE; ведомости нужны для сверки по сотрудникам."
            ),
            blocks="Payroll person-level SoT",
            evidence_path="live/marts/recon_payroll_multi.csv",
            decision_needed="N",
        )

    # watch priority
    watch = MART / "cost_identity_review_priority.csv"
    n_watch = 0
    if watch.exists():
        n_watch = max(0, sum(1 for _ in open(watch, encoding="utf-8")) - 1)
    if n_watch:
        add(
            action_id="A-FIN-WATCH-01",
            priority="P3",
            owner_hint="Производство / финансы",
            category="SKU_IDENTITY",
            title=f"Ревью {n_watch} SKU с unit≫BOM (без автофикса)",
            detail="Высокий ratio при похожем имени — возможен неполный BOM или неверный FILE cost.",
            blocks="Качество маржи TSUM/IM",
            evidence_path="live/marts/cost_identity_review_priority.csv",
        )

    add(
        action_id="A-TSUM-RATE-01",
        priority="P3",
        owner_hint="Юридический / финансы",
        category="DATA_REQUEST",
        title="Агентский % ЦУМ из договора",
        detail="Сейчас net-rate median 0.4668 (эвристика). Договорной % → почти бухгалтерская сверка.",
        blocks="TSUM model SoT",
        evidence_path="live/marts/recon_tsum_net_model.csv",
        decision_needed="N",
    )

    return actions


def build_b2b_open_request() -> list[dict]:
    rows = []
    for s in csv.DictReader(open(W4 / "settlements.csv", encoding="utf-8")):
        if s.get("channel") != "B2B" or s.get("bank_payment_id"):
            continue
        rows.append(
            {
                "settlement_id": s.get("settlement_id"),
                "period_month": s.get("period_month"),
                "sale_date": s.get("sale_date"),
                "buyer": s.get("buyer"),
                "document": s.get("document"),
                "revenue_rub": s.get("revenue_rub"),
                "request": "Найти платёж / акт взаимозачёта / подтвердить отсутствие в выписке",
            }
        )
    rows.sort(key=lambda r: -(fnum(r.get("revenue_rub")) or 0))
    return rows


def build_finance_evidence() -> list[dict]:
    target = {"0-2497", "0-2496", "0-2493A"}
    rows = []
    for s in csv.DictReader(open(W4 / "sales_lines.csv", encoding="utf-8")):
        if s.get("canonical_sku") not in target:
            continue
        if s.get("dq_exclude_from_margin") == "Y":
            continue
        rev = fnum(s.get("revenue_rub")) or 0
        qty = fnum(s.get("qty")) or 0
        cogs = fnum(s.get("cogs_rub"))
        rows.append(
            {
                "sales_line_id": s.get("sales_line_id"),
                "canonical_sku": s.get("canonical_sku"),
                "channel": s.get("channel"),
                "period_month": s.get("period_month"),
                "sku_name": (s.get("sku_name") or "")[:100],
                "qty": s.get("qty"),
                "revenue_rub": s.get("revenue_rub"),
                "unit_price_rub": round(rev / qty, 2) if qty else "",
                "cogs_rub": s.get("cogs_rub"),
                "unit_cost_rub": round(cogs / qty, 2) if cogs is not None and qty else "",
                "margin_rub": s.get("margin_rub"),
                "cogs_source": s.get("cogs_source"),
            }
        )
    return rows


def build_sku_alias_candidates() -> list[dict]:
    """Агрегат коллизий H9 → кандидат в master alias registry."""
    path = H9 / "cost_collisions_fixed.csv"
    if not path.exists():
        return []
    by = defaultdict(
        lambda: {
            "lines": 0,
            "sale_names": set(),
            "cost_names": set(),
            "fix_types": set(),
            "channels": set(),
            "cost_version_ids": set(),
        }
    )
    for r in csv.DictReader(open(path, encoding="utf-8")):
        sku = r.get("canonical_sku") or ""
        if not sku:
            continue
        b = by[sku]
        b["lines"] += 1
        if r.get("sale_name"):
            b["sale_names"].add(r["sale_name"][:80])
        if r.get("bom_name"):
            b["cost_names"].add(r["bom_name"][:80])
        if r.get("fix_type"):
            b["fix_types"].add(r["fix_type"])
        if r.get("channel"):
            b["channels"].add(r["channel"])
        if r.get("cost_version_id"):
            b["cost_version_ids"].add(r["cost_version_id"])

    rows = []
    for sku, v in sorted(by.items(), key=lambda x: -x[1]["lines"]):
        sale_sample = next(iter(v["sale_names"]), "")
        cost_sample = next(iter(v["cost_names"]), "")
        rows.append(
            {
                "canonical_sku": sku,
                "fixed_lines": v["lines"],
                "channels": ",".join(sorted(v["channels"])),
                "fix_types": "|".join(sorted(v["fix_types"])),
                "sale_name_sample": sale_sample,
                "wrong_or_alt_cost_name_sample": cost_sample,
                "preferred_cost_version_ids": "|".join(sorted(v["cost_version_ids"])),
                "registry_status": "CANDIDATE_ALIAS_RISK",
                "proposed_rule": "Match cost by name/category before unit_cost; prefer W3 BOM over H5 stock on mismatch",
                "so_t": "N",
                "note": "Кандидат в master; Accept только после RACI/производства",
            }
        )
    return rows


def write_owner_xlsx(actions: list[dict], b2b_open: list[dict], evidence: list[dict], aliases: list[dict]):
    wb = Workbook()
    ws = wb.active
    ws.title = "00_Actions"
    ws["A1"] = "YANINA Owner Action Pack (H17)"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A2"] = NOW
    ws["A3"] = "Не SoT. Не заполняет ACCEPT за вас — только чеклист."
    header = list(actions[0].keys()) if actions else []
    ws.append([])
    ws.append(header)
    for r in actions:
        ws.append([r.get(h, "") for h in header])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(40, max(12, len(str(col[0].value or "")) + 2))

    def sheet(name, rows):
        w = wb.create_sheet(name)
        if not rows:
            w["A1"] = "empty"
            return
        w.append(list(rows[0].keys()))
        for r in rows:
            w.append(list(r.values()))

    sheet("B2B_Open_Request", b2b_open)
    sheet("Finance_B2B_Evidence", evidence)
    sheet("SKU_Alias_Candidates", aliases)

    # controls snapshot
    ctrl_path = MART / "controls_summary.csv"
    if ctrl_path.exists():
        ctrl = list(csv.DictReader(open(ctrl_path, encoding="utf-8")))
        sheet("Controls_Now", ctrl)

    path = OUT / "YANINA_OWNER_ACTION_PACK_H17.xlsx"
    wb.save(path)
    wb.save(EV / "YANINA_OWNER_ACTION_PACK_H17.xlsx")
    wb.save(ROOT / "live/YANINA_OWNER_ACTION_PACK_H17.xlsx")
    return path


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)
    MART.mkdir(parents=True, exist_ok=True)

    actions = build_owner_actions()
    b2b_open = build_b2b_open_request()
    evidence = build_finance_evidence()
    aliases = build_sku_alias_candidates()

    write_csv(MART / "owner_actions.csv", actions, list(actions[0].keys()) if actions else ["action_id"])
    write_csv(OUT / "owner_actions.csv", actions, list(actions[0].keys()) if actions else ["action_id"])
    write_csv(MART / "data_request_b2b_open.csv", b2b_open, list(b2b_open[0].keys()) if b2b_open else ["settlement_id"])
    write_csv(MART / "finance_b2b_loss_evidence.csv", evidence, list(evidence[0].keys()) if evidence else ["sales_line_id"])
    write_csv(MART / "sku_alias_candidates.csv", aliases, list(aliases[0].keys()) if aliases else ["canonical_sku"])
    write_csv(OUT / "sku_alias_candidates.csv", aliases, list(aliases[0].keys()) if aliases else ["canonical_sku"])

    xlsx = write_owner_xlsx(actions, b2b_open, evidence, aliases)

    # markdown checklist for quick read
    md = ["# Owner Action Pack (H17)", "", f"Updated: {NOW}", "", "Не SoT. Чеклист решений и запросов данных.", ""]
    for a in actions:
        md.append(f"## [{a['priority']}] {a['action_id']} — {a['title']}")
        md.append(f"- Owner hint: **{a['owner_hint']}**")
        md.append(f"- Category: {a['category']}")
        md.append(f"- {a['detail']}")
        md.append(f"- Blocks: {a['blocks']}")
        if a.get("evidence_path"):
            md.append(f"- Evidence: `{a['evidence_path']}`")
        md.append(f"- Status: `{a['status']}`")
        md.append("")
    (OUT / "OWNER_ACTIONS.md").write_text("\n".join(md), encoding="utf-8")
    (ROOT / "live/OWNER_ACTIONS.md").write_text("\n".join(md), encoding="utf-8")
    shutil.copy2(OUT / "OWNER_ACTIONS.md", EV / "OWNER_ACTIONS.md")

    summary = {
        "wave": "H17",
        "generated_at": NOW,
        "path_choice": "Owner action pack + SKU alias candidates — next value without new bank files/RACI autofill",
        "finding": (
            f"H17: {len(actions)} owner actions; B2B open request {len(b2b_open)}; "
            f"finance evidence lines {len(evidence)}; SKU alias candidates {len(aliases)}."
        ),
        "actions_n": len(actions),
        "by_priority": {
            p: sum(1 for a in actions if a["priority"] == p) for p in ("P0", "P1", "P2", "P3")
        },
        "b2b_open_n": len(b2b_open),
        "b2b_open_rev": round(sum(fnum(r.get("revenue_rub")) or 0 for r in b2b_open), 2),
        "finance_evidence_n": len(evidence),
        "sku_alias_candidates_n": len(aliases),
        "xlsx": str(xlsx.relative_to(ROOT)),
        "not_sot": True,
    }
    (OUT / "h17_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    (EV / "h17_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    for name in (
        "owner_actions.csv",
        "sku_alias_candidates.csv",
        "h17_summary.json",
    ):
        shutil.copy2(OUT / name, EV / name)
    shutil.copy2(MART / "data_request_b2b_open.csv", EV / "data_request_b2b_open.csv")
    shutil.copy2(MART / "finance_b2b_loss_evidence.csv", EV / "finance_b2b_loss_evidence.csv")

    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
