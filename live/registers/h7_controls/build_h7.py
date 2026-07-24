#!/usr/bin/env python3
"""
H7: контроли и срезы без RACI.

1) Card spend categories (Мамушкина / VTB StatementFull)
2) Bank outflow by legal entity (IP / DEKOR / Salon) + Sber/card
3) Пересчёт bank↔DDS с учётом новых источников (с пометкой entity mix)
4) Margin anomaly controls (negative SKU, weak months)

Не SoT.
"""
from __future__ import annotations

import csv
import hashlib
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[3]
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/h7_controls_20260724"
MART = ROOT / "live/marts"
W1 = ROOT / "live/registers/w1_bank_cash"
W4 = ROOT / "live/registers/w4_sales_settle"
H3 = ROOT / "live/registers/h3_new_docs"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")

# категории карт: порядок важен (первое совпадение)
CARD_RULES = [
    ("TRAVEL", re.compile(r"aviasales|kupibilet|aviakassa|airline|aeroflot|s7 |отель|hotel|booking|airbnb", re.I)),
    ("MARKETPLACE", re.compile(r"ozon|wildberries|wb\.|yandex\.market|ym\*", re.I)),
    ("OFFICE_SUPPLY", re.compile(r"komus|канц|office", re.I)),
    ("FABRIC_SUPPLY", re.compile(r"welltex|ткан|textile", re.I)),
    ("LOGISTICS_CUSTOMS", re.compile(r"ast-intern|impeks|импекс|тамож|brok|логист|dhl|cdek|boxberry", re.I)),
    ("TAXI_DELIVERY", re.compile(r"yandex(?!\.market)|taxi|яндекс|delivery|dostav|uber", re.I)),
    ("FOOD", re.compile(r"ресторан|cafe|coffee|okey|вкусвилл|перекрёст|perekrest|mcdonald|kfc|zdorovaya|вода", re.I)),
    ("IT_DIGITAL", re.compile(r"apple|google|microsoft|adobe|domain|hosting", re.I)),
    ("BANK_FEE", re.compile(r"комисси|плата за|obsluzh|обслуживан", re.I)),
    ("OTHER", re.compile(r".*", re.I)),
]


def sha16(*parts) -> str:
    raw = "|".join("" if p is None else str(p) for p in parts)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def fnum(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def categorize_purpose(purpose: str) -> str:
    for name, rx in CARD_RULES:
        if rx.search(purpose or ""):
            return name
    return "OTHER"


# ── Card spend ────────────────────────────────────────────────────
def build_card_spend() -> tuple[list[dict], list[dict], list[dict], dict]:
    src = H3 / "card_payments.csv"
    if not src.exists():
        # fallback from bank
        pays = [
            p
            for p in csv.DictReader(open(W1 / "bank_payments.csv", encoding="utf-8"))
            if p.get("source_bank") == "VTB_CARD"
        ]
    else:
        pays = list(csv.DictReader(open(src, encoding="utf-8")))

    rows = []
    for p in pays:
        if p.get("direction") != "out":
            continue
        cat = categorize_purpose(p.get("purpose") or "")
        rows.append(
            {
                "card_spend_id": "CS-" + sha16(p.get("bank_payment_id")),
                "bank_payment_id": p.get("bank_payment_id"),
                "payment_date": p.get("payment_date"),
                "period_month": p.get("period_month"),
                "amount_rub": p.get("amount"),
                "card_holder": p.get("card_holder") or "",
                "category": cat,
                "purpose": (p.get("purpose") or "")[:160],
                "source_file_id": p.get("source_file_id", ""),
            }
        )

    by_cat = defaultdict(lambda: {"amount": 0.0, "n": 0})
    by_m_cat = defaultdict(lambda: {"amount": 0.0, "n": 0})
    for r in rows:
        amt = fnum(r["amount_rub"]) or 0
        by_cat[r["category"]]["amount"] += amt
        by_cat[r["category"]]["n"] += 1
        by_m_cat[(r["period_month"], r["category"])]["amount"] += amt
        by_m_cat[(r["period_month"], r["category"])]["n"] += 1

    cat_total = [
        {
            "category": c,
            "payments": v["n"],
            "amount_rub": round(v["amount"], 2),
            "share_pct": 0.0,
        }
        for c, v in sorted(by_cat.items(), key=lambda x: -x[1]["amount"])
    ]
    total = sum(c["amount_rub"] for c in cat_total) or 1
    for c in cat_total:
        c["share_pct"] = round(c["amount_rub"] / total * 100, 1)

    month_cat = [
        {
            "period_month": m,
            "category": c,
            "payments": v["n"],
            "amount_rub": round(v["amount"], 2),
        }
        for (m, c), v in sorted(by_m_cat.items())
    ]

    stats = {
        "card_out_payments": len(rows),
        "card_out_rub": round(total if cat_total else 0, 2),
        "by_category": cat_total,
        "holders": dict(Counter(r["card_holder"] or "UNKNOWN" for r in rows)),
    }
    return rows, cat_total, month_cat, stats


# ── Entity bank summary ───────────────────────────────────────────
def build_entity_bank() -> tuple[list[dict], list[dict], dict]:
    pays = list(csv.DictReader(open(W1 / "bank_payments.csv", encoding="utf-8")))
    by = defaultdict(lambda: {"out": 0.0, "in": 0.0, "n_out": 0, "n_in": 0})
    by_src = defaultdict(lambda: {"out": 0.0, "in": 0.0, "n": 0})

    for p in pays:
        if p.get("is_internal") == "Y":
            continue
        le = p.get("legal_entity_id") or "UNKNOWN"
        src = p.get("source_bank") or "UNKNOWN"
        amt = fnum(p.get("amount")) or 0
        if p.get("direction") == "out":
            by[le]["out"] += amt
            by[le]["n_out"] += 1
            by_src[src]["out"] += amt
        else:
            by[le]["in"] += amt
            by[le]["n_in"] += 1
            by_src[src]["in"] += amt
        by_src[src]["n"] += 1

    entity = [
        {
            "legal_entity_id": le,
            "payments_out": v["n_out"],
            "payments_in": v["n_in"],
            "out_rub": round(v["out"], 2),
            "in_rub": round(v["in"], 2),
            "net_rub": round(v["in"] - v["out"], 2),
        }
        for le, v in sorted(by.items())
    ]
    source = [
        {
            "source_bank": s,
            "payments": v["n"],
            "out_rub": round(v["out"], 2),
            "in_rub": round(v["in"], 2),
        }
        for s, v in sorted(by_src.items(), key=lambda x: -x[1]["out"])
    ]
    stats = {
        "entities": len(entity),
        "sources": len(source),
        "total_out": round(sum(e["out_rub"] for e in entity), 2),
        "total_in": round(sum(e["in_rub"] for e in entity), 2),
    }
    return entity, source, stats


# ── Bank vs DDS extended ──────────────────────────────────────────
def recon_bank_dds_extended() -> tuple[list[dict], dict]:
    """Operating bank out (ex-internal) by month vs DDS ledger B; split IP-perimeter vs Salon."""
    pays = list(csv.DictReader(open(W1 / "bank_payments.csv", encoding="utf-8")))
    dds = defaultdict(float)
    if (W1 / "cash_lines.csv").exists():
        for c in csv.DictReader(open(W1 / "cash_lines.csv", encoding="utf-8")):
            if c.get("ledger") != "B":
                continue
            # amount_rub preferred
            amt = fnum(c.get("amount_rub"))
            if amt is None:
                amt = fnum(c.get("amount"))
            if amt is None:
                continue
            # DDS expenses typically positive outflows in B/Nal extracts — use abs of negatives if signed
            dds[c["period_month"]] += abs(amt) if float(amt) < 0 else float(amt)

    # Heuristic: many cash_lines store expense as positive already; if totals explode, keep as-is from W1 recon style
    # Prefer existing recon DDS column if available as baseline
    old = {}
    if (W1 / "recon_bank_vs_dds_month.csv").exists():
        for r in csv.DictReader(open(W1 / "recon_bank_vs_dds_month.csv", encoding="utf-8")):
            old[r["period_month"]] = fnum(r.get("dds_b_bn_rub")) or 0.0

    bank_all = defaultdict(float)
    bank_core = defaultdict(float)  # exclude Salon + card (card is detail of VTB account)
    bank_salon = defaultdict(float)
    bank_card = defaultdict(float)

    for p in pays:
        if p.get("direction") != "out" or p.get("is_internal") == "Y":
            continue
        amt = fnum(p.get("amount")) or 0
        m = p["period_month"]
        bank_all[m] += amt
        if p.get("legal_entity_id") == "LE-OOO-SALON-YANINA" or p.get("source_bank") == "SBER":
            bank_salon[m] += amt
        elif p.get("source_bank") == "VTB_CARD":
            bank_card[m] += amt
            bank_core[m] += amt  # card is IP perimeter
        else:
            bank_core[m] += amt

    rows = []
    for m in sorted(set(bank_all) | set(old) | set(dds)):
        dds_v = old.get(m)
        if dds_v is None:
            dds_v = dds.get(m, 0.0)
        core = bank_core.get(m, 0.0)
        salon = bank_salon.get(m, 0.0)
        card = bank_card.get(m, 0.0)
        delta = core - dds_v
        ratio = abs(delta) / max(core, dds_v, 1)
        if core and dds_v:
            if ratio <= 0.05:
                status = "CLOSE"
            elif ratio <= 0.15:
                status = "SOFT_GAP"
            elif ratio <= 0.30:
                status = "WIDE_GAP"
            else:
                status = "GAP"
        elif core and not dds_v:
            status = "BANK_ONLY"
        elif dds_v and not core:
            status = "DDS_ONLY"
        else:
            status = "N/A"
        rows.append(
            {
                "period_month": m,
                "bank_core_out_rub": round(core, 2),
                "bank_card_out_rub": round(card, 2),
                "bank_salon_out_rub": round(salon, 2),
                "bank_all_out_rub": round(bank_all.get(m, 0.0), 2),
                "dds_b_bn_rub": round(dds_v, 2),
                "delta_core_vs_dds": round(delta, 2),
                "status_core_vs_dds": status,
                "note": "core=Alfa+VTB+card (IP/DEKOR perimeter); Salon Sber separated — not in classic DDS",
            }
        )

    st = Counter(r["status_core_vs_dds"] for r in rows if r["status_core_vs_dds"] != "N/A")
    stats = {
        "months": len(rows),
        "status": dict(st),
        "close_soft": [r["period_month"] for r in rows if r["status_core_vs_dds"] in ("CLOSE", "SOFT_GAP")],
        "salon_out_total": round(sum(bank_salon.values()), 2),
        "card_out_total": round(sum(bank_card.values()), 2),
    }
    return rows, stats


# ── Margin anomalies ──────────────────────────────────────────────
def margin_controls() -> tuple[list[dict], list[dict], dict]:
    anomalies = []
    # months
    for r in csv.DictReader(open(MART / "margin_channel_month.csv", encoding="utf-8")):
        pct = fnum(r.get("margin_pct"))
        cov = fnum(r.get("cogs_coverage")) or 0
        flags = []
        if r.get("status") == "WEAK":
            flags.append("WEAK_COGS_COVERAGE")
        if pct is not None and pct < 0:
            flags.append("NEGATIVE_MARGIN")
        if pct is not None and 0 <= pct < 15 and cov >= 0.8:
            flags.append("LOW_MARGIN")
        if flags:
            anomalies.append(
                {
                    "scope": "CHANNEL_MONTH",
                    "channel": r["channel"],
                    "period_month": r["period_month"],
                    "canonical_sku": "",
                    "revenue_rub": r.get("revenue_rub", ""),
                    "margin_pct": r.get("margin_pct", ""),
                    "cogs_coverage": r.get("cogs_coverage", ""),
                    "flags": "|".join(flags),
                    "priority": "HIGH" if "NEGATIVE_MARGIN" in flags else "MED",
                }
            )

    sku_neg = []
    if (MART / "margin_sku_bottom40.csv").exists():
        for r in csv.DictReader(open(MART / "margin_sku_bottom40.csv", encoding="utf-8")):
            pct = fnum(r.get("margin_pct"))
            if pct is not None and pct < 0:
                sku_neg.append(
                    {
                        "scope": "SKU",
                        "channel": r.get("channels", ""),
                        "period_month": "",
                        "canonical_sku": r.get("canonical_sku", ""),
                        "revenue_rub": r.get("revenue_rub", ""),
                        "margin_pct": r.get("margin_pct", ""),
                        "cogs_coverage": "",
                        "flags": "NEGATIVE_MARGIN",
                        "priority": "HIGH",
                        "name": r.get("name", "")[:100],
                        "margin_rub": r.get("margin_rub", ""),
                    }
                )
                anomalies.append(
                    {
                        "scope": "SKU",
                        "channel": r.get("channels", ""),
                        "period_month": "",
                        "canonical_sku": r.get("canonical_sku", ""),
                        "revenue_rub": r.get("revenue_rub", ""),
                        "margin_pct": r.get("margin_pct", ""),
                        "cogs_coverage": "",
                        "flags": "NEGATIVE_MARGIN",
                        "priority": "HIGH",
                    }
                )

    # also scan all sales for negative margin lines aggregation already in bottom40
    stats = {
        "anomaly_rows": len(anomalies),
        "negative_sku": len(sku_neg),
        "channel_month_flags": sum(1 for a in anomalies if a["scope"] == "CHANNEL_MONTH"),
        "high_priority": sum(1 for a in anomalies if a["priority"] == "HIGH"),
    }
    return anomalies, sku_neg, stats


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)
    MART.mkdir(exist_ok=True)

    card_rows, cat_total, month_cat, card_stats = build_card_spend()
    entity, source, ent_stats = build_entity_bank()
    recon, recon_stats = recon_bank_dds_extended()
    anomalies, sku_neg, anom_stats = margin_controls()

    write_csv(
        OUT / "card_spend_lines.csv",
        card_rows,
        list(card_rows[0].keys()) if card_rows else ["card_spend_id"],
    )
    write_csv(OUT / "card_spend_by_category.csv", cat_total, list(cat_total[0].keys()) if cat_total else ["category"])
    write_csv(OUT / "card_spend_month_category.csv", month_cat, list(month_cat[0].keys()) if month_cat else ["period_month"])
    write_csv(MART / "card_spend_by_category.csv", cat_total, list(cat_total[0].keys()) if cat_total else ["category"])

    write_csv(OUT / "bank_by_entity.csv", entity, list(entity[0].keys()) if entity else ["legal_entity_id"])
    write_csv(OUT / "bank_by_source.csv", source, list(source[0].keys()) if source else ["source_bank"])
    write_csv(MART / "bank_by_entity.csv", entity, list(entity[0].keys()) if entity else ["legal_entity_id"])

    write_csv(OUT / "recon_bank_dds_extended.csv", recon, list(recon[0].keys()) if recon else ["period_month"])
    write_csv(W1 / "recon_bank_vs_dds_extended.csv", recon, list(recon[0].keys()) if recon else ["period_month"])
    write_csv(MART / "recon_bank_dds_extended.csv", recon, list(recon[0].keys()) if recon else ["period_month"])

    write_csv(
        OUT / "margin_anomalies.csv",
        anomalies,
        ["scope", "channel", "period_month", "canonical_sku", "revenue_rub", "margin_pct", "cogs_coverage", "flags", "priority"],
    )
    write_csv(
        MART / "margin_anomalies.csv",
        anomalies,
        ["scope", "channel", "period_month", "canonical_sku", "revenue_rub", "margin_pct", "cogs_coverage", "flags", "priority"],
    )
    write_csv(
        MART / "margin_negative_skus.csv",
        sku_neg,
        list(sku_neg[0].keys()) if sku_neg else ["canonical_sku"],
    )

    summary = {
        "generated_at": NOW,
        "wave": "H7",
        "card_spend": card_stats,
        "entity_bank": ent_stats,
        "entity_rows": entity,
        "source_rows": source,
        "recon_extended": recon_stats,
        "margin_controls": anom_stats,
        "finding": (
            f"H7: card out {card_stats['card_out_rub']:.0f} RUB "
            f"({', '.join(c['category']+':'+str(c['share_pct'])+'%' for c in cat_total[:4])}); "
            f"core↔DDS CLOSE/SOFT {len(recon_stats['close_soft'])}; "
            f"Salon out {recon_stats['salon_out_total']:.0f}; "
            f"margin anomalies {anom_stats['anomaly_rows']} (neg SKU {anom_stats['negative_sku']})."
        ),
        "next": "RACI ACCEPT; review negative-margin SKUs with finance",
        "not_sot": True,
    }
    json.dump(summary, open(OUT / "h7_summary.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    json.dump(summary, open(EV / "h7_summary.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    json.dump(summary, open(MART / "h7_summary.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)

    wb = Workbook()
    ws = wb.active
    ws.title = "00_Summary"
    for i, (k, v) in enumerate(summary.items(), 1):
        ws.cell(i, 1, k)
        ws.cell(i, 2, json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v)
    fill = PatternFill("solid", fgColor="1F4E79")

    def add(name, rows):
        w = wb.create_sheet(name)
        if not rows:
            return
        headers = list(rows[0].keys())
        for c, h in enumerate(headers, 1):
            cell = w.cell(1, c, h)
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
        for ri, row in enumerate(rows[:4000], 2):
            for ci, h in enumerate(headers, 1):
                v = row.get(h, "")
                if isinstance(v, (dict, list)):
                    v = json.dumps(v, ensure_ascii=False)
                w.cell(ri, ci, v)

    add("01_Card_Categories", cat_total)
    add("02_Entity_Bank", entity)
    add("03_Bank_Source", source)
    add("04_Recon_Ext", recon)
    add("05_Anomalies", anomalies)
    add("06_Neg_SKU", sku_neg)
    wb.save(EV / "YANINA_H7_CONTROLS_EVIDENCE.xlsx")

    (OUT / "README.md").write_text(
        f"""# H7 Controls

{NOW}

{summary['finding']}

Marts: `../../marts/`
Evidence: `../../evidence/h7_controls_20260724/`
""",
        encoding="utf-8",
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
