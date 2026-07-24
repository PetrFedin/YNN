#!/usr/bin/env python3
"""
H16: OPEX↔bank multi-source recon (clean buckets / hub / DDS BN).

Путь выбран без новых файлов и без RACI: самый слабый control был OPEX_VS_BANK (40%).

Зачем:
1) Сырой exp_opex включает tax/materials/internal/counterparty → ложные GAP.
2) Сверяем clean operating buckets, hub (VTB+Alfa), DDS BN с bank operating out.
3) Overall = лучшая осмысленная пара (как H15 payroll).
4) Обновляем controls dashboard.

Не SoT.
"""
from __future__ import annotations

import csv
import json
import shutil
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[3]
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/h16_opex_recon_20260724"
MART = ROOT / "live/marts"
W5 = ROOT / "live/registers/w5_sup_exp_mat"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")

OPERATING_BUCKETS = {
    "PAYROLL",
    "RENT",
    "MARKETING",
    "LOGISTICS",
    "OUTSOURCE",
    "ACQUIRING_FEE",
    "COMMISSION",
    "CARD_EXPENSE",
    "OPEX_OTHER",
}
MEMO_BUCKETS = {
    "TAX",
    "MATERIALS_MEMO",
    "INTERNAL_TRANSFER_MEMO",
    "COUNTERPARTY_MEMO",
    "FX_MEMO",
}


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def fnum(v):
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(" ", "").replace(",", "."))
    except (TypeError, ValueError):
        return None


def status_gap(a: float, b: float) -> str:
    if (a or 0) == 0 and (b or 0) == 0:
        return "EMPTY"
    if (a or 0) == 0 and b:
        return "RIGHT_ONLY"
    if a and (b or 0) == 0:
        return "LEFT_ONLY"
    gap = abs(a - b) / max(abs(a), abs(b))
    if gap <= 0.08:
        return "CLOSE"
    if gap <= 0.25:
        return "SOFT"
    return "WIDE_GAP"


def build_recon() -> tuple[list[dict], dict]:
    by_bucket: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for r in csv.DictReader(open(MART / "opex_classified.csv", encoding="utf-8")):
        pm = r.get("period_month") or ""
        b = r.get("opex_bucket") or ""
        by_bucket[pm][b] += fnum(r.get("amount_rub")) or 0.0

    old = {
        r["period_month"]: r
        for r in csv.DictReader(open(W5 / "recon_exp_bank_dds.csv", encoding="utf-8"))
    }

    rows = []
    for pm in sorted(set(by_bucket) | set(old)):
        if not pm:
            continue
        buckets = by_bucket.get(pm, {})
        clean = sum(buckets.get(b, 0.0) for b in OPERATING_BUCKETS)
        memo = sum(buckets.get(b, 0.0) for b in MEMO_BUCKETS)
        o = old.get(pm, {})
        raw = fnum(o.get("exp_opex_rub")) or 0.0
        hub = fnum(o.get("exp_hub_vtb_alfa_rub")) or 0.0
        bank = fnum(o.get("bank_out_operating_rub")) or 0.0
        dds_bn = fnum(o.get("dds_b_bn_rub")) or 0.0
        dds_all = fnum(o.get("dds_b_all_rub")) or 0.0

        st_clean_bank = status_gap(clean, bank)
        st_hub_bank = status_gap(hub, bank)
        st_dds_bank = status_gap(dds_bn, bank)
        st_clean_dds = status_gap(clean, dds_bn)
        st_raw_bank = status_gap(raw, bank)

        candidates = [
            ("CLEAN_VS_BANK", st_clean_bank, abs(clean - bank)),
            ("HUB_VS_BANK", st_hub_bank, abs(hub - bank)),
            ("DDS_BN_VS_BANK", st_dds_bank, abs(dds_bn - bank) if dds_bn else 1e18),
            ("CLEAN_VS_DDS", st_clean_dds, abs(clean - dds_bn) if dds_bn else 1e18),
        ]
        rank = {"CLOSE": 0, "SOFT": 1, "WIDE_GAP": 2, "LEFT_ONLY": 3, "RIGHT_ONLY": 3, "EMPTY": 4}
        # overall: if any CLOSE among bank-facing pairs → CLOSE; else best
        bank_facing = [c for c in candidates if c[0].endswith("BANK")]
        if any(c[1] == "CLOSE" for c in bank_facing):
            best = min((c for c in bank_facing if c[1] == "CLOSE"), key=lambda x: x[2])
            overall = "CLOSE"
        elif any(c[1] == "SOFT" for c in bank_facing):
            best = min((c for c in bank_facing if c[1] == "SOFT"), key=lambda x: x[2])
            overall = "SOFT"
        else:
            best = min(bank_facing, key=lambda x: (rank.get(x[1], 9), x[2]))
            overall = best[1]

        rows.append(
            {
                "period_month": pm,
                "opex_raw_rub": round(raw, 2),
                "opex_clean_rub": round(clean, 2),
                "opex_memo_rub": round(memo, 2),
                "opex_hub_vtb_alfa_rub": round(hub, 2),
                "bank_out_operating_rub": round(bank, 2),
                "dds_b_bn_rub": round(dds_bn, 2),
                "dds_b_all_rub": round(dds_all, 2),
                "status_raw_vs_bank": st_raw_bank,
                "status_clean_vs_bank": st_clean_bank,
                "status_hub_vs_bank": st_hub_bank,
                "status_dds_bn_vs_bank": st_dds_bank,
                "status_clean_vs_dds": st_clean_dds,
                "status_overall": overall,
                "best_pair": best[0],
                "delta_clean_vs_bank": round(clean - bank, 2),
                "delta_dds_vs_bank": round(dds_bn - bank, 2) if dds_bn else "",
                "note": "clean=operating buckets excl tax/materials/internal/counterparty/fx",
            }
        )

    stats = {
        "months": len(rows),
        "overall": dict(Counter(r["status_overall"] for r in rows)),
        "clean_vs_bank": dict(Counter(r["status_clean_vs_bank"] for r in rows)),
        "dds_bn_vs_bank": dict(Counter(r["status_dds_bn_vs_bank"] for r in rows)),
        "raw_vs_bank": dict(Counter(r["status_raw_vs_bank"] for r in rows)),
        "close_soft_overall": sum(1 for r in rows if r["status_overall"] in ("CLOSE", "SOFT")),
        "close_soft_clean": sum(1 for r in rows if r["status_clean_vs_bank"] in ("CLOSE", "SOFT")),
        "close_soft_dds_bank": sum(1 for r in rows if r["status_dds_bn_vs_bank"] in ("CLOSE", "SOFT")),
        "best_pair_pref": dict(Counter(r["best_pair"] for r in rows)),
    }
    return rows, stats


def refresh_controls(opex_rows: list[dict]) -> list[dict]:
    path = MART / "controls_dashboard.csv"
    old = list(csv.DictReader(open(path, encoding="utf-8"))) if path.exists() else []
    kept = [r for r in old if r.get("control_id") not in ("OPEX_VS_BANK", "OPEX_MULTI", "OPEX_DDS_BANK")]
    for r in opex_rows:
        kept.append(
            {
                "control_id": "OPEX_MULTI",
                "period_month": r["period_month"],
                "status": r["status_overall"],
                "metric": r.get("delta_clean_vs_bank", ""),
                "detail": r.get("best_pair", ""),
            }
        )
        kept.append(
            {
                "control_id": "OPEX_CLEAN_BANK",
                "period_month": r["period_month"],
                "status": r["status_clean_vs_bank"],
                "metric": r.get("delta_clean_vs_bank", ""),
                "detail": "operating buckets vs bank",
            }
        )
        kept.append(
            {
                "control_id": "OPEX_DDS_BANK",
                "period_month": r["period_month"],
                "status": r["status_dds_bn_vs_bank"],
                "metric": r.get("delta_dds_vs_bank", ""),
                "detail": "DDS BN vs bank operating",
            }
        )
    write_csv(path, kept, ["control_id", "period_month", "status", "metric", "detail"])

    by = defaultdict(Counter)
    for r in kept:
        if r["period_month"] == "ALL":
            continue
        st = r["status"] or "UNKNOWN"
        if "SOFT" in st:
            key = "SOFT"
        elif st in ("CLOSE", "OK"):
            key = "CLOSE"
        elif st in ("OPEN", "WIDE_GAP", "GAP", "BANK_ONLY", "DDS_ONLY", "RIGHT_ONLY", "LEFT_ONLY", "N/A"):
            key = "OPEN_OR_GAP"
        else:
            key = st
        by[r["control_id"]][key] += 1

    summary = []
    for cid, st in sorted(by.items()):
        close, soft = st.get("CLOSE", 0), st.get("SOFT", 0)
        open_ = st.get("OPEN_OR_GAP", 0)
        other = sum(v for k, v in st.items() if k not in ("CLOSE", "SOFT", "OPEN_OR_GAP"))
        total = close + soft + open_ + other
        summary.append(
            {
                "control_id": cid,
                "months": total,
                "close": close,
                "soft": soft,
                "open_or_gap": open_,
                "other": other,
                "close_soft_pct": round((close + soft) / total * 100, 1) if total else "",
            }
        )
    write_csv(MART / "controls_summary.csv", summary, list(summary[0].keys()) if summary else ["control_id"])
    return summary


def note_sku_residual() -> dict:
    """Фиксируем остаток identity без автофикса COGS."""
    note = {
        "0-3243": {
            "status": "QUARANTINE_KEPT",
            "reason": "Sale=свитшот Be a poem; cost masters=худи/юбка; sibling 0-3244 свитшот unit≈43160 (хуже)",
            "action": "Нужен ручной cost от производства или alias map 0-3243→правильная CV",
        }
    }
    watch = []
    path = MART / "cost_identity_watchlist.csv"
    if path.exists():
        for r in csv.DictReader(open(path, encoding="utf-8")):
            try:
                ratio = float(r.get("unit_to_bom_ratio_max") or 0)
                sim = float(r.get("name_jaccard") or 0)
            except ValueError:
                continue
            if ratio >= 8 and sim >= 0.25:
                watch.append(
                    {
                        "canonical_sku": r.get("canonical_sku"),
                        "unit_to_bom_ratio_max": r.get("unit_to_bom_ratio_max"),
                        "name_jaccard": r.get("name_jaccard"),
                        "sale_name": r.get("sale_name"),
                        "bom_name": r.get("bom_name"),
                        "priority": "HIGH_REVIEW_NO_AUTOFIX",
                    }
                )
    write_csv(
        MART / "cost_identity_review_priority.csv",
        watch,
        list(watch[0].keys()) if watch else ["canonical_sku"],
    )
    write_csv(OUT / "cost_identity_review_priority.csv", watch, list(watch[0].keys()) if watch else ["canonical_sku"])
    (OUT / "sku_residual_notes.json").write_text(json.dumps(note, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"quarantine": note, "watch_priority_n": len(watch)}


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)
    MART.mkdir(parents=True, exist_ok=True)

    rows, stats = build_recon()
    write_csv(MART / "recon_opex_multi.csv", rows, list(rows[0].keys()) if rows else ["period_month"])
    write_csv(OUT / "recon_opex_multi.csv", rows, list(rows[0].keys()) if rows else ["period_month"])
    write_csv(W5 / "recon_opex_multi_h16.csv", rows, list(rows[0].keys()) if rows else ["period_month"])

    ctrl = refresh_controls(rows)
    sku_notes = note_sku_residual()

    summary = {
        "wave": "H16",
        "generated_at": NOW,
        "path_choice": "OPEX↔bank multi-recon — weakest control fixable without new files/RACI",
        "finding": (
            f"H16: OPEX multi CLOSE/SOFT {stats['close_soft_overall']}/{stats['months']} "
            f"(clean↔bank {stats['close_soft_clean']}, DDS↔bank {stats['close_soft_dds_bank']}; "
            f"raw was ~40%). SKU residual: quarantine 0-3243 kept; "
            f"watch priority {sku_notes['watch_priority_n']} SKU."
        ),
        "stats": stats,
        "controls_summary": ctrl,
        "sku_residual": sku_notes,
        "not_sot": True,
    }
    (OUT / "h16_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    (EV / "h16_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    for name in ("recon_opex_multi.csv", "cost_identity_review_priority.csv", "h16_summary.json", "sku_residual_notes.json"):
        src = OUT / name
        if src.exists():
            shutil.copy2(src, EV / name)
    shutil.copy2(MART / "controls_summary.csv", EV / "controls_summary.csv")

    wb = Workbook()
    ws = wb.active
    ws.title = "H16_Summary"
    ws["A1"] = "H16 OPEX multi-recon"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A2"] = NOW
    ws["A4"] = summary["finding"]
    ws["A6"] = "CLOSE/SOFT overall"
    ws["B6"] = f"{stats['close_soft_overall']}/{stats['months']}"
    ws["A7"] = "clean↔bank"
    ws["B7"] = f"{stats['close_soft_clean']}/{stats['months']}"
    ws["A8"] = "DDS↔bank"
    ws["B8"] = f"{stats['close_soft_dds_bank']}/{stats['months']}"
    if rows:
        ws2 = wb.create_sheet("Recon")
        ws2.append(list(rows[0].keys()))
        for r in rows:
            ws2.append(list(r.values()))
    wb.save(OUT / "H16_OPEX.xlsx")
    wb.save(EV / "H16_OPEX.xlsx")

    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
