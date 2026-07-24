#!/usr/bin/env python3
"""
H1: усиление горизонтальных связей после W1–W6 (Controlled Staging).

1) SUP.inn → bank.counterparty_inn — P2P идентификация поставщика в платежах
2) SETTLE.document № → bank.purpose — O2C по номеру расходной накладной

Не SoT: не заменяет RACI. Не выдумывает owners.
"""
from __future__ import annotations

import csv
import hashlib
import json
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[3]
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/h1_spine_links_20260724"
W1 = ROOT / "live/registers/w1_bank_cash"
W4 = ROOT / "live/registers/w4_sales_settle"
W5 = ROOT / "live/registers/w5_sup_exp_mat"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")


def sha16(*parts) -> str:
    raw = "|".join("" if p is None else str(p) for p in parts)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def nfc(s: str) -> str:
    return unicodedata.normalize("NFC", s or "")


# ── SUP ↔ BANK by INN ─────────────────────────────────────────────
def link_sup_bank() -> tuple[list[dict], list[dict], dict]:
    suppliers = list(csv.DictReader(open(W5 / "suppliers.csv", encoding="utf-8")))
    bank = list(csv.DictReader(open(W1 / "bank_payments.csv", encoding="utf-8")))

    by_inn: dict[str, dict] = {}
    for s in suppliers:
        inn = (s.get("inn") or "").strip()
        if inn:
            by_inn[inn] = s

    bank_by_inn: dict[str, list] = defaultdict(list)
    for p in bank:
        if p.get("is_internal") == "Y":
            continue
        inn = (p.get("counterparty_inn") or "").strip()
        if inn:
            bank_by_inn[inn].append(p)

    links = []
    for inn, sup in by_inn.items():
        pays = bank_by_inn.get(inn, [])
        out_pays = [p for p in pays if p.get("direction") == "out"]
        in_pays = [p for p in pays if p.get("direction") == "in"]
        out_amt = sum(float(p["amount"] or 0) for p in out_pays)
        in_amt = sum(float(p["amount"] or 0) for p in in_pays)
        status = "MATCHED" if pays else "SUP_INN_NO_BANK"
        links.append(
            {
                "link_id": "SL-" + sha16(inn, sup["counterparty_id"]),
                "counterparty_id": sup["counterparty_id"],
                "supplier_name": sup["name"],
                "inn": inn,
                "status": status,
                "bank_payment_count": len(pays),
                "bank_out_count": len(out_pays),
                "bank_in_count": len(in_pays),
                "bank_out_rub": round(out_amt, 2),
                "bank_in_rub": round(in_amt, 2),
                "sample_bank_payment_id": (out_pays or in_pays or [{"bank_payment_id": ""}])[0]["bank_payment_id"],
                "sample_counterparty_raw": (out_pays or in_pays or [{"counterparty_raw": ""}])[0].get("counterparty_raw", ""),
            }
        )

    # payment-level edges for matched OUT
    edges = []
    for inn, pays in bank_by_inn.items():
        if inn not in by_inn:
            continue
        sup = by_inn[inn]
        for p in pays:
            if p.get("direction") != "out":
                continue
            edges.append(
                {
                    "edge_id": "SE-" + sha16(p["bank_payment_id"], sup["counterparty_id"]),
                    "counterparty_id": sup["counterparty_id"],
                    "supplier_name": sup["name"],
                    "inn": inn,
                    "bank_payment_id": p["bank_payment_id"],
                    "payment_date": p["payment_date"],
                    "period_month": p["period_month"],
                    "amount": p["amount"],
                    "purpose": (p.get("purpose") or "")[:120],
                    "source_bank": p.get("source_bank", ""),
                }
            )

    stats = {
        "suppliers_total": len(suppliers),
        "suppliers_with_inn": len(by_inn),
        "matched_inn": sum(1 for r in links if r["status"] == "MATCHED"),
        "unmatched_inn": sum(1 for r in links if r["status"] == "SUP_INN_NO_BANK"),
        "payment_edges": len(edges),
        "matched_out_rub": round(sum(r["bank_out_rub"] for r in links if r["status"] == "MATCHED"), 2),
    }
    return links, edges, stats


# ── SETTLE ↔ BANK by document number ──────────────────────────────
DOC_NUM_RE = re.compile(
    r"(?:расходн\w*\s+накладн\w*|накладн\w*|рн|р/н)\s*[№#]?\s*(\d{1,6})",
    re.I,
)
PURPOSE_DOC_RE = re.compile(
    r"(?:накладн\w*|рн|р/н|сч\.?|счет|счёт|invoice|инв\.?)\s*[№#]?\s*(\d{1,6})",
    re.I,
)
# also bare "по сч. 122" already common; for B2B sales use накладная numbers


def extract_doc_num(document: str) -> str:
    m = DOC_NUM_RE.search(nfc(document))
    if m:
        return m.group(1).lstrip("0") or m.group(1)
    # fallback: last standalone number before "от"
    m = re.search(r"(\d{1,6})\s+от\s+\d", nfc(document))
    if m:
        return m.group(1).lstrip("0") or m.group(1)
    return ""


def _norm_name(s: str) -> str:
    s = nfc(s).lower()
    s = re.sub(r'[\"«»\'`]', "", s)
    s = re.sub(r"\b(ооо|ип|зао|оао|ао|пао)\b", " ", s)
    s = re.sub(r"[^a-zа-я0-9]+", " ", s, flags=re.I)
    return re.sub(r"\s+", " ", s).strip()


def _name_overlap(a: str, b: str) -> float:
    ta = set(_norm_name(a).split())
    tb = set(_norm_name(b).split())
    if not ta or not tb:
        return 0.0
    return len(ta & tb) / max(len(ta), len(tb))


def link_settle_bank() -> tuple[list[dict], dict]:
    settles = list(csv.DictReader(open(W4 / "settlements.csv", encoding="utf-8")))
    bank = [
        p
        for p in csv.DictReader(open(W1 / "bank_payments.csv", encoding="utf-8"))
        if p.get("direction") == "in" and p.get("is_internal") != "Y"
    ]

    by_doc: dict[str, list] = defaultdict(list)
    for p in bank:
        purpose = p.get("purpose") or ""
        for m in PURPOSE_DOC_RE.finditer(purpose):
            num = m.group(1).lstrip("0") or m.group(1)
            by_doc[num].append(p)

    by_month: dict[str, list] = defaultdict(list)
    for p in bank:
        by_month[p["period_month"]].append(p)

    matches = []
    used_payments = set()

    def add_match(st, p, method, conf, extra=None):
        if p["bank_payment_id"] in used_payments:
            return False
        used_payments.add(p["bank_payment_id"])
        rev = float(st.get("revenue_rub") or 0)
        row = {
            "match_id": "SM-" + sha16(st["settlement_id"], p["bank_payment_id"], method),
            "settlement_id": st["settlement_id"],
            "document": st["document"],
            "doc_num": extract_doc_num(st.get("document") or ""),
            "buyer": st.get("buyer", ""),
            "revenue_rub": st.get("revenue_rub", ""),
            "period_month": st["period_month"],
            "bank_payment_id": p["bank_payment_id"],
            "payment_date": p["payment_date"],
            "bank_amount": p["amount"],
            "counterparty_raw": p.get("counterparty_raw", ""),
            "purpose": (p.get("purpose") or "")[:140],
            "amount_delta": round(rev - float(p["amount"] or 0), 2),
            "match_method": method,
            "confidence": conf,
        }
        if extra:
            row.update(extra)
        matches.append(row)
        return True

    # Pass 1: doc number in purpose
    for st in settles:
        doc_num = extract_doc_num(st.get("document") or "")
        if not doc_num:
            continue
        rev = float(st.get("revenue_rub") or 0)
        scored = []
        for p in by_doc.get(doc_num, []):
            if p["bank_payment_id"] in used_payments:
                continue
            amt = float(p["amount"] or 0)
            same_month = p["period_month"] == st["period_month"]
            ratio = abs(amt - rev) / max(amt, rev, 1)
            conf = "HIGH" if same_month and ratio <= 0.02 else (
                "MED" if same_month and ratio <= 0.10 else (
                    "LOW" if ratio <= 0.02 else None
                )
            )
            if conf:
                scored.append((conf, ratio, p))
        if not scored:
            continue
        order = {"HIGH": 0, "MED": 1, "LOW": 2}
        scored.sort(key=lambda x: (order[x[0]], x[1]))
        conf, _, p = scored[0]
        add_match(st, p, "doc_num_in_purpose", conf)

    # Pass 2: same month + amount ±2% + buyer name overlap
    for st in settles:
        rev = float(st.get("revenue_rub") or 0)
        if rev <= 0:
            continue
        buyer = st.get("buyer") or ""
        best = None
        for p in by_month.get(st["period_month"], []):
            if p["bank_payment_id"] in used_payments:
                continue
            amt = float(p["amount"] or 0)
            ratio = abs(amt - rev) / max(amt, rev, 1)
            if ratio > 0.02:
                continue
            ov = _name_overlap(buyer, p.get("counterparty_raw") or "")
            if ov < 0.4:
                continue
            cand = (ov, ratio, p)
            if best is None or cand > best:
                best = cand
        if best:
            ov, _, p = best
            conf = "HIGH" if ov >= 0.6 else "MED"
            add_match(st, p, "amount_name_same_month", conf, {"name_overlap": round(ov, 3)})

    stats = {
        "settlements": len(settles),
        "with_doc_num": sum(1 for st in settles if extract_doc_num(st.get("document") or "")),
        "bank_in_with_doc_like": sum(len(v) for v in by_doc.values()),
        "matches_total": len(matches),
        "by_confidence": dict(Counter(m["confidence"] for m in matches)),
        "by_method": dict(Counter(m["match_method"] for m in matches)),
        "prev_exact_amount_matches": 31,
    }
    return matches, stats


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)

    links, edges, sup_stats = link_sup_bank()
    settle_matches, settle_stats = link_settle_bank()

    write_csv(
        OUT / "sup_bank_by_inn.csv",
        links,
        [
            "link_id", "counterparty_id", "supplier_name", "inn", "status",
            "bank_payment_count", "bank_out_count", "bank_in_count",
            "bank_out_rub", "bank_in_rub", "sample_bank_payment_id", "sample_counterparty_raw",
        ],
    )
    write_csv(
        OUT / "sup_bank_payment_edges.csv",
        edges,
        [
            "edge_id", "counterparty_id", "supplier_name", "inn", "bank_payment_id",
            "payment_date", "period_month", "amount", "purpose", "source_bank",
        ],
    )
    write_csv(
        OUT / "settle_bank_by_doc.csv",
        settle_matches,
        [
            "match_id", "settlement_id", "document", "doc_num", "buyer", "revenue_rub",
            "period_month", "bank_payment_id", "payment_date", "bank_amount",
            "counterparty_raw", "purpose", "amount_delta", "match_method", "confidence",
            "name_overlap",
        ],
    )

    summary = {
        "generated_at": NOW,
        "wave": "H1",
        "purpose": "spine connectivity hardening after W1-W6",
        "sup_bank": sup_stats,
        "settle_bank": settle_stats,
        "finding": (
            f"H1: SUP.inn↔bank {sup_stats['matched_inn']}/{sup_stats['suppliers_with_inn']} INN, "
            f"{sup_stats['payment_edges']} out edges (~{sup_stats['matched_out_rub']:.0f} RUB); "
            f"SETTLE↔bank: {settle_stats['matches_total']} "
            f"({settle_stats['by_method']}, conf {settle_stats['by_confidence']})."
        ),
        "next": "Owner Packet RACI (ST24-G01) — required for SoT; then domain data requests",
        "not_sot": True,
    }
    json.dump(summary, open(OUT / "h1_summary.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    json.dump(summary, open(EV / "h1_summary.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)

    wb = Workbook()
    ws = wb.active
    ws.title = "00_Summary"
    for i, (k, v) in enumerate(summary.items(), 1):
        ws.cell(i, 1, k)
        ws.cell(i, 2, json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v)
    fill = PatternFill("solid", fgColor="1F4E79")

    def add(name, rows):
        w = wb.create_sheet(name)
        if not rows:
            return
        headers = list(rows[0].keys())
        for c, h in enumerate(headers, 1):
            cell = w.cell(1, c, h)
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
        for ri, row in enumerate(rows[:5000], 2):
            for ci, h in enumerate(headers, 1):
                w.cell(ri, ci, row.get(h, ""))

    add("01_SUP_BANK_INN", links)
    add("02_SETTLE_DOC", settle_matches)
    add("03_SUP_EDGES", edges[:3000])
    wb.save(EV / "YANINA_H1_SPINE_LINKS_EVIDENCE.xlsx")

    (OUT / "README.md").write_text(
        f"""# H1 Spine Links (post W1–W6)

Generated: {NOW}

Усиление связей без претензии на SoT.

- `sup_bank_by_inn.csv` — {sup_stats['matched_inn']}/{sup_stats['suppliers_with_inn']} SUP с ИНН найдены в банке
- `sup_bank_payment_edges.csv` — {sup_stats['payment_edges']} OUT-платежей
- `settle_bank_by_doc.csv` — {settle_stats['matches_total']} settle↔bank по № документа

Evidence: `../../evidence/h1_spine_links_20260724/`

Next: заполнить Owner Packet RACI.
""",
        encoding="utf-8",
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
