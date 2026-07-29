#!/usr/bin/env python3
"""H84: mat_movement document dates ↔ fabric inventory aging (P2).

Движение тканей.xlsx содержит документные строки с датами — H4/H67 брали
только SKU-итоги. H84 строит aging по last in/out и кроссит с остатками 31.05.2026.
Не SoT. Не auto-Accept. Leaf qty/amount в движении ambiguous — деньги из snapshot.
"""
from __future__ import annotations

import csv
import json
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[3]
REG = Path(__file__).resolve().parent
MARTS = ROOT / "live/marts"
DL = Path("/Users/petr/Downloads/YANINA документы")

AS_OF = datetime(2026, 7, 29)  # today / analysis date
STOCK_AS_OF = "2026-05-31"

DATE_RE = re.compile(r"от\s+(\d{2}\.\d{2}\.\d{4})")
DATE_ANY = re.compile(r"(\d{2}\.\d{2}\.\d{4})")
CODE_RE = re.compile(r"(\d{1,3}-\d{3,}[A-Za-zА-Яа-я]?)\s*$")

DOC_IN_KEYS = ("оприход", "приходн", "поступлен")
DOC_OUT_KEYS = ("списан", "расходн", "реализа")
# перемещение: направление по колонкам приход/расход


def resolve(substr: str) -> Path:
    for r in csv.DictReader(open(ROOT / "live/registers/00_SOURCE_CATALOG_107.csv")):
        if substr.lower() in r["file_name"].lower():
            p = Path(r["path"])
            if not p.exists():
                p = DL / r["file_name"]
            if p.exists():
                return p
    raise FileNotFoundError(substr)


def parse_date(s: str) -> datetime | None:
    m = DATE_RE.search(s) or DATE_ANY.search(s)
    if not m:
        return None
    try:
        return datetime.strptime(m.group(1), "%d.%m.%Y")
    except ValueError:
        return None


def article_hint(name: str) -> str:
    m = CODE_RE.search(name.strip())
    return m.group(1) if m else ""


def is_warehouse(name: str) -> bool:
    low = name.lower()
    if "склад" not in low:
        return False
    # textile names rarely start with Склад
    return name.startswith("Склад") or name.startswith("Основной склад")


def is_doc(name: str) -> bool:
    low = name.lower()
    if DATE_RE.search(name) or DATE_ANY.search(name):
        return True
    return any(k in low for k in ("оприход", "приходн", "перемещ", "списан", "расходн", "отчет производ"))


def aging_band(days: int | None) -> str:
    if days is None:
        return "NO_DATE"
    if days <= 90:
        return "0_90_HOT"
    if days <= 180:
        return "91_180_WARM"
    if days <= 365:
        return "181_365_COOL"
    return "365PLUS_STALE"


def days_between(d: datetime | None) -> int | None:
    if not d:
        return None
    return max(0, (AS_OF - d).days)


def wcsv(path: Path, rows: list[dict]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)


def parse_inventory(path: Path) -> list[dict]:
    """Остатки ткани с warehouse (h64 flatten warehouse потерял)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    warehouse = "Основной склад"
    out: list[dict] = []
    for i, r in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if not r or r[0] is None:
            continue
        name = str(r[0]).strip()
        if not name:
            continue
        if name.startswith("Основной склад") or name.startswith("Склад "):
            warehouse = name
            continue
        if name in ("Структурная единица", "Номенклатура"):
            continue
        try:
            qty = float(r[1]) if r[1] not in (None, "") else None
            amount = float(r[2]) if r[2] not in (None, "") else None
            unit = float(r[3]) if r[3] not in (None, "") else None
        except (TypeError, ValueError):
            continue
        if amount is None and qty is None:
            continue
        # skip warehouse subtotal-like huge rows without code
        code = article_hint(name)
        if not code and amount is not None and amount > 500_000:
            continue
        out.append(
            {
                "warehouse": warehouse,
                "name": name[:180],
                "article_hint": code,
                "qty": qty if qty is not None else 0.0,
                "amount_rub": amount if amount is not None else 0.0,
                "unit_cost_rub": unit if unit is not None else "",
                "row": i,
            }
        )
    wb.close()
    return out


def abc_band_for_amount(rows: list[dict]) -> dict[tuple[str, str], str]:
    """ABC by amount within full inventory snapshot (warehouse, name_key)."""
    items = sorted(rows, key=lambda x: -float(x["amount_rub"]))
    total = sum(float(x["amount_rub"]) for x in items) or 1.0
    cum = 0.0
    out: dict[tuple[str, str], str] = {}
    for r in items:
        cum += float(r["amount_rub"])
        share = cum / total
        band = "A" if share <= 0.8 else ("B" if share <= 0.95 else "C")
        key = (r["warehouse"], re.sub(r"\s+", " ", r["name"]).strip().lower())
        out[key] = band
        if r["article_hint"]:
            out[(r["warehouse"], "ART:" + r["article_hint"])] = band
    return out


def parse_movements(path: Path) -> tuple[list[dict], dict]:
    """SKU-level aging signals from document lines."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    warehouse = "Основной склад"
    current: dict | None = None
    skus: list[dict] = []
    docs_n = 0
    wh_set = set()

    def flush():
        nonlocal current
        if current:
            skus.append(current)
            current = None

    for i, r in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        if not r:
            continue
        raw = r[0]
        # qty sub-row under sku/doc — skip (no name)
        if raw is None or str(raw).strip() == "":
            continue
        name = str(raw).strip()
        if name in ("Структурная единица", "Номенклатура", "Документ движения"):
            continue

        in_v = r[2] if len(r) > 2 else None
        out_v = r[3] if len(r) > 3 else None
        end_v = r[4] if len(r) > 4 else None

        if is_warehouse(name):
            flush()
            warehouse = name
            wh_set.add(warehouse)
            continue

        if is_doc(name):
            if not current:
                continue
            docs_n += 1
            dt = parse_date(name)
            low = name.lower()
            has_in = isinstance(in_v, (int, float)) and float(in_v) != 0
            has_out = isinstance(out_v, (int, float)) and float(out_v) != 0
            direction = "UNKNOWN"
            if any(k in low for k in DOC_IN_KEYS) or (has_in and not has_out):
                direction = "IN"
            elif any(k in low for k in DOC_OUT_KEYS) or (has_out and not has_in):
                direction = "OUT"
            elif "перемещ" in low:
                direction = "IN" if has_in and not has_out else ("OUT" if has_out else "TRANSFER")
            current["docs_n"] += 1
            current["doc_types"][direction] += 1
            if dt:
                if current["first_date"] is None or dt < current["first_date"]:
                    current["first_date"] = dt
                if current["last_date"] is None or dt > current["last_date"]:
                    current["last_date"] = dt
                if direction == "IN":
                    if current["last_in"] is None or dt > current["last_in"]:
                        current["last_in"] = dt
                elif direction in ("OUT", "TRANSFER"):
                    if current["last_out"] is None or dt > current["last_out"]:
                        current["last_out"] = dt
                # sample last doc
                current["last_doc"] = name[:120]
            continue

        # SKU header row
        flush()
        current = {
            "warehouse": warehouse,
            "name": name[:180],
            "article_hint": article_hint(name),
            "row_sku": i,
            "move_in_leaf": float(in_v) if isinstance(in_v, (int, float)) else None,
            "move_out_leaf": float(out_v) if isinstance(out_v, (int, float)) else None,
            "move_end_leaf": float(end_v) if isinstance(end_v, (int, float)) else None,
            "docs_n": 0,
            "doc_types": defaultdict(int),
            "first_date": None,
            "last_date": None,
            "last_in": None,
            "last_out": None,
            "last_doc": "",
        }

    flush()
    wb.close()
    stats = {
        "sku_rows": len(skus),
        "doc_rows": docs_n,
        "warehouses": sorted(wh_set),
        "with_dates": sum(1 for s in skus if s["last_date"]),
        "with_article": sum(1 for s in skus if s["article_hint"]),
    }
    return skus, stats


def main() -> None:
    REG.mkdir(parents=True, exist_ok=True)
    MARTS.mkdir(parents=True, exist_ok=True)

    mov_path = resolve("Движение тка")
    inv_path = resolve("Остатки ткани")
    skus, mov_stats = parse_movements(mov_path)
    inv_rows = parse_inventory(inv_path)
    abc_map = abc_band_for_amount(inv_rows)

    inv_by_wh_art: dict[tuple[str, str], list[dict]] = defaultdict(list)
    inv_by_wh_name: dict[tuple[str, str], list[dict]] = defaultdict(list)
    for r in inv_rows:
        if r["article_hint"]:
            inv_by_wh_art[(r["warehouse"], r["article_hint"])].append(r)
        inv_by_wh_name[(r["warehouse"], re.sub(r"\s+", " ", r["name"]).strip().lower())].append(r)

    aging_rows = []
    for s in skus:
        days = days_between(s["last_date"])
        band = aging_band(days)
        wh = s["warehouse"]
        hits: list[dict] = []
        match = "NO_INV"
        if s["article_hint"] and inv_by_wh_art.get((wh, s["article_hint"])):
            hits = inv_by_wh_art[(wh, s["article_hint"])]
            match = "WH_ART"
        else:
            nk = re.sub(r"\s+", " ", s["name"]).strip().lower()
            if inv_by_wh_name.get((wh, nk)):
                hits = inv_by_wh_name[(wh, nk)]
                match = "WH_NAME"

        amount = round(sum(float(h["amount_rub"]) for h in hits), 2)
        qty = round(sum(float(h["qty"]) for h in hits), 4)
        abc = ""
        if hits:
            abc = abc_map.get((wh, "ART:" + s["article_hint"]), "") if s["article_hint"] else ""
            if not abc:
                abc = abc_map.get((wh, re.sub(r"\s+", " ", hits[0]["name"]).strip().lower()), "")

        if amount > 0 and band == "365PLUS_STALE":
            risk = "DEAD_STOCK"
        elif amount > 0 and band == "181_365_COOL":
            risk = "AGING_WATCH"
        elif amount > 0 and s["docs_n"] == 0:
            risk = "STOCK_NO_MOVE_DOCS"
        elif amount <= 0 and band == "365PLUS_STALE":
            risk = "CLEARED_OLD"
        elif match == "NO_INV" and (s["move_end_leaf"] or 0) not in (None, 0):
            risk = "MOVE_WITHOUT_SNAPSHOT"
        else:
            risk = "OK_ACTIVE" if band in ("0_90_HOT", "91_180_WARM") else "REVIEW"

        aging_rows.append(
            {
                "warehouse": wh,
                "name": s["name"],
                "article_hint": s["article_hint"],
                "docs_n": s["docs_n"],
                "docs_in_n": s["doc_types"].get("IN", 0),
                "docs_out_n": s["doc_types"].get("OUT", 0) + s["doc_types"].get("TRANSFER", 0),
                "first_move_date": s["first_date"].strftime("%Y-%m-%d") if s["first_date"] else "",
                "last_move_date": s["last_date"].strftime("%Y-%m-%d") if s["last_date"] else "",
                "last_in_date": s["last_in"].strftime("%Y-%m-%d") if s["last_in"] else "",
                "last_out_date": s["last_out"].strftime("%Y-%m-%d") if s["last_out"] else "",
                "days_since_last_move": days if days is not None else "",
                "aging_band": band,
                "last_doc": s["last_doc"],
                "inv_match": match,
                "inv_qty": qty,
                "inv_amount_rub": amount,
                "inv_abc_band": abc,
                "inv_as_of": STOCK_AS_OF if hits else "",
                "risk_flag": risk,
                "move_end_leaf_ambiguous": s["move_end_leaf"] if s["move_end_leaf"] is not None else "",
                "as_of_analysis": AS_OF.strftime("%Y-%m-%d"),
                "do_not_auto_accept": "YES",
                "so_t": "N",
            }
        )

    # rollups
    band_rows = []
    for band in ("0_90_HOT", "91_180_WARM", "181_365_COOL", "365PLUS_STALE", "NO_DATE"):
        sub = [r for r in aging_rows if r["aging_band"] == band]
        band_rows.append(
            {
                "aging_band": band,
                "sku_n": len(sub),
                "with_inv_n": sum(1 for r in sub if r["inv_match"] != "NO_INV"),
                "inv_amount_rub": round(sum(float(r["inv_amount_rub"]) for r in sub), 2),
                "dead_stock_n": sum(1 for r in sub if r["risk_flag"] == "DEAD_STOCK"),
                "note": {
                    "0_90_HOT": "движение ≤90д",
                    "91_180_WARM": "91–180д",
                    "181_365_COOL": "181–365д — watch",
                    "365PLUS_STALE": ">365д — stale/dead если есть остаток",
                    "NO_DATE": "нет дат в документах движения",
                }[band],
            }
        )

    # ABC × aging cross (only matched inv)
    cross = defaultdict(lambda: {"sku_n": 0, "amount": 0.0})
    for r in aging_rows:
        if r["inv_match"] == "NO_INV":
            continue
        key = (r["inv_abc_band"] or "?", r["aging_band"])
        cross[key]["sku_n"] += 1
        cross[key]["amount"] += float(r["inv_amount_rub"])
    cross_rows = [
        {
            "abc_band": a,
            "aging_band": b,
            "sku_n": v["sku_n"],
            "inv_amount_rub": round(v["amount"], 2),
        }
        for (a, b), v in sorted(cross.items(), key=lambda x: -x[1]["amount"])
    ]

    # dead stock top
    dead = [r for r in aging_rows if r["risk_flag"] == "DEAD_STOCK"]
    dead.sort(key=lambda x: -float(x["inv_amount_rub"]))
    dead_top = dead[:40]

    # warehouse rollup
    wh_rows = []
    for wh in sorted({r["warehouse"] for r in aging_rows}):
        sub = [r for r in aging_rows if r["warehouse"] == wh]
        wh_rows.append(
            {
                "warehouse": wh,
                "sku_n": len(sub),
                "inv_amount_rub": round(sum(float(r["inv_amount_rub"]) for r in sub), 2),
                "stale_amount_rub": round(
                    sum(float(r["inv_amount_rub"]) for r in sub if r["aging_band"] == "365PLUS_STALE"),
                    2,
                ),
                "dead_stock_n": sum(1 for r in sub if r["risk_flag"] == "DEAD_STOCK"),
                "hot_n": sum(1 for r in sub if r["aging_band"] == "0_90_HOT"),
            }
        )

    stale_amt = next(r["inv_amount_rub"] for r in band_rows if r["aging_band"] == "365PLUS_STALE")
    dead_amt = round(sum(float(r["inv_amount_rub"]) for r in dead), 2)

    actions = [
        {
            "priority": "P1",
            "action_id": "H84-A1",
            "what": f"Разбор DEAD_STOCK: {len(dead)} SKU / ~{round(dead_amt/1e6, 2)}M ₽ остатка при last move >365д",
            "who": "Мокеева + Дендерина",
            "evidence": "fabric_dead_stock_top.csv",
            "why": "замороженный WC; кандидаты write-off / перекрой / возврат",
            "gate_delta": "0 (Stage2 WC; не gate score)",
        },
        {
            "priority": "P1",
            "action_id": "H84-A2",
            "what": f"Контроль STALE band: ~{round(stale_amt/1e6, 2)}M ₽ inventory в 365PLUS",
            "who": "Мокеева",
            "evidence": "fabric_aging_band_rollup.csv",
            "why": "приоритет G6 fabrics после RACI Prod Accept",
            "gate_delta": "0",
        },
        {
            "priority": "P2",
            "action_id": "H84-A3",
            "what": "ABC-A × STALE: дорогие позиции без свежего движения",
            "who": "Коновалова (Product cand.) + Мокеева",
            "evidence": "fabric_aging_abc_cross.csv",
            "why": "не резать дешёвый хвост раньше дорогого stale",
            "gate_delta": "0",
        },
        {
            "priority": "P2",
            "action_id": "H84-A4",
            "what": "Не трактовать leaf qty движения как ₽ — деньги только из Остатки 31.05",
            "who": "Сливяк / Data",
            "evidence": "h84_meta.json note ambiguous",
            "why": "H4 warning; ломает WC если смешать",
            "gate_delta": "0",
        },
    ]

    summary = [
        {"metric": "movement_sku_n", "value": mov_stats["sku_rows"], "note": "parsed SKU headers"},
        {"metric": "movement_doc_n", "value": mov_stats["doc_rows"], "note": "document lines with/without dates"},
        {"metric": "sku_with_dates", "value": mov_stats["with_dates"], "note": "aging possible"},
        {"metric": "inv_matched_n", "value": sum(1 for r in aging_rows if r["inv_match"] != "NO_INV"), "note": "art|name"},
        {"metric": "dead_stock_n", "value": len(dead), "note": "amount>0 & >365д"},
        {"metric": "dead_stock_amount_rub", "value": dead_amt, "note": "indicative from snapshot"},
        {"metric": "stale_band_amount_rub", "value": stale_amt, "note": "365PLUS all matched"},
        {"metric": "as_of_analysis", "value": AS_OF.strftime("%Y-%m-%d"), "note": "days_since base"},
        {"metric": "stock_snapshot", "value": STOCK_AS_OF, "note": "amount source"},
    ]

    wcsv(MARTS / "fabric_aging_by_sku.csv", aging_rows)
    wcsv(MARTS / "fabric_aging_band_rollup.csv", band_rows)
    wcsv(MARTS / "fabric_aging_abc_cross.csv", cross_rows)
    wcsv(MARTS / "fabric_aging_warehouse_rollup.csv", wh_rows)
    wcsv(MARTS / "fabric_dead_stock_top.csv", dead_top)
    wcsv(MARTS / "fabric_aging_owner_actions.csv", actions)
    wcsv(MARTS / "fabric_aging_summary.csv", summary)

    for name in [
        "fabric_aging_by_sku.csv",
        "fabric_aging_band_rollup.csv",
        "fabric_aging_abc_cross.csv",
        "fabric_aging_warehouse_rollup.csv",
        "fabric_dead_stock_top.csv",
        "fabric_aging_owner_actions.csv",
        "fabric_aging_summary.csv",
    ]:
        (REG / name).write_text((MARTS / name).read_text(encoding="utf-8"), encoding="utf-8")

    meta = {
        "hypothesis": "H84",
        "title": "mat_movement_fabric_aging",
        "do_not_auto_accept": True,
        "not_sot": True,
        "gate": "18/30",
        "sources": {
            "movement": mov_path.name,
            "inventory": inv_path.name,
        },
        "movement_stats": mov_stats,
        "inventory_n": len(inv_rows),
        "inventory_amount_rub": round(sum(float(r["amount_rub"]) for r in inv_rows), 2),
        "summary": {r["metric"]: r["value"] for r in summary},
        "method": {
            "aging": "days since last document date in Движение тканей",
            "money": "inventory snapshot amount_rub matched by warehouse+article|name",
            "leaf_columns": "AMBIGUOUS — not used as RUB",
            "no_double_count": "warehouse-scoped match (fix vs H84-v1 art-only)",
        },
    }
    (MARTS / "h84_meta.json").write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
    (REG / "h84_meta.json").write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")

    print(json.dumps(meta["summary"], ensure_ascii=False, indent=2))
    print("bands:")
    for r in band_rows:
        print(f"  {r['aging_band']}: sku={r['sku_n']} amt={r['inv_amount_rub']}")
    print("top dead:")
    for r in dead_top[:8]:
        print(f"  {r['article_hint'] or '-'} | {r['inv_amount_rub']} | {r['last_move_date']} | {r['name'][:50]}")


if __name__ == "__main__":
    main()
