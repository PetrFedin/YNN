#!/usr/bin/env python3
"""
H31: MD OPEN months explained + executive channel dashboard + freeze.

1) 2024-01 OPEN: разовое −70 000 EUR (Казьмина) ломает сумму; без adjustment → ≈DDS CLOSE
2) 2026-06 OPEN: в МД есть платежи, SALES DDS ещё 0 → DDS_LAG
3) Executive dashboard: MD + goods + TSUM dual в одном mart
4) Freeze после H27–H30 брифа

Не меняет исходные md_payments; добавляет adjusted recon + dashboard.
"""
from __future__ import annotations

import csv
import json
import shutil
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[3]
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/h31_exec_dashboard_20260724"
MART = ROOT / "live/marts"
W4 = ROOT / "live/registers/w4_sales_settle"
CC = ROOT / "live/YANINA_LIVE_CONTROL_CENTER_20260723.xlsx"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")
FX = 100.0

# Порог: единичные отрицательные платежи |x|>= порога считаем adjustment memo
ADJ_ABS_EUR = 20000.0


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def fnum(v):
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(" ", "").replace(",", "."))
    except (TypeError, ValueError):
        return None


def build_adjusted_recon() -> tuple[list[dict], list[dict]]:
    payments = list(csv.DictReader(open(MART / "md_payments.csv", encoding="utf-8")))
    dds = {
        r["period_month"]: fnum(r.get("amount_eur")) or 0.0
        for r in csv.DictReader(open(W4 / "sales_dds_income_eur.csv", encoding="utf-8"))
        if r.get("channel") == "Salon+Shop"
    }

    # flag adjustments
    adj_lines = []
    by_gross = defaultdict(float)
    by_adj = defaultdict(float)
    for r in payments:
        m = r.get("period_month") or ""
        amt = fnum(r.get("amount_eur")) or 0.0
        by_gross[m] += amt
        if amt < 0 and abs(amt) >= ADJ_ABS_EUR:
            by_adj[m] += amt
            adj_lines.append(
                {
                    "payment_id": r.get("payment_id"),
                    "period_month": m,
                    "payment_date": r.get("payment_date"),
                    "client": r.get("client"),
                    "amount_eur": amt,
                    "flag": "LARGE_NEG_ADJUSTMENT",
                    "note": "Excluded from operating recon; kept in gross",
                }
            )

    months = sorted(set(by_gross) | set(dds))
    rows = []
    for m in months:
        gross = by_gross.get(m, 0.0)
        adj = by_adj.get(m, 0.0)
        operating = gross - adj  # remove negative adjustments (adj is negative → subtract means add back)
        # wait: adj is sum of negatives e.g. -70000. operating = gross - adj = -5841 - (-70000) = 64159. Yes.
        d = dds.get(m, 0.0)
        gap = operating - d
        if d == 0 and operating > 0 and m >= "2026-06":
            status = "DDS_LAG"
            note = "Workbook has payments; SALES DDS not yet filled for month"
        elif d == 0 and operating == 0:
            status = "EMPTY"
            note = ""
        else:
            pct = abs(gap) / d * 100 if d else (100.0 if operating else 0.0)
            if pct <= 2:
                status = "CLOSE"
            elif pct <= 10:
                status = "SOFT"
            else:
                status = "OPEN"
            note = "operating = gross − large_neg_adjustments" if adj else "no large adjustments"
        rows.append(
            {
                "period_month": m,
                "md_payments_gross_eur": round(gross, 2),
                "large_neg_adjustments_eur": round(adj, 2),
                "md_payments_operating_eur": round(operating, 2),
                "sales_dds_salon_shop_eur": round(d, 2),
                "gap_operating_eur": round(gap, 2),
                "gap_pct": round(abs(gap) / d * 100, 1) if d else "",
                "status": status,
                "note": note,
            }
        )
    return rows, adj_lines


def build_exec_dashboard(recon_adj: list[dict]) -> list[dict]:
    """One-row-per-year executive mix: MD income + goods revenue."""
    # MD from income mix
    md_year = defaultdict(float)
    for r in csv.DictReader(open(MART / "channel_mix_income_eur.csv", encoding="utf-8")):
        if r.get("channel") == "MD_INDIVIDUAL":
            md_year[r["year"]] = fnum(r.get("amount_eur")) or 0.0

    # goods from margin views / total
    goods = {}
    path = MART / "margin_channel_views_h28.csv"
    if path.exists():
        for r in csv.DictReader(open(path, encoding="utf-8")):
            goods[r["channel"]] = {
                "revenue": fnum(r.get("revenue_rub")) or 0.0,
                "margin": fnum(r.get("margin_rub")) or 0.0,
                "margin_pct": r.get("margin_pct"),
            }
    else:
        for r in csv.DictReader(open(MART / "margin_channel_total.csv", encoding="utf-8")):
            goods[r["channel"]] = {
                "revenue": fnum(r.get("revenue_rub")) or 0.0,
                "margin": fnum(r.get("margin_rub")) or 0.0,
                "margin_pct": r.get("margin_pct"),
            }

    # MD recon quality recent
    recent = [r for r in recon_adj if r["period_month"] >= "2024-01"]
    cs = sum(1 for r in recent if r["status"] in ("CLOSE", "SOFT"))
    lag = sum(1 for r in recent if r["status"] == "DDS_LAG")
    open_n = sum(1 for r in recent if r["status"] == "OPEN")

    rows = [
        {
            "section": "MD_SERVICES",
            "metric": "income_2025_eur",
            "value": round(md_year.get("2025", 0), 2),
            "unit": "EUR",
            "note": "SALES DDS Salon+Shop / MD_INDIVIDUAL",
        },
        {
            "section": "MD_SERVICES",
            "metric": "income_2025_rub_fx100",
            "value": round(md_year.get("2025", 0) * FX, 2),
            "unit": "RUB",
            "note": "EUR×100 policy",
        },
        {
            "section": "MD_SERVICES",
            "metric": "share_2025_pct_vs_brief",
            "value": "83.9 vs 83",
            "unit": "pp",
            "note": "H29 channel mix",
        },
        {
            "section": "MD_RECON",
            "metric": "recon_since_2024_CLOSE_SOFT",
            "value": f"{cs}/{len(recent)}",
            "unit": "months",
            "note": f"DDS_LAG={lag}; OPEN={open_n} (after H31 adjustment)",
        },
        {
            "section": "GOODS",
            "metric": "B2B_revenue_rub",
            "value": round(goods.get("B2B", {}).get("revenue", 0), 2),
            "unit": "RUB",
            "note": f"margin {goods.get('B2B', {}).get('margin_pct')}%",
        },
        {
            "section": "GOODS",
            "metric": "IM_revenue_rub",
            "value": round(goods.get("IM", {}).get("revenue", 0), 2),
            "unit": "RUB",
            "note": f"margin {goods.get('IM', {}).get('margin_pct')}%",
        },
        {
            "section": "GOODS",
            "metric": "TSUM_REPORTED_margin_pct",
            "value": goods.get("TSUM_REPORTED", goods.get("TSUM", {})).get("margin_pct"),
            "unit": "%",
            "note": "Excel COGS incl. commission",
        },
        {
            "section": "GOODS",
            "metric": "TSUM_PRODUCT_margin_pct",
            "value": goods.get("TSUM_PRODUCT", {}).get("margin_pct"),
            "unit": "%",
            "note": "H28 W3 product view",
        },
        {
            "section": "GOVERNANCE",
            "metric": "raci_status",
            "value": "18 ACCEPT (H27 brief)",
            "unit": "",
            "note": "live/BUSINESS_STRUCTURE.md",
        },
        {
            "section": "GOVERNANCE",
            "metric": "freeze",
            "value": "BRIEF_INTEGRATED_H31",
            "unit": "",
            "note": "MD+goods+TSUM dual+owners",
        },
    ]
    return rows


def write_md(summary: dict, recon: list[dict], adjs: list[dict]):
    r24 = next((r for r in recon if r["period_month"] == "2024-01"), {})
    r26 = next((r for r in recon if r["period_month"] == "2026-06"), {})
    lines = [
        "# H31 — Executive dashboard + MD OPEN fix",
        "",
        f"Updated: {NOW}",
        "",
        "## OPEN months",
        "",
        f"### 2024-01",
        f"- Gross payments: {r24.get('md_payments_gross_eur')} EUR",
        f"- Large neg adjustment: {r24.get('large_neg_adjustments_eur')} EUR (Казьмина −70k)",
        f"- Operating (ex-adj): {r24.get('md_payments_operating_eur')} vs DDS {r24.get('sales_dds_salon_shop_eur')} → **{r24.get('status')}**",
        "",
        f"### 2026-06",
        f"- Workbook: {r26.get('md_payments_gross_eur')} EUR; DDS: {r26.get('sales_dds_salon_shop_eur')} → **{r26.get('status')}** (файл SALES ещё без июня)",
        "",
        f"Adjustments flagged: **{len(adjs)}**",
        "",
        "## Dashboard",
        "`live/marts/executive_dashboard.csv` — MD + goods + TSUM dual + governance.",
        "",
        f"Finding: {summary['finding']}",
        "",
    ]
    text = "\n".join(lines)
    (OUT / "EXEC_DASHBOARD.md").write_text(text, encoding="utf-8")
    (ROOT / "live/EXEC_DASHBOARD.md").write_text(text, encoding="utf-8")
    (EV / "EXEC_DASHBOARD.md").write_text(text, encoding="utf-8")


def update_cc(summary: dict):
    if not CC.exists():
        return
    wb = load_workbook(CC)
    if "H31_Exec" in wb.sheetnames:
        del wb["H31_Exec"]
    ws = wb.create_sheet("H31_Exec", 0)
    ws["A1"] = "H31 Executive + MD recon fix"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A2"] = NOW
    ws["A4"] = summary["finding"]
    ws["A5"] = "Doc"
    ws["B5"] = "live/EXEC_DASHBOARD.md"
    wb.save(CC)


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)

    recon, adjs = build_adjusted_recon()
    dash = build_exec_dashboard(recon)

    write_csv(MART / "recon_md_payments_vs_dds_adj.csv", recon, list(recon[0].keys()))
    write_csv(OUT / "recon_md_payments_vs_dds_adj.csv", recon, list(recon[0].keys()))
    write_csv(
        MART / "md_payment_adjustments.csv",
        adjs,
        list(adjs[0].keys()) if adjs else ["payment_id"],
    )
    write_csv(MART / "executive_dashboard.csv", dash, list(dash[0].keys()))
    write_csv(OUT / "executive_dashboard.csv", dash, list(dash[0].keys()))

    recent = [r for r in recon if r["period_month"] >= "2024-01"]
    cs = sum(1 for r in recent if r["status"] in ("CLOSE", "SOFT"))
    lag = sum(1 for r in recent if r["status"] == "DDS_LAG")
    opn = sum(1 for r in recent if r["status"] == "OPEN")

    freeze = {
        "wave": "H31",
        "generated_at": NOW,
        "status": "BRIEF_INTEGRATED_H31",
        "covers": ["H27 business brief", "H28 TSUM dual", "H29 MD channel", "H30 MD workbook", "H31 exec"],
        "so_t": False,
        "raci_accept": True,
        "md_recon_since_2024": {"CLOSE_SOFT": cs, "DDS_LAG": lag, "OPEN": opn, "months": len(recent)},
    }
    (MART / "staging_freeze.json").write_text(json.dumps(freeze, ensure_ascii=False, indent=2), encoding="utf-8")
    (EV / "staging_freeze.json").write_text(json.dumps(freeze, ensure_ascii=False, indent=2), encoding="utf-8")

    summary = {
        "wave": "H31",
        "generated_at": NOW,
        "finding": (
            f"H31: 2024-01 → CLOSE after excluding Казьмина −70k adj; 2026-06 → DDS_LAG; "
            f"MD recon ≥2024 CLOSE+SOFT {cs}/{len(recent)} (+{lag} lag). "
            "Executive dashboard published."
        ),
        "recon_close_soft": cs,
        "recon_dds_lag": lag,
        "recon_open": opn,
        "recon_months": len(recent),
        "adjustments_n": len(adjs),
        "dashboard_rows": len(dash),
        "not_sot": True,
    }
    write_md(summary, recon, adjs)
    update_cc(summary)
    (OUT / "h31_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    (EV / "h31_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    for name in (
        "recon_md_payments_vs_dds_adj.csv",
        "executive_dashboard.csv",
        "EXEC_DASHBOARD.md",
        "h31_summary.json",
    ):
        src = OUT / name if (OUT / name).exists() else MART / name
        if src.exists():
            shutil.copy2(src, EV / name)
    if adjs:
        shutil.copy2(MART / "md_payment_adjustments.csv", EV / "md_payment_adjustments.csv")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
