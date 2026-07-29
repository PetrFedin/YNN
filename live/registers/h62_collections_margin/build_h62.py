#!/usr/bin/env python3
"""H62: Parse collection/showroom result workbooks → margin marts + MD crosswalk.

Not SoT / not company P&L. Dual contour: showroom/MD commercial results.
"""
from __future__ import annotations

import csv
import hashlib
import json
import re
import shutil
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[3]
if not (ROOT / "live/registers/00_SOURCE_CATALOG_107.csv").exists():
    ROOT = Path.cwd()
REG = Path(__file__).resolve().parent
MARTS = ROOT / "live/marts"
MAPS = ROOT / "live/maps"
EV = ROOT / "live/evidence/h62_collections_margin_20260729"
WAVE_B = ROOT / "live/client_pack/execution_wave_b"


def write_csv(path: Path, fields: list[str], rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)


def norm_art(a) -> str:
    if a is None:
        return ""
    return str(a).strip().upper().replace(" ", "").replace("Ё", "Е")


def surname(client) -> str:
    if not client:
        return ""
    s = str(client).strip().upper().replace("Ё", "Е")
    return re.split(r"\s+", s)[0]


def parse_fx(header_cells) -> int:
    blob = " ".join(str(c) for c in header_cells if c is not None)
    m = re.search(r"(?:курс|РУБЛИ)[^\d]*(\d{2,3})", blob, re.I)
    if m:
        return int(m.group(1))
    m = re.search(r"(\d{2,3})\s*р", blob, re.I)
    if m:
        return int(m.group(1))
    return 100


def collection_meta(fname: str) -> tuple[str, str, str]:
    f = fname.lower()
    mapping = [
        ("кол-я 43", "COL43", "collection", "2024-01"),
        ("кол-я 44", "COL44", "collection", "2024-06"),
        ("кол-я 45", "COL45", "collection", "2025-01"),
        ("кол-я 46", "COL46", "collection", "2025-07"),
        ("кол-я 47", "COL47", "collection", "2026-01"),
        ("капсула 2024", "CAP2024", "capsule", "2024"),
        ("капсула 2025", "CAP2025", "capsule", "2025"),
        ("нг 2024", "NY2024", "nye", "2024"),
        ("нг 24.11.25", "NY2025", "nye", "2025"),
        ("круиз 22.04.24", "CRUISE2024", "cruise", "2024-04"),
        ("круиз 09.04.25", "CRUISE2025", "cruise", "2025-04"),
        ("круиз 14.04.26", "CRUISE2026", "cruise", "2026-04"),
    ]
    for key, cid, ctype, season in mapping:
        if key in f:
            return cid, ctype, season
    return "UNK", "other", ""


def find_col(header_upper: list[str], *keys: str) -> int | None:
    for i, cell in enumerate(header_upper):
        for k in keys:
            if k in cell:
                return i
    return None


def to_float(v):
    if v is None or str(v).strip() == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().lower().replace(",", ".")
    if s in {"подарок", "gift"}:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return None


def main() -> None:
    for d in (REG, MARTS, MAPS, EV, WAVE_B):
        d.mkdir(parents=True, exist_ok=True)

    cat = list(csv.DictReader((ROOT / "live/registers/00_SOURCE_CATALOG_107.csv").open()))
    col_files = [r for r in cat if r["category"] == "collection_result"]

    lines: list[dict] = []
    for r in col_files:
        p = Path(r["path"])
        if not p.exists():
            p = ROOT / "documents" / r["file_name"]
        cid, ctype, season = collection_meta(r["file_name"])
        wb = load_workbook(p, read_only=True, data_only=True)
        if "прайс" not in wb.sheetnames:
            wb.close()
            continue
        ws = wb["прайс"]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            wb.close()
            continue
        header = rows[0]
        fx = parse_fx(header)
        h = [str(c).upper() if c is not None else "" for c in header]
        i_model = find_col(h, "МОДЕЛ")
        i_desc = find_col(h, "ОПИС")
        i_cost = find_col(h, "СЕБ")
        i_price = find_col(h, "ЦЕНА В КОЛ", "ЦЕНА ЗА ЕД, ЕВРО")
        if i_price is None:
            i_price = find_col(h, "ЕВРО")
        i_qty = find_col(h, "КОЛ-ВО")
        i_client = find_col(h, "КЛИЕНТ")
        i_sale = find_col(h, "ЗАКАЗЫ, ЕВРО", "ПРОДАЖА, ЕВРО")
        i_date = len(header) - 1

        cur_model = ""
        cur_cost = None
        cur_price = None
        cur_desc = ""
        cost_unit = "EUR" if ctype == "collection" else "RUB"

        for ri, row in enumerate(rows[1:], start=2):
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue

            def cell(idx):
                if idx is None or idx >= len(row):
                    return None
                return row[idx]

            model, desc, cost, price = cell(i_model), cell(i_desc), cell(i_cost), cell(i_price)
            qty, client, sale, dt = cell(i_qty), cell(i_client), cell(i_sale), cell(i_date)

            if model not in (None, ""):
                cur_model = str(model).strip()
            if desc not in (None, ""):
                cur_desc = str(desc).strip()
            cf = to_float(cost)
            if cf is not None:
                cur_cost = cf
            pf = to_float(price)
            if pf is not None:
                cur_price = pf

            sale_eur = to_float(sale)
            qty_n = to_float(qty)
            is_order = client not in (None, "") and (
                sale_eur is not None or (qty_n is not None and qty_n > 0)
            )
            if not is_order:
                continue

            cost_eur = cost_rub = None
            if cur_cost is not None:
                if cost_unit == "EUR":
                    cost_eur = cur_cost
                    cost_rub = cur_cost * fx
                else:
                    cost_rub = cur_cost
                    cost_eur = cur_cost / fx if fx else None

            sale_date = ""
            if isinstance(dt, datetime):
                sale_date = dt.date().isoformat()
            elif isinstance(dt, str) and dt.strip():
                sale_date = dt.strip()

            q = qty_n if qty_n is not None else 1.0
            sale_rub = sale_eur * fx if sale_eur is not None else None
            gm_eur = None
            if sale_eur is not None and cost_eur is not None:
                gm_eur = sale_eur - cost_eur * q

            line_id = "COL-" + hashlib.sha1(
                f"{cid}|{cur_model}|{client}|{sale_eur}|{sale_date}|{ri}".encode()
            ).hexdigest()[:16]

            lines.append(
                {
                    "collection_line_id": line_id,
                    "collection_id": cid,
                    "collection_type": ctype,
                    "season_hint": season,
                    "source_file": r["file_name"],
                    "source_file_id": r["source_file_id"],
                    "master_file_id": r["master_file_id"],
                    "source_row": ri,
                    "article": cur_model,
                    "article_norm": norm_art(cur_model),
                    "description": cur_desc,
                    "client_raw": str(client).strip(),
                    "client_surname": surname(client),
                    "qty": q,
                    "list_price_eur": cur_price if cur_price is not None else "",
                    "sale_eur": sale_eur if sale_eur is not None else "",
                    "sale_rub_fx": round(sale_rub, 2) if sale_rub is not None else "",
                    "fx_rate_used": fx,
                    "cost_amount_raw": cur_cost if cur_cost is not None else "",
                    "cost_unit": cost_unit,
                    "cost_eur": round(cost_eur, 2) if cost_eur is not None else "",
                    "cost_rub": round(cost_rub, 2) if cost_rub is not None else "",
                    "gm_eur": round(gm_eur, 2) if gm_eur is not None else "",
                    "sale_date": sale_date,
                    "lifecycle_status": r["lifecycle_status"],
                    "md_link_type": "",
                    "md_line_id": "",
                    "so_t": "N",
                    "note": "showroom/collection workbook — not company P&L",
                }
            )
        wb.close()

    salon = list(csv.DictReader((MARTS / "md_salon_orders.csv").open()))
    shop = list(csv.DictReader((MARTS / "md_shop_sales.csv").open()))
    salon_by_art: dict[str, list] = defaultdict(list)
    shop_by_art: dict[str, list] = defaultdict(list)
    for s in salon:
        a = norm_art(s.get("article"))
        if a:
            salon_by_art[a].append(s)
    for s in shop:
        a = norm_art(s.get("article"))
        if a:
            shop_by_art[a].append(s)

    link_rows = []
    for L in lines:
        art, sur = L["article_norm"], L["client_surname"]
        candidates = []
        for s in salon_by_art.get(art, []):
            score = 50
            if sur and sur in surname(s.get("client")):
                score += 40
            if L["sale_date"] and (s.get("order_date") or "")[:7] == L["sale_date"][:7]:
                score += 10
            candidates.append(
                (
                    "SALON",
                    s.get("order_line_id"),
                    s.get("client"),
                    s.get("total_amount"),
                    s.get("cost_amount"),
                    score,
                    s.get("period_month"),
                )
            )
        for s in shop_by_art.get(art, []):
            score = 45
            if sur and sur in surname(s.get("client")):
                score += 40
            candidates.append(
                (
                    "SHOP",
                    s.get("shop_line_id"),
                    s.get("client"),
                    s.get("total_amount"),
                    s.get("cost_amount"),
                    score,
                    s.get("period_month"),
                )
            )
        candidates.sort(key=lambda x: -x[5])
        best = candidates[0] if candidates else None
        if best and best[5] >= 90:
            link_type = "STRONG"
        elif best and best[5] >= 50:
            link_type = "WEAK_ARTICLE_ONLY"
        elif best:
            link_type = "WEAK"
        else:
            link_type = "NONE"
        L["md_link_type"] = link_type
        L["md_line_id"] = best[1] if best else ""
        link_rows.append(
            {
                "collection_line_id": L["collection_line_id"],
                "collection_id": L["collection_id"],
                "article": L["article"],
                "client_surname": sur,
                "sale_eur": L["sale_eur"],
                "link_type": link_type,
                "md_channel": best[0] if best else "",
                "md_line_id": best[1] if best else "",
                "md_client": best[2] if best else "",
                "md_total_amount": best[3] if best else "",
                "md_cost_amount": best[4] if best else "",
                "md_period_month": best[6] if best else "",
                "score": best[5] if best else 0,
                "candidates_n": len(candidates),
                "do_not_auto_accept": "YES",
            }
        )

    def fnum(v):
        return float(v) if v not in ("", None) else None

    by_c: dict[str, list] = defaultdict(list)
    for L in lines:
        by_c[L["collection_id"]].append(L)

    sum_rows = []
    for cid, rs in sorted(by_c.items()):
        sales = [fnum(x["sale_eur"]) for x in rs]
        sales = [v for v in sales if v is not None]
        costs = [fnum(x["cost_eur"]) for x in rs if fnum(x["cost_eur"]) is not None]
        gms = [fnum(x["gm_eur"]) for x in rs if fnum(x["gm_eur"]) is not None]
        strong = sum(1 for x in rs if x["md_link_type"] == "STRONG")
        art_only = sum(1 for x in rs if x["md_link_type"] == "WEAK_ARTICLE_ONLY")
        none = sum(1 for x in rs if x["md_link_type"] == "NONE")
        sale_sum = round(sum(sales), 2)
        cost_sum = round(sum(costs), 2) if costs else ""
        gm_sum = round(sum(gms), 2) if gms else ""
        gm_pct = round(100 * sum(gms) / sale_sum, 1) if gms and sale_sum else ""
        sum_rows.append(
            {
                "collection_id": cid,
                "collection_type": rs[0]["collection_type"],
                "season_hint": rs[0]["season_hint"],
                "source_file": rs[0]["source_file"],
                "order_lines_n": len(rs),
                "sale_eur": sale_sum,
                "sale_rub_fx100ish": round(sale_sum * float(rs[0]["fx_rate_used"]), 2),
                "fx_rate_used": rs[0]["fx_rate_used"],
                "lines_with_cost_n": len(costs),
                "cost_eur_on_costed_lines": cost_sum,
                "gm_eur_on_costed_lines": gm_sum,
                "gm_pct_on_costed_lines": gm_pct,
                "md_strong_links_n": strong,
                "md_article_only_links_n": art_only,
                "md_no_link_n": none,
                "md_strong_link_pct": round(100 * strong / len(rs), 1) if rs else 0,
                "lifecycle_status": rs[0]["lifecycle_status"],
                "so_t": "N",
                "contour": "SHOWROOM_MD — not goods company P&L",
            }
        )

    # top models by sale
    by_art: dict[str, dict] = {}
    for L in lines:
        key = f"{L['collection_id']}|{L['article_norm']}"
        if key not in by_art:
            by_art[key] = {
                "collection_id": L["collection_id"],
                "article": L["article"],
                "description": L["description"],
                "orders_n": 0,
                "sale_eur": 0.0,
                "clients": set(),
            }
        by_art[key]["orders_n"] += 1
        se = fnum(L["sale_eur"])
        if se is not None:
            by_art[key]["sale_eur"] += se
        if L["client_surname"]:
            by_art[key]["clients"].add(L["client_surname"])
    top_models = []
    for v in by_art.values():
        top_models.append(
            {
                "collection_id": v["collection_id"],
                "article": v["article"],
                "description": v["description"],
                "orders_n": v["orders_n"],
                "sale_eur": round(v["sale_eur"], 2),
                "unique_clients_n": len(v["clients"]),
            }
        )
    top_models.sort(key=lambda x: -x["sale_eur"])
    top_models = top_models[:40]

    # overall
    all_sale = sum(fnum(x["sale_eur"]) or 0 for x in lines)
    meta = {
        "horizon": "H62",
        "date": str(date.today()),
        "title": "Collections/showroom → margin lines + MD crosswalk",
        "files_n": len(col_files),
        "order_lines_n": len(lines),
        "sale_eur_total": round(all_sale, 2),
        "md_strong_links": sum(1 for x in link_rows if x["link_type"] == "STRONG"),
        "no_fake_accept": True,
        "so_t": False,
        "note": "Closes layer2-optional gap for 12 collection_result files",
    }
    (REG / "meta.json").write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")

    line_fields = list(lines[0].keys()) if lines else []
    write_csv(REG / "collection_order_lines.csv", line_fields, lines)
    write_csv(REG / "collection_md_links.csv", list(link_rows[0].keys()), link_rows)
    write_csv(REG / "collection_margin_by_collection.csv", list(sum_rows[0].keys()), sum_rows)
    write_csv(REG / "collection_top40_models.csv", list(top_models[0].keys()), top_models)

    for name in [
        "collection_order_lines.csv",
        "collection_md_links.csv",
        "collection_margin_by_collection.csv",
        "collection_top40_models.csv",
        "meta.json",
    ]:
        src = REG / name
        if name == "meta.json":
            shutil.copy2(src, MARTS / "h62_meta.json")
            shutil.copy2(src, MAPS / "h62_meta.json")
            shutil.copy2(src, EV / "meta.json")
        else:
            shutil.copy2(src, MARTS / name)
            shutil.copy2(src, MAPS / name)
            shutil.copy2(src, EV / name)

    shutil.copy2(REG / "collection_margin_by_collection.csv", WAVE_B / "12_collection_margin_by_collection.csv")
    shutil.copy2(REG / "collection_top40_models.csv", WAVE_B / "13_collection_top40_models.csv")
    shutil.copy2(REG / "collection_md_links.csv", WAVE_B / "14_collection_md_links.csv")

    print(json.dumps(meta, ensure_ascii=False, indent=2))
    for s in sum_rows:
        print(
            s["collection_id"],
            "lines",
            s["order_lines_n"],
            "sale€",
            s["sale_eur"],
            "strong%",
            s["md_strong_link_pct"],
            "gm%",
            s["gm_pct_on_costed_lines"],
        )


if __name__ == "__main__":
    main()
