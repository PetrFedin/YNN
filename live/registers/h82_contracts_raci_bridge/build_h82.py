#!/usr/bin/env python3
"""H82: contracts + org_raci + штатка ↔ formal RACI / signoff (P2 ops).

Не SoT. Не auto-Accept. Усиливает исполнение после sign session:
кто в «Финансы и платежи» ↔ кто в штатке ↔ OPEN/ACCEPT RACI ↔ scope Stage1.
"""
from __future__ import annotations

import csv
import json
import re
import zipfile
import warnings
from pathlib import Path

from openpyxl import load_workbook

warnings.filterwarnings("ignore")

ROOT = Path(__file__).resolve().parents[3]
DL = Path("/Users/petr/Downloads/YANINA документы")
REG = Path(__file__).resolve().parent
MARTS = ROOT / "live/marts"
CLIENT = ROOT / "live/client_pack"
MAPS = ROOT / "live/maps"
WAVE = ROOT / "live/wave_b_ops"  # ops cards рядом с client pack, без коллизии номеров


def resolve_catalog(substr: str) -> Path:
    rows = list(csv.DictReader(open(ROOT / "live/registers/00_SOURCE_CATALOG_107.csv")))
    for r in rows:
        if substr.lower() in r["file_name"].lower():
            p = Path(r["path"])
            if not p.exists():
                p = DL / r["file_name"]
            if p.exists():
                return p
    raise FileNotFoundError(substr)


def docx_paras(path: Path) -> list[str]:
    with zipfile.ZipFile(path) as z:
        xml = z.read("word/document.xml").decode("utf-8", errors="ignore")
    paras: list[str] = []
    for block in re.split(r"</w:p>", xml):
        bits = re.findall(r"<w:t[^>]*>(.*?)</w:t>", block)
        if bits:
            t = re.sub(r"\s+", " ", " ".join(bits)).strip()
            if t:
                paras.append(t)
    return paras


def norm_fio(s: str) -> str:
    s = re.sub(r"\s+", " ", (s or "").strip())
    s = s.replace("ё", "е").replace("Ё", "Е")
    return s.upper()


def fio_key(s: str) -> str:
    """Фамилия + инициал имени для fuzzy match."""
    parts = norm_fio(s).replace(".", " ").split()
    if not parts:
        return ""
    if len(parts) == 1:
        return parts[0]
    return f"{parts[0]}|{parts[1][:1]}"


# --- domain blocks from «Финансы и платежи» ---
# Ручная разметка по тексту документа (source of truth = docx, не Accept).
FINANCE_BLOCKS: list[dict] = [
    {
        "domain_raw": "CASH_CASHLESS",
        "raci_domain": "CASH",
        "text": "Финансы и платежи",
        "people": [
            ("Мамушкина Елена", "исполнительный директор", "наличные", "R"),
            ("Сливяк Галина", "главный бухгалтер", "безналичные", "R"),
        ],
    },
    {
        "domain_raw": "ACCOUNTING_TAX_1C",
        "raci_domain": "TAX",
        "text": "Бухгалтерия и налоги, 1С, подготовка выгрузок",
        "people": [("Сливяк Галина", "главный бухгалтер", "1С/выгрузки", "R")],
    },
    {
        "domain_raw": "PAYROLL_FOT",
        "raci_domain": "PAYROLL",
        "text": "ФОТ",
        "people": [
            ("Богдашкина Евгения", "начальник вышивального цеха", "вышивальщицы", "R"),
            ("Мокеева Анна", "производственный директор", "мастера и конструктора", "R"),
            ("Жукова Анна", "руководитель проектного отдела", "мастера/конструктора отдела", "R"),
            ("Сливяк Галина", "главный бухгалтер", "окладники, свод, ведомости", "A"),
            ("Мамушкина Елена", "исполнительный директор", "выдача", "A"),
        ],
    },
    {
        "domain_raw": "PROCUREMENT",
        "raci_domain": "COST",  # ближайший formal domain; закупки не отдельный ACCEPT
        "text": "Закупки",
        "people": [
            ("Дендерина Ирина", "снабженец", "основной снабженец", "R"),
            ("Меркушина Татьяна", "руководитель проектного отдела", "аутсорс частично", "R"),
            ("Богдашкина Евгения", "начальник вышивального цеха", "фурнитура вышивки", "R"),
        ],
    },
    {
        "domain_raw": "WAREHOUSE_MATERIALS",
        "raci_domain": "PRODUCT",
        "text": "Склад и материалы",
        "people": [
            ("Дендерина Ирина", "снабженец", "замещает кладовщика", "R"),
            ("Коновалова Анна", "товаровед", "замещает кладовщика", "R"),
        ],
        "vacancy": "Кладовщик",
    },
    {
        "domain_raw": "PRODUCTION_MD",
        "raci_domain": "PRODUCTION",
        "text": "Производство Модный дом",
        "people": [
            ("Богдашкина Евгения", "начальник вышивального цеха", "вышивальщицы", "R"),
            ("Мокеева Анна", "производственный директор", "мастера и конструктора", "R"),
            ("Шалагинова Татьяна", "технолог", "технолог", "C"),
        ],
    },
    {
        "domain_raw": "CHANNELS_IM",
        "raci_domain": "PRODUCT",
        "text": "Площадки/ИМ",
        "people": [
            ("Меркушина Татьяна", "руководитель проектного отдела", "площадки/ИМ", "R"),
            ("Жукова Анна", "руководитель проектного отдела", "МД/ИМ", "R"),
        ],
    },
    {
        "domain_raw": "COGS_COSTING",
        "raci_domain": "COST",
        "text": "Себестоимость",
        "people": [
            ("Богдашкина Евгения", "начальник вышивального цеха", "вышивка", "R"),
            ("Мокеева Анна", "производственный директор", "производство", "R"),
            ("Меркушина Татьяна", "руководитель проектного отдела", "проекты", "R"),
            ("Жукова Анна", "руководитель проектного отдела", "проекты", "R"),
        ],
    },
    {
        "domain_raw": "B2B",
        "raci_domain": "B2B",
        "text": "B2B",
        "people": [("Коптева Марина", "операционный директор", "контролирует; менеджера нет", "A")],
        "gap": "менеджера нет",
    },
    {
        "domain_raw": "B2C",
        "raci_domain": "",  # нет DOM-B2C в H23 formal — только ops evidence
        "text": "B2C",
        "people": [("Лимачева Инна", "менеджер по работе с клиентами", "B2C", "R")],
    },
]


def load_shtatka(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    people: list[dict] = []
    section = "UNKNOWN"
    for row in ws.iter_rows(values_only=True, max_col=6):
        cells = list(row) + [None] * 6
        a, b, c, d = cells[0], cells[1], cells[2], cells[3]
        if a is None and isinstance(b, str) and b.strip():
            section = b.strip()
            continue
        if isinstance(a, (int, float)) and b:
            fio = str(b).strip()
            role = str(c).strip() if c else ""
            note = str(d).strip() if d else ""
            unformal = "не оформл" in note.lower()
            people.append(
                {
                    "row_n": int(a),
                    "fio": fio,
                    "fio_norm": norm_fio(fio),
                    "fio_key": fio_key(fio),
                    "role": role,
                    "note": note,
                    "section": section,
                    "employment_flag": "UNFORMAL" if unformal else ("OK" if not note else "NOTE"),
                }
            )
    wb.close()
    return people


def extract_contract_scope(paras: list[str]) -> dict:
    joined = "\n".join(paras)
    in_scope = []
    out_scope = []
    # known clauses from Stage1 diagnostic contract
    if "диагностик" in joined.lower():
        in_scope.append("комплексная диагностика бизнеса (аналитика/консультации)")
        in_scope.append("диагностический отчет + карта ограничений + рекомендации + презентация")
    if "не предусматривает внедрение" in joined.lower() or "не предусматривает" in joined.lower():
        out_scope.append("внедрение рекомендаций")
        out_scope.append("постановка управленческого учета")
        out_scope.append("разработка регламентов / сопровождение изменений / аудит")
    # scan for explicit bullets
    for p in paras:
        pl = p.lower()
        if "не предусматривает" in pl and len(p) < 500:
            out_scope.append(p[:240])
        if "результатом работ является" in pl:
            in_scope.append(p[:240])
    # dedupe
    def uniq(xs):
        seen = set()
        out = []
        for x in xs:
            k = x[:80]
            if k not in seen:
                seen.add(k)
                out.append(x)
        return out

    return {"in_scope": uniq(in_scope), "out_scope": uniq(out_scope), "paras_n": len(paras)}


def match_staff(fio: str, staff_by_key: dict[str, dict], staff_by_norm: dict[str, dict]) -> dict | None:
    n = norm_fio(fio)
    if n in staff_by_norm:
        return staff_by_norm[n]
    k = fio_key(fio)
    return staff_by_key.get(k)


def main() -> None:
    REG.mkdir(parents=True, exist_ok=True)
    MARTS.mkdir(parents=True, exist_ok=True)
    WAVE.mkdir(parents=True, exist_ok=True)

    fin_path = resolve_catalog("Финансы и платежи")
    diag_path = resolve_catalog("диагностики")
    plan_path = resolve_catalog("план работы")
    nda_path = resolve_catalog("NDA")
    sht_path = resolve_catalog("Штатка")

    fin_paras = docx_paras(fin_path)
    diag_paras = docx_paras(diag_path)
    plan_paras = docx_paras(plan_path)
    scope = extract_contract_scope(diag_paras)
    staff = load_shtatka(sht_path)
    staff_by_norm = {p["fio_norm"]: p for p in staff}
    staff_by_key: dict[str, dict] = {}
    for p in staff:
        staff_by_key.setdefault(p["fio_key"], p)

    # formal RACI
    formal = list(csv.DictReader(open(MARTS / "raci_formal_vs_candidate_map.csv")))
    formal_by_dom = {r["domain"]: r for r in formal}
    signoff = list(csv.DictReader(open(MARTS / "raci_yanina_signoff_sheet.csv")))

    # --- bridge rows: finance people × staff × formal ---
    bridge_rows = []
    for block in FINANCE_BLOCKS:
        formal_row = formal_by_dom.get(block["raci_domain"], {})
        for fio, title, note, raci_letter in block["people"]:
            st = match_staff(fio, staff_by_key, staff_by_norm)
            bridge_rows.append(
                {
                    "source_doc": fin_path.name,
                    "domain_raw": block["domain_raw"],
                    "raci_domain": block["raci_domain"],
                    "block_text": block["text"],
                    "fio": fio,
                    "title_in_raci_doc": title,
                    "note_in_raci_doc": note,
                    "implied_raci_letter": raci_letter,
                    "in_shtatka": "YES" if st else "NO",
                    "shtatka_role": st["role"] if st else "",
                    "shtatka_section": st["section"] if st else "",
                    "employment_flag": st["employment_flag"] if st else "MISSING_IN_SHTATKA",
                    "h23_formal_status": formal_row.get("h23_formal_status", "NO_FORMAL_DOMAIN"),
                    "h23_owner": formal_row.get("h23_owner", ""),
                    "h27_candidate": formal_row.get("h27_candidate_owner", ""),
                    "gap_vs_formal": (
                        "MATCH_OWNER"
                        if formal_row.get("h23_owner")
                        and fio_key(formal_row.get("h23_owner", "")) == fio_key(fio)
                        else (
                            "MATCH_CANDIDATE"
                            if formal_row.get("h27_candidate_owner")
                            and fio_key(formal_row.get("h27_candidate_owner", "")) == fio_key(fio)
                            else (
                                "SUPPORTING_ROLE"
                                if formal_row
                                else "NO_FORMAL_LINK"
                            )
                        )
                    ),
                    "vacancy_in_block": block.get("vacancy", ""),
                    "gap_flag": block.get("gap", ""),
                    "do_not_auto_accept": "YES",
                }
            )

    # domain rollup
    domain_rows = []
    for block in FINANCE_BLOCKS:
        formal_row = formal_by_dom.get(block["raci_domain"], {})
        people = [p[0] for p in block["people"]]
        in_staff = sum(1 for f in people if match_staff(f, staff_by_key, staff_by_norm))
        unformal = sum(
            1
            for f in people
            if (st := match_staff(f, staff_by_key, staff_by_norm))
            and st["employment_flag"] == "UNFORMAL"
        )
        domain_rows.append(
            {
                "domain_raw": block["domain_raw"],
                "raci_domain": block["raci_domain"],
                "block_text": block["text"],
                "people_n": len(people),
                "people": " | ".join(people),
                "in_shtatka_n": in_staff,
                "unformal_n": unformal,
                "vacancy": block.get("vacancy", ""),
                "gap": block.get("gap", ""),
                "h23_formal_status": formal_row.get("h23_formal_status", "NO_FORMAL_DOMAIN"),
                "h23_owner": formal_row.get("h23_owner", ""),
                "h27_candidate": formal_row.get("h27_candidate_owner", ""),
                "owner_action": (
                    "KEEP_ACCEPTED"
                    if formal_row.get("h23_formal_status") == "ACCEPTED"
                    else (
                        "SIGN_CANDIDATE_ON_YANINA_SHEET"
                        if formal_row.get("h23_formal_status") == "OPEN_NEEDS_OWNER"
                        else "MAP_ONLY"
                    )
                ),
                "do_not_auto_accept": "YES",
            }
        )

    # staff coverage vs RACI names
    raci_names = {norm_fio(r["fio"]) for r in bridge_rows}
    staff_cov = []
    for p in staff:
        hit = p["fio_norm"] in raci_names or any(
            fio_key(p["fio"]) == fio_key(r["fio"]) for r in bridge_rows
        )
        staff_cov.append(
            {
                **{k: p[k] for k in ("row_n", "fio", "role", "section", "note", "employment_flag")},
                "mentioned_in_finance_raci": "YES" if hit else "NO",
                "priority_if_open_owner": (
                    "HIGH"
                    if hit and p["employment_flag"] == "UNFORMAL"
                    else ("MED" if hit else "LOW")
                ),
            }
        )

    # signoff enrichment: is candidate in finance doc + shtatka?
    sign_enrich = []
    for s in signoff:
        fio = s.get("candidate_fio", "")
        st = match_staff(fio, staff_by_key, staff_by_norm)
        in_fin = any(fio_key(fio) == fio_key(r["fio"]) for r in bridge_rows)
        sign_enrich.append(
            {
                "rank": s.get("rank", ""),
                "role_id": s.get("role_id", ""),
                "candidate_fio": fio,
                "in_finance_raci_doc": "YES" if in_fin else "NO",
                "in_shtatka": "YES" if st else "NO",
                "shtatka_role": st["role"] if st else "",
                "employment_flag": st["employment_flag"] if st else "MISSING",
                "evidence_ready_for_sign": (
                    "YES" if in_fin and st and st["employment_flag"] != "UNFORMAL" else "SOFT"
                ),
                "why_now": s.get("why_now", ""),
                "unlocks": s.get("unlocks", ""),
                "decision": s.get("decision", ""),
                "do_not_auto_accept": "YES",
            }
        )

    # contract perimeter cards
    contract_files = [
        {
            "file": diag_path.name,
            "role": "SCOPE_STAGE1",
            "sha_in_catalog": "YES",
            "in_scope_summary": " | ".join(scope["in_scope"][:3]),
            "out_scope_summary": " | ".join(scope["out_scope"][:4]),
            "gate_link": "не двигает score; задаёт границы обещаний",
            "owner_action": "напоминать на встрече: Stage1 = диагностика, не UE/регламенты",
        },
        {
            "file": nda_path.name,
            "role": "NDA",
            "sha_in_catalog": "YES",
            "in_scope_summary": "конфиденциальность данных диагностики",
            "out_scope_summary": "не заменяет RACI / SoT Accept",
            "gate_link": "режим данных",
            "owner_action": "уже в пакете; не блокирует gate",
        },
        {
            "file": plan_path.name,
            "role": "WORK_PLAN",
            "sha_in_catalog": "YES",
            "in_scope_summary": "пилот 3–4 нед → комплексная диагностика; пилот может зачесться в бюджет",
            "out_scope_summary": plan_paras[0][:180] if plan_paras else "",
            "gate_link": "календарь, не SoT",
            "owner_action": "сверить ожидания сроков с фактом Stage1",
        },
        {
            "file": fin_path.name,
            "role": "OPERATING_RACI_DRAFT",
            "sha_in_catalog": "YES",
            "in_scope_summary": "операционный контур: cash/bank/FOT/закупки/склад/производство/каналы/B2B/B2C",
            "out_scope_summary": "не formal H23 Accept; vacancy кладовщик; B2B менеджера нет",
            "gate_link": "evidence для OPEN roles на yanina signoff",
            "owner_action": "сверить с 01_SIGN + raci_yanina_signoff_sheet",
        },
        {
            "file": sht_path.name,
            "role": "STAFF_ROSTER",
            "sha_in_catalog": "YES",
            "in_scope_summary": f"штатка: {len(staff)} строк; unformal={sum(1 for p in staff if p['employment_flag']=='UNFORMAL')}",
            "out_scope_summary": "не payroll SoT; не заменяет ведомости",
            "gate_link": "проверка что candidate owners существуют в орг",
            "owner_action": "закрыть UNFORMAL для ключевых owners до Wave B",
        },
    ]

    # owner actions (P0-linked only where RACI unlocks gate)
    actions = [
        {
            "priority": "P0",
            "action_id": "H82-A1",
            "what": "На встрече: OPEN Product/Cost/Prod/B2B по signoff + finance doc evidence",
            "who": "Янина",
            "evidence": "h82_signoff_evidence.csv + 01_SIGN_CHECKBOXES",
            "unlocks": "RACI completeness → Wave B owners",
            "gate_delta": "косвенно (исполнение), score только после human sign",
        },
        {
            "priority": "P0",
            "action_id": "H82-A2",
            "what": "Зафиксировать scope: Stage1 ≠ внедрение UE/регламентов (договор)",
            "who": "Янина + Петр",
            "evidence": "h82_contract_perimeter.csv",
            "unlocks": "нет расползания ожиданий на встрече",
            "gate_delta": "0 (governance)",
        },
        {
            "priority": "P1",
            "action_id": "H82-A3",
            "what": "Закрыть vacancy «Кладовщик» или Accept временного R: Дендерина+Коновалова",
            "who": "Мамушкина / Коптева",
            "evidence": "domain WAREHOUSE_MATERIALS",
            "unlocks": "G6/WC stewardship яснее",
            "gate_delta": "0",
        },
        {
            "priority": "P1",
            "action_id": "H82-A4",
            "what": "B2B: либо назначить менеджера, либо Accept «контроль = Коптева» как Owner DOM-B2B",
            "who": "Янина",
            "evidence": "domain B2B gap=менеджера нет",
            "unlocks": "E01 call-script execution owner",
            "gate_delta": "0 напрямую; разблокирует Wave B B2B",
        },
        {
            "priority": "P2",
            "action_id": "H82-A5",
            "what": f"Разбор UNFORMAL в штатке ({sum(1 for p in staff if p['employment_flag']=='UNFORMAL')} чел) — приоритет hit RACI",
            "who": "Сливяк + Мамушкина",
            "evidence": "h82_shtatka_coverage.csv priority HIGH",
            "unlocks": "HR↔payroll контур (следующий P2)",
            "gate_delta": "0",
        },
    ]

    def wcsv(path: Path, rows: list[dict]):
        if not rows:
            path.write_text("")
            return
        with open(path, "w", newline="", encoding="utf-8") as f:
            wr = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            wr.writeheader()
            wr.writerows(rows)

    # marts
    wcsv(MARTS / "contracts_raci_people_bridge.csv", bridge_rows)
    wcsv(MARTS / "contracts_raci_domain_rollup.csv", domain_rows)
    wcsv(MARTS / "contracts_shtatka_coverage.csv", staff_cov)
    wcsv(MARTS / "contracts_signoff_evidence.csv", sign_enrich)
    wcsv(MARTS / "contracts_perimeter_files.csv", contract_files)
    wcsv(MARTS / "contracts_raci_owner_actions.csv", actions)

    # register copies
    for name in [
        "contracts_raci_people_bridge.csv",
        "contracts_raci_domain_rollup.csv",
        "contracts_shtatka_coverage.csv",
        "contracts_signoff_evidence.csv",
        "contracts_perimeter_files.csv",
        "contracts_raci_owner_actions.csv",
    ]:
        (REG / name).write_text((MARTS / name).read_text(encoding="utf-8"), encoding="utf-8")

    meta = {
        "hypothesis": "H82",
        "title": "contracts_raci_shtatka_bridge",
        "do_not_auto_accept": True,
        "not_sot": True,
        "gate": "18/30",
        "sources": {
            "finance_raci": fin_path.name,
            "diagnostic_contract": diag_path.name,
            "nda": nda_path.name,
            "plan": plan_path.name,
            "shtatka": sht_path.name,
        },
        "counts": {
            "bridge_rows": len(bridge_rows),
            "domains": len(domain_rows),
            "shtatka_people": len(staff),
            "shtatka_unformal": sum(1 for p in staff if p["employment_flag"] == "UNFORMAL"),
            "shtatka_in_raci": sum(1 for p in staff_cov if p["mentioned_in_finance_raci"] == "YES"),
            "signoff_evidence_yes": sum(1 for s in sign_enrich if s["evidence_ready_for_sign"] == "YES"),
            "signoff_soft": sum(1 for s in sign_enrich if s["evidence_ready_for_sign"] == "SOFT"),
            "vacancy_kladovshik": True,
            "b2b_no_manager": True,
        },
        "scope": scope,
        "finance_doc_chars": sum(len(p) for p in fin_paras),
    }
    (MARTS / "h82_meta.json").write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
    (REG / "h82_meta.json").write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")

    # verify finance text still contains key names (drift guard)
    fin_blob = " ".join(fin_paras)
    missing = [r["fio"] for r in bridge_rows if r["fio"].split()[0] not in fin_blob]
    meta["drift_missing_in_docx"] = sorted(set(missing))
    (MARTS / "h82_meta.json").write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")

    print(json.dumps(meta["counts"], ensure_ascii=False, indent=2))
    print("drift", meta["drift_missing_in_docx"])
    print("sign YES", meta["counts"]["signoff_evidence_yes"], "SOFT", meta["counts"]["signoff_soft"])


if __name__ == "__main__":
    main()
