#!/usr/bin/env python3
"""
H3: ингест 14 новых документов + учёт в архитектуре.

Новое:
- Сбер ООО «Салон Юлия Янина» (ИНН 7715219770) — третий LE
- VTB card StatementFull (держатель Мамушкина Е.А.)
- Справочник номенклатуры → SKU master
- Продажи B2B/ИМ/ЦУМ до 06.2026
- Движения тканей/товаров (+себестоимость)
- Финансы и платежи.docx → RACI кандидаты (Мамушкина / Сливяк)

Не SoT: decision_ACCEPT в RACI остаётся за владельцем.
"""
from __future__ import annotations

import csv
import hashlib
import json
import re
import shutil
import tempfile
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[3]
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/h3_new_docs_20260724"
DOCS = ROOT / "documents"
W1 = ROOT / "live/registers/w1_bank_cash"
W4 = ROOT / "live/registers/w4_sales_settle"
CAT93 = ROOT / "live/registers/00_SOURCE_CATALOG_93.csv"
CAT = ROOT / "live/registers/00_SOURCE_CATALOG.csv"
OWNER = ROOT / "live/YANINA_OWNER_PACKET_RACI_AND_REQUESTS.xlsx"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")

NEW_FILES = [
    ("StatementFull_0400-C-869430_2024-01-01_2024-12-31_2026-07-24_09-31-20.xls", "bank_card", "2024", "ИП_ЯНИНА", "Банк / карточный счёт", "FILE-094"),
    ("StatementFull_0400-C-869430_2025-01-01_2025-12-31_2026-07-24_09-30-49.xls", "bank_card", "2025", "ИП_ЯНИНА", "Банк / карточный счёт", "FILE-095"),
    ("StatementFull_0400-C-869430_2026-01-01_2026-06-30_2026-07-24_09-31-18.xls", "bank_card", "2026", "ИП_ЯНИНА", "Банк / карточный счёт", "FILE-096"),
    ("Движение тканей.xlsx", "mat_movement", "2024-2026", "GROUP", "Материалы / ткани", "FILE-097"),
    ("Движение товара по складам.xlsx", "stock_movement", "2024-2026", "GROUP", "Склад / товар", "FILE-098"),
    ("Движение товаров с себестоимостью.xlsx", "stock_cost_movement", "2024-2026", "GROUP", "Склад / себестоимость", "FILE-099"),
    ("Продажи B2B 2024-06.2026.xlsx", "sales_b2b", "2024-2026", "GROUP", "Продажи B2B", "FILE-100"),
    ("Продажи ИМ 2024-06.2026.xlsx", "sales_im", "2024-2026", "GROUP", "Продажи ИМ", "FILE-101"),
    ("Продажи ЦУМ 2024-06.2026.xlsx", "sales_tsum", "2024-2026", "GROUP", "Продажи ЦУМ", "FILE-102"),
    ("СберБизнес. Выписка за 2024.01.01-2024.12.31 счёт 40702810638040103938.xlsx", "bank_sber", "2024", "ООО_САЛОН_ЯНИНА", "Банк / Сбер", "FILE-103"),
    ("СберБизнес. Выписка за 2025.01.01-2025.12.31 счёт 40702810638040103938.xlsx", "bank_sber", "2025", "ООО_САЛОН_ЯНИНА", "Банк / Сбер", "FILE-104"),
    ("СберБизнес. Выписка за 2026.01.01-2026.06.30 счёт 40702810638040103938.xlsx", "bank_sber", "2026", "ООО_САЛОН_ЯНИНА", "Банк / Сбер", "FILE-105"),
    ("Справочник номенклатуры товара.xls", "sku_master", "2026", "GROUP", "Продукт / номенклатура", "FILE-106"),
    ("Финансы и платежи.docx", "org_raci", "2026", "GROUP", "Governance / RACI", "FILE-107"),
]


def nfc(s: str) -> str:
    return unicodedata.normalize("NFC", s or "")


def sha16(*parts) -> str:
    raw = "|".join("" if p is None else str(p) for p in parts)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def resolve(name: str) -> Path | None:
    t = nfc(name)
    for p in DOCS.iterdir():
        if nfc(p.name) == t:
            return p
    return None


def money(v) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    s = re.sub(r"[^\d.\-]", "", s)
    try:
        return float(s) if s else None
    except ValueError:
        return None


# ── Catalog ───────────────────────────────────────────────────────
def build_catalog() -> list[dict]:
    old = list(csv.DictReader(open(CAT93, encoding="utf-8-sig")))
    rows = list(old)
    for fname, cat, years, ent, domain, fid in NEW_FILES:
        path = resolve(fname)
        if not path:
            continue
        digest = sha256_file(path)
        rows.append(
            {
                "source_file_id": "SRC-" + digest[:12].upper(),
                "file_name": path.name,
                "file_name_nfc": nfc(path.name),
                "ext": path.suffix.lower(),
                "size_bytes": path.stat().st_size,
                "sha256": digest,
                "category": cat,
                "entities": ent,
                "years": years,
                "status": "Active",
                "path": str(path),
                "master_file_id": fid,
                "domain": domain,
                "period": years,
                "entity_scope": ent,
                "lifecycle_status": "Active source candidate",
                "risk": "Средний" if cat.startswith("bank") else "Низкий",
            }
        )
    fields = list(old[0].keys())
    write_csv(CAT, rows, fields)
    write_csv(ROOT / "live/registers/00_SOURCE_CATALOG_107.csv", rows, fields)
    # keep 93 as historical snapshot; also refresh pointer note in README later
    return rows


# ── Legal ─────────────────────────────────────────────────────────
def update_legal():
    p = W1 / "legal.csv"
    rows = list(csv.DictReader(open(p, encoding="utf-8")))
    if not any(r["legal_entity_id"] == "LE-OOO-SALON-YANINA" for r in rows):
        rows.append(
            {
                "legal_entity_id": "LE-OOO-SALON-YANINA",
                "inn": "7715219770",
                "name": 'ООО "Салон Юлия Янина"',
                "entity_type": "OOO",
                "valid_from": "2024-01-01",
                "valid_to": "",
                "source_evidence": "Sber statement header FILE-103..105",
                "notes": "Третье юрлицо; счёт 40702810638040103938 Сбер",
            }
        )
        write_csv(p, rows, list(rows[0].keys()))
    return rows


# ── Sber parser ───────────────────────────────────────────────────
def parse_sber(path: Path, source_file_id: str) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    out = []
    for r in range(12, (ws.max_row or 0) + 1):
        dt = ws.cell(r, 2).value
        if not isinstance(dt, datetime):
            continue
        debit_block = str(ws.cell(r, 5).value or "")
        credit_block = str(ws.cell(r, 9).value or "")
        amt_deb = money(ws.cell(r, 10).value)
        amt_cred = money(ws.cell(r, 14).value)
        doc_no = ws.cell(r, 15).value
        # purpose often further right — scan
        purpose = ""
        for c in range(16, min(24, (ws.max_column or 16) + 1)):
            v = ws.cell(r, c).value
            if v and isinstance(v, str) and len(v) > len(purpose):
                purpose = v.strip()

        direction = "out" if (amt_deb or 0) > 0 else "in"
        amount = amt_deb if direction == "out" else amt_cred
        if amount is None:
            continue

        # counterparty from the other side block
        cp_raw = credit_block if direction == "out" else debit_block
        lines = [x.strip() for x in cp_raw.split("\n") if x.strip()]
        cp_inn = ""
        cp_name = ""
        if len(lines) >= 3:
            cp_inn = re.sub(r"\D", "", lines[1])[:12]
            cp_name = lines[2]
        elif len(lines) == 2:
            cp_name = lines[1]

        pid = sha16(source_file_id, dt.isoformat(), doc_no, amount, direction)
        out.append(
            {
                "bank_payment_id": pid,
                "bank_account_id": "40702810638040103938",
                "legal_entity_id": "LE-OOO-SALON-YANINA",
                "payment_date": dt.strftime("%Y-%m-%d"),
                "period_month": dt.strftime("%Y-%m"),
                "doc_no": str(doc_no or ""),
                "direction": direction,
                "amount": round(float(amount), 2),
                "debit": round(float(amt_deb), 2) if amt_deb else "",
                "credit": round(float(amt_cred), 2) if amt_cred else "",
                "currency": "RUB",
                "counterparty_raw": cp_name or cp_raw.replace("\n", " | ")[:160],
                "counterparty_inn": cp_inn,
                "counterparty_id": "",
                "purpose": purpose[:200],
                "doc_type": "Sber statement",
                "source_file_id": source_file_id,
                "source_row_id": f"r{r}",
                "match_status": "UNMATCHED",
                "cash_line_id": "",
                "source_bank": "SBER",
                "is_internal": "N",
            }
        )
    wb.close()
    return out


# ── Card StatementFull ────────────────────────────────────────────
def parse_card(path: Path, source_file_id: str) -> list[dict]:
    tmp = Path(tempfile.gettempdir()) / f"card_{sha16(path.name)}.xlsx"
    shutil.copy2(path, tmp)
    wb = load_workbook(tmp, data_only=True)
    ws = wb.active
    header = str(ws.cell(1, 2).value or "")
    m_acc = re.search(r"счёту:\s*(\d{20})", header)
    account = m_acc.group(1) if m_acc else "40802810404000000049"
    out = []
    for r in range(3, (ws.max_row or 0) + 1):
        card = ws.cell(r, 2).value
        holder = ws.cell(r, 3).value
        dt = ws.cell(r, 4).value
        if not dt:
            continue
        if isinstance(dt, datetime):
            d = dt
        else:
            try:
                d = datetime.strptime(str(dt)[:19], "%Y-%m-%d %H:%M:%S")
            except ValueError:
                continue
        purpose = str(ws.cell(r, 7).value or "")
        income = money(ws.cell(r, 11).value) or 0.0
        expense = money(ws.cell(r, 12).value) or 0.0
        # expense cells often negative
        if expense < 0:
            expense = abs(expense)
        if income < 0:
            income = abs(income)
        if expense > 0:
            direction, amount = "out", expense
        elif income > 0:
            direction, amount = "in", income
        else:
            continue
        auth = ws.cell(r, 6).value
        pid = sha16(source_file_id, d.isoformat(), auth, amount, direction)
        out.append(
            {
                "bank_payment_id": pid,
                "bank_account_id": account,
                "legal_entity_id": "LE-IP-YANINA",
                "payment_date": d.strftime("%Y-%m-%d"),
                "period_month": d.strftime("%Y-%m"),
                "doc_no": str(auth or ""),
                "direction": direction,
                "amount": round(float(amount), 2),
                "debit": round(float(amount), 2) if direction == "out" else "",
                "credit": round(float(amount), 2) if direction == "in" else "",
                "currency": "RUB",
                "counterparty_raw": purpose[:80],
                "counterparty_inn": "",
                "counterparty_id": "",
                "purpose": purpose[:200],
                "doc_type": "VTB card StatementFull",
                "source_file_id": source_file_id,
                "source_row_id": f"r{r}",
                "match_status": "UNMATCHED",
                "cash_line_id": "",
                "source_bank": "VTB_CARD",
                "is_internal": "N",
                "card_mask": str(card or ""),
                "card_holder": str(holder or ""),
            }
        )
    wb.close()
    return out


# ── SKU master ────────────────────────────────────────────────────
def parse_sku_master(path: Path, source_file_id: str) -> list[dict]:
    sh = xlrd.open_workbook(path).sheet_by_index(0)
    out = []
    for r in range(1, sh.nrows):
        name = str(sh.cell_value(r, 0) or "").strip()
        unit = str(sh.cell_value(r, 2) or "").strip()
        article = str(sh.cell_value(r, 4) or "").strip()
        code = str(sh.cell_value(r, 5) or "").strip()
        composition = str(sh.cell_value(r, 6) or "").strip()
        barcode = str(sh.cell_value(r, 7) or "").strip()
        if not code and not name:
            continue
        out.append(
            {
                "sku_id": "SKU-" + sha16(code or article or name),
                "canonical_sku": article or code,
                "article": article,
                "code_1c": code,
                "name": name,
                "unit": unit,
                "composition": composition[:120],
                "barcode": barcode,
                "source_file_id": source_file_id,
                "status": "CANDIDATE_MASTER",
            }
        )
    return out


# ── Sales extended ────────────────────────────────────────────────
def parse_sales_tree(path: Path, channel: str, source_file_id: str) -> list[dict]:
    """1C tree: SKU row → buyer → document."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    out = []
    cur_sku = cur_buyer = None
    for i, r in enumerate(rows[3:], start=4):  # skip headers
        if not r:
            continue
        c0 = r[0]
        # SKU line: code like 00-0000
        if isinstance(c0, str) and re.match(r"^\d{2}-\d+", c0):
            cur_sku = {
                "code": c0,
                "article": r[1],
                "name": r[2],
                "qty": money(r[3]),
                "price": money(r[4]),
                "revenue": money(r[5]),
                "cogs": money(r[9]) if len(r) > 9 else None,
            }
            cur_buyer = None
            continue
        if cur_sku and isinstance(c0, str) and c0 and not str(c0).startswith("Расход") and "Чек" not in str(c0) and "Отчет" not in str(c0) and "Наклад" not in str(c0):
            # buyer row often has name in col0 and totals
            if not re.match(r"^\d{2}-\d+", c0):
                cur_buyer = c0
                continue
        doc = None
        if isinstance(c0, str) and any(k in c0 for k in ("Расходная", "Чек", "Отчет", "Накладная", "Реализация")):
            doc = c0
        if cur_sku and doc:
            # date from doc
            m = re.search(r"от\s+(\d{2}\.\d{2}\.\d{4})", doc)
            sale_date = ""
            period = ""
            if m:
                d = datetime.strptime(m.group(1), "%d.%m.%Y")
                sale_date = d.strftime("%Y-%m-%d")
                period = d.strftime("%Y-%m")
            rev = money(r[5]) if len(r) > 5 else cur_sku.get("revenue")
            cogs = money(r[9]) if len(r) > 9 else cur_sku.get("cogs")
            qty = money(r[3]) if len(r) > 3 else cur_sku.get("qty")
            out.append(
                {
                    "sales_line_id": "SL-" + sha16(source_file_id, doc, cur_sku["code"], i),
                    "channel": channel,
                    "code_1c": cur_sku["code"],
                    "article": cur_sku.get("article") or "",
                    "sku_name": cur_sku.get("name") or "",
                    "buyer": cur_buyer or "",
                    "document": doc,
                    "sale_date": sale_date,
                    "period_month": period,
                    "qty": qty if qty is not None else "",
                    "revenue_rub": rev if rev is not None else "",
                    "cogs_rub": cogs if cogs is not None else "",
                    "source_file_id": source_file_id,
                    "source_row_id": f"r{i}",
                }
            )
    return out


# ── Fabric movement summary ───────────────────────────────────────
def summarize_fabric(path: Path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    # row4 has warehouse totals in cols
    rows = []
    for i, r in enumerate(ws.iter_rows(max_row=20, values_only=True)):
        rows.append(r)
    wb.close()
    # from probe: row index 3 (0-based) Основной склад amounts
    wh = None
    for r in rows:
        if r and r[0] and "склад" in str(r[0]).lower():
            wh = {
                "warehouse": r[0],
                "in_rub": money(r[2]),
                "out_rub": money(r[3]),
                "end_rub": money(r[4]),
            }
            break
    return {"fabric_file": path.name, "warehouse_totals": wh, "approx_rows": 19937}


def summarize_stock_cost(path: Path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    n = 0
    in_cost = 0.0
    for i, r in enumerate(ws.iter_rows(min_row=4, values_only=True)):
        n += 1
        if r and r[0] and re.match(r"^[A-Za-z0-9\-]+$", str(r[0])) and money(r[4]):
            in_cost += money(r[4]) or 0
        if n > 200000:
            break
    wb.close()
    return {"stock_cost_file": path.name, "approx_rows": 5935, "sample_in_cost_sum_first_pass": round(in_cost, 2)}


# ── RACI from docx ────────────────────────────────────────────────
def apply_raci_from_docx():
    path = resolve("Финансы и платежи.docx")
    if not path or not OWNER.exists():
        return []
    try:
        from docx import Document
    except ImportError:
        return []
    text = "\n".join(p.text for p in Document(str(path)).paragraphs)
    candidates = []
    if "Мамушкина" in text:
        candidates.append(
            {
                "fio": "Мамушкина Елена",
                "role_hint": "Исполнительный директор / Owner CASH (наличные)",
                "domains": ["Cash", "SRC-CTRL-04 bank/cash", "Payroll cash side"],
            }
        )
    if "Сливяк" in text:
        candidates.append(
            {
                "fio": "Сливяк Галина",
                "role_hint": "Главный бухгалтер / Owner BANK (безналичные)",
                "domains": ["Bank", "Tax", "SRC-CTRL-04", "BUD/TAX"],
            }
        )

    wb = load_workbook(OWNER)
    # fill RACI named_person where empty for matching roles — ONLY suggest, leave decision blank
    if "RACI" in wb.sheetnames:
        ws = wb["RACI"]
        headers = [c.value for c in ws[1]]
        # map columns
        col = {h: i + 1 for i, h in enumerate(headers) if h}

        def set_if_empty(row, key, val):
            c = col.get(key)
            if not c:
                return
            if ws.cell(row, c).value in (None, ""):
                ws.cell(row, c, val)

        for r in range(2, ws.max_row + 1):
            item = str(ws.cell(r, col.get("item", 2)).value or "")
            role = str(ws.cell(r, col.get("role", 3)).value or "")
            low = item.lower()
            # cash / налич
            if any(k in low for k in ("cash", "касс", "налич")) or "карта" in low:
                if role == "Owner":
                    set_if_empty(r, "named_person_FIO", "Мамушкина Елена")
            # bank / безнал / tax
            if any(k in low for k in ("bank", "банк", "безнал", "tax", "налог", "payroll", "зп", "budget", "бюджет")):
                if role == "Owner":
                    set_if_empty(r, "named_person_FIO", "Сливяк Галина")
            # bank perimeter specifically
            if "bank perimeter" in low or "vtb" in low or "выписк" in low:
                if role == "Owner":
                    set_if_empty(r, "named_person_FIO", "Сливяк Галина")
                if role == "Approver":
                    set_if_empty(r, "approver_FIO", "Мамушкина Елена")

    # candidates sheet
    name = "CANDIDATES_FROM_DOCS"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, 1)
    fill = PatternFill("solid", fgColor="1F4E79")
    hdr = Font(color="FFFFFF", bold=True)
    warn = PatternFill("solid", fgColor="FFF2CC")
    ws["A1"] = "Кандидаты из новых документов — decision_ACCEPT_REJECT всё ещё за вами"
    ws["A1"].font = Font(bold=True, size=12, color="1F4E79")
    ws["A2"] = f"Обновлено H3 {NOW}. Источник: Финансы и платежи.docx + card StatementFull (Мамушкина Е.А.)"
    headers = ["fio", "role_hint", "domains", "source", "status", "action"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(4, c, h)
        cell.fill = fill
        cell.font = hdr
    rows_c = [
        ("Мамушкина Елена", "Исполнительный директор / CASH Owner", "Cash, card, наличные", "Финансы и платежи.docx + VTB StatementFull holder", "PREFILLED_RACI_DRAFT", "Проставить ACCEPT/REJECT в RACI"),
        ("Сливяк Галина", "Главный бухгалтер / BANK+TAX Owner", "Bank, Tax, Budget, безнал", "Финансы и платежи.docx", "PREFILLED_RACI_DRAFT", "Проставить ACCEPT/REJECT в RACI"),
        ("ЯНИН ЕВГЕНИЙ НИКОЛАЕВИЧ", "Tax signer (ДЕКОР)", "Tax DEKOR", "tax PDF H2", "CANDIDATE", "Подтвердить роль"),
        ("ЯНИНА ЮЛИЯ ФЕДОРОВНА", "Tax signer / собственник ИП", "Tax IP, Approver", "tax PDF H2", "CANDIDATE", "Подтвердить роль Approver"),
    ]
    for ri, row in enumerate(rows_c, 5):
        for ci, v in enumerate(row, 1):
            cell = ws.cell(ri, ci, v)
            if ci == 1:
                cell.fill = warn
    if "README" in wb.sheetnames:
        wb["README"]["A7"] = f"H3 {NOW}: 14 новых файлов; RACI draft prefilled Мамушкина/Сливяк — нужен ACCEPT"
    wb.save(OWNER)
    return candidates


def merge_bank(payments: list[dict]):
    """Append new SBER/CARD payments into W1 bank_payments (backup once)."""
    bak = W1 / "bank_payments_pre_h3.csv"
    src = W1 / "bank_payments.csv"
    if not bak.exists():
        shutil.copy2(src, bak)
    existing = list(csv.DictReader(open(src, encoding="utf-8")))
    ids = {r["bank_payment_id"] for r in existing}
    fields = list(existing[0].keys())
    added = 0
    for p in payments:
        if p["bank_payment_id"] in ids:
            continue
        row = {k: p.get(k, "") for k in fields}
        existing.append(row)
        ids.add(p["bank_payment_id"])
        added += 1
    write_csv(src, existing, fields)
    # accounts
    acc = list(csv.DictReader(open(W1 / "bank_accounts.csv", encoding="utf-8")))
    have = {a["bank_account_id"] for a in acc}
    if "40702810638040103938" not in have:
        acc.append(
            {
                "bank_account_id": "40702810638040103938",
                "legal_entity_id": "LE-OOO-SALON-YANINA",
                "bank_name": "ПАО Сбербанк",
                "currency": "RUB",
                "status": "ACTIVE_PARSED_SBER_H3",
                "notes": "FILE-103..105 | ООО Салон Юлия Янина ИНН 7715219770",
            }
        )
    # card account already exists as VTB PDF — update notes
    for a in acc:
        if a["bank_account_id"] == "40802810404000000049":
            a["notes"] = (a.get("notes") or "") + " | +StatementFull card H3 (Мамушкина Е.А.)"
            a["status"] = "ACTIVE_PARSED_VTB_PDF_PLUS_CARD"
    write_csv(W1 / "bank_accounts.csv", acc, list(acc[0].keys()))
    return added, len(existing)


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)

    catalog = build_catalog()
    legal = update_legal()
    raci = apply_raci_from_docx()

    # map file id
    fid = {nfc(fn): fid for fn, _, _, _, _, fid in NEW_FILES}

    sber_all = []
    for fname, cat, years, ent, domain, file_id in NEW_FILES:
        path = resolve(fname)
        if not path:
            continue
        if cat == "bank_sber":
            sber_all.extend(parse_sber(path, file_id))

    card_all = []
    for fname, cat, years, ent, domain, file_id in NEW_FILES:
        path = resolve(fname)
        if path and cat == "bank_card":
            card_all.extend(parse_card(path, file_id))

    sku = []
    path = resolve("Справочник номенклатуры товара.xls")
    if path:
        sku = parse_sku_master(path, "FILE-106")

    sales = []
    for fname, channel, fid_ in [
        ("Продажи B2B 2024-06.2026.xlsx", "B2B", "FILE-100"),
        ("Продажи ИМ 2024-06.2026.xlsx", "IM", "FILE-101"),
        ("Продажи ЦУМ 2024-06.2026.xlsx", "TSUM", "FILE-102"),
    ]:
        p = resolve(fname)
        if p:
            sales.extend(parse_sales_tree(p, channel, fid_))

    fabric = summarize_fabric(resolve("Движение тканей.xlsx")) if resolve("Движение тканей.xlsx") else {}
    stock_cost = summarize_stock_cost(resolve("Движение товаров с себестоимостью.xlsx")) if resolve("Движение товаров с себестоимостью.xlsx") else {}

    bank_fields = [
        "bank_payment_id", "bank_account_id", "legal_entity_id", "payment_date", "period_month",
        "doc_no", "direction", "amount", "debit", "credit", "currency", "counterparty_raw",
        "counterparty_inn", "counterparty_id", "purpose", "doc_type", "source_file_id",
        "source_row_id", "match_status", "cash_line_id", "source_bank", "is_internal",
    ]
    write_csv(OUT / "sber_payments.csv", sber_all, bank_fields)
    write_csv(
        OUT / "card_payments.csv",
        card_all,
        bank_fields + ["card_mask", "card_holder"],
    )
    write_csv(
        OUT / "sku_master.csv",
        sku,
        ["sku_id", "canonical_sku", "article", "code_1c", "name", "unit", "composition", "barcode", "source_file_id", "status"],
    )
    write_csv(
        OUT / "sales_extended_2024_2026.csv",
        sales,
        [
            "sales_line_id", "channel", "code_1c", "article", "sku_name", "buyer", "document",
            "sale_date", "period_month", "qty", "revenue_rub", "cogs_rub", "source_file_id", "source_row_id",
        ],
    )

    # merge bank
    added, total_bank = merge_bank(sber_all + [{k: p.get(k, "") for k in bank_fields} for p in card_all])

    # sales by channel stats
    by_ch = Counter(s["channel"] for s in sales)
    rev_by_ch = defaultdict(float)
    for s in sales:
        rev_by_ch[s["channel"]] += float(s["revenue_rub"] or 0)
    months = sorted({s["period_month"] for s in sales if s["period_month"]})

    # sku overlap with articles that have values
    with_art = sum(1 for s in sku if s["article"])

    summary = {
        "generated_at": NOW,
        "wave": "H3",
        "new_files": 14,
        "catalog_total": len(catalog),
        "legal_entities": [r["legal_entity_id"] for r in legal],
        "sber_payments": len(sber_all),
        "sber_out_rub": round(sum(p["amount"] for p in sber_all if p["direction"] == "out"), 2),
        "sber_in_rub": round(sum(p["amount"] for p in sber_all if p["direction"] == "in"), 2),
        "card_payments": len(card_all),
        "card_holders": sorted({p.get("card_holder", "") for p in card_all if p.get("card_holder")}),
        "card_out_rub": round(sum(p["amount"] for p in card_all if p["direction"] == "out"), 2),
        "bank_payments_added": added,
        "bank_payments_total": total_bank,
        "sku_master_rows": len(sku),
        "sku_with_article": with_art,
        "sales_lines": len(sales),
        "sales_by_channel": dict(by_ch),
        "sales_revenue_by_channel": {k: round(v, 2) for k, v in rev_by_ch.items()},
        "sales_months": months[:3] + ["..."] + months[-3:] if len(months) > 6 else months,
        "fabric": fabric,
        "stock_cost": stock_cost,
        "raci_docx": raci,
        "finding": (
            f"H3: +14 files → catalog {len(catalog)}; "
            f"NEW LE Салон Янина; Sber {len(sber_all)} pays; "
            f"card {len(card_all)} (holder Мамушкина); "
            f"SKU master {len(sku)}; sales ext {len(sales)}; "
            f"RACI draft Мамушкина/Сливяк."
        ),
        "next": "Owner ACCEPT in RACI; optional rebuild W4 from sales_extended; deep fabric/stock parsers",
        "not_sot": True,
    }
    json.dump(summary, open(OUT / "h3_summary.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    json.dump(summary, open(EV / "h3_summary.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)

    wb = Workbook()
    ws = wb.active
    ws.title = "00_Summary"
    for i, (k, v) in enumerate(summary.items(), 1):
        ws.cell(i, 1, k)
        ws.cell(i, 2, json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v)
    fill = PatternFill("solid", fgColor="1F4E79")

    def add(name, rows):
        w = wb.create_sheet(name)
        if not rows:
            return
        headers = list(rows[0].keys())
        for c, h in enumerate(headers, 1):
            cell = w.cell(1, c, h)
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
        for ri, row in enumerate(rows[:4000], 2):
            for ci, h in enumerate(headers, 1):
                w.cell(ri, ci, row.get(h, ""))

    add("01_Sber", sber_all[:3000])
    add("02_Card", card_all)
    add("03_SKU", sku[:3000])
    add("04_Sales", sales[:3000])
    add("05_New_Catalog", [r for r in catalog if r["master_file_id"] >= "FILE-094"])
    wb.save(EV / "YANINA_H3_NEW_DOCS_EVIDENCE.xlsx")

    (OUT / "README.md").write_text(
        f"""# H3 New Documents Ingest

Generated: {NOW}

- Catalog: `../00_SOURCE_CATALOG_107.csv` ({len(catalog)} files)
- Sber: {len(sber_all)} payments → merged into W1
- Card StatementFull: {len(card_all)}
- SKU master: {len(sku)}
- Sales extended: {len(sales)}
- RACI draft from docx: Мамушкина / Сливяк (нужен ACCEPT)

Evidence: `../../evidence/h3_new_docs_20260724/`
""",
        encoding="utf-8",
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
