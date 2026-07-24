#!/usr/bin/env python3
"""
H20: Finance exception recommendations (гипотезы, не SoT).

Выбран автономный путь #2: без RACI/новых файлов дать конкретные
рекомендации по P1 exceptions, чтобы Owners могли ACCEPT одним решением.

SKU:
- 0-2493A / 0-2496 / 0-2497 — B2B @ ~10 000 vs stock cost 12–13.5K
- 0-3243 — quarantine (свитшот ≠ худи)

НЕ меняет COGS/маржу/RACI. Только recommendations + packet sheet.
"""
from __future__ import annotations

import csv
import json
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[3]
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/h20_finance_recs_20260724"
MART = ROOT / "live/marts"
W3 = ROOT / "live/registers/w3_sku_cost"
W4 = ROOT / "live/registers/w4_sales_settle"
PACKET = ROOT / "live/YANINA_OWNER_PACKET_RACI_AND_REQUESTS.xlsx"
CC = ROOT / "live/YANINA_LIVE_CONTROL_CENTER_20260723.xlsx"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF")


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def fnum(v):
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(" ", "").replace(",", "."))
    except (TypeError, ValueError):
        return None


def sibling_channel_prices() -> dict:
    """Цены соседних 0-249* по каналам — контекст wholesale ladder."""
    out = {}
    for r in csv.DictReader(open(W4 / "sales_lines.csv", encoding="utf-8")):
        sku = r.get("canonical_sku") or ""
        if not sku.startswith("0-249") or r.get("dq_exclude_from_margin") == "Y":
            continue
        qty = fnum(r.get("qty")) or 0
        rev = fnum(r.get("revenue_rub")) or 0
        if qty <= 0:
            continue
        ch = r.get("channel") or ""
        out.setdefault(sku, {}).setdefault(ch, []).append(rev / qty)
    # median-ish
    summary = {}
    for sku, chs in out.items():
        summary[sku] = {
            ch: round(sorted(vals)[len(vals) // 2], 2) for ch, vals in chs.items() if vals
        }
    return summary


def b2b_10k_pattern() -> dict:
    n = 0
    n10 = 0
    for r in csv.DictReader(open(W4 / "sales_lines.csv", encoding="utf-8")):
        if r.get("channel") != "B2B" or r.get("dq_exclude_from_margin") == "Y":
            continue
        n += 1
        qty = fnum(r.get("qty")) or 1
        up = (fnum(r.get("revenue_rub")) or 0) / qty
        if 9999 <= up <= 10001:
            n10 += 1
    return {"b2b_lines": n, "unit_price_approx_10000": n10}


def build_recommendations() -> list[dict]:
    fin = {
        r["canonical_sku"]: r
        for r in csv.DictReader(open(MART / "finance_neg_sku_review.csv", encoding="utf-8"))
    }
    sib = sibling_channel_prices()
    pat = b2b_10k_pattern()
    rows = []

    # --- B2B wholesale trio ---
    for sku in ("0-2497", "0-2496", "0-2493A"):
        f = fin.get(sku, {})
        sib_2493 = sib.get("0-2493", {})
        evidence = (
            f"B2B unit≈{f.get('unit_price_rub')} vs cost≈{f.get('unit_cost_rub')} "
            f"(gap {f.get('unit_gap_rub')}); cost source FILE/H5 DERIVED_STOCK_MOVEMENT; "
            f"sibling 0-2493 ladder B2B={sib_2493.get('B2B')} / IM={sib_2493.get('IM')} / "
            f"TSUM={sib_2493.get('TSUM')}; B2B lines with ~10k price: "
            f"{pat['unit_price_approx_10000']}/{pat['b2b_lines']}."
        )
        rows.append(
            {
                "rec_id": f"REC-{sku}",
                "action_id": f"A-FIN-{sku}",
                "canonical_sku": sku,
                "topic": f.get("name", "")[:90],
                "channel": "B2B",
                "impact_revenue_rub": f.get("revenue_rub"),
                "impact_margin_rub": f.get("margin_rub"),
                "proposed_decision": "OK_COMMERCIAL_LOSS",
                "confidence": "HIGH",
                "rationale": (
                    "Цена ровно ~10 000 на всех строках = оптовый/партнёрский прайс, "
                    "не случайная ошибка ввода. У соседнего 0-2493 B2B≈10K при IM/TSUM "
                    "в разы выше — каналная лестница. Автоправка cost запрещена: "
                    "cost из остатков (DERIVED), не BOM; риск сломать маржу."
                ),
                "do_not": "Не подменять cost на BOM без версии; не исключать из продаж",
                "optional_followup": "Если есть BOM FULL по артикулу — сверить unit_cost vs stock",
                "margin_treatment_if_accepted": "FLAG_WHOLESALE_EXCEPTION (оставить в revenue, пометить margin)",
                "evidence": evidence,
                "owner_decision_ACCEPT_REJECT": "",  # fills later
                "so_t": "N",
                "status": "PROPOSED",
            }
        )

    # --- 0-3243 quarantine ---
    f = fin.get("0-3243", {})
    # confirm no Be a poem cost
    poem_costs = 0
    hoodie_cost = ""
    for r in csv.DictReader(open(W3 / "cost_versions.csv", encoding="utf-8")):
        if r.get("canonical_sku") == "0-3243":
            hoodie_cost = f"{r.get('name')}|{r.get('unit_cost_rub')}"
        name = (r.get("name") or "").lower()
        if "poem" in name or "be a poem" in name:
            poem_costs += 1
    rows.append(
        {
            "rec_id": "REC-0-3243",
            "action_id": "A-FIN-0-3243",
            "canonical_sku": "0-3243",
            "topic": f.get("name", "")[:90],
            "channel": "IM",
            "impact_revenue_rub": f.get("revenue_rub"),
            "impact_margin_rub": "",
            "proposed_decision": "KEEP_QUARANTINE_NEED_COST_VERSION",
            "confidence": "HIGH",
            "rationale": (
                "Продажа = свитшот «Be a poem»; единственный cost 0-3243 = худи "
                f"({hoodie_cost}). 0-3244 = юбка ~43160 — ещё хуже. "
                f"Cost с 'poem' в имени: {poem_costs}. Релинк запрещён."
            ),
            "do_not": "Не брать cost 0-3244; не возвращать худи-COGS на свитшот",
            "optional_followup": "Запросить у производства cost version свитшота 0-3243 / Be a poem",
            "margin_treatment_if_accepted": "KEEP_EXCLUDED_FROM_COGS_UNTIL_COST_EXISTS",
            "evidence": (
                "quarantine SL-ed909efc193b9f4c; old_cogs 33555.87 (hoodie) removed; "
                "revenue 30250 kept in sales, COGS blank"
            ),
            "owner_decision_ACCEPT_REJECT": "",
            "so_t": "N",
            "status": "PROPOSED",
        }
    )

    # portfolio note
    loss = sum(fnum(r["impact_margin_rub"]) or 0 for r in rows if r["canonical_sku"] != "0-3243")
    rows.append(
        {
            "rec_id": "REC-PORTFOLIO",
            "action_id": "A-FIN-WATCH-01",
            "canonical_sku": "PORTFOLIO",
            "topic": "Суммарный эффект B2B wholesale exceptions",
            "channel": "B2B",
            "impact_revenue_rub": "160000.09",
            "impact_margin_rub": round(loss, 2),
            "proposed_decision": "ACCEPT_AS_POLICY_EXCEPTION_SET",
            "confidence": "HIGH",
            "rationale": (
                f"Суммарная отриц. маржа тройки ≈ {loss:,.0f} ₽ на ~160K выручки. "
                "На overall GM ~53% влияние точечное. Лучше политика исключения, "
                "чем «чинить» cost."
            ),
            "do_not": "Не раздувать в системный rewrite cost master",
            "optional_followup": "После RACI — завести margin_exception register",
            "margin_treatment_if_accepted": "POLICY_FLAG",
            "evidence": "finance_neg_sku_review + sales_lines B2B May/Jun 2025-2026",
            "owner_decision_ACCEPT_REJECT": "",
            "so_t": "N",
            "status": "PROPOSED",
        }
    )
    return rows


def write_md(recs: list[dict]):
    lines = [
        "# Finance Recommendations (H20)",
        "",
        f"Updated: {NOW}",
        "",
        "Гипотезы для Owners. **Не SoT. COGS не менялись.**",
        "",
    ]
    for r in recs:
        lines.append(f"## {r['rec_id']} — `{r['canonical_sku']}`")
        lines.append(f"- Proposed: **{r['proposed_decision']}** (confidence {r['confidence']})")
        lines.append(f"- Action: `{r['action_id']}`")
        lines.append(f"- Impact: revenue {r['impact_revenue_rub']} / margin {r['impact_margin_rub']}")
        lines.append(f"- Why: {r['rationale']}")
        lines.append(f"- Do not: {r['do_not']}")
        lines.append(f"- If accepted: {r['margin_treatment_if_accepted']}")
        lines.append(f"- Evidence: {r['evidence']}")
        lines.append("")
    lines.extend(
        [
            "## Как принять",
            "",
            "1. Открыть Owner Packet → лист `RECOMMENDATIONS_H20`",
            "2. Колонка `owner_decision_ACCEPT_REJECT` = ACCEPT или REJECT",
            "3. Прислать файл в чат — после ACCEPT можно пометить exceptions в marts",
            "",
        ]
    )
    text = "\n".join(lines)
    (OUT / "FINANCE_RECOMMENDATIONS.md").write_text(text, encoding="utf-8")
    (ROOT / "live/FINANCE_RECOMMENDATIONS.md").write_text(text, encoding="utf-8")
    (EV / "FINANCE_RECOMMENDATIONS.md").write_text(text, encoding="utf-8")


def update_packet(recs: list[dict]):
    if not PACKET.exists():
        return
    bak = EV / "YANINA_OWNER_PACKET_RACI_AND_REQUESTS.pre_h20.xlsx"
    if not bak.exists():
        shutil.copy2(PACKET, bak)
    wb = load_workbook(PACKET)
    name = "RECOMMENDATIONS_H20"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, 3)  # near RACI
    ws["A1"] = "H20 Finance Recommendations — PROPOSED (не применено)"
    ws["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws["A2"] = f"{NOW} | Заполните owner_decision_ACCEPT_REJECT. AI не проставляет ACCEPT."
    fields = list(recs[0].keys())
    ws.append([])
    ws.append(fields)
    for c in range(1, len(fields) + 1):
        cell = ws.cell(4, c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
    for r in recs:
        ws.append([r.get(f, "") for f in fields])
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        ws.column_dimensions[letter].width = min(42, max(12, len(str(col[0].value or "")) + 2))
    # README bump
    if "README" in wb.sheetnames:
        readme = wb["README"]
        readme["A13"] = (
            f"H20 {NOW}: лист RECOMMENDATIONS_H20 — proposed OK_COMMERCIAL_LOSS "
            "для 0-2493A/2496/2497; KEEP_QUARANTINE для 0-3243."
        )
    wb.save(PACKET)
    shutil.copy2(PACKET, EV / "YANINA_OWNER_PACKET_RACI_AND_REQUESTS.xlsx")


def update_cc(summary: dict):
    if not CC.exists():
        return
    wb = load_workbook(CC)
    if "H20_FinanceRec" in wb.sheetnames:
        del wb["H20_FinanceRec"]
    ws = wb.create_sheet("H20_FinanceRec", 0)
    ws["A1"] = "H20 Finance Recommendations"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A2"] = NOW
    ws["A4"] = summary["finding"]
    ws["A5"] = "Doc"
    ws["B5"] = "live/FINANCE_RECOMMENDATIONS.md"
    ws["A6"] = "Packet sheet"
    ws["B6"] = "RECOMMENDATIONS_H20"
    ws["A7"] = "Applied to COGS?"
    ws["B7"] = "NO"
    wb.save(CC)


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)

    recs = build_recommendations()
    fields = list(recs[0].keys())
    write_csv(MART / "finance_recommendations.csv", recs, fields)
    write_csv(OUT / "finance_recommendations.csv", recs, fields)
    write_md(recs)
    update_packet(recs)

    b2b = [r for r in recs if r["canonical_sku"] not in ("0-3243", "PORTFOLIO")]
    summary = {
        "wave": "H20",
        "generated_at": NOW,
        "path_choice": "Option 2 — finance exception recommendations without RACI/new files",
        "finding": (
            "H20: proposed OK_COMMERCIAL_LOSS for 0-2493A/2496/2497 "
            f"(combined margin impact {sum(fnum(r['impact_margin_rub']) or 0 for r in b2b):,.0f} ₽); "
            "KEEP_QUARANTINE for 0-3243. COGS unchanged."
        ),
        "recommendations_n": len(recs),
        "proposed_ok_commercial_loss": [r["canonical_sku"] for r in b2b],
        "proposed_keep_quarantine": ["0-3243"],
        "cogs_changed": False,
        "raci_changed": False,
        "not_sot": True,
    }
    (OUT / "h20_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    (EV / "h20_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    shutil.copy2(MART / "finance_recommendations.csv", EV / "finance_recommendations.csv")
    update_cc(summary)
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
