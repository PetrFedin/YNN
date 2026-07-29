#!/usr/bin/env python3
"""H63: Budget 2025H2 / 2026 plan vs fact extract + opex crosswalk.

Budgets already contain Plan/Fact — we normalize to marts and map articles to opex buckets.
Units: treated as EUR (workbook scale); RUB view = EUR * fx (default 100). Not audited SoT.
"""
from __future__ import annotations

import csv
import json
import re
import shutil
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[3]
if not (ROOT / "live/registers/00_SOURCE_CATALOG_107.csv").exists():
    ROOT = Path.cwd()
REG = Path(__file__).resolve().parent
MARTS = ROOT / "live/marts"
MAPS = ROOT / "live/maps"
EV = ROOT / "live/evidence/h63_budget_vs_fact_20260729"
WAVE_C = ROOT / "live/client_pack/execution_wave_c"

FX_DEFAULT = 100.0

ARTICLE_TO_OPEX = [
    (r"аренда", "RENT"),
    (r"помещен", "RENT"),
    (r"аутсорсинг|производств", "OUTSOURCE"),
    (r"материал|ткан|фурнитур", "MATERIALS_MEMO"),
    (r"заработн|зарплат|карты|наличн", "PAYROLL"),
    (r"ндфл|страх\.|налоги с зарп", "TAX"),
    (r"налоги с доход", "TAX"),
    (r"транспорт|логист|команд", "LOGISTICS"),
    (r"эквайр|услуг.*банк", "ACQUIRING_FEE"),
    (r"сайт|соцсет|маркетинг|съемк|стилист|мероприят|подарк|представитель", "MARKETING"),
    (r"офисн", "OPEX_OTHER"),
    (r"юридич", "OPEX_OTHER"),
    (r"благотвор", "OPEX_OTHER"),
    (r"собственник", "INTERNAL_TRANSFER_MEMO"),
]


def write_csv(path: Path, fields, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)


def map_opex(article: str) -> str:
    a = (article or "").lower()
    for pat, bucket in ARTICLE_TO_OPEX:
        if re.search(pat, a, re.I):
            return bucket
    return "UNMAPPED"


def to_float(v):
    if v is None or str(v).strip() == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", ".").replace(" ", ""))
    except ValueError:
        return None


def month_key(dt) -> str:
    if isinstance(dt, datetime):
        return dt.strftime("%Y-%m")
    return ""


def parse_h2_2025(path: Path, meta: dict) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    # row1: months at cols 1,3,5,7,9,11
    months = []
    header = rows[1]
    for i in range(1, 12, 2):
        months.append(month_key(header[i]))
    # cols: for month idx m: plan=1+2m, fact=2+2m
    out = []
    for ri, row in enumerate(rows[3:], start=4):
        article = row[0]
        if not article or not str(article).strip():
            continue
        article = str(article).strip()
        if article.lower().startswith("итого") and "план" in article.lower():
            continue
        for mi, pm in enumerate(months):
            if not pm:
                continue
            plan = to_float(row[1 + 2 * mi])
            fact = to_float(row[2 + 2 * mi])
            if plan is None and fact is None:
                continue
            var = None if plan is None or fact is None else fact - plan
            var_pct = None if plan in (None, 0) or fact is None else round(100 * (fact - plan) / abs(plan), 1)
            out.append(
                {
                    "budget_id": "BUD2025H2",
                    "period_month": pm,
                    "article": article,
                    "section": "expense" if "поступ" not in article.lower() else "income",
                    "plan_eur": plan if plan is not None else "",
                    "fact_eur": fact if fact is not None else "",
                    "variance_eur": round(var, 2) if var is not None else "",
                    "variance_pct": var_pct if var_pct is not None else "",
                    "plan_rub_fx100": round(plan * FX_DEFAULT, 2) if plan is not None else "",
                    "fact_rub_fx100": round(fact * FX_DEFAULT, 2) if fact is not None else "",
                    "opex_bucket_map": map_opex(article),
                    "source_file": meta["file_name"],
                    "source_file_id": meta["source_file_id"],
                    "unit_assumption": "EUR_LIKE",
                    "so_t": "N",
                }
            )
    return out


def parse_2026(path: Path, meta: dict) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    # row1: months start col 2,4,6... possibly through Dec — read until no date
    months = []
    header = rows[1]
    for i in range(2, len(header), 2):
        mk = month_key(header[i])
        if not mk:
            break
        months.append(mk)
    out = []
    section = "unknown"
    for ri, row in enumerate(rows[3:], start=4):
        article = row[1] if len(row) > 1 else None
        if article is None or str(article).strip() == "":
            continue
        article = str(article).strip()
        if article.lower().startswith("пояснен"):
            break
        if article.lower() in {"поступления"}:
            section = "income"
        elif article.lower() in {"расходы"}:
            section = "expense"
        for mi, pm in enumerate(months):
            plan = to_float(row[2 + 2 * mi]) if 2 + 2 * mi < len(row) else None
            fact = to_float(row[3 + 2 * mi]) if 3 + 2 * mi < len(row) else None
            if plan is None and fact is None:
                continue
            var = None if plan is None or fact is None else fact - plan
            var_pct = None if plan in (None, 0) or fact is None else round(100 * (fact - plan) / abs(plan), 1)
            out.append(
                {
                    "budget_id": "BUD2026",
                    "period_month": pm,
                    "article": article,
                    "section": section,
                    "plan_eur": plan if plan is not None else "",
                    "fact_eur": fact if fact is not None else "",
                    "variance_eur": round(var, 2) if var is not None else "",
                    "variance_pct": var_pct if var_pct is not None else "",
                    "plan_rub_fx100": round(plan * FX_DEFAULT, 2) if plan is not None else "",
                    "fact_rub_fx100": round(fact * FX_DEFAULT, 2) if fact is not None else "",
                    "opex_bucket_map": map_opex(article),
                    "source_file": meta["file_name"],
                    "source_file_id": meta["source_file_id"],
                    "unit_assumption": "EUR_LIKE",
                    "so_t": "N",
                }
            )
    return out


def main():
    for d in (REG, MARTS, MAPS, EV, WAVE_C):
        d.mkdir(parents=True, exist_ok=True)

    cat = list(csv.DictReader((ROOT / "live/registers/00_SOURCE_CATALOG_107.csv").open()))
    lines = []
    for r in cat:
        if r["category"] != "budget":
            continue
        p = Path(r["path"])
        if not p.exists():
            p = ROOT / "documents" / r["file_name"]
        if "2026" in r["file_name"]:
            lines.extend(parse_2026(p, r))
        else:
            lines.extend(parse_h2_2025(p, r))

    # monthly totals income/expense
    month_tot = defaultdict(lambda: {"plan_in": 0.0, "fact_in": 0.0, "plan_ex": 0.0, "fact_ex": 0.0, "n": 0})
    for L in lines:
        k = (L["budget_id"], L["period_month"])
        month_tot[k]["n"] += 1
        pe, fe = L["plan_eur"], L["fact_eur"]
        pe = float(pe) if pe != "" else None
        fe = float(fe) if fe != "" else None
        # only top-level income/expense headers and leaf? Use section; skip rollup names for double count
        art = L["article"].lower()
        if art in {"поступления", "расходы"}:
            if art == "поступления":
                if pe is not None:
                    month_tot[k]["plan_in"] = pe
                if fe is not None:
                    month_tot[k]["fact_in"] = fe
            else:
                if pe is not None:
                    month_tot[k]["plan_ex"] = pe
                if fe is not None:
                    month_tot[k]["fact_ex"] = fe

    month_rows = []
    for (bid, pm), v in sorted(month_tot.items()):
        net_plan = v["plan_in"] - v["plan_ex"] if v["plan_in"] or v["plan_ex"] else ""
        net_fact = v["fact_in"] - v["fact_ex"] if v["fact_in"] or v["fact_ex"] else ""
        if isinstance(net_plan, float):
            net_plan = round(net_plan, 2)
        if isinstance(net_fact, float):
            net_fact = round(net_fact, 2)
        month_rows.append(
            {
                "budget_id": bid,
                "period_month": pm,
                "plan_income_eur": round(v["plan_in"], 2) if v["plan_in"] else "",
                "fact_income_eur": round(v["fact_in"], 2) if v["fact_in"] else "",
                "plan_expense_eur": round(v["plan_ex"], 2) if v["plan_ex"] else "",
                "fact_expense_eur": round(v["fact_ex"], 2) if v["fact_ex"] else "",
                "plan_net_eur": net_plan,
                "fact_net_eur": net_fact,
                "income_var_eur": round(v["fact_in"] - v["plan_in"], 2) if v["fact_in"] and v["plan_in"] else "",
                "expense_var_eur": round(v["fact_ex"] - v["plan_ex"], 2) if v["fact_ex"] and v["plan_ex"] else "",
                "lines_n": v["n"],
                "so_t": "N",
            }
        )

    # top variances (absolute) expense-like
    var_rows = []
    for L in lines:
        if L["variance_eur"] == "" or L["article"].lower() in {"поступления", "расходы"}:
            continue
        var_rows.append(L)
    var_rows.sort(key=lambda x: -abs(float(x["variance_eur"])))
    top_var = var_rows[:40]

    # opex fact by bucket-month for overlap
    opex = list(csv.DictReader((MARTS / "opex_classified.csv").open()))
    opex_bm = defaultdict(float)
    for r in opex:
        opex_bm[(r["period_month"], r["opex_bucket"])] += float(r["amount_rub"] or 0)

    bridge = []
    # aggregate budget fact by bucket-month
    bud_bm = defaultdict(lambda: {"plan": 0.0, "fact": 0.0})
    for L in lines:
        if L["section"] != "expense":
            continue
        if L["article"].lower() in {"расходы", "поступления"}:
            continue
        b = L["opex_bucket_map"]
        if b == "UNMAPPED":
            continue
        pe = float(L["plan_eur"]) if L["plan_eur"] != "" else 0.0
        fe = float(L["fact_eur"]) if L["fact_eur"] != "" else 0.0
        bud_bm[(L["period_month"], b)]["plan"] += pe
        bud_bm[(L["period_month"], b)]["fact"] += fe

    for (pm, bucket), v in sorted(bud_bm.items()):
        opex_rub = opex_bm.get((pm, bucket), 0.0)
        fact_rub_fx = v["fact"] * FX_DEFAULT
        ratio = round(opex_rub / fact_rub_fx, 3) if fact_rub_fx else ""
        bridge.append(
            {
                "period_month": pm,
                "opex_bucket": bucket,
                "budget_plan_eur": round(v["plan"], 2),
                "budget_fact_eur": round(v["fact"], 2),
                "budget_fact_rub_fx100": round(fact_rub_fx, 2),
                "opex_classified_rub": round(opex_rub, 2),
                "opex_vs_budget_fact_ratio": ratio,
                "note": "Indicative only — budget EUR_LIKE vs opex RUB; not Accept",
                "so_t": "N",
            }
        )

    # flags: missing fact months
    flags = []
    for L in lines:
        if L["plan_eur"] != "" and L["fact_eur"] == "":
            flags.append(
                {
                    "budget_id": L["budget_id"],
                    "period_month": L["period_month"],
                    "article": L["article"],
                    "issue": "PLAN_WITHOUT_FACT",
                    "plan_eur": L["plan_eur"],
                }
            )
    # June 2026 income fact empty known
    june_income = [
        L
        for L in lines
        if L["period_month"] == "2026-06" and L["article"].lower() == "поступления" and L["fact_eur"] == ""
    ]

    meta = {
        "horizon": "H63",
        "date": str(date.today()),
        "title": "Budget plan/fact normalize + opex bridge",
        "lines_n": len(lines),
        "months_n": len({(L["budget_id"], L["period_month"]) for L in lines}),
        "top_variance_n": len(top_var),
        "bridge_rows": len(bridge),
        "plan_without_fact_n": len(flags),
        "fx_view": FX_DEFAULT,
        "no_fake_accept": True,
        "so_t": False,
        "june_2026_income_fact_missing": bool(june_income),
    }
    (REG / "meta.json").write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")

    write_csv(REG / "budget_plan_fact_lines.csv", list(lines[0].keys()), lines)
    write_csv(REG / "budget_month_totals.csv", list(month_rows[0].keys()), month_rows)
    write_csv(REG / "budget_top40_variances.csv", list(top_var[0].keys()), top_var)
    write_csv(REG / "budget_opex_bridge.csv", list(bridge[0].keys()) if bridge else ["period_month"], bridge)
    write_csv(
        REG / "budget_plan_without_fact.csv",
        list(flags[0].keys()) if flags else ["budget_id", "period_month", "article", "issue", "plan_eur"],
        flags,
    )

    for name in [
        "budget_plan_fact_lines.csv",
        "budget_month_totals.csv",
        "budget_top40_variances.csv",
        "budget_opex_bridge.csv",
        "budget_plan_without_fact.csv",
        "meta.json",
    ]:
        src = REG / name
        if name == "meta.json":
            shutil.copy2(src, MARTS / "h63_meta.json")
            shutil.copy2(src, MAPS / "h63_meta.json")
            shutil.copy2(src, EV / "meta.json")
        else:
            shutil.copy2(src, MARTS / name)
            shutil.copy2(src, MAPS / name)
            shutil.copy2(src, EV / name)

    shutil.copy2(REG / "budget_month_totals.csv", WAVE_C / "11_budget_month_totals.csv")
    shutil.copy2(REG / "budget_top40_variances.csv", WAVE_C / "12_budget_top40_variances.csv")
    shutil.copy2(REG / "budget_opex_bridge.csv", WAVE_C / "13_budget_opex_bridge.csv")

    print(json.dumps(meta, ensure_ascii=False, indent=2))
    print("TOP variances:")
    for r in top_var[:10]:
        print(r["period_month"], r["article"][:40], "var", r["variance_eur"], "plan", r["plan_eur"], "fact", r["fact_eur"])


if __name__ == "__main__":
    main()
