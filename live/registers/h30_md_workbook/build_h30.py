#!/usr/bin/env python3
"""
H30: Parse «МД — копия.xlsx» → line-level MD orders/payments.

Зачем: H29 дал Salon+Shop income; нужен operational detail заказов/платежей.
Сверка: финансы 2025 ≈ 2.320M EUR vs SALES DDS Salon+Shop 2.326M EUR (~0.3%).

Листы:
- финансы → md_payments (EUR)
- салон → md_salon_orders
- маг → md_shop_sales

Не SoT. Не меняет W4 goods sales_lines.
"""
from __future__ import annotations

import csv
import hashlib
import json
import shutil
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[3]
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/h30_md_workbook_20260724"
MART = ROOT / "live/marts"
W4 = ROOT / "live/registers/w4_sales_settle"
DOCS = ROOT / "documents"
MD_FILE = DOCS / "МД — копия.xlsx"
CC = ROOT / "live/YANINA_LIVE_CONTROL_CENTER_20260723.xlsx"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")
FX = 100.0
SOURCE_ID = "FILE-MD"  # catalog may have SRC; keep stable label


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def fnum(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(" ", "").replace(",", "."))
    except (TypeError, ValueError):
        return None


def sha16(*parts) -> str:
    h = hashlib.sha256("|".join("" if p is None else str(p) for p in parts).encode())
    return h.hexdigest()[:16]


def month_key(d) -> str:
    if isinstance(d, datetime):
        return d.strftime("%Y-%m")
    return ""


def parse_finance(wb) -> list[dict]:
    ws = wb["финансы"]
    rows = list(ws.iter_rows(values_only=True))
    out = []
    last_client = ""
    for i, r in enumerate(rows[2:], start=3):
        if not r:
            continue
        client = r[0]
        if client:
            last_client = str(client).strip()
        d = r[1]
        if not isinstance(d, datetime):
            continue
        cash = fnum(r[2]) or 0.0
        term = fnum(r[3]) or 0.0
        euro_acc = fnum(r[4]) or 0.0
        total = fnum(r[5])
        if total is None:
            total = cash + term + euro_acc
        shop_calc = fnum(r[6])
        salon_calc = fnum(r[7])
        calc_tot = fnum(r[8])
        balance = fnum(r[9])
        term_dekor_rub = fnum(r[10])
        term_ip_rub = fnum(r[11])
        if total == 0 and not any([cash, term, euro_acc, shop_calc, salon_calc]):
            continue
        out.append(
            {
                "payment_id": "MDP-" + sha16(SOURCE_ID, "fin", i, d, last_client, total),
                "period_month": month_key(d),
                "payment_date": d.date().isoformat(),
                "client": last_client,
                "amount_eur": round(total, 2),
                "cash_eur": round(cash, 2),
                "terminal_eur": round(term, 2),
                "euro_account_eur": round(euro_acc, 2),
                "calc_shop_eur": shop_calc if shop_calc is not None else "",
                "calc_salon_eur": salon_calc if salon_calc is not None else "",
                "calc_total_eur": calc_tot if calc_tot is not None else "",
                "balance_eur": balance if balance is not None else "",
                "terminal_dekor_rub": term_dekor_rub if term_dekor_rub is not None else "",
                "terminal_ip_rub": term_ip_rub if term_ip_rub is not None else "",
                "amount_rub_fx100": round(total * FX, 2),
                "source_file": MD_FILE.name,
                "source_sheet": "финансы",
                "source_row": i,
                "channel": "MD_INDIVIDUAL",
                "so_t": "N",
            }
        )
    return out


def parse_salon(wb) -> list[dict]:
    ws = wb["салон"]
    rows = list(ws.iter_rows(values_only=True))
    out = []
    last_client = ""
    for i, r in enumerate(rows[1:], start=2):
        if not r:
            continue
        if r[0]:
            last_client = str(r[0]).strip()
        article = str(r[1]).strip() if r[1] else ""
        desc = str(r[2]).strip() if r[2] else ""
        order_d = r[3] if isinstance(r[3], datetime) else None
        fit_d = r[4] if isinstance(r[4], datetime) else None
        redo_d = r[5] if isinstance(r[5], datetime) else None
        delivery_d = r[6] if isinstance(r[6], datetime) else None
        # period: delivery > order
        period_d = delivery_d or order_d
        if period_d is None and not desc and not article:
            continue
        cost = fnum(r[7])
        price = fnum(r[8])
        discount = fnum(r[9])
        total = fnum(r[10])
        if total is None and price is None:
            continue
        out.append(
            {
                "order_line_id": "MDS-" + sha16(SOURCE_ID, "salon", i, last_client, article, desc, period_d),
                "period_month": month_key(period_d) if period_d else "",
                "client": last_client,
                "article": article,
                "description": desc[:200],
                "order_date": order_d.date().isoformat() if order_d else "",
                "fitting_date": fit_d.date().isoformat() if fit_d else "",
                "redo_date": redo_d.date().isoformat() if redo_d else "",
                "delivery_date": delivery_d.date().isoformat() if delivery_d else "",
                "cost_amount": cost if cost is not None else "",
                "price_amount": price if price is not None else "",
                "discount_pct": discount if discount is not None else "",
                "total_amount": total if total is not None else "",
                "amount_unit_hint": "EUR_LIKELY_POST2020",
                "subchannel": "salon",
                "source_file": MD_FILE.name,
                "source_sheet": "салон",
                "source_row": i,
                "channel": "MD_INDIVIDUAL",
                "so_t": "N",
            }
        )
    return out


def parse_shop(wb) -> list[dict]:
    ws = wb["маг"]
    rows = list(ws.iter_rows(values_only=True))
    out = []
    for i, r in enumerate(rows[1:], start=2):
        if not r:
            continue
        article = str(r[0]).strip() if r[0] else ""
        name = str(r[1]).strip() if r[1] else ""
        size = str(r[2]).strip() if r[2] else ""
        in_d = r[3] if isinstance(r[3], datetime) else None
        sale_d = r[4] if isinstance(r[4], datetime) else None
        client = str(r[5]).strip() if r[5] else ""
        cost = fnum(r[6])
        price = fnum(r[7])
        discount = fnum(r[8])
        total = fnum(r[9])
        if sale_d is None and total is None and price is None:
            continue
        out.append(
            {
                "shop_line_id": "MDM-" + sha16(SOURCE_ID, "mag", i, article, client, sale_d, total),
                "period_month": month_key(sale_d) if sale_d else "",
                "article": article,
                "name": name[:200],
                "size": size,
                "intake_date": in_d.date().isoformat() if in_d else "",
                "sale_date": sale_d.date().isoformat() if sale_d else "",
                "client": client,
                "cost_amount": cost if cost is not None else "",
                "price_amount": price if price is not None else "",
                "discount_pct": discount if discount is not None else "",
                "total_amount": total if total is not None else "",
                "amount_unit_hint": "EUR_LIKELY_POST2020",
                "subchannel": "shop",
                "source_file": MD_FILE.name,
                "source_sheet": "маг",
                "source_row": i,
                "channel": "MD_INDIVIDUAL",
                "so_t": "N",
            }
        )
    return out


def month_sum(rows: list[dict], amount_key: str, month_key_name: str = "period_month") -> dict[str, float]:
    by = defaultdict(float)
    for r in rows:
        m = r.get(month_key_name) or ""
        if not m:
            continue
        v = fnum(r.get(amount_key))
        if v is None:
            continue
        by[m] += v
    return dict(by)


def build_recon(payments: list[dict]) -> list[dict]:
    pay = month_sum(payments, "amount_eur")
    dds = {}
    for r in csv.DictReader(open(W4 / "sales_dds_income_eur.csv", encoding="utf-8")):
        if r.get("channel") != "Salon+Shop":
            continue
        dds[r["period_month"]] = fnum(r.get("amount_eur")) or 0.0
    months = sorted(set(pay) | set(dds))
    rows = []
    for m in months:
        p = pay.get(m, 0.0)
        d = dds.get(m, 0.0)
        gap = p - d
        pct = abs(gap) / d * 100 if d else (100.0 if p else 0.0)
        if pct <= 2:
            status = "CLOSE"
        elif pct <= 10:
            status = "SOFT"
        else:
            status = "OPEN"
        rows.append(
            {
                "period_month": m,
                "md_payments_eur": round(p, 2),
                "sales_dds_salon_shop_eur": round(d, 2),
                "gap_eur": round(gap, 2),
                "gap_pct": round(pct, 1) if d or p else "",
                "status": status,
                "note": "MD workbook финансы vs SALES DDS Salon+Shop",
            }
        )
    return rows


def write_md(summary: dict, recon: list[dict]):
    c2025 = [r for r in recon if r["period_month"].startswith("2025")]
    close = sum(1 for r in c2025 if r["status"] in ("CLOSE", "SOFT"))
    lines = [
        "# H30 — МД workbook (заказы / платежи)",
        "",
        f"Updated: {NOW}",
        "",
        f"Source: `{MD_FILE.name}`",
        "",
        f"- Payments: **{summary['payments_n']}**",
        f"- Salon order lines: **{summary['salon_n']}**",
        f"- Shop sale lines: **{summary['shop_n']}**",
        f"- Recon 2025 CLOSE+SOFT: **{close}/{len(c2025)}**",
        f"- 2025 payments EUR: **{summary['pay_2025_eur']:,.0f}** vs DDS **{summary['dds_2025_eur']:,.0f}** (gap {summary['gap_2025_eur']:,.0f})",
        "",
        "## Политика",
        "1. Платежи (`финансы`) — cash/income proxy для MD_INDIVIDUAL.",
        "2. Салон/маг — operational detail; валюта исторически смешанная → hint EUR_LIKELY_POST2020.",
        "3. Не джойнить в W4 goods COGS (остатков МД нет).",
        "",
        "Files: `md_payments.csv`, `md_salon_orders.csv`, `md_shop_sales.csv`, `recon_md_payments_vs_dds.csv`",
        "",
    ]
    text = "\n".join(lines)
    (OUT / "MD_WORKBOOK.md").write_text(text, encoding="utf-8")
    (ROOT / "live/MD_WORKBOOK.md").write_text(text, encoding="utf-8")
    # append to MD_CHANNEL
    md_ch = ROOT / "live/MD_CHANNEL.md"
    if md_ch.exists():
        prev = md_ch.read_text(encoding="utf-8")
        if "H30" not in prev:
            md_ch.write_text(
                prev.rstrip()
                + f"\n\n---\n\n## H30 workbook parse ({NOW})\n\n"
                + f"Payments/salon/shop parsed. 2025 gap vs DDS: {summary['gap_2025_eur']:,.0f} EUR. "
                + "See `live/MD_WORKBOOK.md`.\n",
                encoding="utf-8",
            )


def update_cc(summary: dict):
    if not CC.exists():
        return
    wb = load_workbook(CC)
    if "H30_MD_Workbook" in wb.sheetnames:
        del wb["H30_MD_Workbook"]
    ws = wb.create_sheet("H30_MD_Workbook", 0)
    ws["A1"] = "H30 MD workbook parse"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A2"] = NOW
    ws["A4"] = summary["finding"]
    ws["A5"] = "2025 gap EUR"
    ws["B5"] = summary["gap_2025_eur"]
    ws["A6"] = "Doc"
    ws["B6"] = "live/MD_WORKBOOK.md"
    wb.save(CC)


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)
    if not MD_FILE.exists():
        raise SystemExit(f"missing {MD_FILE}")

    wb = load_workbook(MD_FILE, read_only=True, data_only=True)
    payments = parse_finance(wb)
    salon = parse_salon(wb)
    shop = parse_shop(wb)
    wb.close()

    recon = build_recon(payments)

    write_csv(MART / "md_payments.csv", payments, list(payments[0].keys()) if payments else ["payment_id"])
    write_csv(OUT / "md_payments.csv", payments, list(payments[0].keys()) if payments else ["payment_id"])
    write_csv(MART / "md_salon_orders.csv", salon, list(salon[0].keys()) if salon else ["order_line_id"])
    write_csv(MART / "md_shop_sales.csv", shop, list(shop[0].keys()) if shop else ["shop_line_id"])
    write_csv(MART / "recon_md_payments_vs_dds.csv", recon, list(recon[0].keys()) if recon else ["period_month"])
    write_csv(OUT / "recon_md_payments_vs_dds.csv", recon, list(recon[0].keys()) if recon else ["period_month"])

    # also register copies under h30
    write_csv(OUT / "md_salon_orders.csv", salon, list(salon[0].keys()) if salon else ["order_line_id"])
    write_csv(OUT / "md_shop_sales.csv", shop, list(shop[0].keys()) if shop else ["shop_line_id"])

    pay_2025 = sum(fnum(r["amount_eur"]) or 0 for r in payments if (r.get("period_month") or "").startswith("2025"))
    dds_2025 = sum(
        fnum(r["amount_eur"]) or 0
        for r in csv.DictReader(open(W4 / "sales_dds_income_eur.csv", encoding="utf-8"))
        if r.get("channel") == "Salon+Shop" and (r.get("period_month") or "").startswith("2025")
    )
    recon_cs = sum(1 for r in recon if r["period_month"].startswith("2024") or r["period_month"].startswith("2025") or r["period_month"].startswith("2026"))
    close_soft = sum(
        1
        for r in recon
        if r["period_month"] >= "2024-01" and r["status"] in ("CLOSE", "SOFT")
    )
    months_recent = sum(1 for r in recon if r["period_month"] >= "2024-01")

    summary = {
        "wave": "H30",
        "generated_at": NOW,
        "finding": (
            f"H30: MD workbook parsed — payments {len(payments)}, salon {len(salon)}, shop {len(shop)}; "
            f"2025 payments {pay_2025:,.0f} EUR vs DDS Salon+Shop {dds_2025:,.0f} "
            f"(gap {pay_2025-dds_2025:,.0f}); recon CLOSE+SOFT {close_soft}/{months_recent} since 2024."
        ),
        "payments_n": len(payments),
        "salon_n": len(salon),
        "shop_n": len(shop),
        "pay_2025_eur": round(pay_2025, 2),
        "dds_2025_eur": round(dds_2025, 2),
        "gap_2025_eur": round(pay_2025 - dds_2025, 2),
        "recon_close_soft_since_2024": close_soft,
        "recon_months_since_2024": months_recent,
        "not_sot": True,
    }
    write_md(summary, recon)
    update_cc(summary)
    (OUT / "h30_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    (EV / "h30_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    for name in (
        "md_payments.csv",
        "md_salon_orders.csv",
        "md_shop_sales.csv",
        "recon_md_payments_vs_dds.csv",
        "MD_WORKBOOK.md",
        "h30_summary.json",
    ):
        src = OUT / name if (OUT / name).exists() else MART / name
        if src.exists():
            shutil.copy2(src, EV / name)
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
