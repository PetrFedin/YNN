#!/usr/bin/env python3
"""
H23: Apply owner ACCEPT (chat command) to RACI draft + confirm H21 flags.

Пользователь явно написал ACCEPT → фиксируем:
1) decision_ACCEPT_REJECT=ACCEPT на RACI-строках с named_person_FIO
2) H21 margin exceptions → OWNER_ACCEPTED (больше не provisional-only)
3) SoT policy draft для принятых доменов (Cash/Bank/Tax/Payroll)
4) Снять freeze AWAITING → RACI_ACCEPTED_STAGING

Не выдумываем ФИО на пустых доменах (Product/Cost/Prod/B2B/Data steward).
"""
from __future__ import annotations

import csv
import json
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[3]
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/h23_raci_accept_20260724"
MART = ROOT / "live/marts"
W4 = ROOT / "live/registers/w4_sales_settle"
PACKET = ROOT / "live/YANINA_OWNER_PACKET_RACI_AND_REQUESTS.xlsx"
CC = ROOT / "live/YANINA_LIVE_CONTROL_CENTER_20260723.xlsx"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")
TODAY = datetime.now().strftime("%Y%m%d")
EVIDENCE = "chat ACCEPT 2026-07-24 (user confirmed RACI draft)"

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF")


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def accept_raci(wb) -> dict:
    ws = wb["RACI"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}

    accepted = []
    skipped = []
    for r in range(2, (ws.max_row or 1) + 1):
        item_id = ws.cell(r, idx["item_id"] + 1).value
        role = ws.cell(r, idx["role"] + 1).value
        fio = ws.cell(r, idx["named_person_FIO"] + 1).value
        if not item_id:
            continue
        if fio and str(fio).strip():
            ws.cell(r, idx["decision_ACCEPT_REJECT"] + 1).value = "ACCEPT"
            ws.cell(r, idx["approval_date_YYYYMMDD"] + 1).value = TODAY
            ws.cell(r, idx["evidence_link_or_file"] + 1).value = EVIDENCE
            ws.cell(r, idx["status"] + 1).value = "ACCEPTED"
            if "notes" in idx:
                ws.cell(r, idx["notes"] + 1).value = "Accepted via chat command ACCEPT; staging→owned domains"
            accepted.append({"item_id": item_id, "role": role, "fio": fio})
        else:
            ws.cell(r, idx["status"] + 1).value = "OPEN_NEEDS_OWNER"
            if "notes" in idx:
                ws.cell(r, idx["notes"] + 1).value = "ACCEPT applied to named draft only; this row still needs FIO"
            skipped.append({"item_id": item_id, "role": role})

    return {"accepted": accepted, "skipped_needs_fio": skipped}


def accept_recommendations(wb) -> int:
    name = "RECOMMENDATIONS_H20"
    if name not in wb.sheetnames:
        return 0
    ws = wb[name]
    header_row = None
    headers = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        vals = [str(c) if c is not None else "" for c in row]
        if "rec_id" in vals:
            header_row = i
            headers = list(row)
            break
    if not header_row:
        return 0
    hidx = {h: i for i, h in enumerate(headers) if h}
    n = 0
    for r in range(header_row + 1, (ws.max_row or header_row) + 1):
        rid = ws.cell(r, 1).value
        if not rid:
            continue
        if "owner_decision_ACCEPT_REJECT" in hidx:
            ws.cell(r, hidx["owner_decision_ACCEPT_REJECT"] + 1).value = "ACCEPT"
        if "status" in hidx:
            ws.cell(r, hidx["status"] + 1).value = "OWNER_ACCEPTED"
        n += 1
    ws["A2"] = f"{NOW} | OWNER ACCEPT (chat). H21 flags upgraded to OWNER_ACCEPTED."
    return n


def upgrade_h21_flags() -> dict:
    """PROVISIONAL_H21 → OWNER_ACCEPTED_H23 on sales_lines + marts."""
    path = W4 / "sales_lines.csv"
    sales = list(csv.DictReader(open(path, encoding="utf-8")))
    n = 0
    for s in sales:
        pol = s.get("margin_exception_policy") or ""
        if pol.startswith("PROVISIONAL_H21:"):
            s["margin_exception_policy"] = pol.replace("PROVISIONAL_H21:", "OWNER_ACCEPTED_H23:", 1)
            n += 1
        elif s.get("margin_exception") == "Y" and "OWNER_ACCEPTED" not in pol:
            # keep as-is if already
            pass
    fields = list(sales[0].keys())
    write_csv(path, sales, fields)

    # margin_exceptions registry
    reg_path = MART / "margin_exceptions.csv"
    if reg_path.exists():
        reg = list(csv.DictReader(open(reg_path, encoding="utf-8")))
        for r in reg:
            r["application"] = "OWNER_ACCEPTED_H23"
            r["owner_accept"] = "ACCEPT"
            r["so_t"] = "DOMAIN_OWNED"  # not full platform SoT
            r["note"] = "Owner ACCEPT 2026-07-24; domain-owned exception policy"
        write_csv(reg_path, reg, list(reg[0].keys()))
        write_csv(OUT / "margin_exceptions.csv", reg, list(reg[0].keys()))

    # recommendations csv
    rec_path = MART / "finance_recommendations.csv"
    if rec_path.exists():
        recs = list(csv.DictReader(open(rec_path, encoding="utf-8")))
        for r in recs:
            r["status"] = "OWNER_ACCEPTED"
            r["owner_decision_ACCEPT_REJECT"] = "ACCEPT"
        write_csv(rec_path, recs, list(recs[0].keys()))

    return {"sales_flags_upgraded": n}


def write_sot_policy(raci_stats: dict) -> Path:
    """Политика после ACCEPT: что стало owned, что ещё нет."""
    lines = [
        "# SoT Policy (после RACI ACCEPT) — H23",
        "",
        f"Updated: {NOW}",
        "",
        "## Статус",
        "",
        "**RACI draft ACCEPTED** для строк с ФИО.",
        "Это **domain-owned staging**, не полный Source of Truth всего контура.",
        "",
        "### Accepted owners",
        "",
        "| Domain / control | Owner |",
        "|------------------|-------|",
        "| Cash / Treasury (DOM-CASH) | Мамушкина Елена |",
        "| Bank perimeter (SRC-CTRL-04 Owner) | Сливяк Галина |",
        "| Bank Approver | Мамушкина Елена |",
        "| Tax (DOM-TAX, SRC-CTRL-05 Owner) | Сливяк Галина |",
        "| Tax Approver | ЯНИНА ЮЛИЯ ФЕДОРОВНА |",
        "| Payroll (DOM-PAYROLL, SRC-CTRL-01 Owner) | Сливяк Галина |",
        "",
        "### Ещё OPEN (нужны ФИО)",
        "",
    ]
    for s in raci_stats.get("skipped_needs_fio", []):
        lines.append(f"- `{s['item_id']}` / {s['role']}")
    lines.extend(
        [
            "",
            "## Политики данных (accepted)",
            "",
            "1. **Margin exceptions** `WHOLESALE_OK_LOSS` (0-2493A/2496/2497) — owner-accepted; "
            "reported margin включает, clean margin исключает.",
            "2. **0-3243** — quarantine до cost version свитшота; не релинковать на худи/юбку.",
            "3. **Release gate** H18 остаётся PROVISIONAL как operational gate; "
            "можно ужесточать пороги с data steward (когда назначен).",
            "4. Регистры W1–W6 + marts — **controlled staging with named owners**, "
            "не audited accounting SoT.",
            "",
            "## Следующие шаги (фаза C)",
            "",
            "1. Назначить Owners на Product / Cost / Production / B2B / Data steward",
            "2. Зафиксировать SKU alias registry (кандидаты H17) под Product Owner",
            "3. Ужесточить release gate (запрет SOFT) после data steward",
            "4. Закрывать BLOCKED months данными из DATA_REQUESTS_NOW",
            "",
            f"Evidence: `{EVIDENCE}`",
            "",
        ]
    )
    text = "\n".join(lines)
    path = ROOT / "live/SOT_POLICY.md"
    path.write_text(text, encoding="utf-8")
    (OUT / "SOT_POLICY.md").write_text(text, encoding="utf-8")
    (EV / "SOT_POLICY.md").write_text(text, encoding="utf-8")
    (ROOT / "SOT_POLICY.md").write_text(text, encoding="utf-8")
    return path


def update_freeze(summary: dict):
    freeze_path = MART / "staging_freeze.json"
    freeze = {
        "wave": "H23",
        "generated_at": NOW,
        "status": "RACI_ACCEPTED_STAGING",
        "previous": "FROZEN_AWAITING_OWNER",
        "so_t": False,
        "domain_owned": True,
        "raci_accept": True,
        "finding": summary["finding"],
        "open_domains": [s["item_id"] for s in summary.get("raci_skipped", [])],
    }
    freeze_path.write_text(json.dumps(freeze, ensure_ascii=False, indent=2), encoding="utf-8")
    (EV / "staging_freeze.json").write_text(json.dumps(freeze, ensure_ascii=False, indent=2), encoding="utf-8")


def update_cc(summary: dict):
    if not CC.exists():
        return
    wb = load_workbook(CC)
    if "H23_Accept" in wb.sheetnames:
        del wb["H23_Accept"]
    ws = wb.create_sheet("H23_Accept", 0)
    ws["A1"] = "H23 RACI ACCEPT"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A2"] = NOW
    ws["A4"] = summary["finding"]
    ws["A5"] = "Accepted rows"
    ws["B5"] = summary["raci_accepted_n"]
    ws["A6"] = "Open needs FIO"
    ws["B6"] = summary["raci_skipped_n"]
    ws["A7"] = "H21 flags"
    ws["B7"] = "OWNER_ACCEPTED"
    ws["A8"] = "SoT?"
    ws["B8"] = "Domain-owned staging (not full SoT)"
    ws["A9"] = "Policy"
    ws["B9"] = "live/SOT_POLICY.md"
    wb.save(CC)


def update_readme(wb):
    if "README" not in wb.sheetnames:
        return
    ws = wb["README"]
    ws["A15"] = (
        f"H23 {NOW}: RACI ACCEPT applied to named draft rows. "
        "Domain-owned staging. See live/SOT_POLICY.md"
    )
    ws["A15"].font = Font(bold=True, color="006600")


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)
    shutil.copy2(PACKET, EV / "YANINA_OWNER_PACKET_RACI_AND_REQUESTS.pre_h23.xlsx")

    wb = load_workbook(PACKET)
    raci = accept_raci(wb)
    rec_n = accept_recommendations(wb)
    update_readme(wb)

    # ACCEPT log sheet
    if "H23_ACCEPT_LOG" in wb.sheetnames:
        del wb["H23_ACCEPT_LOG"]
    ws = wb.create_sheet("H23_ACCEPT_LOG", 2)
    ws["A1"] = "H23 ACCEPT log"
    ws["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws["A2"] = f"{NOW} | {EVIDENCE}"
    ws.append([])
    ws.append(["item_id", "role", "fio", "decision"])
    for c in range(1, 5):
        ws.cell(4, c).fill = HDR_FILL
        ws.cell(4, c).font = HDR_FONT
    for a in raci["accepted"]:
        ws.append([a["item_id"], a["role"], a["fio"], "ACCEPT"])
    ws.append([])
    ws.append(["Still OPEN_NEEDS_OWNER:"])
    for s in raci["skipped_needs_fio"]:
        ws.append([s["item_id"], s["role"], "", "PENDING_FIO"])

    wb.save(PACKET)
    shutil.copy2(PACKET, EV / "YANINA_OWNER_PACKET_RACI_AND_REQUESTS.xlsx")
    shutil.copy2(PACKET, OUT / "YANINA_OWNER_PACKET_RACI_AND_REQUESTS.xlsx")

    h21 = upgrade_h21_flags()
    write_sot_policy(raci)

    # owners mart
    owners = [
        {"domain": "CASH", "owner": "Мамушкина Елена", "status": "ACCEPTED", "source": "DOM-CASH"},
        {"domain": "BANK", "owner": "Сливяк Галина", "status": "ACCEPTED", "source": "SRC-CTRL-04/DOM"},
        {"domain": "TAX", "owner": "Сливяк Галина", "status": "ACCEPTED", "source": "DOM-TAX"},
        {"domain": "TAX_APPROVER", "owner": "ЯНИНА ЮЛИЯ ФЕДОРОВНА", "status": "ACCEPTED", "source": "SRC-CTRL-05"},
        {"domain": "PAYROLL", "owner": "Сливяк Галина", "status": "ACCEPTED", "source": "DOM-PAYROLL"},
        {"domain": "PRODUCT", "owner": "", "status": "OPEN_NEEDS_OWNER", "source": "DOM-PRODUCT"},
        {"domain": "COST", "owner": "", "status": "OPEN_NEEDS_OWNER", "source": "DOM-COST"},
        {"domain": "PRODUCTION", "owner": "", "status": "OPEN_NEEDS_OWNER", "source": "DOM-PROD"},
        {"domain": "B2B", "owner": "", "status": "OPEN_NEEDS_OWNER", "source": "DOM-B2B"},
        {"domain": "DATA_STEWARD", "owner": "", "status": "OPEN_NEEDS_OWNER", "source": "DOM-DATA"},
    ]
    write_csv(MART / "sot_owners.csv", owners, list(owners[0].keys()))
    write_csv(OUT / "sot_owners.csv", owners, list(owners[0].keys()))

    summary = {
        "wave": "H23",
        "generated_at": NOW,
        "command": "ACCEPT",
        "finding": (
            f"H23: RACI ACCEPT on {len(raci['accepted'])} named rows; "
            f"{len(raci['skipped_needs_fio'])} still OPEN_NEEDS_OWNER; "
            f"H21 flags upgraded ({h21['sales_flags_upgraded']} lines); "
            f"recommendations ACCEPT={rec_n}. Domain-owned staging — not full SoT."
        ),
        "raci_accepted_n": len(raci["accepted"]),
        "raci_skipped_n": len(raci["skipped_needs_fio"]),
        "raci_skipped": raci["skipped_needs_fio"],
        "recommendations_accepted": rec_n,
        "h21_flags_upgraded": h21["sales_flags_upgraded"],
        "full_sot": False,
        "domain_owned": True,
        "policy": "live/SOT_POLICY.md",
    }
    update_freeze(summary)
    update_cc(summary)
    (OUT / "h23_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    (EV / "h23_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    shutil.copy2(MART / "sot_owners.csv", EV / "sot_owners.csv")

    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
