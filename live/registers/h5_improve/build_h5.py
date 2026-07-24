#!/usr/bin/env python3
"""
H5: улучшения данных без RACI ACCEPT.

1) Движение товаров с себестоимостью → stock_cost_lines + derived unit costs
2) Усиление alias (латиница/кириллица + strip size) → перелинковка W4↔W3
3) Движение по складам → stock_by_warehouse summary
4) Sber tax-like для LE-OOO-SALON-YANINA

Не SoT.
"""
from __future__ import annotations

import csv
import hashlib
import json
import re
import shutil
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[3]
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/h5_improve_20260724"
DOCS = ROOT / "documents"
W1 = ROOT / "live/registers/w1_bank_cash"
W3 = ROOT / "live/registers/w3_sku_cost"
W4 = ROOT / "live/registers/w4_sales_settle"
W5 = ROOT / "live/registers/w5_sup_exp_mat"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")

CY2LAT = str.maketrans({
    "А": "A", "В": "B", "Е": "E", "К": "K", "М": "M", "Н": "H",
    "О": "O", "Р": "P", "С": "C", "Т": "T", "Х": "X",
    "а": "A", "в": "B", "е": "E", "к": "K", "м": "M", "н": "H",
    "о": "O", "р": "P", "с": "C", "т": "T", "х": "X",
    "Ё": "E", "ё": "E",
})


def nfc(s: str) -> str:
    return unicodedata.normalize("NFC", s or "")


def sha16(*parts) -> str:
    raw = "|".join("" if p is None else str(p) for p in parts)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def resolve(name: str) -> Path | None:
    t = nfc(name)
    for p in DOCS.iterdir():
        if nfc(p.name) == t:
            return p
    return None


def normalize_article(raw) -> str:
    if raw is None:
        return ""
    s = nfc(str(raw)).strip().upper().replace("Ё", "Е")
    s = s.translate(CY2LAT)
    s = s.replace("Т-", "T-")
    s = s.replace(" ", "").replace("\xa0", "").replace("–", "-").replace("—", "-")
    return s


def canonical_sku(raw) -> str:
    s = normalize_article(raw)
    s = re.sub(r"/\d+[A-Z]?$", "", s)  # size suffix
    return s


def alias_keys(raw) -> list[str]:
    """Набор ключей для нечёткого матча cost↔sales."""
    base = canonical_sku(raw)
    if not base:
        return []
    keys = {base}
    # без хвостовой буквы цвета (0-1984F → 0-1984)
    keys.add(re.sub(r"[A-Z]$", "", base))
    # T-2401B → T-2401
    keys.add(re.sub(r"([A-Z])$", "", base))
    # убрать все буквы после цифр в хвосте: 0-2032N → 0-2032
    keys.add(re.sub(r"(\d)[A-Z]+$", r"\1", base))
    return [k for k in keys if k]


# ── Stock cost movements ──────────────────────────────────────────
def parse_stock_cost() -> tuple[list[dict], list[dict], dict]:
    path = resolve("Движение товаров с себестоимостью.xlsx")
    if not path:
        return [], [], {"error": "missing"}
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    lines = []
    derived = []
    warehouse = ""
    pending_name = ""
    for i, r in enumerate(ws.iter_rows(min_row=4, max_col=40, values_only=True), start=4):
        if not r or r[0] is None:
            continue
        a0 = str(r[0]).strip()
        low = a0.lower()
        # warehouse / subdivision totals
        if any(k in a0 for k in ("Подразделение", "Aldo Coppola", "ДЕМИ", "ЦУМ", "Остатки")) and not re.match(r"^[0-9A-Za-z]", a0):
            warehouse = a0
            continue
        if "подразделение" in low or (a0.endswith(")") and "склад" not in low and not re.match(r"^[0-9A-Za-zА-Я]", a0[:1])):
            if any(x in a0 for x in ("Aldo", "ДЕМИ", "ЦУМ", "Мокеев", "Жуков", "Меркуш")):
                warehouse = a0
                continue
        # article row
        if re.match(r"^[0-9A-Za-zА-Яа-я][0-9A-Za-zА-Яа-я\-/]{2,}$", a0) and len(a0) < 40 and "," not in a0:
            art = a0
            open_qty, open_cost = r[1], r[2]
            in_qty, in_cost = r[3], r[4]
            out_qty, out_cost = r[21], r[22]
            # unit cost preference: inbound, else outbound, else opening
            unit = None
            src = ""
            for q, c, label in (
                (in_qty, in_cost, "IN"),
                (out_qty, out_cost, "OUT"),
                (open_qty, open_cost, "OPEN"),
            ):
                if isinstance(q, (int, float)) and q and isinstance(c, (int, float)) and c:
                    unit = round(float(c) / float(q), 2)
                    src = label
                    break
            can = canonical_sku(art)
            row = {
                "stock_cost_id": "SC-" + sha16(art, warehouse, i),
                "article_raw": art,
                "canonical_sku": can,
                "warehouse": warehouse,
                "name": pending_name,
                "open_qty": open_qty if isinstance(open_qty, (int, float)) else "",
                "open_cost_rub": open_cost if isinstance(open_cost, (int, float)) else "",
                "in_qty": in_qty if isinstance(in_qty, (int, float)) else "",
                "in_cost_rub": in_cost if isinstance(in_cost, (int, float)) else "",
                "out_qty": out_qty if isinstance(out_qty, (int, float)) else "",
                "out_cost_rub": out_cost if isinstance(out_cost, (int, float)) else "",
                "unit_cost_rub": unit if unit is not None else "",
                "unit_cost_source": src,
                "source_file_id": "FILE-099",
                "source_row_id": f"r{i}",
            }
            lines.append(row)
            if unit is not None:
                derived.append(
                    {
                        "cost_version_id": "CV-H5-" + sha16(can, unit, src),
                        "canonical_sku": can,
                        "article_raw": art,
                        "name": pending_name[:80],
                        "unit_cost_rub": unit,
                        "completeness": "DERIVED_STOCK_MOVEMENT",
                        "source_file_id": "FILE-099",
                        "channel_hint": warehouse[:40],
                        "owner_file": "Движение товаров с себестоимостью",
                    }
                )
            pending_name = ""
            continue
        # name row under article
        if not re.match(r"^[0-9A-Za-z]", a0) and len(a0) > 10:
            pending_name = a0
            # also attach to last line if just created
            if lines and not lines[-1]["name"]:
                lines[-1]["name"] = a0[:160]
                if derived and derived[-1]["article_raw"] == lines[-1]["article_raw"]:
                    derived[-1]["name"] = a0[:80]
    wb.close()

    # collapse derived: one best unit per canonical (prefer IN)
    best = {}
    for d in derived:
        can = d["canonical_sku"]
        prev = best.get(can)
        rank = {"IN": 0, "OUT": 1, "OPEN": 2}.get(d.get("unit_cost_source") or lines_source(d), 9)
        # unit_cost_source not on derived — use completeness path; store from matching
        if prev is None:
            best[can] = d
        else:
            # keep first (IN preferred because we append in that order per article)
            pass
    # rebuild best properly
    best = {}
    for d in derived:
        can = d["canonical_sku"]
        # find source from lines
        src = next((L["unit_cost_source"] for L in lines if L["canonical_sku"] == can and L["unit_cost_rub"] == d["unit_cost_rub"]), "IN")
        d = dict(d)
        d["unit_cost_source"] = src
        prev = best.get(can)
        rank = {"IN": 0, "OUT": 1, "OPEN": 2}.get(src, 9)
        if prev is None or rank < {"IN": 0, "OUT": 1, "OPEN": 2}.get(prev.get("unit_cost_source"), 9):
            best[can] = d
    derived_best = list(best.values())

    stats = {
        "stock_cost_lines": len(lines),
        "with_unit_cost": sum(1 for L in lines if L["unit_cost_rub"] != ""),
        "derived_canonical": len(derived_best),
        "warehouses": sorted({L["warehouse"] for L in lines if L["warehouse"]})[:12],
    }
    return lines, derived_best, stats


def lines_source(d):
    return "IN"


# ── Warehouse stock by location ───────────────────────────────────
def parse_stock_by_warehouse() -> tuple[list[dict], dict]:
    path = resolve("Движение товара по складам.xlsx")
    if not path:
        return [], {"error": "missing"}
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    # header warehouses at row1 cols
    header = next(ws.iter_rows(min_row=1, max_row=1, max_col=30, values_only=True))
    # each warehouse block: 4 cols start at 2,6,10,...
    wh_cols = []
    for c, name in enumerate(header):
        if name and c >= 2:
            wh_cols.append((c, str(name)[:60]))
    # actually from probe: col2 Aldo, col6 DEMI, col10 TSUM — each 4 qty cols
    blocks = []
    for c, name in wh_cols:
        blocks.append({"name": name, "end_col": c + 3})  # кон. остаток = start+3

    rows = []
    for i, r in enumerate(ws.iter_rows(min_row=4, max_col=30, values_only=True), start=4):
        if not r or not r[0]:
            continue
        a0 = str(r[0]).strip()
        if "," not in a0:
            continue
        # "0-1075, Блуза..."
        art, _, name = a0.partition(",")
        art = art.strip()
        name = name.strip()
        if not re.match(r"^[0-9A-Za-zА-Яа-я]", art):
            continue
        for b in blocks:
            end = r[b["end_col"]] if len(r) > b["end_col"] else None
            if not isinstance(end, (int, float)) or end == 0:
                continue
            rows.append(
                {
                    "stock_wh_id": "SW-" + sha16(art, b["name"]),
                    "article_raw": art,
                    "canonical_sku": canonical_sku(art),
                    "name": name[:120],
                    "warehouse": b["name"],
                    "qty_end": round(float(end), 3),
                    "source_file_id": "FILE-098",
                    "source_row_id": f"r{i}",
                }
            )
    wb.close()
    stats = {
        "stock_wh_rows": len(rows),
        "skus": len({r["canonical_sku"] for r in rows}),
        "warehouses": sorted({r["warehouse"] for r in rows}),
        "qty_end_sum": round(sum(r["qty_end"] for r in rows), 2),
    }
    return rows, stats


# ── Improve cost index + relink W4 ────────────────────────────────
def build_cost_index(derived: list[dict]) -> dict:
    """canonical + aliases → best cost row."""
    index = {}  # key -> {unit_cost, cost_version_id, canonical_sku, source}

    def add(key, unit, cvid, can, source, completeness):
        if not key or unit in (None, ""):
            return
        unit = float(unit)
        prev = index.get(key)
        rank = 0 if completeness == "FULL" else (1 if completeness == "DERIVED_STOCK_MOVEMENT" else 2)
        if prev is None or rank < prev["rank"]:
            index[key] = {
                "unit_cost_rub": unit,
                "cost_version_id": cvid,
                "canonical_sku": can,
                "source": source,
                "rank": rank,
            }

    if (W3 / "cost_versions.csv").exists():
        for c in csv.DictReader(open(W3 / "cost_versions.csv", encoding="utf-8")):
            can = canonical_sku(c.get("canonical_sku") or c.get("article_raw"))
            unit = c.get("unit_cost_rub")
            if not unit:
                continue
            for k in alias_keys(c.get("article_raw") or can):
                add(k, unit, c.get("cost_version_id"), can, "W3", c.get("completeness") or "")
            for k in alias_keys(can):
                add(k, unit, c.get("cost_version_id"), can, "W3", c.get("completeness") or "")

    for d in derived:
        can = d["canonical_sku"]
        for k in alias_keys(d.get("article_raw") or can):
            add(k, d["unit_cost_rub"], d["cost_version_id"], can, "H5_STOCK", "DERIVED_STOCK_MOVEMENT")

    return index


def relink_w4(cost_index: dict) -> dict:
    sales = list(csv.DictReader(open(W4 / "sales_lines.csv", encoding="utf-8")))
    before = sum(1 for s in sales if s.get("w3_cost_version_id"))
    linked_new = 0
    for s in sales:
        keys = alias_keys(s.get("article_raw") or s.get("canonical_sku"))
        hit = None
        for k in keys:
            if k in cost_index:
                hit = cost_index[k]
                break
        if not hit:
            continue
        prev = s.get("w3_cost_version_id") or ""
        s["canonical_sku"] = hit["canonical_sku"] or s.get("canonical_sku")
        s["w3_unit_cost"] = hit["unit_cost_rub"]
        s["w3_cost_version_id"] = hit["cost_version_id"]
        # refresh cogs if missing or was empty
        qty = float(s["qty"] or 0) if s.get("qty") not in (None, "") else 0
        if s.get("cogs_source") != "FILE" and qty and hit["unit_cost_rub"]:
            cogs = round(qty * float(hit["unit_cost_rub"]), 2)
            s["cogs_rub"] = cogs
            s["cogs_source"] = "W3_H5" if hit["source"] == "W3" else "H5_STOCK"
            rev = float(s["revenue_rub"] or 0)
            s["margin_rub"] = round(rev - cogs, 2)
        if not prev:
            linked_new += 1

    fields = list(sales[0].keys())
    write_csv(W4 / "sales_lines.csv", sales, fields)
    write_csv(W4 / "sales_lines_1c.csv", [s for s in sales if s["channel"] in ("B2B", "IM")], fields)
    after = sum(1 for s in sales if s.get("w3_cost_version_id"))
    with_cogs = sum(1 for s in sales if s.get("cogs_rub") not in (None, ""))
    return {
        "sales_lines": len(sales),
        "with_cost_before": before,
        "with_cost_after": after,
        "newly_linked": linked_new,
        "delta": after - before,
        "with_cogs": with_cogs,
    }


def refresh_sku_coverage(cost_index: dict) -> dict:
    master = list(csv.DictReader(open(W3 / "sku_master.csv", encoding="utf-8")))
    sales_cans = set()
    if (W4 / "sales_lines.csv").exists():
        for s in csv.DictReader(open(W4 / "sales_lines.csv", encoding="utf-8")):
            for k in alias_keys(s.get("article_raw") or s.get("canonical_sku")):
                sales_cans.add(k)
    cost_keys = set(cost_index.keys())
    rows = []
    for m in master:
        keys = alias_keys(m.get("article") or m.get("canonical_sku") or m.get("code_1c"))
        in_cost = "Y" if any(k in cost_keys for k in keys) else "N"
        in_sales = "Y" if any(k in sales_cans for k in keys) else "N"
        rows.append({**m, "in_cost": in_cost, "in_sales": in_sales, "in_intersection": "Y" if in_cost == "Y" and in_sales == "Y" else "N"})
    write_csv(W3 / "sku_master.csv", rows, list(rows[0].keys()))
    write_csv(OUT / "sku_coverage_h5.csv", rows, list(rows[0].keys()))
    return {
        "master": len(rows),
        "in_cost": sum(1 for r in rows if r["in_cost"] == "Y"),
        "in_sales": sum(1 for r in rows if r["in_sales"] == "Y"),
        "intersection": sum(1 for r in rows if r["in_intersection"] == "Y"),
    }


# ── Sber tax for Salon ────────────────────────────────────────────
def sber_tax_recon() -> tuple[list[dict], list[dict], dict]:
    pays = [
        p
        for p in csv.DictReader(open(W1 / "bank_payments.csv", encoding="utf-8"))
        if p.get("source_bank") == "SBER" and p.get("legal_entity_id") == "LE-OOO-SALON-YANINA"
    ]
    cp_ok = re.compile(r"казначейство|уфк\b|фнс|осфр|пенсион|социальн", re.I)
    tax = []
    for p in pays:
        blob = f"{p.get('counterparty_raw','')} {p.get('purpose','')}"
        if p.get("direction") == "out" and cp_ok.search(blob):
            tax.append(p)
    by_m = defaultdict(float)
    for p in tax:
        by_m[p["period_month"]] += float(p["amount"] or 0)
    recon = [
        {
            "period_month": m,
            "sber_tax_out_rub": round(v, 2),
            "legal_entity_id": "LE-OOO-SALON-YANINA",
            "note": "budget payments from Sber Salon account",
        }
        for m, v in sorted(by_m.items())
    ]
    stats = {
        "sber_payments_total": len(pays),
        "sber_tax_like": len(tax),
        "sber_tax_out_rub": round(sum(by_m.values()), 2),
        "months": len(recon),
    }
    return tax, recon, stats


def append_derived_costs(derived: list[dict]) -> int:
    """Добавить derived costs в cost_versions, не дублируя canonical с FULL."""
    path = W3 / "cost_versions.csv"
    existing = list(csv.DictReader(open(path, encoding="utf-8")))
    have = {canonical_sku(c.get("canonical_sku") or c.get("article_raw")) for c in existing}
    fields = list(existing[0].keys())
    added = 0
    bak = W3 / "cost_versions_pre_h5.csv"
    if not bak.exists():
        shutil.copy2(path, bak)
    for d in derived:
        can = d["canonical_sku"]
        if can in have:
            continue
        row = {k: "" for k in fields}
        row.update(
            {
                "cost_version_id": d["cost_version_id"],
                "canonical_sku": can,
                "article_raw": d.get("article_raw", ""),
                "name": d.get("name", ""),
                "unit_cost_rub": d["unit_cost_rub"],
                "completeness": "DERIVED_STOCK_MOVEMENT",
                "source_file_id": "FILE-099",
                "owner_file": d.get("owner_file", ""),
                "channel_hint": d.get("channel_hint", ""),
            }
        )
        existing.append(row)
        have.add(can)
        added += 1
    write_csv(path, existing, fields)
    return added


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)

    stock_lines, derived, stock_stats = parse_stock_cost()
    wh_rows, wh_stats = parse_stock_by_warehouse()
    added_cost = append_derived_costs(derived)
    cost_index = build_cost_index(derived)
    w4_link = relink_w4(cost_index)
    sku_cov = refresh_sku_coverage(cost_index)
    tax_pays, tax_recon, tax_stats = sber_tax_recon()

    write_csv(
        OUT / "stock_cost_lines.csv",
        stock_lines,
        list(stock_lines[0].keys()) if stock_lines else ["stock_cost_id"],
    )
    write_csv(W5 / "stock_cost_lines.csv", stock_lines, list(stock_lines[0].keys()) if stock_lines else ["stock_cost_id"])
    write_csv(
        OUT / "derived_unit_costs.csv",
        derived,
        list(derived[0].keys()) if derived else ["cost_version_id"],
    )
    write_csv(
        OUT / "stock_by_warehouse.csv",
        wh_rows,
        list(wh_rows[0].keys()) if wh_rows else ["stock_wh_id"],
    )
    write_csv(W5 / "stock_by_warehouse.csv", wh_rows, list(wh_rows[0].keys()) if wh_rows else ["stock_wh_id"])
    write_csv(
        OUT / "sber_salon_tax_payments.csv",
        [
            {
                "bank_payment_id": p["bank_payment_id"],
                "period_month": p["period_month"],
                "payment_date": p["payment_date"],
                "amount": p["amount"],
                "counterparty_raw": p.get("counterparty_raw", ""),
                "purpose": (p.get("purpose") or "")[:140],
            }
            for p in tax_pays
        ],
        ["bank_payment_id", "period_month", "payment_date", "amount", "counterparty_raw", "purpose"],
    )
    write_csv(OUT / "sber_salon_tax_recon.csv", tax_recon, list(tax_recon[0].keys()) if tax_recon else ["period_month"])

    summary = {
        "generated_at": NOW,
        "wave": "H5",
        "stock_cost": stock_stats,
        "derived_costs_added_to_w3": added_cost,
        "stock_by_warehouse": wh_stats,
        "w4_relink": w4_link,
        "sku_coverage": sku_cov,
        "sber_salon_tax": tax_stats,
        "finding": (
            f"H5: stock_cost {stock_stats.get('stock_cost_lines')} lines / "
            f"+{added_cost} derived costs; W4 cost links {w4_link['with_cost_before']}→{w4_link['with_cost_after']} "
            f"(+{w4_link['delta']}); SKU∩ {sku_cov['intersection']}; "
            f"Sber Salon tax {tax_stats['sber_tax_like']} pays ~{tax_stats['sber_tax_out_rub']} RUB."
        ),
        "next": "RACI ACCEPT; optional deep BOM from fabric issues",
        "not_sot": True,
    }
    json.dump(summary, open(OUT / "h5_summary.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    json.dump(summary, open(EV / "h5_summary.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)

    # update w4 summary snippet
    if (W4 / "w4_summary.json").exists():
        w4s = json.load(open(W4 / "w4_summary.json"))
        w4s["h5_relink"] = w4_link
        w4s["generated_at"] = NOW
        w4s["finding"] = (
            f"W4+H5: {w4_link['sales_lines']} sales; cost-linked {w4_link['with_cost_after']}; "
            f"cogs {w4_link['with_cogs']}."
        )
        json.dump(w4s, open(W4 / "w4_summary.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)

    wb = Workbook()
    ws = wb.active
    ws.title = "00_Summary"
    for i, (k, v) in enumerate(summary.items(), 1):
        ws.cell(i, 1, k)
        ws.cell(i, 2, json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v)
    fill = PatternFill("solid", fgColor="1F4E79")

    def add(name, rows):
        w = wb.create_sheet(name)
        if not rows:
            return
        headers = list(rows[0].keys())
        for c, h in enumerate(headers, 1):
            cell = w.cell(1, c, h)
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
        for ri, row in enumerate(rows[:4000], 2):
            for ci, h in enumerate(headers, 1):
                v = row.get(h, "")
                if isinstance(v, (dict, list)):
                    v = json.dumps(v, ensure_ascii=False)
                w.cell(ri, ci, v)

    add("01_Stock_Cost", stock_lines[:3000])
    add("02_Derived_Costs", derived)
    add("03_Stock_WH", wh_rows[:3000])
    add("04_Sber_Tax", tax_recon)
    wb.save(EV / "YANINA_H5_IMPROVE_EVIDENCE.xlsx")

    (OUT / "README.md").write_text(
        f"""# H5 Improvements

{NOW}

{summary['finding']}

Evidence: `../../evidence/h5_improve_20260724/`
""",
        encoding="utf-8",
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
