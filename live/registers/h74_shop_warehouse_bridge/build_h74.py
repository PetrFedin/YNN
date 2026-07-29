#!/usr/bin/env python3
"""H74 — Цех ЗП (портн/сметка) ↔ collections / person-cost / stock warehouses.

Зачем:
  5× ЗП_ЦЕХ_0N.26.xlsx — пофамильные блоки портных с артикулами и «К выплате».
  Закрывает G7 цех ↔ G5 коллекции ↔ G6 склады без fake Accept.

Правила:
  - портн: блоки по ФИО + Итого col15
  - сметка: часовые работы (другой контур оплаты)
  - stock hit ожидаемо редкий (collection-style ≠ goods SKU) — это сигнал, не баг
  - do_not_auto_accept=YES · so_t=N
"""
from __future__ import annotations

import csv
import json
import re
import shutil
import warnings
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

warnings.filterwarnings("ignore")

ROOT = Path(__file__).resolve().parents[3]
REG = Path(__file__).resolve().parent
MARTS = ROOT / "live/marts"
MAPS = ROOT / "live/maps"
EV = ROOT / "live/evidence/h74_shop_warehouse_bridge_20260729"
WAVE_B = ROOT / "live/client_pack/execution_wave_b"
DOWNLOADS = Path("/Users/petr/Downloads/YANINA документы")


def write_csv(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)


def fnum(x):
    try:
        if x in (None, ""):
            return None
        return float(x)
    except (TypeError, ValueError):
        return None


def norm_art(raw) -> str:
    if raw is None or raw == "":
        return ""
    s = str(raw).strip().upper().replace(" ", "").replace("Ё", "Е")
    s = re.sub(r"^К(?=\d)", "", s)
    if s in ("Б/Н", "БН", "NONE", "-", "Н/Д"):
        return ""
    m = re.search(r"(\d{1,2}-\d{2,4}[A-ZА-Я]?)", s)
    if m:
        return m.group(1)
    m = re.search(r"(0-\d+[A-ZА-Я]?)", s)
    return m.group(1) if m else ""


def is_person(s) -> bool:
    if not isinstance(s, str):
        return False
    s = s.strip()
    if "Итого" in s or "№" in s or len(s) < 5:
        return False
    return bool(re.match(r"^[А-ЯЁA-Z][а-яёa-z\-]+ [А-ЯЁA-Zа-яё]", s))


def month_from_name(name: str) -> str:
    m = re.search(r"(\d{2})\.(\d{2})", name)
    return f"20{m.group(2)}-{m.group(1)}" if m else ""


def catalog() -> list[dict]:
    out = []
    with (ROOT / "live/registers/00_SOURCE_CATALOG_107.csv").open(encoding="utf-8") as f:
        for r in csv.DictReader(f):
            if r["category"] != "payroll_shop":
                continue
            p = Path(r["path"])
            if not p.exists():
                p = DOWNLOADS / r["file_name"]
            out.append({**r, "_path": p})
    return out


def parse_portn(path: Path, meta: dict, period: str) -> tuple[list[dict], list[dict]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    sn = [s for s in wb.sheetnames if "портн" in s.lower()]
    if not sn:
        wb.close()
        return [], []
    rows = list(wb[sn[0]].iter_rows(values_only=True))
    wb.close()
    lines: list[dict] = []
    pays: list[dict] = []
    person_starts = []
    for i, row in enumerate(rows):
        b = row[1] if row and len(row) > 1 else None
        if not is_person(b):
            continue
        ok = False
        for j in range(i + 1, min(i + 4, len(rows))):
            if rows[j] and str(rows[j][1] or "").startswith("№"):
                ok = True
                break
        if ok:
            person_starts.append((i, b.strip()))

    for idx, (start, name) in enumerate(person_starts):
        end = person_starts[idx + 1][0] if idx + 1 < len(person_starts) else len(rows)
        section = "UNKNOWN"
        pay_total = None
        for i in range(start + 1, end):
            row = rows[i]
            if not row:
                continue
            b1 = str(row[1] or "").strip()
            b2 = str(row[2] or "").strip()
            if b1.startswith("№"):
                continue
            low = b2.lower()
            if low in ("отшивы", "переделки", "аксес") or low.startswith("по час"):
                section = b2[:40]
                continue
            if low == "итого":
                pay_total = fnum(row[15] if len(row) > 15 else None)
                continue
            art = norm_art(row[4] if len(row) > 4 else None)
            if not b2 and not art:
                continue
            base = fnum(row[8] if len(row) > 8 else None)
            ue_cost = fnum(row[11] if len(row) > 11 else None)
            price = fnum(row[13] if len(row) > 13 else None)
            pay = fnum(row[15] if len(row) > 15 else None)
            hours = fnum(row[10] if len(row) > 10 else None)
            lines.append(
                {
                    "line_id": f"PORTN-{period}-{name.split()[0]}-{i}",
                    "period_month": period,
                    "sheet_kind": "portn",
                    "section": section,
                    "tailor_full": name,
                    "tailor_surname": name.split()[0],
                    "product_name": b2[:80],
                    "fabric_note": str(row[3] or "")[:60] if len(row) > 3 else "",
                    "article_raw": str(row[4] or "").strip() if len(row) > 4 else "",
                    "article_norm": art,
                    "constructor": str(row[5] or "").strip() if len(row) > 5 else "",
                    "client_or_project": str(row[6] or "").strip() if len(row) > 6 else "",
                    "base_cost_rub": base if base is not None else "",
                    "ue_cost_rub": ue_cost if ue_cost is not None else "",
                    "price_rub": price if price is not None else "",
                    "pay_rub": pay if pay is not None else "",
                    "qty_or_hours": hours if hours is not None else "",
                    "source_file": meta["file_name"],
                    "source_file_id": meta["source_file_id"],
                    "so_t": "N",
                }
            )
        pays.append(
            {
                "pay_id": f"PAY-{period}-{name.split()[0]}",
                "period_month": period,
                "sheet_kind": "portn",
                "tailor_full": name,
                "tailor_surname": name.split()[0],
                "pay_total_rub": pay_total if pay_total is not None else "",
                "source_file": meta["file_name"],
                "so_t": "N",
            }
        )
    return lines, pays


def parse_smetka(path: Path, meta: dict, period: str) -> tuple[list[dict], list[dict]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    sn = [s for s in wb.sheetnames if "смет" in s.lower()]
    if not sn:
        wb.close()
        return [], []
    rows = list(wb[sn[0]].iter_rows(values_only=True))
    wb.close()
    lines: list[dict] = []
    pays: list[dict] = []
    person_starts = []
    for i, row in enumerate(rows):
        a0 = row[0] if row else None
        if is_person(a0):
            person_starts.append((i, a0.strip()))
    for idx, (start, name) in enumerate(person_starts):
        end = person_starts[idx + 1][0] if idx + 1 < len(person_starts) else len(rows)
        pay_total = None
        for i in range(start + 1, end):
            row = rows[i]
            if not row:
                continue
            if str(row[0] or "").startswith("№") or str(row[1] or "").startswith("№"):
                continue
            art = norm_art(row[3] if len(row) > 3 else None)
            pname = str(row[1] or "").strip()
            if pname.lower() in ("итого", "") and not art:
                for j in range(len(row)):
                    if isinstance(row[j], (int, float)) and j >= 7:
                        pay_total = float(row[j])
                continue
            if not pname and not art:
                continue
            hours = fnum(row[7] if len(row) > 7 else None)
            cost = fnum(row[9] if len(row) > 9 else None)
            lines.append(
                {
                    "line_id": f"SMET-{period}-{name.split()[0]}-{i}",
                    "period_month": period,
                    "sheet_kind": "smetka",
                    "section": "smetka_hours",
                    "tailor_full": name,
                    "tailor_surname": name.split()[0],
                    "product_name": pname[:80],
                    "fabric_note": str(row[2] or "")[:60] if len(row) > 2 else "",
                    "article_raw": str(row[3] or "").strip() if len(row) > 3 else "",
                    "article_norm": art,
                    "constructor": str(row[4] or "").strip() if len(row) > 4 else "",
                    "client_or_project": str(row[5] or "").strip() if len(row) > 5 else "",
                    "base_cost_rub": "",
                    "ue_cost_rub": "",
                    "price_rub": cost if cost is not None else "",
                    "pay_rub": cost if cost is not None else "",
                    "qty_or_hours": hours if hours is not None else "",
                    "source_file": meta["file_name"],
                    "source_file_id": meta["source_file_id"],
                    "so_t": "N",
                }
            )
        pays.append(
            {
                "pay_id": f"PAYS-{period}-{name.split()[0]}",
                "period_month": period,
                "sheet_kind": "smetka",
                "tailor_full": name,
                "tailor_surname": name.split()[0],
                "pay_total_rub": pay_total if pay_total is not None else "",
                "source_file": meta["file_name"],
                "so_t": "N",
            }
        )
    return lines, pays


def build_stock_index():
    by = defaultdict(lambda: {"cost": 0.0, "qty": 0.0, "wh": set(), "n": 0})
    with (MARTS / "stock_cost_articles.csv").open(encoding="utf-8") as f:
        for r in csv.DictReader(f):
            a = r["article_norm"]
            by[a]["n"] += 1
            try:
                by[a]["cost"] += float(r["cost_rub"] or 0)
            except ValueError:
                pass
            try:
                by[a]["qty"] += float(r["qty_in_move"] or 0)
            except ValueError:
                pass
            by[a]["wh"].add(r["warehouse"])
    return by


def stock_lookup(art: str, stock_by: dict):
    if not art:
        return None
    if art in stock_by:
        return stock_by[art]
    hits = []
    for k, v in stock_by.items():
        base = k.split("/")[0]
        if k.startswith(art) or art.startswith(base) or base.startswith(art):
            hits.append(v)
    if not hits:
        return None
    m = {"cost": 0.0, "qty": 0.0, "wh": set(), "n": 0}
    for v in hits[:30]:
        m["cost"] += v["cost"]
        m["qty"] += v["qty"]
        m["wh"] |= v["wh"]
        m["n"] += v["n"]
    return m


def main() -> dict:
    files = catalog()
    all_lines: list[dict] = []
    all_pays: list[dict] = []
    for meta in files:
        period = month_from_name(meta["file_name"])
        a, b = parse_portn(meta["_path"], meta, period)
        c, d = parse_smetka(meta["_path"], meta, period)
        all_lines.extend(a)
        all_lines.extend(c)
        all_pays.extend(b)
        all_pays.extend(d)

    col = {
        r["article_norm"]: r
        for r in csv.DictReader(open(MARTS / "collection_sku_stock_bridge.csv", encoding="utf-8"))
    }
    pc = {
        r["article_norm"]: r
        for r in csv.DictReader(open(MARTS / "person_cost_collection_bridge.csv", encoding="utf-8"))
    }
    stock_by = build_stock_index()

    bridge = []
    for r in all_lines:
        a = r["article_norm"]
        c = col.get(a) if a else None
        p = pc.get(a) if a else None
        s = stock_lookup(a, stock_by) if a else None
        flags = []
        if c:
            flags.append("COLLECTION")
        if p:
            flags.append("PERSON_COST")
        if s:
            flags.append("STOCK_WH")
        if flags:
            bv = "|".join(flags)
        elif not a:
            bv = "NO_ART"
        else:
            bv = "ORPHAN"
        bridge.append(
            {
                **r,
                "in_collection": "Y" if c else "N",
                "collection_sale_eur": c["collection_sale_eur"] if c else "",
                "in_person_cost": "Y" if p else "N",
                "in_stock": "Y" if s else "N",
                "stock_cost_rub": round(s["cost"], 2) if s else "",
                "stock_qty": round(s["qty"], 2) if s else "",
                "warehouses": "|".join(sorted(s["wh"])) if s else "",
                "bridge_value": bv,
                "do_not_auto_accept": "YES",
            }
        )

    # tailor × month summary (portn pay)
    by_tm = defaultdict(
        lambda: {
            "lines_n": 0,
            "arts": set(),
            "col_n": 0,
            "pc_n": 0,
            "stk_n": 0,
            "pay_lines": 0.0,
            "pay_total": "",
        }
    )
    pay_map = {
        (r["period_month"], r["tailor_surname"], r["sheet_kind"]): r["pay_total_rub"]
        for r in all_pays
    }
    for r in bridge:
        key = (r["period_month"], r["tailor_surname"], r["sheet_kind"])
        b = by_tm[key]
        b["lines_n"] += 1
        if r["article_norm"]:
            b["arts"].add(r["article_norm"])
        if r["in_collection"] == "Y":
            b["col_n"] += 1
        if r["in_person_cost"] == "Y":
            b["pc_n"] += 1
        if r["in_stock"] == "Y":
            b["stk_n"] += 1
        if r["pay_rub"] != "":
            b["pay_lines"] += float(r["pay_rub"])
        b["pay_total"] = pay_map.get(key, "")

    tailor_month = []
    for (period, surname, kind), b in sorted(by_tm.items()):
        tailor_month.append(
            {
                "period_month": period,
                "tailor_surname": surname,
                "sheet_kind": kind,
                "lines_n": b["lines_n"],
                "unique_articles_n": len(b["arts"]),
                "collection_lines_n": b["col_n"],
                "person_cost_lines_n": b["pc_n"],
                "stock_lines_n": b["stk_n"],
                "pay_from_lines_rub": round(b["pay_lines"], 2),
                "pay_total_rub": b["pay_total"],
                "so_t": "N",
            }
        )

    # article aggregate
    by_art = defaultdict(
        lambda: {
            "lines_n": 0,
            "tailors": set(),
            "constructors": set(),
            "pay": 0.0,
            "months": set(),
            "col": "",
            "pc": "",
            "stk": "",
            "wh": set(),
            "sale": "",
        }
    )
    for r in bridge:
        a = r["article_norm"] or "NO_ART"
        b = by_art[a]
        b["lines_n"] += 1
        b["tailors"].add(r["tailor_surname"])
        if r["constructor"]:
            b["constructors"].add(r["constructor"].split()[0])
        if r["pay_rub"] != "":
            b["pay"] += float(r["pay_rub"])
        b["months"].add(r["period_month"])
        b["col"] = r["in_collection"]
        b["pc"] = r["in_person_cost"]
        b["stk"] = r["in_stock"]
        if r["warehouses"]:
            b["wh"] |= set(r["warehouses"].split("|"))
        if r["collection_sale_eur"] != "":
            b["sale"] = r["collection_sale_eur"]

    art_sum = []
    for a, b in sorted(by_art.items(), key=lambda x: -x[1]["lines_n"]):
        if a == "NO_ART":
            continue
        flags = []
        if b["col"] == "Y":
            flags.append("COLLECTION")
        if b["pc"] == "Y":
            flags.append("PERSON_COST")
        if b["stk"] == "Y":
            flags.append("STOCK_WH")
        art_sum.append(
            {
                "article_norm": a,
                "lines_n": b["lines_n"],
                "tailors": "|".join(sorted(b["tailors"])),
                "constructors": "|".join(sorted(b["constructors"])),
                "months_n": len(b["months"]),
                "pay_lines_sum_rub": round(b["pay"], 2),
                "in_collection": b["col"],
                "collection_sale_eur": b["sale"],
                "in_person_cost": b["pc"],
                "in_stock": b["stk"],
                "warehouses": "|".join(sorted(b["wh"]))[:200],
                "bridge_value": "|".join(flags) if flags else "ORPHAN",
                "do_not_auto_accept": "YES",
                "so_t": "N",
            }
        )

    # warehouse ABC from shop-linked stock hits
    wh_abc = Counter()
    for r in bridge:
        if r["warehouses"]:
            for w in r["warehouses"].split("|"):
                wh_abc[w] += 1
    wh_rows = [
        {
            "warehouse": w,
            "shop_line_hits_n": n,
            "note": "frequency of цех lines that also appear in stock_cost warehouses",
            "so_t": "N",
        }
        for w, n in wh_abc.most_common()
    ]

    month_tot = []
    for m in sorted({r["period_month"] for r in all_pays}):
        portn = [r for r in all_pays if r["period_month"] == m and r["sheet_kind"] == "portn"]
        month_tot.append(
            {
                "period_month": m,
                "portn_tailors_n": len(portn),
                "portn_pay_sum_rub": round(
                    sum(float(r["pay_total_rub"]) for r in portn if r["pay_total_rub"] != ""), 2
                ),
                "lines_n": sum(1 for r in bridge if r["period_month"] == m),
                "lines_with_art_n": sum(
                    1 for r in bridge if r["period_month"] == m and r["article_norm"]
                ),
                "collection_lines_n": sum(
                    1 for r in bridge if r["period_month"] == m and r["in_collection"] == "Y"
                ),
                "stock_lines_n": sum(
                    1 for r in bridge if r["period_month"] == m and r["in_stock"] == "Y"
                ),
                "so_t": "N",
            }
        )

    uniq_arts = {r["article_norm"] for r in bridge if r["article_norm"]}
    meta = {
        "horizon": "H74",
        "date": str(date.today()),
        "title": "Цех ЗП ↔ collections / person-cost / stock warehouses",
        "files_n": len(files),
        "lines_n": len(bridge),
        "lines_with_art_n": sum(1 for r in bridge if r["article_norm"]),
        "unique_articles_n": len(uniq_arts),
        "collection_lines_n": sum(1 for r in bridge if r["in_collection"] == "Y"),
        "person_cost_lines_n": sum(1 for r in bridge if r["in_person_cost"] == "Y"),
        "stock_lines_n": sum(1 for r in bridge if r["in_stock"] == "Y"),
        "unique_arts_collection_n": sum(1 for a in uniq_arts if a in col),
        "unique_arts_stock_n": sum(1 for a in uniq_arts if stock_lookup(a, stock_by)),
        "portn_pay_sum_rub": round(
            sum(
                float(r["pay_total_rub"])
                for r in all_pays
                if r["sheet_kind"] == "portn" and r["pay_total_rub"] != ""
            ),
            2,
        ),
        "collection_sale_eur_touched": round(
            sum(
                float(r["collection_sale_eur"])
                for r in art_sum
                if r["collection_sale_eur"] != "" and r["in_collection"] == "Y"
            ),
            2,
        ),
        "no_fake_accept": True,
        "so_t": False,
        "note": "Stock hit rare by design (H64): collection-style arts rarely = goods SKU warehouses",
    }

    REG.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)
    (REG / "meta.json").write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
    write_csv(REG / "shop_work_lines.csv", all_lines)
    write_csv(REG / "shop_pay_totals.csv", all_pays)
    write_csv(REG / "shop_collection_stock_bridge.csv", bridge)
    write_csv(REG / "shop_tailor_month_summary.csv", tailor_month)
    write_csv(REG / "shop_article_summary.csv", art_sum)
    if wh_rows:
        write_csv(REG / "shop_warehouse_hits.csv", wh_rows)
    write_csv(REG / "shop_month_totals.csv", month_tot)

    copies = [
        "shop_work_lines.csv",
        "shop_pay_totals.csv",
        "shop_collection_stock_bridge.csv",
        "shop_tailor_month_summary.csv",
        "shop_article_summary.csv",
        "shop_warehouse_hits.csv",
        "shop_month_totals.csv",
        "meta.json",
    ]
    for name in copies:
        src = REG / name
        if not src.exists():
            continue
        if name == "meta.json":
            shutil.copy2(src, MARTS / "h74_meta.json")
            shutil.copy2(src, MAPS / "h74_meta.json")
            shutil.copy2(src, EV / "meta.json")
        else:
            shutil.copy2(src, MARTS / name)
            shutil.copy2(src, MAPS / name)
            shutil.copy2(src, EV / name)

    shutil.copy2(REG / "shop_collection_stock_bridge.csv", WAVE_B / "38_shop_collection_stock_bridge.csv")
    shutil.copy2(REG / "shop_article_summary.csv", WAVE_B / "39_shop_article_summary.csv")
    shutil.copy2(REG / "shop_tailor_month_summary.csv", WAVE_B / "40_shop_tailor_month_summary.csv")
    shutil.copy2(REG / "shop_warehouse_hits.csv", WAVE_B / "41_shop_warehouse_hits.csv")
    shutil.copy2(REG / "shop_month_totals.csv", WAVE_B / "42_shop_month_totals.csv")

    print(json.dumps(meta, ensure_ascii=False, indent=2))
    return meta


if __name__ == "__main__":
    main()
