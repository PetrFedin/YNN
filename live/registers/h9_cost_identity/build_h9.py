#!/usr/bin/env python3
"""
H9: identity-aware COGS + returns mart + finance review packet.

Проблема: один canonical_sku ≠ один продукт (коллизии артикулов в 1С/остатках).
Пример: T-3178 sales = майка «Перец», H5/FILE cost = рубашка «горох» → ложный убыток.

1) Сравнить токены имени sales vs cost; при плохом match и наличии W3 BOM — переключить COGS
2) Mart возвратов (RETURN_OR_CREDIT)
3) Finance packet по остаточным neg SKU (коммерческий убыток / нужна проверка)
4) Пересобрать margin marts

Не SoT. Не трогает RACI.
"""
from __future__ import annotations

import csv
import json
import re
import shutil
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[3]
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/h9_cost_identity_20260724"
MART = ROOT / "live/marts"
W3 = ROOT / "live/registers/w3_sku_cost"
W4 = ROOT / "live/registers/w4_sales_settle"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")

STOP = {
    "женский",
    "женская",
    "женское",
    "из",
    "с",
    "и",
    "для",
    "на",
    "в",
    "по",
    "the",
    "a",
    "of",
}


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def fnum(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def tokens(name: str) -> set[str]:
    if not name:
        return set()
    raw = re.findall(r"[A-Za-zА-Яа-яЁё0-9]+", name.lower())
    return {t for t in raw if len(t) >= 3 and t not in STOP}


def jaccard(a: set[str], b: set[str]) -> float:
    if not a or not b:
        return 0.0
    return len(a & b) / len(a | b)


def load_w3_bom() -> dict[str, dict]:
    """Лучший W3 BOM unit_cost по SKU (предпочитаем не-H5 id)."""
    best: dict[str, dict] = {}
    path = W3 / "cost_versions.csv"
    if not path.exists():
        return best
    for r in csv.DictReader(open(path, encoding="utf-8")):
        can = r.get("canonical_sku") or ""
        unit = fnum(r.get("unit_cost_rub"))
        if not can or unit is None or unit <= 0:
            continue
        cid = r.get("cost_version_id") or ""
        # предпочитаем исходный W3 (не CV-H5-*)
        score = 2 if not cid.startswith("CV-H5-") else 1
        cur = best.get(can)
        if cur is None or score > cur["_score"] or (score == cur["_score"] and unit < cur["unit_cost"]):
            best[can] = {
                "unit_cost": unit,
                "name": r.get("name") or "",
                "cost_version_id": cid,
                "direction": r.get("direction") or "",
                "_score": score,
            }
    return best


def apply_identity_fix(sales: list[dict], bom: dict[str, dict]) -> tuple[list[dict], dict]:
    switched = 0
    flagged = 0
    collisions = []

    for s in sales:
        if s.get("dq_exclude_from_margin") == "Y":
            continue
        can = s.get("canonical_sku") or ""
        if not can or can not in bom:
            continue
        rev = fnum(s.get("revenue_rub"))
        qty = fnum(s.get("qty")) or 0
        cogs = fnum(s.get("cogs_rub"))
        if rev is None or rev <= 0 or qty <= 0 or cogs is None:
            continue

        sale_name = s.get("sku_name") or ""
        # текущее «имя» cost часто не хранится — сравним с bom.name и эвристикой unit
        bom_row = bom[can]
        sale_tok = tokens(sale_name)
        bom_tok = tokens(bom_row["name"])
        sim = jaccard(sale_tok, bom_tok)

        unit_cur = cogs / qty if qty else None
        unit_bom = bom_row["unit_cost"]
        src = s.get("cogs_source") or ""

        # коллизия: bom похож на sales, а текущий unit >> bom (×3+) и маржа отрицательная
        margin = rev - cogs
        likely_collision = (
            sim >= 0.25
            and unit_cur is not None
            and unit_cur > unit_bom * 3
            and margin < 0
            and src in ("FILE", "H5_STOCK", "W3_H5", "W3")
        )
        # или sales и bom совпадают по ключевому слову продукта, а cost source FILE с unit≈H5 collision
        product_hit = bool(sale_tok & bom_tok & {"майка", "худи", "свитшот", "платье", "юбка", "рубашка"})
        if product_hit and unit_cur and unit_cur > unit_bom * 3 and margin < 0:
            likely_collision = True

        if not likely_collision:
            # мягкий флаг: neg margin + bom даёт положительную маржу
            if margin < 0 and (rev - unit_bom * qty) > 0 and sim >= 0.2:
                flags = (s.get("dq_flags") or "").split("|") if s.get("dq_flags") else []
                if "COST_IDENTITY_REVIEW" not in flags:
                    flags.append("COST_IDENTITY_REVIEW")
                    s["dq_flags"] = "|".join(x for x in flags if x)
                    flagged += 1
            continue

        new_cogs = round(unit_bom * qty, 2)
        collisions.append(
            {
                "sales_line_id": s.get("sales_line_id"),
                "canonical_sku": can,
                "channel": s.get("channel"),
                "sale_name": sale_name[:120],
                "bom_name": bom_row["name"][:120],
                "name_jaccard": round(sim, 3),
                "old_cogs": round(cogs, 2),
                "old_source": src,
                "new_cogs": new_cogs,
                "new_source": "W3_BOM_H9",
                "cost_version_id": bom_row["cost_version_id"],
                "old_margin": round(margin, 2),
                "new_margin": round(rev - new_cogs, 2),
            }
        )
        s["cogs_rub"] = new_cogs
        s["cogs_source"] = "W3_BOM_H9"
        s["w3_unit_cost"] = unit_bom
        s["w3_cost_version_id"] = bom_row["cost_version_id"]
        s["margin_rub"] = round(rev - new_cogs, 2)
        flags = (s.get("dq_flags") or "").split("|") if s.get("dq_flags") else []
        flags = [x for x in flags if x and x != "INFLATED_FILE_COGS"]
        flags.append("COST_SKU_COLLISION_FIXED")
        s["dq_flags"] = "|".join(flags)
        switched += 1

    stats = {
        "switched_lines": switched,
        "review_flagged": flagged,
        "collision_rows": len(collisions),
    }
    return collisions, stats


def rebuild_marts(sales: list[dict]) -> dict:
    by_cm = defaultdict(
        lambda: {
            "revenue": 0.0,
            "revenue_costed": 0.0,
            "cogs": 0.0,
            "qty": 0.0,
            "lines": 0,
            "lines_with_cogs": 0,
        }
    )
    by_sku = defaultdict(
        lambda: {"revenue": 0.0, "cogs": 0.0, "qty": 0.0, "lines": 0, "name": "", "channels": set(), "costed_rev": 0.0}
    )
    returns_cm = defaultdict(lambda: {"lines": 0, "revenue": 0.0, "cogs": 0.0})

    for s in sales:
        flags = s.get("dq_flags") or ""
        if "RETURN_OR_CREDIT" in flags:
            ch = s.get("channel") or ""
            pm = s.get("period_month") or ""
            returns_cm[(ch, pm)]["lines"] += 1
            returns_cm[(ch, pm)]["revenue"] += fnum(s.get("revenue_rub")) or 0.0
            returns_cm[(ch, pm)]["cogs"] += fnum(s.get("cogs_rub")) or 0.0

        if s.get("dq_exclude_from_margin") == "Y":
            continue
        ch = s.get("channel") or ""
        pm = s.get("period_month") or ""
        rev = fnum(s.get("revenue_rub")) or 0.0
        cogs = fnum(s.get("cogs_rub"))
        qty = fnum(s.get("qty")) or 0.0
        by_cm[(ch, pm)]["revenue"] += rev
        by_cm[(ch, pm)]["qty"] += qty
        by_cm[(ch, pm)]["lines"] += 1
        if cogs is not None:
            by_cm[(ch, pm)]["cogs"] += cogs
            by_cm[(ch, pm)]["revenue_costed"] += rev
            by_cm[(ch, pm)]["lines_with_cogs"] += 1

        can = s.get("canonical_sku") or ""
        if can:
            by_sku[can]["revenue"] += rev
            by_sku[can]["qty"] += qty
            by_sku[can]["lines"] += 1
            if cogs is not None:
                by_sku[can]["cogs"] += cogs
                by_sku[can]["costed_rev"] += rev
            if not by_sku[can]["name"]:
                by_sku[can]["name"] = (s.get("sku_name") or "")[:100]
            by_sku[can]["channels"].add(ch)

    channel_month = []
    for (ch, pm), v in sorted(by_cm.items()):
        margin = v["revenue_costed"] - v["cogs"] if v["lines_with_cogs"] else None
        pct = (margin / v["revenue_costed"] * 100) if margin is not None and v["revenue_costed"] else None
        cov = v["lines_with_cogs"] / v["lines"] if v["lines"] else 0
        channel_month.append(
            {
                "channel": ch,
                "period_month": pm,
                "lines": v["lines"],
                "lines_with_cogs": v["lines_with_cogs"],
                "cogs_coverage": round(cov, 3),
                "qty": round(v["qty"], 2),
                "revenue_rub": round(v["revenue"], 2),
                "revenue_costed_rub": round(v["revenue_costed"], 2),
                "cogs_rub": round(v["cogs"], 2) if v["lines_with_cogs"] else "",
                "margin_rub": round(margin, 2) if margin is not None else "",
                "margin_pct": round(pct, 1) if pct is not None else "",
                "status": "OK" if cov >= 0.85 else ("PARTIAL" if cov >= 0.5 else "WEAK"),
            }
        )

    by_ch = defaultdict(lambda: {"revenue": 0.0, "revenue_costed": 0.0, "cogs": 0.0, "lines": 0, "lines_with_cogs": 0})
    for r in channel_month:
        ch = r["channel"]
        by_ch[ch]["revenue"] += r["revenue_rub"]
        by_ch[ch]["lines"] += r["lines"]
        if r["cogs_rub"] != "":
            by_ch[ch]["cogs"] += float(r["cogs_rub"])
            by_ch[ch]["revenue_costed"] += float(r["revenue_costed_rub"])
            by_ch[ch]["lines_with_cogs"] += r["lines_with_cogs"]

    channel_total = []
    for ch, v in sorted(by_ch.items()):
        margin = v["revenue_costed"] - v["cogs"] if v["lines_with_cogs"] else None
        pct = (margin / v["revenue_costed"] * 100) if margin is not None and v["revenue_costed"] else None
        cov = v["lines_with_cogs"] / v["lines"] if v["lines"] else 0
        channel_total.append(
            {
                "channel": ch,
                "lines": v["lines"],
                "cogs_coverage": round(cov, 3),
                "revenue_rub": round(v["revenue"], 2),
                "revenue_costed_rub": round(v["revenue_costed"], 2),
                "cogs_rub": round(v["cogs"], 2),
                "margin_rub": round(margin, 2) if margin is not None else "",
                "margin_pct": round(pct, 1) if pct is not None else "",
            }
        )

    sku_rows = []
    for can, v in by_sku.items():
        if not v["costed_rev"]:
            continue
        margin = v["costed_rev"] - v["cogs"]
        pct = margin / v["costed_rev"] * 100 if v["costed_rev"] else None
        sku_rows.append(
            {
                "canonical_sku": can,
                "name": v["name"],
                "channels": ",".join(sorted(v["channels"])),
                "lines": v["lines"],
                "qty": round(v["qty"], 2),
                "revenue_rub": round(v["revenue"], 2),
                "cogs_rub": round(v["cogs"], 2),
                "margin_rub": round(margin, 2),
                "margin_pct": round(pct, 1) if pct is not None else "",
            }
        )
    sku_rows.sort(key=lambda x: x["margin_rub"], reverse=True)

    neg = [r for r in sku_rows if r["margin_rub"] < 0]
    top40 = sku_rows[:40]
    bottom40 = list(reversed(sku_rows[-40:]))

    returns_rows = []
    for (ch, pm), v in sorted(returns_cm.items()):
        returns_rows.append(
            {
                "channel": ch,
                "period_month": pm,
                "lines": v["lines"],
                "revenue_rub": round(v["revenue"], 2),
                "cogs_rub": round(v["cogs"], 2),
            }
        )

    write_csv(
        MART / "margin_channel_month.csv",
        channel_month,
        [
            "channel",
            "period_month",
            "lines",
            "lines_with_cogs",
            "cogs_coverage",
            "qty",
            "revenue_rub",
            "revenue_costed_rub",
            "cogs_rub",
            "margin_rub",
            "margin_pct",
            "status",
        ],
    )
    write_csv(
        MART / "margin_channel_total.csv",
        channel_total,
        [
            "channel",
            "lines",
            "cogs_coverage",
            "revenue_rub",
            "revenue_costed_rub",
            "cogs_rub",
            "margin_rub",
            "margin_pct",
        ],
    )
    write_csv(
        MART / "margin_sku_top40.csv",
        top40,
        ["canonical_sku", "name", "channels", "lines", "qty", "revenue_rub", "cogs_rub", "margin_rub", "margin_pct"],
    )
    write_csv(
        MART / "margin_sku_bottom40.csv",
        bottom40,
        ["canonical_sku", "name", "channels", "lines", "qty", "revenue_rub", "cogs_rub", "margin_rub", "margin_pct"],
    )
    write_csv(
        MART / "margin_negative_skus.csv",
        [
            {
                "scope": "SKU",
                "channel": r["channels"],
                "period_month": "",
                "canonical_sku": r["canonical_sku"],
                "revenue_rub": r["revenue_rub"],
                "margin_pct": r["margin_pct"],
                "cogs_coverage": "",
                "flags": "NEGATIVE_MARGIN",
                "priority": "HIGH",
                "name": r["name"],
                "margin_rub": r["margin_rub"],
            }
            for r in neg
        ],
        [
            "scope",
            "channel",
            "period_month",
            "canonical_sku",
            "revenue_rub",
            "margin_pct",
            "cogs_coverage",
            "flags",
            "priority",
            "name",
            "margin_rub",
        ],
    )
    write_csv(
        MART / "returns_by_channel_month.csv",
        returns_rows,
        ["channel", "period_month", "lines", "revenue_rub", "cogs_rub"],
    )

    tot_rev = sum(r["revenue_costed_rub"] for r in channel_total)
    tot_cogs = sum(r["cogs_rub"] for r in channel_total)
    tot_m = tot_rev - tot_cogs
    return {
        "negative_skus": len(neg),
        "overall_revenue_costed": round(tot_rev, 2),
        "overall_margin": round(tot_m, 2),
        "overall_margin_pct": round(tot_m / tot_rev * 100, 1) if tot_rev else None,
        "returns_lines": sum(r["lines"] for r in returns_rows),
        "returns_revenue_sum": round(sum(r["revenue_rub"] for r in returns_rows), 2),
    }


def finance_packet(sales: list[dict], bom: dict[str, dict]) -> list[dict]:
    by_sku = defaultdict(
        lambda: {
            "revenue": 0.0,
            "cogs": 0.0,
            "qty": 0.0,
            "lines": 0,
            "name": "",
            "channels": set(),
            "sources": set(),
        }
    )
    for s in sales:
        if s.get("dq_exclude_from_margin") == "Y":
            continue
        can = s.get("canonical_sku") or ""
        cogs = fnum(s.get("cogs_rub"))
        rev = fnum(s.get("revenue_rub"))
        if not can or cogs is None or rev is None:
            continue
        by_sku[can]["revenue"] += rev
        by_sku[can]["cogs"] += cogs
        by_sku[can]["qty"] += fnum(s.get("qty")) or 0
        by_sku[can]["lines"] += 1
        by_sku[can]["sources"].add(s.get("cogs_source") or "")
        by_sku[can]["channels"].add(s.get("channel") or "")
        if not by_sku[can]["name"]:
            by_sku[can]["name"] = (s.get("sku_name") or "")[:120]

    rows = []
    for can, v in by_sku.items():
        margin = v["revenue"] - v["cogs"]
        if margin >= 0:
            continue
        bom_u = bom.get(can, {}).get("unit_cost")
        bom_name = bom.get(can, {}).get("name", "")
        alt_m = None
        if bom_u and v["qty"]:
            alt_m = round(v["revenue"] - bom_u * v["qty"], 2)
        reason = "WHOLESALE_BELOW_STOCK_COST"
        if bom_name and jaccard(tokens(v["name"]), tokens(bom_name)) < 0.15:
            reason = "POSSIBLE_SKU_COLLISION"
        elif alt_m is not None and alt_m > 0:
            reason = "BOM_WOULD_FLIP_POSITIVE"
        rows.append(
            {
                "canonical_sku": can,
                "name": v["name"],
                "channels": ",".join(sorted(v["channels"])),
                "lines": v["lines"],
                "qty": round(v["qty"], 2),
                "revenue_rub": round(v["revenue"], 2),
                "cogs_rub": round(v["cogs"], 2),
                "margin_rub": round(margin, 2),
                "margin_pct": round(margin / v["revenue"] * 100, 1) if v["revenue"] else "",
                "cogs_sources": "|".join(sorted(v["sources"])),
                "bom_unit_cost": bom_u if bom_u else "",
                "bom_name": bom_name[:100],
                "alt_margin_if_bom": alt_m if alt_m is not None else "",
                "review_reason": reason,
                "action": "FINANCE_CONFIRM_COST_OR_PRICE",
            }
        )
    rows.sort(key=lambda x: x["margin_rub"])
    return rows


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)
    MART.mkdir(parents=True, exist_ok=True)

    src = W4 / "sales_lines.csv"
    bak = W4 / "sales_lines_pre_h9.csv"
    if not bak.exists():
        shutil.copy2(src, bak)

    sales = list(csv.DictReader(open(src, encoding="utf-8")))
    before = rebuild_marts(sales)  # snapshot metrics path will overwrite — compute manually
    # quick before metrics
    neg_before = before["negative_skus"]
    pct_before = before["overall_margin_pct"]

    bom = load_w3_bom()
    collisions, stats = apply_identity_fix(sales, bom)

    # rewrite sales
    fields = list(sales[0].keys()) if sales else []
    write_csv(src, sales, fields)
    write_csv(OUT / "cost_collisions_fixed.csv", collisions, list(collisions[0].keys()) if collisions else ["canonical_sku"])

    after = rebuild_marts(sales)
    packet = finance_packet(sales, bom)
    write_csv(
        MART / "finance_neg_sku_review.csv",
        packet,
        list(packet[0].keys()) if packet else ["canonical_sku"],
    )
    write_csv(OUT / "finance_neg_sku_review.csv", packet, list(packet[0].keys()) if packet else ["canonical_sku"])

    # anomalies refresh
    anomalies = []
    for r in packet:
        anomalies.append(
            {
                "scope": "SKU",
                "channel": r["channels"],
                "period_month": "",
                "canonical_sku": r["canonical_sku"],
                "revenue_rub": r["revenue_rub"],
                "margin_pct": r["margin_pct"],
                "cogs_coverage": "",
                "flags": r["review_reason"],
                "priority": "HIGH",
            }
        )
    write_csv(
        MART / "margin_anomalies.csv",
        anomalies,
        ["scope", "channel", "period_month", "canonical_sku", "revenue_rub", "margin_pct", "cogs_coverage", "flags", "priority"],
    )

    summary = {
        "wave": "H9",
        "generated_at": NOW,
        "finding": (
            f"H9: fixed {stats['switched_lines']} cost-identity collisions; "
            f"neg SKUs {neg_before}→{after['negative_skus']}; "
            f"margin {pct_before}%→{after['overall_margin_pct']}%; "
            f"returns {after['returns_lines']} lines / {after['returns_revenue_sum']} ₽."
        ),
        "stats": stats,
        "margin_before": {"negative_skus": neg_before, "margin_pct": pct_before},
        "margin_after": after,
        "finance_packet_rows": len(packet),
        "not_sot": True,
    }
    (OUT / "h9_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    (EV / "h9_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    shutil.copy2(OUT / "cost_collisions_fixed.csv", EV / "cost_collisions_fixed.csv")
    shutil.copy2(MART / "finance_neg_sku_review.csv", EV / "finance_neg_sku_review.csv")
    shutil.copy2(MART / "returns_by_channel_month.csv", EV / "returns_by_channel_month.csv")

    # evidence xlsx
    wb = Workbook()
    ws = wb.active
    ws.title = "H9_Summary"
    ws["A1"] = "H9 Cost Identity"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A2"] = NOW
    ws["A4"] = summary["finding"]
    ws["A6"] = "Switched lines"
    ws["B6"] = stats["switched_lines"]
    ws["A7"] = "Neg SKUs"
    ws["B7"] = f"{neg_before} → {after['negative_skus']}"
    ws["A8"] = "Margin %"
    ws["B8"] = f"{pct_before} → {after['overall_margin_pct']}"
    ws["A9"] = "Returns rev sum"
    ws["B9"] = after["returns_revenue_sum"]
    ws2 = wb.create_sheet("Collisions")
    if collisions:
        ws2.append(list(collisions[0].keys()))
        for r in collisions:
            ws2.append(list(r.values()))
    ws3 = wb.create_sheet("FinanceReview")
    if packet:
        ws3.append(list(packet[0].keys()))
        for r in packet:
            ws3.append(list(r.values()))
    wb.save(EV / "H9_COST_IDENTITY.xlsx")
    wb.save(OUT / "H9_COST_IDENTITY.xlsx")

    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
