#!/usr/bin/env python3
"""
H14: Operating bridge (gross margin → opex/tax → result) + controls dashboard.

Зачем:
1) Свести уже посчитанную маржу продаж с opex/tax/payroll в один месячный мост.
2) Дашборд статусов сверки (IM/TSUM/B2B/DDS/tax/opex/payroll) — одна точка контроля.
3) Материалы/внутренние переводы — memo, не в operating (анти-двойной счёт с COGS).

Не SoT. RACI не трогаем.
"""
from __future__ import annotations

import csv
import json
import re
import shutil
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[3]
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/h14_operating_bridge_20260724"
MART = ROOT / "live/marts"
W2 = ROOT / "live/registers/w2_payroll"
W4 = ROOT / "live/registers/w4_sales_settle"
W5 = ROOT / "live/registers/w5_sup_exp_mat"
W6 = ROOT / "live/registers/w6_tax_bud"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def fnum(v):
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(" ", "").replace(",", "."))
    except (TypeError, ValueError):
        return None


def expense_bucket(name: str, counterparty_surnames: set[str] | None = None) -> str:
    """Классификация статей opex для bridge (не бухгалтерский SoT)."""
    n = (name or "").lower().replace("ё", "е")
    raw = name or ""

    if any(x in n for x in ("оплата труда", "зарплат", "больнич", "отпускн", "преми")):
        return "PAYROLL"
    if any(x in n for x in ("налог", "усн", "ндфл", "взнос", "енп", "единый налоговый", "пени по налог")):
        return "TAX"
    if any(x in n for x in ("аренд",)):
        return "RENT"
    if any(x in n for x in ("реклам", "маркет", "таргет", "smm", "контекст", "вебинар")):
        return "MARKETING"
    if any(x in n for x in ("сдэк", "доставк", "логист", "транспорт", "курьер")):
        return "LOGISTICS"
    if any(x in n for x in ("ткан", "фурнитур", "материал", "расходник")):
        return "MATERIALS_MEMO"
    if any(x in n for x in ("тинькофф", "тбанк", "ссылка в", "счет в тинь", "между счета")):
        return "INTERNAL_TRANSFER_MEMO"
    if "валют" in n or "покупка/продажа" in n:
        return "FX_MEMO"
    if "эквайр" in n or "комиссия эквайр" in n:
        return "ACQUIRING_FEE"
    if "корп.карт" in n or "расходы по карте" in n or "по карте" in n:
        return "CARD_EXPENSE"
    if "агентск" in n or "комисси" in n:
        return "COMMISSION"
    if any(x in n for x in ("аутсорс",)):
        return "OUTSOURCE"

    # платежи контрагентам/байерам, попавшие в expense как ФИО/город — не operating
    if counterparty_surnames:
        tokens = set(re.findall(r"[A-Za-zА-Яа-яЁё]{4,}", raw.upper().replace("Ё", "Е")))
        if tokens & counterparty_surnames:
            return "COUNTERPARTY_MEMO"

    return "OPEX_OTHER"


def load_counterparty_surnames() -> set[str]:
    """Фамилии B2B buyers из settlements — чтобы не тащить их в operating opex."""
    surs: set[str] = set()
    path = W4 / "settlements.csv"
    if not path.exists():
        return surs
    for r in csv.DictReader(open(path, encoding="utf-8")):
        if r.get("channel") != "B2B":
            continue
        buyer = (r.get("buyer") or "").upper().replace("Ё", "Е")
        buyer = re.sub(r"\b(ООО|ИП|АО|ПАО)\b", " ", buyer)
        toks = re.findall(r"[A-ZА-Я]{4,}", buyer)
        if toks:
            surs.add(toks[0])
    return surs


def build_opex_classified() -> tuple[list[dict], dict]:
    rows_out = []
    by_bucket = defaultdict(float)
    src = W5 / "expense_opex_only.csv"
    surnames = load_counterparty_surnames()
    for r in csv.DictReader(open(src, encoding="utf-8")):
        amt = fnum(r.get("amount_rub")) or 0.0
        b = expense_bucket(r.get("article_name") or "", surnames)
        rows_out.append(
            {
                **{k: r.get(k) for k in ("expense_line_id", "period_month", "article_name", "legal_entity_id", "amount_rub")},
                "opex_bucket": b,
            }
        )
        by_bucket[b] += amt
    stats = {k: round(v, 2) for k, v in sorted(by_bucket.items(), key=lambda x: -x[1])}
    return rows_out, stats


def build_operating_bridge(opex_rows: list[dict]) -> tuple[list[dict], dict]:
    # sales margin by month (all channels)
    sales_m = defaultdict(lambda: {"revenue": 0.0, "revenue_costed": 0.0, "cogs": 0.0, "margin": 0.0})
    for r in csv.DictReader(open(MART / "margin_channel_month.csv", encoding="utf-8")):
        pm = r.get("period_month") or ""
        if not pm:
            continue
        sales_m[pm]["revenue"] += fnum(r.get("revenue_rub")) or 0.0
        sales_m[pm]["revenue_costed"] += fnum(r.get("revenue_costed_rub")) or 0.0
        if r.get("cogs_rub") != "":
            sales_m[pm]["cogs"] += fnum(r.get("cogs_rub")) or 0.0
        if r.get("margin_rub") != "":
            sales_m[pm]["margin"] += fnum(r.get("margin_rub")) or 0.0

    # opex by month/bucket
    opex_m = defaultdict(lambda: defaultdict(float))
    for r in opex_rows:
        pm = r.get("period_month") or ""
        if not pm:
            continue
        opex_m[pm][r["opex_bucket"]] += fnum(r.get("amount_rub")) or 0.0

    # tax_cash curated (W6) — предпочтительнее expense TAX bucket для контроля
    tax_m = defaultdict(float)
    for r in csv.DictReader(open(W6 / "tax_cash_lines.csv", encoding="utf-8")):
        tax_m[r.get("period_month") or ""] += fnum(r.get("amount_rub")) or 0.0

    # card spend (optional operating)
    card_m = defaultdict(float)
    card_path = MART / "card_spend_by_category.csv"
    # prefer monthly if exists
    card_month_path = ROOT / "live/registers/h7_controls/card_spend_month_category.csv"
    if card_month_path.exists():
        for r in csv.DictReader(open(card_month_path, encoding="utf-8")):
            card_m[r.get("period_month") or ""] += fnum(r.get("amount_rub") or r.get("out_rub")) or 0.0

    months = sorted(set(sales_m) | set(opex_m) | set(tax_m))
    bridge = []
    for pm in months:
        if not pm:
            continue
        s = sales_m.get(pm, {})
        o = opex_m.get(pm, {})
        payroll = o.get("PAYROLL", 0.0)
        rent = o.get("RENT", 0.0)
        marketing = o.get("MARKETING", 0.0)
        logistics = o.get("LOGISTICS", 0.0)
        outsource = o.get("OUTSOURCE", 0.0)
        acquiring = o.get("ACQUIRING_FEE", 0.0)
        commission = o.get("COMMISSION", 0.0)
        card_exp = o.get("CARD_EXPENSE", 0.0)
        other = o.get("OPEX_OTHER", 0.0)
        tax_exp = o.get("TAX", 0.0)
        tax_cash = tax_m.get(pm, 0.0)
        tax_use = tax_cash if tax_cash else tax_exp
        materials_memo = o.get("MATERIALS_MEMO", 0.0)
        internal_memo = o.get("INTERNAL_TRANSFER_MEMO", 0.0)
        cp_memo = o.get("COUNTERPARTY_MEMO", 0.0)
        fx_memo = o.get("FX_MEMO", 0.0)
        card = card_m.get(pm, 0.0)

        gross = s.get("margin", 0.0)
        opex_core = payroll + rent + marketing + logistics + outsource + acquiring + commission + card_exp + other
        operating = gross - opex_core - tax_use

        rev_c = s.get("revenue_costed", 0.0)
        bridge.append(
            {
                "period_month": pm,
                "revenue_rub": round(s.get("revenue", 0.0), 2),
                "revenue_costed_rub": round(rev_c, 2),
                "cogs_rub": round(s.get("cogs", 0.0), 2),
                "gross_margin_rub": round(gross, 2),
                "gross_margin_pct": round(gross / rev_c * 100, 1) if rev_c else "",
                "payroll_rub": round(payroll, 2),
                "rent_rub": round(rent, 2),
                "marketing_rub": round(marketing, 2),
                "logistics_rub": round(logistics, 2),
                "outsource_rub": round(outsource, 2),
                "acquiring_fee_rub": round(acquiring, 2),
                "commission_rub": round(commission, 2),
                "card_expense_rub": round(card_exp, 2),
                "opex_other_rub": round(other, 2),
                "opex_core_rub": round(opex_core, 2),
                "tax_cash_rub": round(tax_cash, 2),
                "tax_expense_rub": round(tax_exp, 2),
                "tax_used_rub": round(tax_use, 2),
                "operating_result_rub": round(operating, 2),
                "operating_margin_pct": round(operating / rev_c * 100, 1) if rev_c else "",
                "memo_materials_rub": round(materials_memo, 2),
                "memo_internal_transfers_rub": round(internal_memo, 2),
                "memo_counterparty_rub": round(cp_memo, 2),
                "memo_fx_rub": round(fx_memo, 2),
                "memo_card_spend_rub": round(card, 2),
                "note": "operating=GM - opex_core - tax; counterparty/materials/internal/fx=memo",
            }
        )

    # totals
    def sumf(key):
        return round(sum(fnum(r[key]) or 0 for r in bridge), 2)

    tot_rev = sumf("revenue_costed_rub")
    tot_op = sumf("operating_result_rub")
    tot_gm = sumf("gross_margin_rub")
    stats = {
        "months": len(bridge),
        "revenue_costed_total": tot_rev,
        "gross_margin_total": tot_gm,
        "gross_margin_pct": round(tot_gm / tot_rev * 100, 1) if tot_rev else None,
        "opex_core_total": sumf("opex_core_rub"),
        "tax_used_total": sumf("tax_used_rub"),
        "operating_result_total": tot_op,
        "operating_margin_pct": round(tot_op / tot_rev * 100, 1) if tot_rev else None,
        "months_operating_positive": sum(1 for r in bridge if (fnum(r["operating_result_rub"]) or 0) > 0),
        "months_operating_negative": sum(1 for r in bridge if (fnum(r["operating_result_rub"]) or 0) < 0),
    }
    return bridge, stats


def build_controls_dashboard() -> tuple[list[dict], dict]:
    """Сводка статусов сверок в один лист."""
    rows = []

    def add(control, period, status, metric, detail=""):
        rows.append(
            {
                "control_id": control,
                "period_month": period,
                "status": status or "",
                "metric": metric,
                "detail": detail,
            }
        )

    # IM combo
    im_path = MART / "recon_im_combo.csv"
    if im_path.exists():
        for r in csv.DictReader(open(im_path, encoding="utf-8")):
            add("IM_ACQ_COMBO", r["period_month"], r["status"], r.get("gap_pct", ""), r.get("combo_source", ""))

    # TSUM net model
    tsum_path = MART / "recon_tsum_net_model.csv"
    if tsum_path.exists():
        for r in csv.DictReader(open(tsum_path, encoding="utf-8")):
            if (fnum(r.get("sales_revenue_rub")) or 0) <= 0:
                continue
            add("TSUM_NET_MODEL", r["period_month"], r.get("status_vs_model"), r.get("gap_vs_model_pct"), f"rate={r.get('net_rate_used')}")

    # bank vs DDS
    dds_path = MART / "recon_bank_dds_extended.csv"
    if dds_path.exists():
        for r in csv.DictReader(open(dds_path, encoding="utf-8")):
            add("BANK_DDS_CORE", r["period_month"], r.get("status_core_vs_dds"), r.get("delta_core_vs_dds"), "")

    # tax
    tax_path = W6 / "recon_tax_cash_bank.csv"
    if tax_path.exists():
        for r in csv.DictReader(open(tax_path, encoding="utf-8")):
            add("TAX_CASH_BANK", r["period_month"], r.get("status"), r.get("delta"), "")

    # opex vs bank
    opex_path = W5 / "recon_exp_bank_dds.csv"
    if opex_path.exists():
        for r in csv.DictReader(open(opex_path, encoding="utf-8")):
            add("OPEX_VS_BANK", r["period_month"], r.get("status_opex_vs_bank"), r.get("delta_opex_vs_bank"), "")

    # payroll
    zp_path = W2 / "recon_zp_dds_bank.csv"
    if zp_path.exists():
        for r in csv.DictReader(open(zp_path, encoding="utf-8")):
            add("PAYROLL_VS_DDS", r["period_month"], r.get("status_zp_vs_dds"), r.get("delta_zp_vs_dds"), "")

    # B2B coverage snapshot (not monthly)
    settles = list(csv.DictReader(open(W4 / "settlements.csv", encoding="utf-8")))
    b2b = [s for s in settles if s.get("channel") == "B2B"]
    linked = sum(1 for s in b2b if s.get("bank_payment_id"))
    open_n = len(b2b) - linked
    add(
        "B2B_SETTLE_BANK",
        "ALL",
        "PARTIAL" if open_n else "CLOSE",
        f"{linked}/{len(b2b)}",
        f"open={open_n}",
    )

    # summary counts by control
    by_ctrl = defaultdict(Counter)
    for r in rows:
        if r["period_month"] == "ALL":
            continue
        st = r["status"] or "UNKNOWN"
        # normalize soft variants
        if "SOFT" in st:
            key = "SOFT"
        elif st in ("CLOSE", "OK"):
            key = "CLOSE"
        elif st in ("OPEN", "WIDE_GAP", "GAP", "BANK_ONLY", "DDS_ONLY", "N/A"):
            key = "OPEN_OR_GAP"
        else:
            key = st
        by_ctrl[r["control_id"]][key] += 1

    stats = {k: dict(v) for k, v in by_ctrl.items()}
    stats["b2b_linked_open"] = f"{linked}/{open_n}"
    return rows, stats


def build_bridge_totals(bridge: list[dict]) -> list[dict]:
    """Горизонтальный total + по годам."""
    rows = []

    def agg(label, subset):
        rev = sum(fnum(r["revenue_costed_rub"]) or 0 for r in subset)
        gm = sum(fnum(r["gross_margin_rub"]) or 0 for r in subset)
        opex = sum(fnum(r["opex_core_rub"]) or 0 for r in subset)
        tax = sum(fnum(r["tax_used_rub"]) or 0 for r in subset)
        op = sum(fnum(r["operating_result_rub"]) or 0 for r in subset)
        rows.append(
            {
                "scope": label,
                "revenue_costed_rub": round(rev, 2),
                "gross_margin_rub": round(gm, 2),
                "gross_margin_pct": round(gm / rev * 100, 1) if rev else "",
                "opex_core_rub": round(opex, 2),
                "tax_used_rub": round(tax, 2),
                "operating_result_rub": round(op, 2),
                "operating_margin_pct": round(op / rev * 100, 1) if rev else "",
                "months": len(subset),
            }
        )

    agg("TOTAL", bridge)
    by_year = defaultdict(list)
    for r in bridge:
        by_year[r["period_month"][:4]].append(r)
    for y in sorted(by_year):
        agg(f"YEAR_{y}", by_year[y])
    return rows


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)
    MART.mkdir(parents=True, exist_ok=True)

    opex_rows, opex_stats = build_opex_classified()
    write_csv(
        OUT / "opex_classified.csv",
        opex_rows,
        list(opex_rows[0].keys()) if opex_rows else ["expense_line_id"],
    )
    write_csv(MART / "opex_classified.csv", opex_rows, list(opex_rows[0].keys()) if opex_rows else ["expense_line_id"])

    bridge, bridge_stats = build_operating_bridge(opex_rows)
    write_csv(MART / "operating_bridge_month.csv", bridge, list(bridge[0].keys()) if bridge else ["period_month"])
    write_csv(OUT / "operating_bridge_month.csv", bridge, list(bridge[0].keys()) if bridge else ["period_month"])

    totals = build_bridge_totals(bridge)
    write_csv(MART / "operating_bridge_totals.csv", totals, list(totals[0].keys()) if totals else ["scope"])
    write_csv(OUT / "operating_bridge_totals.csv", totals, list(totals[0].keys()) if totals else ["scope"])

    controls, ctrl_stats = build_controls_dashboard()
    write_csv(MART / "controls_dashboard.csv", controls, list(controls[0].keys()) if controls else ["control_id"])
    write_csv(OUT / "controls_dashboard.csv", controls, list(controls[0].keys()) if controls else ["control_id"])

    # control summary compact
    ctrl_sum = []
    for cid, st in sorted(ctrl_stats.items()):
        if cid == "b2b_linked_open":
            continue
        close = st.get("CLOSE", 0)
        soft = st.get("SOFT", 0)
        open_ = st.get("OPEN_OR_GAP", 0)
        other = sum(v for k, v in st.items() if k not in ("CLOSE", "SOFT", "OPEN_OR_GAP"))
        total = close + soft + open_ + other
        ctrl_sum.append(
            {
                "control_id": cid,
                "months": total,
                "close": close,
                "soft": soft,
                "open_or_gap": open_,
                "other": other,
                "close_soft_pct": round((close + soft) / total * 100, 1) if total else "",
            }
        )
    write_csv(MART / "controls_summary.csv", ctrl_sum, list(ctrl_sum[0].keys()) if ctrl_sum else ["control_id"])

    summary = {
        "wave": "H14",
        "generated_at": NOW,
        "finding": (
            f"H14: operating bridge {bridge_stats['months']}m; "
            f"gross margin {bridge_stats['gross_margin_pct']}% → "
            f"operating {bridge_stats['operating_margin_pct']}% "
            f"({bridge_stats['operating_result_total']:,.0f} ₽); "
            f"opex buckets classified; controls dashboard {len(ctrl_sum)} series."
        ),
        "opex_buckets": opex_stats,
        "bridge": bridge_stats,
        "controls": ctrl_stats,
        "not_sot": True,
        "caveats": [
            "materials/internal/counterparty/fx excluded from operating (memo only)",
            "card_expense in opex_core; h7 card mart is separate memo (possible overlap)",
            "tax_used prefers W6 tax_cash over expense TAX articles",
            "OPEX from 1C expense register — not audited P&L",
        ],
    }
    (OUT / "h14_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    (EV / "h14_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    for name in (
        "opex_classified.csv",
        "operating_bridge_month.csv",
        "operating_bridge_totals.csv",
        "controls_dashboard.csv",
        "h14_summary.json",
    ):
        src = OUT / name
        if src.exists():
            shutil.copy2(src, EV / name)
    shutil.copy2(MART / "controls_summary.csv", EV / "controls_summary.csv")

    wb = Workbook()
    ws = wb.active
    ws.title = "H14_Summary"
    ws["A1"] = "H14 Operating Bridge"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A2"] = NOW
    ws["A4"] = summary["finding"]
    ws["A6"] = "Gross margin %"
    ws["B6"] = bridge_stats["gross_margin_pct"]
    ws["A7"] = "Operating margin %"
    ws["B7"] = bridge_stats["operating_margin_pct"]
    ws["A8"] = "Operating result"
    ws["B8"] = bridge_stats["operating_result_total"]
    ws["A9"] = "+ months / − months"
    ws["B9"] = f"{bridge_stats['months_operating_positive']}/{bridge_stats['months_operating_negative']}"
    ws2 = wb.create_sheet("Totals")
    if totals:
        ws2.append(list(totals[0].keys()))
        for r in totals:
            ws2.append(list(r.values()))
    ws3 = wb.create_sheet("Controls")
    if ctrl_sum:
        ws3.append(list(ctrl_sum[0].keys()))
        for r in ctrl_sum:
            ws3.append(list(r.values()))
    wb.save(OUT / "H14_OPERATING_BRIDGE.xlsx")
    wb.save(EV / "H14_OPERATING_BRIDGE.xlsx")

    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
