#!/usr/bin/env python3
"""H64: SKU master + stock cost + fabric inventory → alias/collection crosswalk + fabric ABC.

Not SoT. Strengthens weak capsule/cruise article links and fabric WC view.
"""
from __future__ import annotations

import csv
import json
import re
import shutil
from collections import defaultdict
from datetime import date
from pathlib import Path

import xlrd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[3]
if not (ROOT / "live/registers/00_SOURCE_CATALOG_107.csv").exists():
    ROOT = Path.cwd()
REG = Path(__file__).resolve().parent
MARTS = ROOT / "live/marts"
MAPS = ROOT / "live/maps"
EV = ROOT / "live/evidence/h64_sku_stock_fabric_20260729"
WAVE_B = ROOT / "live/client_pack/execution_wave_b"


def write_csv(path: Path, fields, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)


def norm_art(a: str) -> str:
    if not a:
        return ""
    s = str(a).strip().upper().replace(" ", "").replace("Ё", "Е")
    return s


def catalog_by_cat(cat: str) -> dict:
    rows = list(csv.DictReader((ROOT / "live/registers/00_SOURCE_CATALOG_107.csv").open()))
    for r in rows:
        if r["category"] == cat:
            p = Path(r["path"])
            if not p.exists():
                p = ROOT / "documents" / r["file_name"]
            return {**r, "_path": p}
    raise KeyError(cat)


def parse_sku_master(path: Path, meta: dict) -> list[dict]:
    book = xlrd.open_workbook(path)
    sh = book.sheet_by_index(0)
    out = []
    for r in range(1, sh.nrows):
        name = str(sh.cell_value(r, 0) or "").strip()
        unit = str(sh.cell_value(r, 2) or "").strip()
        article = str(sh.cell_value(r, 4) or "").strip()
        code = str(sh.cell_value(r, 5) or "").strip()
        composition = str(sh.cell_value(r, 6) or "").strip()
        if not name and not article:
            continue
        out.append(
            {
                "sku_row_id": f"SKU-{r}",
                "name": name,
                "article": article,
                "article_norm": norm_art(article),
                "code_1c": code,
                "unit": unit,
                "composition": composition[:120],
                "source_file": meta["file_name"],
                "source_file_id": meta["source_file_id"],
                "so_t": "N",
            }
        )
    return out


def parse_stock_cost(path: Path, meta: dict) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    out = []
    cur_wh = ""
    i = 3
    while i < len(rows):
        row = rows[i]
        a0 = row[0] if row else None
        if a0 and isinstance(a0, str) and "подразделение" in a0.lower():
            cur_wh = a0.strip()
            i += 1
            continue
        # article row: looks like 0-1984F
        if a0 and re.match(r"^[0-9A-Za-zА-Яа-я./\-]+$", str(a0).strip()) and not str(a0).startswith("Футболка") and len(str(a0)) < 40:
            art = str(a0).strip()
            # qty/cost often on same or next name row
            qty = row[3]
            cost = row[4]
            name = ""
            if i + 1 < len(rows) and rows[i + 1] and rows[i + 1][0] and not re.match(r"^[0-9]", str(rows[i + 1][0])):
                name = str(rows[i + 1][0]).strip()
                if rows[i + 1][3] not in (None, "") and qty in (None, ""):
                    qty = rows[i + 1][3]
                if rows[i + 1][4] not in (None, "") and cost in (None, ""):
                    cost = rows[i + 1][4]
                i += 2
            else:
                i += 1
            try:
                qty_f = float(qty) if qty not in (None, "") else None
            except Exception:
                qty_f = None
            try:
                cost_f = float(cost) if cost not in (None, "") else None
            except Exception:
                cost_f = None
            if qty_f is None and cost_f is None and not name:
                continue
            out.append(
                {
                    "stock_line_id": f"STK-{len(out)+1}",
                    "warehouse": cur_wh,
                    "article": art,
                    "article_norm": norm_art(art),
                    "name": name,
                    "qty_in_move": qty_f if qty_f is not None else "",
                    "cost_rub": cost_f if cost_f is not None else "",
                    "unit_cost_rub": round(cost_f / qty_f, 2) if qty_f and cost_f and qty_f else (cost_f if cost_f else ""),
                    "source_file": meta["file_name"],
                    "source_file_id": meta["source_file_id"],
                    "so_t": "N",
                }
            )
            continue
        i += 1
    return out


def _is_fabric_subheader(name: str) -> bool:
    """Exclude warehouse headers / Итого so ABC is not inflated."""
    n = name.strip().lower()
    if n in {"основной склад", "номенклатура", "итого"}:
        return True
    if n.startswith("склад ") or n.startswith("итого"):
        return True
    return False


def parse_fabric_inventory(path: Path, meta: dict) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    out = []
    for i, row in enumerate(rows[3:], start=4):
        name = row[0]
        if not name or not str(name).strip():
            continue
        name = str(name).strip()
        if _is_fabric_subheader(name):
            continue
        try:
            qty = float(row[1]) if row[1] not in (None, "") else None
            amount = float(row[2]) if row[2] not in (None, "") else None
            unit = float(row[3]) if row[3] not in (None, "") else None
        except Exception:
            continue
        if amount is None and qty is None:
            continue
        # extract fabric code like 20-006 / 6-015
        m = re.search(r"(\d{1,2}-\d{3,4}[A-Za-zА-Яа-я]?)\s*$", name)
        code = m.group(1) if m else ""
        # bare large amounts without code are usually subtotals
        if not code and amount is not None and amount > 500_000:
            continue
        out.append(
            {
                "fabric_id": f"FAB-{i}",
                "name": name,
                "fabric_code": code,
                "qty": qty if qty is not None else "",
                "amount_rub": amount if amount is not None else "",
                "unit_cost_rub": unit if unit is not None else "",
                "source_file": meta["file_name"],
                "source_file_id": meta["source_file_id"],
                "as_of": "2026-05-31",
                "so_t": "N",
            }
        )
    return out


def article_family(art: str) -> str:
    if art.startswith("0-") or art.startswith("ИМ-"):
        return "GOODS_STYLE"
    if re.match(r"^\d{2}-", art):
        return "COLLECTION_STYLE"
    return "OTHER"


def abc_rank(rows: list[dict], value_key: str) -> list[dict]:
    items = []
    for r in rows:
        try:
            v = float(r[value_key]) if r[value_key] != "" else 0.0
        except Exception:
            v = 0.0
        items.append((v, r))
    items.sort(key=lambda x: -x[0])
    total = sum(v for v, _ in items) or 1.0
    cum = 0.0
    out = []
    for i, (v, r) in enumerate(items, 1):
        cum += v
        share = cum / total
        band = "A" if share <= 0.8 else ("B" if share <= 0.95 else "C")
        out.append(
            {
                **r,
                "abc_rank": i,
                "abc_value_rub": round(v, 2),
                "abc_cum_share_pct": round(100 * share, 2),
                "abc_band": band,
            }
        )
    return out


def main():
    for d in (REG, MARTS, MAPS, EV, WAVE_B):
        d.mkdir(parents=True, exist_ok=True)

    sku_meta = catalog_by_cat("sku_master")
    stock_meta = catalog_by_cat("stock_cost_movement")
    fab_meta = catalog_by_cat("inventory")

    sku = parse_sku_master(sku_meta["_path"], sku_meta)
    stock = parse_stock_cost(stock_meta["_path"], stock_meta)
    fabrics = parse_fabric_inventory(fab_meta["_path"], fab_meta)
    fabric_abc = abc_rank(fabrics, "amount_rub")

    # aggregate stock by article
    stock_by = defaultdict(lambda: {"name": "", "qty": 0.0, "cost": 0.0, "wh": set()})
    for s in stock:
        a = s["article_norm"]
        if not a:
            continue
        if s["name"]:
            stock_by[a]["name"] = s["name"]
        if s["qty_in_move"] != "":
            stock_by[a]["qty"] += float(s["qty_in_move"])
        if s["cost_rub"] != "":
            stock_by[a]["cost"] += float(s["cost_rub"])
        if s["warehouse"]:
            stock_by[a]["wh"].add(s["warehouse"])

    # alias
    alias = {norm_art(r["canonical_sku"]): r for r in csv.DictReader((MARTS / "sku_alias_master.csv").open())}

    # collection articles
    col_arts = defaultdict(lambda: {"types": set(), "sale_eur": 0.0, "orders": 0})
    for r in csv.DictReader((MARTS / "collection_order_lines.csv").open()):
        a = r["article_norm"]
        if not a:
            continue
        col_arts[a]["types"].add(r["collection_type"])
        col_arts[a]["orders"] += 1
        if r["sale_eur"] != "":
            col_arts[a]["sale_eur"] += float(r["sale_eur"])

    sku_by_art = {r["article_norm"]: r for r in sku if r["article_norm"]}

    bridge = []
    for art, info in sorted(col_arts.items(), key=lambda x: -x[1]["sale_eur"]):
        st = stock_by.get(art)
        sk = sku_by_art.get(art)
        al = alias.get(art)
        in_stock = "Y" if st else "N"
        in_sku = "Y" if sk else "N"
        in_alias = "Y" if al else "N"
        bridge.append(
            {
                "article_norm": art,
                "collection_types": "|".join(sorted(info["types"])),
                "collection_orders_n": info["orders"],
                "collection_sale_eur": round(info["sale_eur"], 2),
                "in_stock_cost_move": in_stock,
                "stock_name": st["name"] if st else "",
                "stock_qty_sum": round(st["qty"], 2) if st else "",
                "stock_cost_rub_sum": round(st["cost"], 2) if st else "",
                "in_sku_master": in_sku,
                "sku_name": sk["name"] if sk else "",
                "in_alias_master": in_alias,
                "alias_status": al.get("alias_status", "") if al else "",
                "link_coverage": (
                    "FULL"
                    if in_stock == "Y" and in_sku == "Y"
                    else (
                        "STOCK_ONLY"
                        if in_stock == "Y"
                        else ("SKU_ONLY" if in_sku == "Y" else ("ALIAS_ONLY" if in_alias == "Y" else "NONE"))
                    )
                ),
                "priority": "HIGH"
                if info["sale_eur"] >= 5000 and in_stock == "N" and in_sku == "N"
                else ("MED" if info["sale_eur"] >= 2000 else "LOW"),
                "so_t": "N",
            }
        )

    # coverage summary
    n = len(bridge)
    cov = {
        "collection_articles_n": n,
        "in_stock_n": sum(1 for r in bridge if r["in_stock_cost_move"] == "Y"),
        "in_sku_n": sum(1 for r in bridge if r["in_sku_master"] == "Y"),
        "in_alias_n": sum(1 for r in bridge if r["in_alias_master"] == "Y"),
        "full_n": sum(1 for r in bridge if r["link_coverage"] == "FULL"),
        "none_n": sum(1 for r in bridge if r["link_coverage"] == "NONE"),
        "high_gap_n": sum(1 for r in bridge if r["priority"] == "HIGH"),
    }

    # fabric ABC summary
    fab_sum = {
        "lines_n": len(fabric_abc),
        "total_rub": round(sum(float(r["abc_value_rub"]) for r in fabric_abc), 2),
        "A_n": sum(1 for r in fabric_abc if r["abc_band"] == "A"),
        "B_n": sum(1 for r in fabric_abc if r["abc_band"] == "B"),
        "C_n": sum(1 for r in fabric_abc if r["abc_band"] == "C"),
        "A_rub": round(sum(float(r["abc_value_rub"]) for r in fabric_abc if r["abc_band"] == "A"), 2),
    }

    # top HIGH gaps for owner
    high_gaps = [r for r in bridge if r["priority"] == "HIGH"][:40]

    # family coverage — explains why COLLECTION_STYLE has ~0 stock match
    fam_rows = []
    for fam in ("GOODS_STYLE", "COLLECTION_STYLE", "OTHER"):
        sub = [r for r in bridge if article_family(r["article_norm"]) == fam]
        fam_rows.append(
            {
                "article_family": fam,
                "articles_n": len(sub),
                "sale_eur": round(sum(float(r["collection_sale_eur"]) for r in sub), 2),
                "in_stock_n": sum(1 for r in sub if r["in_stock_cost_move"] == "Y"),
                "in_sku_n": sum(1 for r in sub if r["in_sku_master"] == "Y"),
                "none_n": sum(1 for r in sub if r["link_coverage"] == "NONE"),
                "note": (
                    "COL43-47 codes are MD-showroom; goods stock uses 0-xxxx — use person-cost/MD links"
                    if fam == "COLLECTION_STYLE"
                    else ""
                ),
            }
        )

    meta = {
        "horizon": "H64",
        "date": str(date.today()),
        "title": "SKU/stock/fabric bridge + fabric ABC",
        "sku_rows": len(sku),
        "sku_with_article": sum(1 for r in sku if r["article_norm"]),
        "stock_lines": len(stock),
        "stock_articles": len(stock_by),
        "fabric_lines": len(fabrics),
        "bridge": cov,
        "fabric_abc": fab_sum,
        "article_families": fam_rows,
        "no_fake_accept": True,
        "so_t": False,
    }
    (REG / "meta.json").write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")

    write_csv(REG / "sku_master_normalized.csv", list(sku[0].keys()), sku)
    write_csv(REG / "stock_cost_articles.csv", list(stock[0].keys()) if stock else ["article"], stock)
    write_csv(REG / "fabric_inventory_abc.csv", list(fabric_abc[0].keys()), fabric_abc)
    write_csv(REG / "collection_sku_stock_bridge.csv", list(bridge[0].keys()), bridge)
    write_csv(REG / "collection_article_gaps_top40.csv", list(high_gaps[0].keys()) if high_gaps else list(bridge[0].keys()), high_gaps)
    write_csv(REG / "collection_article_family_coverage.csv", list(fam_rows[0].keys()), fam_rows)

    for name in [
        "sku_master_normalized.csv",
        "stock_cost_articles.csv",
        "fabric_inventory_abc.csv",
        "collection_sku_stock_bridge.csv",
        "collection_article_gaps_top40.csv",
        "collection_article_family_coverage.csv",
        "meta.json",
    ]:
        src = REG / name
        if name == "meta.json":
            shutil.copy2(src, MARTS / "h64_meta.json")
            shutil.copy2(src, MAPS / "h64_meta.json")
            shutil.copy2(src, EV / "meta.json")
        else:
            shutil.copy2(src, MARTS / name)
            shutil.copy2(src, MAPS / name)
            shutil.copy2(src, EV / name)

    shutil.copy2(REG / "collection_sku_stock_bridge.csv", WAVE_B / "15_collection_sku_stock_bridge.csv")
    shutil.copy2(REG / "collection_article_gaps_top40.csv", WAVE_B / "16_collection_article_gaps_top40.csv")
    shutil.copy2(REG / "fabric_inventory_abc.csv", WAVE_B / "17_fabric_inventory_abc.csv")
    shutil.copy2(REG / "collection_article_family_coverage.csv", WAVE_B / "18_collection_article_family_coverage.csv")

    print(json.dumps(meta, ensure_ascii=False, indent=2))
    print("HIGH gaps sample:")
    for r in high_gaps[:8]:
        print(r["article_norm"], r["collection_sale_eur"], r["collection_types"], r["link_coverage"])


if __name__ == "__main__":
    main()
