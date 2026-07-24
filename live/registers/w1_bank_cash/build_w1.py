#!/usr/bin/env python3
"""
W1: наполнение REG-LEGAL / REG-BANK / REG-CASH из локальных исходников.

Зачем:
- денежный скелет архитектуры (разблокирует O2C/P2P/H2P cash);
- стабильные provisional ID (bank_payment_id, cash_line_id) без выдуманных SoT;
- evidence сверки банк↔ДДС по месяцам.

Ограничение W1:
- BANK пока только из Alfa xlsx (3 файла). VTB PDF — отдельный парсер (отмечен GAP).
- CASH из листов data B / data D (зерно строки ДДС), не из сводных pivot.
"""
from __future__ import annotations

import csv
import hashlib
import json
import re
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[3]
DOCS = ROOT / "documents"
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/w1_bank_cash_20260723"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")


def sha16(*parts) -> str:
    raw = "|".join("" if p is None else str(p) for p in parts)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]


def parse_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


def to_float(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def month_key(d: date | None) -> str | None:
    return d.strftime("%Y-%m") if d else None


def load_catalog():
    path = ROOT / "live/registers/00_SOURCE_CATALOG_93.csv"
    rows = list(csv.DictReader(open(path, encoding="utf-8-sig")))
    return {r["file_name"]: r for r in rows}


def write_csv(path: Path, rows: list[dict], fieldnames: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)


# ── LEGAL ────────────────────────────────────────────────────────────
def build_legal() -> list[dict]:
    # Из выписок Alfa (ИП) + налоговые формы (ДЕКОР). INN ДЕКОР — из каталога/типичных форм.
    return [
        {
            "legal_entity_id": "LE-IP-YANINA",
            "inn": "770701688220",
            "name": "ИП Янина Юлия Федоровна",
            "entity_type": "IP",
            "valid_from": "2024-01-01",
            "valid_to": "",
            "source_evidence": "Выписка Alfa 40802810202620002686 header",
            "notes": "подтверждено в bank xlsx",
        },
        {
            "legal_entity_id": "LE-OOO-DEKOR",
            "inn": "",  # заполнить из налоговой PDF при парсинге
            "name": "ООО ДЕКОР",
            "entity_type": "OOO",
            "valid_from": "2024-01-01",
            "valid_to": "",
            "source_evidence": "tax forms FILE-001.. naming",
            "notes": "INN TBD из 6-НДФЛ/РСВ PDF — GAP-W1-LEGAL-01",
        },
    ]


# ── BANK accounts + payments ─────────────────────────────────────────
BANK_XLSX = [
    (
        "Выписка_40802810202620002686_01.01.2024–31.12.2024.xlsx",
        "40802810202620002686",
        "FILE-013",
    ),
    (
        "Выписка_40802810202620002686_01.01.2025–31.12.2025.xlsx",
        "40802810202620002686",
        "FILE-014",
    ),
    (
        "Выписка_40802810202620002686_01.01.2026–23.06.2026.xlsx",
        "40802810202620002686",
        "FILE-015",
    ),
]

VTB_PDF_GAP = [
    ("FILE-008", "40702810804000008482", "VTB PDF — парсер не в W1"),
    ("FILE-009", "40802810404000000049", "VTB PDF — парсер не в W1"),
    ("FILE-010", "40802810500000006196", "VTB PDF — парсер не в W1"),
]


def parse_bank_xlsx(path: Path, account_id: str, source_file_id: str) -> tuple[dict, list[dict]]:
    # Alfa export ломает read_only (отдаёт 1 строку) — только обычный режим.
    wb = load_workbook(path, read_only=False, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    meta = {
        "bank_account_id": account_id,
        "account_raw": None,
        "owner_name": None,
        "owner_inn": None,
        "bik": None,
        "currency": "RUB",
        "bank_name": "АО Альфа-Банк (xlsx выписка)",
        "legal_entity_id": "LE-IP-YANINA",
        "source_file_id": source_file_id,
        "period_from": None,
        "period_to": None,
        "incoming_balance": None,
        "outgoing_balance": None,
        "turnover_debit": None,
        "turnover_credit": None,
    }

    # header scrape (rows may be short / ragged)
    for r in rows[:12]:
        if not r:
            continue
        a0 = str(r[0]).strip() if r[0] else ""
        a1 = r[1] if len(r) > 1 else None
        if a0.startswith("Выписка"):
            meta["account_raw"] = str(a1).replace(" ", "") if a1 else account_id
        elif a0.startswith("За период"):
            m = re.search(r"(\d{2}\.\d{2}\.\d{4}).*?(\d{2}\.\d{2}\.\d{4})", str(a1) if a1 else "")
            if m:
                meta["period_from"] = parse_date(m.group(1)).isoformat() if parse_date(m.group(1)) else None
                meta["period_to"] = parse_date(m.group(2)).isoformat() if parse_date(m.group(2)) else None
        elif a0.startswith("Владелец"):
            meta["owner_name"] = str(a1).strip() if a1 else None
        elif a0.startswith("ИНН"):
            meta["owner_inn"] = str(a1).strip() if a1 else None
        elif a0 == "БИК":
            meta["bik"] = str(a1).strip() if a1 else None
        elif a0.startswith("Остаток входящий"):
            meta["incoming_balance"] = to_float(a1)
            meta["turnover_debit"] = to_float(r[4]) if len(r) > 4 else None
        elif a0.startswith("Остаток исходящий"):
            meta["outgoing_balance"] = to_float(a1)
            meta["turnover_credit"] = to_float(r[4]) if len(r) > 4 else None

    # find header row "Дата"
    start = None
    for i, r in enumerate(rows):
        if not r or r[0] is None:
            continue
        if str(r[0]).strip() == "Дата" and len(r) > 2 and str(r[2] or "").strip() == "Дебет":
            # next row is subheader (Наименование/ИНН...); data starts after
            start = i + 2
            break
        if str(r[0]).strip() == "Дата" and len(r) > 1 and "документ" in str(r[1] or "").lower():
            start = i + 2
            break
    payments = []
    if start is None:
        return meta, payments

    def cell(row, i):
        return row[i] if len(row) > i else None

    for idx, r in enumerate(rows[start:], start=start + 1):
        if not r or r[0] is None:
            continue
        d = parse_date(r[0])
        if d is None:
            continue
        doc_no = str(cell(r, 1)).strip() if cell(r, 1) is not None else ""
        debit = to_float(cell(r, 2))
        credit = to_float(cell(r, 3))
        cp_name = str(cell(r, 4)).strip() if cell(r, 4) else ""
        cp_inn = str(cell(r, 5)).strip() if cell(r, 5) else ""
        purpose = str(cell(r, 10)).strip() if cell(r, 10) else ""
        doc_type = str(cell(r, 12)).strip() if cell(r, 12) else ""
        amount = credit if credit else (debit if debit else 0.0)
        direction = "in" if credit else "out"
        bank_payment_id = sha16(account_id, d.isoformat(), doc_no, debit, credit, purpose[:40])
        payments.append(
            {
                "bank_payment_id": bank_payment_id,
                "bank_account_id": account_id,
                "legal_entity_id": "LE-IP-YANINA",
                "payment_date": d.isoformat(),
                "period_month": month_key(d),
                "doc_no": doc_no,
                "direction": direction,
                "amount": amount,
                "debit": debit if debit is not None else "",
                "credit": credit if credit is not None else "",
                "currency": "RUB",
                "counterparty_raw": cp_name,
                "counterparty_inn": cp_inn,
                "counterparty_id": "",  # W5
                "purpose": purpose,
                "doc_type": doc_type,
                "source_file_id": source_file_id,
                "source_row_id": f"r{idx}",
                "match_status": "UNMATCHED",
                "cash_line_id": "",
            }
        )
    return meta, payments


def build_bank(catalog: dict) -> tuple[list[dict], list[dict], list[dict]]:
    accounts = []
    payments = []
    gaps = []

    # one logical account (same number across years)
    accounts.append(
        {
            "bank_account_id": "40802810202620002686",
            "legal_entity_id": "LE-IP-YANINA",
            "bank_name": "АО Альфа-Банк",
            "currency": "RUB",
            "status": "ACTIVE_PARSED",
            "notes": "parsed from 3 yearly xlsx",
        }
    )
    for fid, acc, note in VTB_PDF_GAP:
        accounts.append(
            {
                "bank_account_id": acc,
                "legal_entity_id": "LE-IP-YANINA" if acc.startswith("408") else "LE-OOO-DEKOR",
                "bank_name": "Банк ВТБ (PDF)",
                "currency": "RUB",
                "status": "GAP_PDF_NOT_PARSED",
                "notes": f"{fid}: {note}",
            }
        )
        gaps.append({"gap_id": f"GAP-W1-BANK-{acc[-4:]}", "file_id": fid, "account": acc, "issue": note})

    for fname, acc, fallback_fid in BANK_XLSX:
        path = DOCS / fname
        meta_cat = catalog.get(fname, {})
        source_file_id = meta_cat.get("master_file_id") or fallback_fid
        if not path.exists():
            gaps.append({"gap_id": "GAP-W1-BANK-FILE", "file_id": source_file_id, "account": acc, "issue": f"missing {fname}"})
            continue
        meta, pays = parse_bank_xlsx(path, acc, source_file_id)
        # store period meta as gap-free account period coverage
        accounts[0][f"cov_{source_file_id}"] = f"{meta.get('period_from')}..{meta.get('period_to')}"
        payments.extend(pays)

    return accounts, payments, gaps


# ── CASH from DDS ────────────────────────────────────────────────────
DDS_FILES = [
    ("ДДС 2024.xlsx", "FILE-017", 2024),
    ("ДДС 2025.xlsx", "FILE-018", 2025),
    ("ДДС 2026.xlxs.xlsx", "FILE-019", 2026),
]


def parse_dds_sheet(ws, sheet_name: str, source_file_id: str, ledger: str) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    # normalize keys
    idx = {h: i for i, h in enumerate(header)}

    def col(*names):
        for n in names:
            if n in idx:
                return idx[n]
        return None

    i_month = col("месяц")
    i_dir = col("Направление деятельности")
    i_art = col("Статья расхода")
    i_comment = col("комментарии")
    i_nal = col("Нал/Бнал")
    i_rub = col("сумма руб.")
    i_eur = col("в евро")
    i_rate = col("курс евро")
    i_coll = col("Коллекция")
    i_owner = col("Ответственный")
    out = []
    for rnum, r in enumerate(rows[1:], start=2):
        if not r or (i_month is not None and r[i_month] is None and (i_art is None or r[i_art] is None)):
            continue
        d = parse_date(r[i_month]) if i_month is not None else None
        art = str(r[i_art]).strip() if i_art is not None and r[i_art] else ""
        if not art and d is None:
            continue
        rub = to_float(r[i_rub]) if i_rub is not None else None
        eur = to_float(r[i_eur]) if i_eur is not None else None
        nal = str(r[i_nal]).strip() if i_nal is not None and r[i_nal] else ""
        # direction: DDS rows are mostly outflows (expenses)
        direction = "out"
        cash_line_id = sha16(source_file_id, sheet_name, rnum, month_key(d), art, rub, eur, nal)
        out.append(
            {
                "cash_line_id": cash_line_id,
                "period_month": month_key(d) or "",
                "ledger": ledger,  # B=business, D=owners
                "article_id": sha16("ART", art.lower()) if art else "",
                "article_name": art,
                "direction": direction,
                "amount_rub": rub if rub is not None else "",
                "amount_eur": eur if eur is not None else "",
                "currency_primary": "EUR",
                "fx_rate": to_float(r[i_rate]) if i_rate is not None else "",
                "cash_type": nal,
                "direction_activity": str(r[i_dir]).strip() if i_dir is not None and r[i_dir] else "",
                "collection": str(r[i_coll]).strip() if i_coll is not None and r[i_coll] else "",
                "responsible": str(r[i_owner]).strip() if i_owner is not None and r[i_owner] else "",
                "comment": str(r[i_comment]).strip() if i_comment is not None and r[i_comment] else "",
                "legal_entity_id": "LE-IP-YANINA",  # operational DDS is IP perimeter (уточнить)
                "bank_payment_id": "",
                "payroll_batch_id": "",
                "sales_line_id": "",
                "source_file_id": source_file_id,
                "source_sheet": sheet_name,
                "source_row_id": f"r{rnum}",
                "match_status": "UNMATCHED",
            }
        )
    return out


def build_cash(catalog: dict) -> tuple[list[dict], list[dict]]:
    lines = []
    gaps = []
    for fname, fallback_fid, year in DDS_FILES:
        path = DOCS / fname
        meta = catalog.get(fname, {})
        fid = meta.get("master_file_id") or fallback_fid
        if not path.exists():
            gaps.append({"gap_id": f"GAP-W1-DDS-{year}", "file_id": fid, "issue": f"missing {fname}"})
            continue
        wb = load_workbook(path, read_only=True, data_only=True)
        for sheet, ledger in (("data B", "B"), ("data D", "D")):
            if sheet not in wb.sheetnames:
                gaps.append({"gap_id": f"GAP-W1-DDS-SHEET-{year}-{ledger}", "file_id": fid, "issue": f"no sheet {sheet}"})
                continue
            lines.extend(parse_dds_sheet(wb[sheet], sheet, fid, ledger))
        wb.close()
    return lines, gaps


# ── Reconciliation bank out vs DDS Б/Нал by month ────────────────────
def reconcile(payments: list[dict], cash: list[dict]) -> list[dict]:
    bank_out = defaultdict(float)
    bank_in = defaultdict(float)
    for p in payments:
        m = p["period_month"]
        if not m:
            continue
        if p["direction"] == "out":
            bank_out[m] += float(p["amount"] or 0)
        else:
            bank_in[m] += float(p["amount"] or 0)

    # DDS: only non-cash bank-like types for comparison with bank statement
    dds_bn = defaultdict(float)  # Б/Нал rub
    dds_all = defaultdict(float)
    dds_payroll = defaultdict(float)
    for c in cash:
        m = c["period_month"]
        if not m or c["ledger"] != "B":
            continue
        rub = float(c["amount_rub"]) if c["amount_rub"] not in ("", None) else 0.0
        dds_all[m] += rub
        ct = (c["cash_type"] or "").lower()
        if "б/нал" in ct or "бнал" in ct.replace(" ", "") or ct in ("б/нал", "бнал"):
            dds_bn[m] += rub
        if "оплата труда" in (c["article_name"] or "").lower():
            dds_payroll[m] += rub

    months = sorted(set(bank_out) | set(bank_in) | set(dds_all))
    rows = []
    for m in months:
        bo = bank_out[m]
        bi = bank_in[m]
        db = dds_bn[m]
        da = dds_all[m]
        # compare bank outflows vs DDS безнал (same currency RUB)
        delta = bo - db if (bo or db) else None
        # tolerance: 1% or 1000 RUB
        status = "N/A"
        if bo and db:
            tol = max(1000.0, 0.02 * max(bo, db))
            if abs(bo - db) <= tol:
                status = "CLOSE"
            elif abs(bo - db) / max(bo, db) <= 0.10:
                status = "SOFT_GAP"
            else:
                status = "GAP"
        elif bo and not db:
            status = "BANK_ONLY"
        elif db and not bo:
            status = "DDS_ONLY"
        rows.append(
            {
                "period_month": m,
                "bank_out_rub": round(bo, 2),
                "bank_in_rub": round(bi, 2),
                "dds_b_bn_rub": round(db, 2),
                "dds_b_all_rub": round(da, 2),
                "dds_b_payroll_rub": round(dds_payroll[m], 2),
                "delta_bank_out_vs_dds_bn": round(delta, 2) if delta is not None else "",
                "status": status,
                "note": "Alfa xlsx only; VTB PDF not in bank hub yet; DDS includes cash/corp that may not hit this account",
            }
        )
    return rows


def soft_match_sample(payments: list[dict], cash: list[dict], month: str, limit: int = 30) -> list[dict]:
    """Пробуем сопоставить строки за месяц: сумма + дата±0 (DDS имеет только месяц)."""
    bank = [p for p in payments if p["period_month"] == month and p["direction"] == "out"]
    dds = [
        c
        for c in cash
        if c["period_month"] == month
        and c["ledger"] == "B"
        and c["amount_rub"] not in ("", None)
        and "б/нал" in (c["cash_type"] or "").lower().replace(" ", "")
    ]
    # amount multiset match
    by_amt = defaultdict(list)
    for c in dds:
        by_amt[round(float(c["amount_rub"]), 2)].append(c)

    matches = []
    used = set()
    for p in bank:
        amt = round(float(p["amount"] or 0), 2)
        cands = [c for c in by_amt.get(amt, []) if c["cash_line_id"] not in used]
        if not cands:
            continue
        c = cands[0]
        used.add(c["cash_line_id"])
        matches.append(
            {
                "period_month": month,
                "bank_payment_id": p["bank_payment_id"],
                "cash_line_id": c["cash_line_id"],
                "amount_rub": amt,
                "payment_date": p["payment_date"],
                "counterparty_raw": p["counterparty_raw"],
                "dds_article": c["article_name"],
                "dds_comment": c["comment"][:80],
                "match_method": "exact_amount_same_month",
                "confidence": "LOW",  # без даты в DDS — только сумма
            }
        )
        if len(matches) >= limit:
            break
    return matches


def build_evidence_xlsx(recon: list[dict], matches: list[dict], summary: dict, gaps: list[dict]):
    EV.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)

    def sheet(name, headers, rows):
        ws = wb.create_sheet(name)
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.fill = fill
            cell.font = font
        for i, row in enumerate(rows, 2):
            for j, h in enumerate(headers, 1):
                ws.cell(i, j, row.get(h, ""))
                ws.cell(i, j).alignment = Alignment(wrap_text=True, vertical="top")
        for i, _ in enumerate(headers, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = 18
        return ws

    ws0 = wb.active
    ws0.title = "00_Summary"
    for i, (k, v) in enumerate(summary.items(), 1):
        ws0.cell(i, 1, k)
        ws0.cell(i, 2, json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v)

    sheet("01_Month_Reconcile", list(recon[0].keys()) if recon else ["period_month"], recon)
    sheet(
        "02_Soft_Matches_Sample",
        list(matches[0].keys()) if matches else ["period_month"],
        matches,
    )
    sheet("03_Gaps", list(gaps[0].keys()) if gaps else ["gap_id"], gaps)

    path = EV / "YANINA_W1_BANK_CASH_EVIDENCE.xlsx"
    wb.save(path)
    return path


def main():
    catalog = load_catalog()
    legal = build_legal()
    accounts, payments, bank_gaps = build_bank(catalog)
    cash, cash_gaps = build_cash(catalog)
    gaps = bank_gaps + cash_gaps

    recon = reconcile(payments, cash)
    # pick pilot month: prefer CLOSE/SOFT, else latest with both sides
    pilot = None
    for pref in ("CLOSE", "SOFT_GAP", "GAP"):
        cands = [r for r in recon if r["status"] == pref]
        if cands:
            pilot = sorted(cands, key=lambda x: x["period_month"], reverse=True)[0]["period_month"]
            break
    if not pilot:
        both = [r for r in recon if r["bank_out_rub"] and r["dds_b_bn_rub"]]
        pilot = both[-1]["period_month"] if both else (recon[-1]["period_month"] if recon else "2026-04")

    matches = soft_match_sample(payments, cash, pilot, limit=50)

    # write registers
    write_csv(
        OUT / "legal.csv",
        legal,
        ["legal_entity_id", "inn", "name", "entity_type", "valid_from", "valid_to", "source_evidence", "notes"],
    )
    # flatten account cov_* into notes only for stable schema
    acc_out = []
    for a in accounts:
        notes = a.get("notes", "")
        extra = {k: v for k, v in a.items() if k.startswith("cov_")}
        if extra:
            notes = (notes + " | " + json.dumps(extra, ensure_ascii=False)).strip(" |")
        acc_out.append(
            {
                "bank_account_id": a["bank_account_id"],
                "legal_entity_id": a["legal_entity_id"],
                "bank_name": a["bank_name"],
                "currency": a["currency"],
                "status": a["status"],
                "notes": notes,
            }
        )
    write_csv(
        OUT / "bank_accounts.csv",
        acc_out,
        ["bank_account_id", "legal_entity_id", "bank_name", "currency", "status", "notes"],
    )
    bank_fields = [
        "bank_payment_id",
        "bank_account_id",
        "legal_entity_id",
        "payment_date",
        "period_month",
        "doc_no",
        "direction",
        "amount",
        "debit",
        "credit",
        "currency",
        "counterparty_raw",
        "counterparty_inn",
        "counterparty_id",
        "purpose",
        "doc_type",
        "source_file_id",
        "source_row_id",
        "match_status",
        "cash_line_id",
    ]
    write_csv(OUT / "bank_payments.csv", payments, bank_fields)

    cash_fields = [
        "cash_line_id",
        "period_month",
        "ledger",
        "article_id",
        "article_name",
        "direction",
        "amount_rub",
        "amount_eur",
        "currency_primary",
        "fx_rate",
        "cash_type",
        "direction_activity",
        "collection",
        "responsible",
        "comment",
        "legal_entity_id",
        "bank_payment_id",
        "payroll_batch_id",
        "sales_line_id",
        "source_file_id",
        "source_sheet",
        "source_row_id",
        "match_status",
    ]
    write_csv(OUT / "cash_lines.csv", cash, cash_fields)
    write_csv(OUT / "recon_bank_vs_dds_month.csv", recon, list(recon[0].keys()) if recon else ["period_month"])
    write_csv(OUT / "soft_matches_pilot.csv", matches, list(matches[0].keys()) if matches else ["period_month"])
    # normalize gap schema
    gap_rows = []
    for g in gaps:
        gap_rows.append(
            {
                "gap_id": g.get("gap_id", ""),
                "file_id": g.get("file_id", ""),
                "account": g.get("account", ""),
                "issue": g.get("issue", ""),
            }
        )
    write_csv(OUT / "gaps.csv", gap_rows, ["gap_id", "file_id", "account", "issue"])

    # status counts
    status_counts = Counter(r["status"] for r in recon)
    summary = {
        "generated_at": NOW,
        "wave": "W1",
        "legal_entities": len(legal),
        "bank_accounts": len(acc_out),
        "bank_payments_parsed": len(payments),
        "cash_lines_parsed": len(cash),
        "cash_B": sum(1 for c in cash if c["ledger"] == "B"),
        "cash_D": sum(1 for c in cash if c["ledger"] == "D"),
        "recon_months": len(recon),
        "recon_status_counts": dict(status_counts),
        "pilot_month": pilot,
        "pilot_soft_matches": len(matches),
        "gaps": len(gaps),
        "principle": "provisional IDs; not SoT until RACI + retests",
        "next": "parse VTB PDF OR soft-match payroll articles; W2 PAYROLL",
    }
    json.dump(summary, open(OUT / "w1_summary.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    EV.mkdir(parents=True, exist_ok=True)
    json.dump(summary, open(EV / "w1_summary.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)

    ev_path = build_evidence_xlsx(recon, matches, summary, gaps)

    # replace stubs with pointers
    write_csv(OUT / "bank_stub.csv", payments[:0], bank_fields)  # keep schema header-only alias
    # actually overwrite stubs as README pointers
    (OUT / "README.md").write_text(
        f"""# W1 BANK / CASH / LEGAL

Generated: {NOW}

## Файлы регистров
- `legal.csv` — 2 юрлица (ИП подтверждён ИНН из выписки; ООО ДЕКОР — INN TBD)
- `bank_accounts.csv` — Alfa parsed + 3 VTB PDF gap
- `bank_payments.csv` — **{len(payments)}** платежей из 3 Alfa xlsx
- `cash_lines.csv` — **{len(cash)}** строк ДДС (data B + data D)
- `recon_bank_vs_dds_month.csv` — помесячная сверка
- `soft_matches_pilot.csv` — пилот {pilot}: exact amount same month
- `gaps.csv` — VTB PDF и пр.

## Evidence
`../../evidence/w1_bank_cash_20260723/YANINA_W1_BANK_CASH_EVIDENCE.xlsx`

## Важно
Это Controlled Staging, не Source of Truth.
Match по сумме+месяцу = LOW confidence (в ДДС нет payment_date).
""",
        encoding="utf-8",
    )

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    print("evidence:", ev_path)


if __name__ == "__main__":
    main()
