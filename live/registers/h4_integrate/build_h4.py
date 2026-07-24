#!/usr/bin/env python3
"""
H4: интеграция новых продаж/SKU/тканей в живые регистры W3/W4/W5.

1) W4: sales_extended (H3) → sales_lines + settlements + cost link + settle↔bank
2) W3: покрытие SKU master (H3) ↔ cost ↔ продажи
3) W5: движения тканей (FILE-097) → material_movements staging

Не SoT. RACI ACCEPT по-прежнему нужен.
"""
from __future__ import annotations

import csv
import hashlib
import json
import re
import shutil
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[3]
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/h4_integrate_20260724"
H3 = ROOT / "live/registers/h3_new_docs"
W1 = ROOT / "live/registers/w1_bank_cash"
W3 = ROOT / "live/registers/w3_sku_cost"
W4 = ROOT / "live/registers/w4_sales_settle"
W5 = ROOT / "live/registers/w5_sup_exp_mat"
DOCS = ROOT / "documents"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")


def nfc(s: str) -> str:
    return unicodedata.normalize("NFC", s or "")


def sha16(*parts) -> str:
    raw = "|".join("" if p is None else str(p) for p in parts)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def resolve(name: str) -> Path | None:
    t = nfc(name)
    for p in DOCS.iterdir():
        if nfc(p.name) == t:
            return p
    return None


def normalize_article(raw) -> str:
    if raw is None:
        return ""
    s = nfc(str(raw)).strip().upper().replace("Ё", "Е")
    s = s.replace("Т-", "T-")
    for a, b in zip("АВЕКМНОРСТХ", "ABEKMHOPCTX"):
        s = s.replace(a, b)
    s = s.replace(" ", "").replace("\xa0", "").replace("–", "-").replace("—", "-")
    return s


def canonical_sku(raw) -> str:
    s = normalize_article(raw)
    # drop trailing /size like /3 /42
    s = re.sub(r"/\d+[A-Z]?$", "", s)
    return s


# ── W4 rebuild from H3 sales ──────────────────────────────────────
def rebuild_w4() -> dict:
    ext = list(csv.DictReader(open(H3 / "sales_extended_2024_2026.csv", encoding="utf-8")))
    # backup once
    for fn in ("sales_lines.csv", "sales_lines_1c.csv", "settlements.csv"):
        src = W4 / fn
        bak = W4 / f"{fn.replace('.csv', '')}_pre_h4.csv"
        if src.exists() and not bak.exists():
            shutil.copy2(src, bak)

    # cost index
    cost_by_can = {}
    if (W3 / "cost_versions.csv").exists():
        for c in csv.DictReader(open(W3 / "cost_versions.csv", encoding="utf-8")):
            can = c.get("canonical_sku") or canonical_sku(c.get("article_raw"))
            if not can:
                continue
            # prefer completeness FULL / latest
            prev = cost_by_can.get(can)
            if prev is None or (c.get("completeness") == "FULL" and prev.get("completeness") != "FULL"):
                cost_by_can[can] = c
            elif prev and c.get("unit_cost_rub") and not prev.get("unit_cost_rub"):
                cost_by_can[can] = c

    sales = []
    settlements = {}
    for r in ext:
        art = r.get("article") or ""
        can = canonical_sku(art)
        doc = r.get("document") or ""
        channel = r.get("channel") or ""
        st_id = "ST-" + sha16(channel, doc) if doc else ""
        cost = cost_by_can.get(can)
        unit_cost = ""
        cvid = ""
        if cost:
            unit_cost = cost.get("unit_cost_rub") or ""
            cvid = cost.get("cost_version_id") or ""
        # margin from file cogs if present else w3
        rev = float(r["revenue_rub"] or 0)
        cogs_file = float(r["cogs_rub"] or 0) if r.get("cogs_rub") not in (None, "") else None
        qty = float(r["qty"] or 0) if r.get("qty") not in (None, "") else 0
        cogs_w3 = float(unit_cost) * qty if unit_cost not in (None, "") and qty else None
        cogs = cogs_file if cogs_file is not None else cogs_w3
        margin = round(rev - cogs, 2) if cogs is not None else ""

        line = {
            "sales_line_id": r["sales_line_id"],
            "settlement_id": st_id,
            "channel": channel,
            "period_month": r.get("period_month") or "",
            "sale_date": r.get("sale_date") or "",
            "article_raw": art,
            "canonical_sku": can,
            "code_1c": r.get("code_1c") or "",
            "sku_name": r.get("sku_name") or "",
            "buyer": r.get("buyer") or "",
            "document": doc,
            "qty": r.get("qty") or "",
            "revenue_rub": r.get("revenue_rub") or "",
            "cogs_rub": round(cogs, 2) if cogs is not None else "",
            "cogs_source": "FILE" if cogs_file is not None else ("W3" if cogs_w3 is not None else ""),
            "margin_rub": margin,
            "w3_unit_cost": unit_cost,
            "w3_cost_version_id": cvid,
            "source_file_id": r.get("source_file_id") or "",
            "source_row_id": r.get("source_row_id") or "",
        }
        sales.append(line)

        if st_id:
            st = settlements.setdefault(
                st_id,
                {
                    "settlement_id": st_id,
                    "channel": channel,
                    "document": doc,
                    "sale_date": r.get("sale_date") or "",
                    "period_month": r.get("period_month") or "",
                    "buyer": r.get("buyer") or "",
                    "revenue_rub": 0.0,
                    "cogs_rub": 0.0,
                    "n_lines": 0,
                    "source_file_id": r.get("source_file_id") or "",
                    "bank_payment_id": "",
                    "status": "OPEN",
                },
            )
            st["revenue_rub"] += rev
            if cogs is not None:
                st["cogs_rub"] += cogs
            st["n_lines"] += 1
            if not st["buyer"] and r.get("buyer"):
                st["buyer"] = r["buyer"]
            if not st["sale_date"] and r.get("sale_date"):
                st["sale_date"] = r["sale_date"]
                st["period_month"] = r.get("period_month") or ""

    settle_rows = []
    for st in settlements.values():
        st["revenue_rub"] = round(st["revenue_rub"], 2)
        st["cogs_rub"] = round(st["cogs_rub"], 2)
        settle_rows.append(st)

    # soft match settle↔bank IN exact amount same month
    bank_in = defaultdict(list)
    for p in csv.DictReader(open(W1 / "bank_payments.csv", encoding="utf-8")):
        if p.get("direction") == "in" and p.get("is_internal") != "Y":
            bank_in[p["period_month"]].append(p)

    matches = []
    used = set()
    for st in settle_rows:
        rev = float(st["revenue_rub"] or 0)
        if rev <= 0:
            continue
        for p in bank_in.get(st["period_month"], []):
            if p["bank_payment_id"] in used:
                continue
            amt = float(p["amount"] or 0)
            if abs(amt - rev) <= 0.01:
                used.add(p["bank_payment_id"])
                st["bank_payment_id"] = p["bank_payment_id"]
                st["status"] = "LINKED_H4_EXACT"
                matches.append(
                    {
                        "period_month": st["period_month"],
                        "settlement_id": st["settlement_id"],
                        "document": st["document"],
                        "buyer": st["buyer"],
                        "revenue_rub": st["revenue_rub"],
                        "bank_payment_id": p["bank_payment_id"],
                        "payment_date": p["payment_date"],
                        "counterparty_raw": p.get("counterparty_raw", ""),
                        "match_method": "exact_amount_same_month",
                        "confidence": "LOW",
                    }
                )
                break

    # also re-apply H1 HIGH/MED if settlement_id matches
    h1p = H3.parent / "h1_spine_links" / "settle_bank_by_doc.csv"
    # path is live/registers/h1_spine_links
    h1 = ROOT / "live/registers/h1_spine_links/settle_bank_by_doc.csv"
    h1_applied = 0
    if h1.exists():
        by_doc = {}
        for m in csv.DictReader(open(h1, encoding="utf-8")):
            if m.get("confidence") in ("HIGH", "MED"):
                by_doc[m["document"]] = m
        for st in settle_rows:
            if st.get("bank_payment_id"):
                continue
            m = by_doc.get(st["document"])
            if not m:
                continue
            st["bank_payment_id"] = m["bank_payment_id"]
            st["status"] = f"LINKED_H1_{m['confidence']}"
            h1_applied += 1

    sales_fields = [
        "sales_line_id", "settlement_id", "channel", "period_month", "sale_date",
        "article_raw", "canonical_sku", "code_1c", "sku_name", "buyer", "document",
        "qty", "revenue_rub", "cogs_rub", "cogs_source", "margin_rub",
        "w3_unit_cost", "w3_cost_version_id", "source_file_id", "source_row_id",
    ]
    settle_fields = [
        "settlement_id", "channel", "document", "sale_date", "period_month", "buyer",
        "revenue_rub", "cogs_rub", "n_lines", "source_file_id", "bank_payment_id", "status",
    ]

    write_csv(W4 / "sales_lines.csv", sales, sales_fields)
    write_csv(W4 / "sales_lines_1c.csv", [s for s in sales if s["channel"] in ("B2B", "IM")], sales_fields)
    write_csv(W4 / "settlements.csv", settle_rows, settle_fields)
    write_csv(W4 / "soft_matches_settle_bank.csv", matches, list(matches[0].keys()) if matches else ["period_month"])
    write_csv(OUT / "sales_lines_h4.csv", sales, sales_fields)
    write_csv(OUT / "settlements_h4.csv", settle_rows, settle_fields)

    with_cost = sum(1 for s in sales if s["w3_cost_version_id"])
    with_cogs = sum(1 for s in sales if s["cogs_rub"] != "")
    linked = sum(1 for s in settle_rows if s.get("bank_payment_id"))
    by_ch = Counter(s["channel"] for s in sales)
    rev = sum(float(s["revenue_rub"] or 0) for s in sales)

    w4_summary = {
        "generated_at": NOW,
        "wave": "W4+H4",
        "sales_lines_total": len(sales),
        "sales_by_channel": dict(by_ch),
        "sales_revenue_rub": round(rev, 2),
        "settlements": len(settle_rows),
        "settlements_linked_bank": linked,
        "soft_exact_matches": len(matches),
        "h1_links_reapplied": h1_applied,
        "lines_with_w3_cost": with_cost,
        "lines_with_cogs": with_cogs,
        "source": "H3 sales_extended FILE-100/101/102",
        "finding": (
            f"H4/W4: {len(sales)} sales ({dict(by_ch)}), {len(settle_rows)} settlements, "
            f"{linked} bank-linked, {with_cost} w3-cost, {with_cogs} with cogs."
        ),
    }
    json.dump(w4_summary, open(W4 / "w4_summary.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    (W4 / "README.md").write_text(
        f"# W4 SALES/SETTLE (H4 refresh)\n\n{NOW}\n\n{w4_summary['finding']}\n",
        encoding="utf-8",
    )
    return w4_summary


# ── SKU coverage ──────────────────────────────────────────────────
def sku_coverage() -> tuple[list[dict], dict]:
    master = list(csv.DictReader(open(H3 / "sku_master.csv", encoding="utf-8")))
    cost_cans = set()
    if (W3 / "cost_versions.csv").exists():
        for c in csv.DictReader(open(W3 / "cost_versions.csv", encoding="utf-8")):
            can = c.get("canonical_sku") or canonical_sku(c.get("article_raw"))
            if can:
                cost_cans.add(can)

    sales_cans = set()
    sales_rev = defaultdict(float)
    if (W4 / "sales_lines.csv").exists():
        for s in csv.DictReader(open(W4 / "sales_lines.csv", encoding="utf-8")):
            can = s.get("canonical_sku") or canonical_sku(s.get("article_raw"))
            if can:
                sales_cans.add(can)
                sales_rev[can] += float(s.get("revenue_rub") or 0)

    rows = []
    for m in master:
        can = canonical_sku(m.get("article") or m.get("canonical_sku") or m.get("code_1c"))
        in_cost = "Y" if can in cost_cans else "N"
        in_sales = "Y" if can in sales_cans else "N"
        rows.append(
            {
                "sku_id": m["sku_id"],
                "canonical_sku": can,
                "article": m.get("article") or "",
                "code_1c": m.get("code_1c") or "",
                "name": (m.get("name") or "")[:120],
                "in_cost": in_cost,
                "in_sales": in_sales,
                "in_intersection": "Y" if in_cost == "Y" and in_sales == "Y" else "N",
                "sales_revenue_rub": round(sales_rev.get(can, 0), 2) if can in sales_cans else "",
                "status": m.get("status") or "CANDIDATE_MASTER",
            }
        )

    # write enhanced master into W3 + H4
    fields = list(rows[0].keys()) if rows else ["canonical_sku"]
    write_csv(OUT / "sku_coverage.csv", rows, fields)
    write_csv(W3 / "sku_master_h3.csv", rows, fields)
    # refresh w3 sku_master as candidate (backup)
    bak = W3 / "sku_master_pre_h4.csv"
    if (W3 / "sku_master.csv").exists() and not bak.exists():
        shutil.copy2(W3 / "sku_master.csv", bak)
    write_csv(W3 / "sku_master.csv", rows, fields)

    stats = {
        "master_rows": len(rows),
        "with_article_canonical": sum(1 for r in rows if r["canonical_sku"]),
        "in_cost": sum(1 for r in rows if r["in_cost"] == "Y"),
        "in_sales": sum(1 for r in rows if r["in_sales"] == "Y"),
        "intersection": sum(1 for r in rows if r["in_intersection"] == "Y"),
        "cost_canonical_total": len(cost_cans),
        "sales_canonical_total": len(sales_cans),
    }
    return rows, stats


# ── Fabric movements ──────────────────────────────────────────────
def parse_fabric_movements() -> tuple[list[dict], dict]:
    path = resolve("Движение тканей.xlsx")
    if not path:
        return [], {"error": "missing"}
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    out = []
    warehouse = "Основной склад"
    warehouse_totals = []
    for i, r in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        if not r or r[0] is None:
            continue
        name = str(r[0]).strip()
        if not name:
            continue
        low = name.lower()
        if "склад" in low and "складк" not in low and not any(k in low for k in ("атлас", "шелк", "ткан", "креп", "шифон")):
            warehouse = name
            # amounts often on same row
            if isinstance(r[2], (int, float)) and float(r[2]) > 1000:
                warehouse_totals.append(
                    {
                        "warehouse": warehouse,
                        "in_rub": round(float(r[2]), 2),
                        "out_rub": round(float(r[3]), 2) if isinstance(r[3], (int, float)) else None,
                        "end_rub": round(float(r[4]), 2) if isinstance(r[4], (int, float)) else None,
                    }
                )
            continue
        if name.startswith("Документ") or name in ("Номенклатура",):
            continue
        in_v = r[2] if len(r) > 2 else None
        out_v = r[3] if len(r) > 3 else None
        end_v = r[4] if len(r) > 4 else None
        if not isinstance(in_v, (int, float)) and not isinstance(out_v, (int, float)) and not isinstance(end_v, (int, float)):
            continue
        if any(k in low for k in ("приходная", "расходная", "перемещ", "списан", "оприход", "отчет")):
            continue
        hint = ""
        m = re.search(r"(\d{2}-\d{3,}[A-Za-zА-Яа-я]?)", name)
        if m:
            hint = m.group(1)
        # только номенклатурные строки (есть артикул-хинт или текстильное слово)
        textile = any(k in low for k in ("атлас", "шелк", "креп", "шифон", "бархат", "трикотаж", "ткань", "кружев", "органза", "фатин", "шерсть", "хлопок", "вискоз", "лён", "лен ", "дюшес", "сатин", "габардин", "подклад"))
        if not hint and not textile:
            continue
        mid = "MAT-" + sha16(name, warehouse)
        # значения в файле часто в метрах/усл.ед., не ₽ — помечаем ambiguous
        out.append(
            {
                "material_movement_id": mid,
                "material_id": mid,
                "name": name[:160],
                "article_hint": hint,
                "qty_in": round(float(in_v), 4) if isinstance(in_v, (int, float)) else "",
                "qty_out": round(float(out_v), 4) if isinstance(out_v, (int, float)) else "",
                "qty_end": round(float(end_v), 4) if isinstance(end_v, (int, float)) else "",
                "value_unit": "AMBIGUOUS_QTY_OR_AMOUNT",
                "snapshot_date": "2026-07-24",
                "warehouse": warehouse,
                "movement_type": "FABRIC_MOVEMENT_H4",
                "source_file_id": "FILE-097",
                "source_row_id": f"r{i}",
            }
        )
    wb.close()

    fields = list(out[0].keys()) if out else ["material_movement_id"]
    write_csv(OUT / "fabric_movements.csv", out, fields)
    write_csv(W5 / "material_movements_fabric.csv", out, fields)

    end_sum = round(sum(w["end_rub"] or 0 for w in warehouse_totals), 2)
    stats = {
        "fabric_skus": len(out),
        "with_article_hint": sum(1 for r in out if r["article_hint"]),
        "warehouse_totals_rub": warehouse_totals,
        "warehouse_end_rub_sum": end_sum,
        "note": "Leaf qty columns ambiguous (m vs RUB); trust warehouse_totals_rub for money",
    }
    meta_path = W5 / "material_meta.json"
    meta = json.load(open(meta_path)) if meta_path.exists() else {}
    meta["h4_fabric_movements"] = stats
    meta["h4_generated_at"] = NOW
    json.dump(meta, open(meta_path, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    return out, stats


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)

    w4 = rebuild_w4()
    sku_rows, sku_stats = sku_coverage()
    fabric_rows, fabric_stats = parse_fabric_movements()

    summary = {
        "generated_at": NOW,
        "wave": "H4",
        "w4": w4,
        "sku_coverage": sku_stats,
        "fabric": fabric_stats,
        "finding": (
            f"H4: W4 refresh {w4['sales_lines_total']} sales / {w4['settlements']} settlements "
            f"({w4['settlements_linked_bank']} bank-linked); "
            f"SKU cover cost {sku_stats['in_cost']}/sales {sku_stats['in_sales']}/∩ {sku_stats['intersection']}; "
            f"fabric SKUs {fabric_stats.get('fabric_skus')} warehouses_end~{fabric_stats.get('warehouse_end_rub_sum')}."
        ),
        "next": "RACI ACCEPT; optional stock_cost_movement deep parse",
        "not_sot": True,
    }
    json.dump(summary, open(OUT / "h4_summary.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    json.dump(summary, open(EV / "h4_summary.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)

    wb = Workbook()
    ws = wb.active
    ws.title = "00_Summary"
    for i, (k, v) in enumerate(summary.items(), 1):
        ws.cell(i, 1, k)
        ws.cell(i, 2, json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v)
    fill = PatternFill("solid", fgColor="1F4E79")

    def add(name, rows):
        w = wb.create_sheet(name)
        if not rows:
            return
        headers = list(rows[0].keys())
        for c, h in enumerate(headers, 1):
            cell = w.cell(1, c, h)
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
        for ri, row in enumerate(rows[:4000], 2):
            for ci, h in enumerate(headers, 1):
                v = row.get(h, "")
                if isinstance(v, (dict, list)):
                    v = json.dumps(v, ensure_ascii=False)
                w.cell(ri, ci, v)

    add("01_W4_Summary", [w4])
    add("02_SKU_Coverage", sku_rows[:3000])
    add("03_Fabric", fabric_rows[:3000])
    wb.save(EV / "YANINA_H4_INTEGRATE_EVIDENCE.xlsx")

    (OUT / "README.md").write_text(
        f"""# H4 Integrate new docs into W3/W4/W5

{NOW}

{summary['finding']}

Evidence: `../../evidence/h4_integrate_20260724/`
""",
        encoding="utf-8",
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
