#!/usr/bin/env python3
"""
W4: REG-SALES + REG-SETTLE staging (O2C skeleton).

Источники:
- Продажи B2B / ИМ (1С nested: SKU → покупатель → документ)
- Факт анализ B2B (помесячная выручка Yanina по Mercury)
- SALES «ДДС - доход» (каналы EUR)
- TSUM lines из W3 (уже есть)

Settlement: provisional settlement_id из расходной накладной / чека.
Не SoT: нет bank_payment_id на уровне инвойса.
"""
from __future__ import annotations

import csv
import hashlib
import json
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[3]
DOCS = ROOT / "documents"
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/w4_sales_settle_20260724"
W1 = ROOT / "live/registers/w1_bank_cash"
W3 = ROOT / "live/registers/w3_sku_cost"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")

DOC_RE = re.compile(
    r"(расходн\w*\s+накладн\w*|чек\s*ккм|отчет о розничных|сч[её]т[а-я]*\s*фактур)",
    re.I,
)
DATE_IN_DOC = re.compile(r"от\s+(\d{2}\.\d{2}\.\d{4})")
CODE_RE = re.compile(r"^\d{2}-[0-9A-Za-zА-Яа-я]+$")


def nfc(s: str) -> str:
    return unicodedata.normalize("NFC", s or "")


def sha16(*parts) -> str:
    raw = "|".join("" if p is None else str(p) for p in parts)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]


def to_float(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    if s in ("-", "—"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def normalize_article(raw) -> str:
    if raw is None:
        return ""
    s = nfc(str(raw)).strip().upper().replace("Ё", "Е")
    s = s.replace("Т-", "T-")
    for a, b in zip("АВЕКМНОРСТХ", "ABEKMHOPCTX"):
        s = s.replace(a, b)
    s = s.replace(" ", "").replace("\xa0", "").replace("–", "-").replace("—", "-")
    # keep size suffix after / for alias, but canonical without trailing size if /N
    return s


def canonical_sku(raw) -> str:
    s = normalize_article(raw)
    # drop /size variant for join to cost: T-2420A/3 → T-2420A
    s = re.sub(r"/\d+$", "", s)
    return s


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def resolve(name: str) -> Path | None:
    t = nfc(name)
    for p in DOCS.iterdir():
        if nfc(p.name) == t:
            return p
    return None


def parse_doc_date(doc: str) -> date | None:
    m = DATE_IN_DOC.search(doc or "")
    if not m:
        return None
    try:
        return datetime.strptime(m.group(1), "%d.%m.%Y").date()
    except ValueError:
        return None


def parse_1c_nested_sales(path: Path, channel: str, source_file_id: str):
    """Парсит 1С-выгрузку: номенклатура → покупатель → документ движения."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    sales = []
    settlements = {}  # settlement_id -> agg
    cur_item = None
    cur_buyer = None

    for rnum, r in enumerate(rows[1:], start=2):  # skip header
        if not r or r[0] is None:
            continue
        a0 = str(r[0]).strip()
        art = r[1]
        name = r[2]
        qty = to_float(r[3])
        price = to_float(r[4])
        revenue = to_float(r[5])
        cogs = to_float(r[9]) if len(r) > 9 else None
        gp = to_float(r[11]) if len(r) > 11 else None

        is_item = art is not None and str(art).strip() != "" and (
            CODE_RE.match(a0) or a0.startswith("00-") or bool(re.match(r"^\d{2}-", a0))
        )
        is_doc = bool(DOC_RE.search(a0))

        if is_item:
            cur_item = {
                "nomenclature_code": a0,
                "article_raw": str(art).strip(),
                "canonical_sku": canonical_sku(art),
                "name": str(name).strip() if name else "",
            }
            cur_buyer = None
            continue

        if is_doc:
            if not cur_item:
                continue
            buyer = cur_buyer or ""
            d = parse_doc_date(a0)
            period_month = d.strftime("%Y-%m") if d else ""
            settlement_id = "ST-" + sha16(channel, a0)
            sales_line_id = "SL-" + sha16(source_file_id, rnum, cur_item["canonical_sku"], a0, buyer)
            line = {
                "sales_line_id": sales_line_id,
                "settlement_id": settlement_id,
                "channel": channel,
                "period_month": period_month,
                "sale_date": d.isoformat() if d else "",
                "article_raw": cur_item["article_raw"],
                "canonical_sku": cur_item["canonical_sku"],
                "nomenclature_code": cur_item["nomenclature_code"],
                "name": cur_item["name"],
                "buyer": buyer,
                "document": a0,
                "qty": qty if qty is not None else "",
                "price": price if price is not None else "",
                "revenue_rub": revenue if revenue is not None else "",
                "cogs_rub": cogs if cogs is not None else "",
                "gross_profit_rub": gp if gp is not None else "",
                "source_file_id": source_file_id,
                "source_row_id": f"r{rnum}",
                "bank_payment_id": "",
                "match_status": "UNMATCHED",
            }
            sales.append(line)
            st = settlements.setdefault(
                settlement_id,
                {
                    "settlement_id": settlement_id,
                    "channel": channel,
                    "document": a0,
                    "sale_date": d.isoformat() if d else "",
                    "period_month": period_month,
                    "buyer": buyer,
                    "revenue_rub": 0.0,
                    "cogs_rub": 0.0,
                    "n_lines": 0,
                    "source_file_id": source_file_id,
                    "bank_payment_id": "",
                    "status": "OPEN",
                },
            )
            st["revenue_rub"] += float(revenue or 0)
            st["cogs_rub"] += float(cogs or 0)
            st["n_lines"] += 1
            if buyer and not st["buyer"]:
                st["buyer"] = buyer
            continue

        # buyer / other grouping row
        if cur_item and not is_item and a0 not in ("Покупатель", "Документ движения"):
            cur_buyer = a0

    # finalize settlement amounts rounded
    settle_list = []
    for st in settlements.values():
        st["revenue_rub"] = round(st["revenue_rub"], 2)
        st["cogs_rub"] = round(st["cogs_rub"], 2)
        settle_list.append(st)
    return sales, settle_list


def parse_b2b_fact_monthly(path: Path, year: int, source_file_id: str) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Лист1"] if "Лист1" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    header = rows[0]
    months = []
    for c in header[1:]:
        if isinstance(c, datetime):
            months.append(c.strftime("%Y-%m"))
        elif isinstance(c, date):
            months.append(c.strftime("%Y-%m"))
        else:
            months.append(None)

    # find Выручка Yanina row
    rev_row = None
    cogs_row = None
    gp_row = None
    for r in rows[1:]:
        if not r or r[0] is None:
            continue
        lab = str(r[0]).strip().lower()
        if "выручка yanina" in lab:
            rev_row = r
        elif "себестоимость проданных" in lab:
            cogs_row = r
        elif "валовая прибыль" in lab:
            gp_row = r

    out = []
    for i, m in enumerate(months):
        if not m:
            continue
        col = i + 1
        rev = to_float(rev_row[col]) if rev_row and len(rev_row) > col else None
        cogs = to_float(cogs_row[col]) if cogs_row and len(cogs_row) > col else None
        gp = to_float(gp_row[col]) if gp_row and len(gp_row) > col else None
        if rev is None and cogs is None:
            continue
        if rev == 0 and (cogs is None or cogs == 0):
            continue
        out.append(
            {
                "period_month": m,
                "channel": "B2B_MERCURY",
                "metric": "yanina_revenue",
                "revenue_rub": round(rev or 0, 2),
                "cogs_rub": round(cogs or 0, 2) if cogs is not None else "",
                "gross_profit_rub": round(gp or 0, 2) if gp is not None else "",
                "source_file_id": source_file_id,
                "year": year,
            }
        )
    return out


def parse_sales_dds_income(path: Path, source_file_id: str) -> list[dict]:
    """Канальные доходы EUR из листа ДДС - доход."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["ДДС - доход"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    month_names = {
        "jan": "01", "feb": "02", "march": "03", "apr": "04", "may": "05", "june": "06",
        "july": "07", "aug": "08", "sept": "09", "oct": "10", "nov": "11", "dec": "12",
    }
    # header row 0
    hdr = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    mon_cols = []
    for i, h in enumerate(hdr):
        for k, v in month_names.items():
            if h.startswith(k):
                mon_cols.append((i, v))
                break

    out = []
    year = None
    for r in rows[1:]:
        if not r:
            continue
        if r[0] and "ДДС" in str(r[0]):
            m = re.search(r"(20\d{2})", str(r[0]))
            year = int(m.group(1)) if m else year
            channel = str(r[1]).strip() if r[1] else "Salon+Shop"
        else:
            channel = str(r[1]).strip() if r[1] else ""
            if not channel or not year:
                continue
        for col, mm in mon_cols:
            if len(r) <= col:
                continue
            val = to_float(r[col])
            if val is None or val == 0:
                continue
            out.append(
                {
                    "period_month": f"{year}-{mm}",
                    "channel": channel,
                    "amount_eur": round(val, 2),
                    "source_file_id": source_file_id,
                    "source": "SALES_DDS_INCOME",
                }
            )
    return out


def load_tsum_as_sales() -> list[dict]:
    p = W3 / "tsum_sales_lines.csv"
    if not p.exists():
        return []
    out = []
    for r in csv.DictReader(open(p, encoding="utf-8")):
        out.append(
            {
                "sales_line_id": r["sales_line_id"],
                "settlement_id": "",  # TSUM settlement missing
                "channel": "TSUM",
                "period_month": "",
                "sale_date": "",
                "article_raw": r.get("article_raw", ""),
                "canonical_sku": r.get("canonical_sku", ""),
                "nomenclature_code": r.get("item_id", ""),
                "name": r.get("name_raw", ""),
                "buyer": "TSUM",
                "document": "",
                "qty": r.get("qty", ""),
                "price": "",
                "revenue_rub": r.get("sales_amount_rub", ""),
                "cogs_rub": "",
                "gross_profit_rub": "",
                "source_file_id": r.get("source_file_id", ""),
                "source_row_id": r.get("source_row_id", ""),
                "bank_payment_id": "",
                "match_status": "NO_SETTLEMENT",
            }
        )
    return out


def recon_monthly(sales_lines, b2b_fact, dds_income):
    """Сводки по месяцам + soft vs bank in."""
    by = defaultdict(lambda: {"revenue": 0.0, "cogs": 0.0, "n": 0, "channels": Counter()})
    for s in sales_lines:
        m = s.get("period_month") or ""
        if not m:
            continue
        by[m]["revenue"] += float(s["revenue_rub"] or 0)
        by[m]["cogs"] += float(s["cogs_rub"] or 0)
        by[m]["n"] += 1
        by[m]["channels"][s["channel"]] += 1

    fact = {r["period_month"]: r for r in b2b_fact}
    # bank in operating
    bank_in = defaultdict(float)
    if (W1 / "bank_payments.csv").exists():
        for p in csv.DictReader(open(W1 / "bank_payments.csv", encoding="utf-8")):
            if p.get("direction") != "in" or p.get("is_internal") == "Y":
                continue
            bank_in[p["period_month"]] += float(p["amount"] or 0)

    # dds income total eur per month
    dds_m = defaultdict(float)
    for r in dds_income:
        dds_m[r["period_month"]] += float(r["amount_eur"] or 0)

    months = sorted(set(by) | set(fact) | set(bank_in) | set(dds_m))
    rows = []
    for m in months:
        b = by.get(m, {"revenue": 0.0, "cogs": 0.0, "n": 0, "channels": Counter()})
        f = fact.get(m)
        rev_1c = b["revenue"]
        rev_fact = float(f["revenue_rub"]) if f else None
        # compare 1C B2B+IM month vs Mercury fact only if both
        status = "N/A"
        delta = ""
        if rev_1c and rev_fact is not None:
            # 1C includes all B2B buyers; fact is Mercury-only — expect 1C >= fact often
            delta = round(rev_1c - rev_fact, 2)
            if rev_fact > 0 and abs(delta) / max(rev_1c, rev_fact) <= 0.15:
                status = "SOFT"
            elif rev_1c >= rev_fact * 0.5:
                status = "PARTIAL_SCOPE"
            else:
                status = "GAP"
        elif rev_1c and rev_fact is None:
            status = "1C_ONLY"
        elif rev_fact is not None and not rev_1c:
            status = "FACT_ONLY"

        rows.append(
            {
                "period_month": m,
                "sales_1c_lines": b["n"],
                "sales_1c_revenue_rub": round(rev_1c, 2),
                "sales_1c_cogs_rub": round(b["cogs"], 2),
                "b2b_mercury_fact_rev_rub": round(rev_fact, 2) if rev_fact is not None else "",
                "delta_1c_vs_mercury_fact": delta,
                "status_1c_vs_fact": status,
                "bank_in_operating_rub": round(bank_in.get(m, 0), 2),
                "sales_dds_income_eur": round(dds_m.get(m, 0), 2),
                "channels_1c": ",".join(f"{k}:{v}" for k, v in b["channels"].most_common()),
                "note": "1C B2B+IM vs Mercury fact = разные периметры; bank_in не = выручка (лаги/смесь)",
            }
        )
    return rows


def soft_match_settlements_to_bank(settlements, limit_per_month=20):
    """Exact amount same month: settlement revenue ↔ bank in (LOW)."""
    if not (W1 / "bank_payments.csv").exists():
        return []
    pays = [
        p
        for p in csv.DictReader(open(W1 / "bank_payments.csv", encoding="utf-8"))
        if p.get("direction") == "in" and p.get("is_internal") != "Y"
    ]
    by_month_amt = defaultdict(lambda: defaultdict(list))
    for p in pays:
        by_month_amt[p["period_month"]][round(float(p["amount"] or 0), 2)].append(p)

    matches = []
    used = set()
    for st in settlements:
        m = st.get("period_month") or ""
        amt = round(float(st.get("revenue_rub") or 0), 2)
        if not m or amt <= 0:
            continue
        cands = [p for p in by_month_amt[m].get(amt, []) if p["bank_payment_id"] not in used]
        if not cands:
            continue
        p = cands[0]
        used.add(p["bank_payment_id"])
        matches.append(
            {
                "period_month": m,
                "settlement_id": st["settlement_id"],
                "document": st["document"],
                "buyer": st["buyer"],
                "revenue_rub": amt,
                "bank_payment_id": p["bank_payment_id"],
                "payment_date": p["payment_date"],
                "counterparty_raw": p["counterparty_raw"],
                "match_method": "exact_amount_same_month",
                "confidence": "LOW",
            }
        )
    return matches


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)

    sales = []
    settlements = []

    b2b_path = resolve("Продажи B2B 2024-06.2026.xlsx")
    if b2b_path:
        s, st = parse_1c_nested_sales(b2b_path, "B2B", "FILE-B2B-1C")
        sales.extend(s)
        settlements.extend(st)

    im_path = resolve("Продажи ИМ 2024-06.2026.xlsx")
    if im_path:
        s, st = parse_1c_nested_sales(im_path, "IM", "FILE-IM-1C")
        sales.extend(s)
        settlements.extend(st)

    tsum = load_tsum_as_sales()
    sales_all = sales + tsum

    b2b_fact = []
    for year, fname, fid in [
        (2025, "Факт анализ продаж B2B 2025.xlsx", "FILE-091"),
        (2026, "Факт анализ продаж B2B 2026.xlsx", "FILE-092"),
    ]:
        p = resolve(fname)
        if p:
            b2b_fact.extend(parse_b2b_fact_monthly(p, year, fid))

    dds_income = []
    sp = resolve("SALES 2024-2026.xlsx")
    if sp:
        dds_income = parse_sales_dds_income(sp, "FILE-007")

    recon = recon_monthly(sales, b2b_fact, dds_income)  # 1C only for monthly grain
    matches = soft_match_settlements_to_bank(settlements)

    # link cost from W3 sku for 1C lines
    cost_best = {}
    if (W3 / "cost_versions.csv").exists():
        by = defaultdict(list)
        for c in csv.DictReader(open(W3 / "cost_versions.csv", encoding="utf-8")):
            if c.get("unit_cost_rub") in ("", None):
                continue
            try:
                if float(c["unit_cost_rub"]) <= 0:
                    continue
            except ValueError:
                continue
            by[c["canonical_sku"]].append(c)
        for k, vs in by.items():
            cost_best[k] = sorted(vs, key=lambda x: (x.get("owner_file") == "MERKUSHINA", x.get("cost_year") or ""), reverse=True)[0]

    cost_hit = 0
    for s in sales:
        b = cost_best.get(s["canonical_sku"])
        if not b:
            s["w3_unit_cost"] = ""
            s["w3_cost_version_id"] = ""
            continue
        cost_hit += 1
        s["w3_unit_cost"] = b["unit_cost_rub"]
        s["w3_cost_version_id"] = b["cost_version_id"]

    sales_fields = [
        "sales_line_id", "settlement_id", "channel", "period_month", "sale_date",
        "article_raw", "canonical_sku", "nomenclature_code", "name", "buyer", "document",
        "qty", "price", "revenue_rub", "cogs_rub", "gross_profit_rub",
        "source_file_id", "source_row_id", "bank_payment_id", "match_status",
        "w3_unit_cost", "w3_cost_version_id",
    ]
    for s in sales_all:
        s.setdefault("w3_unit_cost", "")
        s.setdefault("w3_cost_version_id", "")

    settle_fields = [
        "settlement_id", "channel", "document", "sale_date", "period_month", "buyer",
        "revenue_rub", "cogs_rub", "n_lines", "source_file_id", "bank_payment_id", "status",
    ]

    write_csv(OUT / "sales_lines.csv", sales_all, sales_fields)
    write_csv(OUT / "sales_lines_1c.csv", sales, sales_fields)
    write_csv(OUT / "settlements.csv", settlements, settle_fields)
    write_csv(OUT / "b2b_mercury_fact_monthly.csv", b2b_fact, list(b2b_fact[0].keys()) if b2b_fact else ["period_month"])
    write_csv(OUT / "sales_dds_income_eur.csv", dds_income, list(dds_income[0].keys()) if dds_income else ["period_month"])
    write_csv(OUT / "recon_sales_monthly.csv", recon, list(recon[0].keys()) if recon else ["period_month"])
    write_csv(OUT / "soft_matches_settle_bank.csv", matches, list(matches[0].keys()) if matches else ["period_month"])

    # stats
    ch_rev = Counter()
    for s in sales:
        ch_rev[s["channel"]] += float(s["revenue_rub"] or 0)
    tsum_rev = sum(float(s["revenue_rub"] or 0) for s in tsum)

    summary = {
        "generated_at": NOW,
        "wave": "W4",
        "sales_lines_1c": len(sales),
        "sales_lines_tsum": len(tsum),
        "sales_lines_total": len(sales_all),
        "settlements": len(settlements),
        "b2b_fact_months": len(b2b_fact),
        "dds_income_rows": len(dds_income),
        "soft_settle_bank_matches": len(matches),
        "1c_lines_with_w3_cost": cost_hit,
        "1c_revenue_by_channel_rub": {k: round(v, 2) for k, v in ch_rev.items()},
        "tsum_revenue_rub": round(tsum_rev, 2),
        "settlements_with_date": sum(1 for s in settlements if s.get("sale_date")),
        "finding": (
            f"W4: {len(sales)} 1C sales lines + {len(tsum)} TSUM; "
            f"{len(settlements)} settlements; "
            f"{len(matches)} settle↔bank exact-amount matches (LOW); "
            f"{cost_hit} 1C lines linked to W3 cost."
        ),
        "next": "W5 SUP/EXP/MAT or harden settlement→bank via payment purpose/invoice no",
    }
    json.dump(summary, open(OUT / "w4_summary.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    json.dump(summary, open(EV / "w4_summary.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)

    wb = Workbook()
    ws = wb.active
    ws.title = "00_Summary"
    for i, (k, v) in enumerate(summary.items(), 1):
        ws.cell(i, 1, k)
        ws.cell(i, 2, json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v)
    fill = PatternFill("solid", fgColor="1F4E79")

    def add(name, rows):
        w = wb.create_sheet(name)
        if not rows:
            return
        headers = list(rows[0].keys())
        for c, h in enumerate(headers, 1):
            cell = w.cell(1, c, h)
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
        for ri, row in enumerate(rows[:4000], 2):
            for ci, h in enumerate(headers, 1):
                w.cell(ri, ci, row.get(h, ""))

    add("01_Recon_Monthly", recon)
    add("02_Soft_Settle_Bank", matches)
    add("03_B2B_Fact", b2b_fact)
    add("04_Settlements", settlements[:2000])
    wb.save(EV / "YANINA_W4_SALES_SETTLE_EVIDENCE.xlsx")

    (OUT / "README.md").write_text(
        f"""# W4 SALES / SETTLE

Generated: {NOW}

- `sales_lines_1c.csv` — {len(sales)} (B2B+IM)
- `sales_lines.csv` — + TSUM from W3 → {len(sales_all)}
- `settlements.csv` — {len(settlements)}
- soft settle↔bank matches: {len(matches)} (LOW)
- 1C↔W3 cost links: {cost_hit}

Evidence: `../../evidence/w4_sales_settle_20260724/`

Controlled Staging — не SoT. Settlement→bank без invoice id в назначении платежа.
""",
        encoding="utf-8",
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
