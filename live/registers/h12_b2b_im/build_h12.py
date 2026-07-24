#!/usr/bin/env python3
"""
H12: B2B buyer-pool allocator + IM best-lag recon.

Зачем:
1) Оставшиеся B2B OPEN часто имеют платежи того же ИП, но:
   - нормализация «ИНДИВИДУАЛЬНЫЙ ПРЕДПРИНИМАТЕЛЬ» vs «ИП»
   - платежи «заняты» соседними накладными / лаг >4 мес при точной сумме
   Нужен аллокатор пула платежей по buyer.
2) IM: эквайринг с лагом ±1 — брать best lag для статуса месяца.

Не SoT.
"""
from __future__ import annotations

import csv
import json
import re
import shutil
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[3]
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/h12_b2b_im_20260724"
MART = ROOT / "live/marts"
W4 = ROOT / "live/registers/w4_sales_settle"
H10 = ROOT / "live/registers/h10_channel_cash"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def fnum(v):
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(" ", "").replace(",", "."))
    except (TypeError, ValueError):
        return None


def month_idx(pm: str) -> int | None:
    if not pm or "-" not in pm:
        return None
    y, m = pm.split("-")[:2]
    try:
        return int(y) * 12 + int(m)
    except ValueError:
        return None


def norm_name(s: str) -> str:
    s = (s or "").upper().replace("Ё", "Е")
    s = re.sub(r"ИНДИВИДУАЛЬН\w*\s+ПРЕДПРИНИМАТЕЛЬ", " ИП ", s)
    s = re.sub(r"[\"«»]", " ", s)
    s = re.sub(
        r"\b(ООО|АО|ЗАО|ПАО|ИП|ОБЩЕСТВО|ОГРАНИЧЕННОЙ|ОТВЕТСТВЕННОСТЬЮ|Г|МОСКВА|САНКТ|ПЕТЕРБУРГ)\b",
        " ",
        s,
    )
    s = re.sub(r"Р/С\s*\d+", " ", s)
    s = re.sub(r"[^A-ZА-Я0-9]+", " ", s)
    return " ".join(t for t in s.split() if len(t) >= 3)


def name_overlap(a: str, b: str) -> float:
    ta = set(norm_name(a).split())
    tb = set(norm_name(b).split())
    if not ta or not tb:
        return 0.0
    return len(ta & tb) / len(ta | tb)


def surname(buyer: str) -> str:
    toks = norm_name(buyer).split()
    return toks[0] if toks else ""


def doc_nums(text: str) -> set[str]:
    u = (text or "").upper().replace("Ё", "Е")
    nums = set(re.findall(r"НАКЛАДН\w*\s*№?\s*(\d{1,6})", u))
    nums |= set(re.findall(r"№\s*(\d{1,6})\b", u))
    out = set()
    for n in nums:
        if n in {"2023", "2024", "2025", "2026"}:
            continue
        if len(n) <= 2 and int(n) <= 31:
            continue
        out.add(n)
    return out


def status_gap(a: float, b: float) -> str:
    if a == 0 and b == 0:
        return "EMPTY"
    if a == 0 or b == 0:
        return "OPEN"
    gap = abs(b - a) / max(abs(a), abs(b))
    if gap <= 0.08:
        return "CLOSE"
    if gap <= 0.25:
        return "SOFT"
    return "OPEN"


def buyer_pool_match(bank_rows: list[dict]) -> tuple[list[dict], dict]:
    settles = list(csv.DictReader(open(W4 / "settlements.csv", encoding="utf-8")))
    bak = W4 / "settlements_pre_h12.csv"
    if not bak.exists():
        shutil.copy2(W4 / "settlements.csv", bak)

    for s in settles:
        for c in ("bank_payment_ids", "linked_amount_rub", "link_coverage"):
            s.setdefault(c, s.get(c, ""))

    used = set()
    for s in settles:
        if s.get("bank_payment_id"):
            used.add(s["bank_payment_id"])
        for x in (s.get("bank_payment_ids") or "").split("|"):
            if x:
                used.add(x)

    # free bank candidates
    bank = []
    for r in bank_rows:
        if r.get("in_class") not in ("OTHER_IN", "B2B_NAMED"):
            continue
        if r.get("bank_payment_id") in used:
            continue
        amt = fnum(r.get("amount")) or 0
        # иногда сумма «сломана» в выписке, а в purpose есть «83 290.00 ОПЛАТА»
        if amt < 1000:
            pur = r.get("purpose") or ""
            m = re.search(r"(\d{1,3}(?:\s\d{3})*|\d+)[.,](\d{2})\s*(?:ОПЛАТ|РУБ)", pur.upper())
            if m:
                amt = fnum(m.group(1).replace(" ", "") + "." + m.group(2)) or amt
        if amt < 500:
            continue
        rr = dict(r)
        rr["_amount_effective"] = amt
        bank.append(rr)

    open_b2b = [
        s
        for s in settles
        if s.get("channel") == "B2B" and not s.get("bank_payment_id") and (fnum(s.get("revenue_rub")) or 0) >= 1000
    ]

    matches = []
    newly_full = newly_partial = 0

    # group open by surname
    by_sur: dict[str, list] = defaultdict(list)
    for s in open_b2b:
        sur = surname(s.get("buyer") or "")
        if len(sur) >= 4:
            by_sur[sur].append(s)

    # index bank by surname stem
    bank_by_sur: dict[str, list] = defaultdict(list)
    for p in bank:
        cn = norm_name(p.get("counterparty_raw") or "")
        pn = norm_name(p.get("purpose") or "")
        # try all open surnames — O(open*bank) ok (~20*400)
        for sur in by_sur:
            if sur in cn or sur in pn or sur[:5] in cn or sur[:5] in pn:
                bank_by_sur[sur].append(p)

    # process larger settlements first
    open_sorted = sorted(open_b2b, key=lambda s: -(fnum(s.get("revenue_rub")) or 0))

    for st in open_sorted:
        if st.get("bank_payment_id"):
            continue
        buyer = st.get("buyer") or ""
        sur = surname(buyer)
        rev = fnum(st.get("revenue_rub")) or 0
        mi = month_idx(st.get("period_month") or "")
        st_docs = doc_nums(st.get("document") or "")

        pool = []
        for p in bank_by_sur.get(sur, []):
            if p["bank_payment_id"] in used:
                continue
            ov = name_overlap(buyer, p.get("counterparty_raw") or "")
            # surname in CP is enough even if ov low (full FIO order / ИП form)
            cn = norm_name(p.get("counterparty_raw") or "")
            if sur not in cn and ov < 0.25:
                # still allow if purpose has surname and amount close
                if sur not in norm_name(p.get("purpose") or ""):
                    continue
                ov = max(ov, 0.3)
            elif sur in cn:
                ov = max(ov, 0.4)
            amt = fnum(p.get("_amount_effective")) or fnum(p.get("amount")) or 0
            pmi = month_idx(p.get("period_month") or "")
            lag = abs(pmi - mi) if mi is not None and pmi is not None else 99
            doc_hit = bool(st_docs & doc_nums(p.get("purpose") or ""))
            pool.append({"p": p, "amt": amt, "ov": ov, "lag": lag, "doc_hit": doc_hit, "ratio": abs(amt - rev) / max(amt, rev)})

        if not pool:
            continue

        chosen = []
        kind = None

        # A) exact single: близкая сумма + сильное имя; лаг до 8 при ov≥0.9
        exact = []
        for c in pool:
            if c["ov"] < 0.35:
                continue
            max_lag = 8 if c["ov"] >= 0.9 else 4
            max_ratio = 0.025 if c["ov"] >= 0.9 else 0.015
            if c["doc_hit"]:
                max_lag = max(max_lag, 6)
                max_ratio = max(max_ratio, 0.02)
            if c["ratio"] <= max_ratio and c["lag"] <= max_lag:
                exact.append(c)
        exact.sort(key=lambda x: (x["ratio"], x["lag"], -x["ov"]))
        if exact:
            chosen = [exact[0]]
            kind = "FULL_EXACT"
        else:
            # B) multi greedy: лаг до 8 при сильном имени (тот же ИП)
            max_lag = 8
            cand = [c for c in pool if c["ov"] >= 0.35 and c["lag"] <= max_lag]
            cand.sort(key=lambda x: (0 if x["ov"] >= 0.9 else 1, x["lag"], -x["amt"]))
            acc = 0.0
            for c in cand:
                if acc >= rev * 0.98:
                    break
                if c["p"]["bank_payment_id"] in used:
                    continue
                if any(x["p"]["bank_payment_id"] == c["p"]["bank_payment_id"] for x in chosen):
                    continue
                if acc + c["amt"] <= rev * 1.03:
                    chosen.append(c)
                    acc += c["amt"]
            if chosen and abs(acc - rev) / rev <= 0.05:
                kind = "FULL_PARTS"
            elif chosen and acc / rev >= 0.4:
                kind = "PARTIAL"
            else:
                chosen = []
                kind = None

        if not chosen:
            continue

        ids = [c["p"]["bank_payment_id"] for c in chosen]
        for i in ids:
            used.add(i)
        linked_amt = round(sum(c["amt"] for c in chosen), 2)
        cov = round(linked_amt / rev, 3)
        primary = max(chosen, key=lambda x: x["amt"])

        st["bank_payment_id"] = primary["p"]["bank_payment_id"]
        st["bank_payment_ids"] = "|".join(ids)
        st["linked_amount_rub"] = linked_amt
        st["link_coverage"] = cov
        st["status"] = f"LINKED_H12_{kind}"
        if kind.startswith("FULL"):
            newly_full += 1
        else:
            newly_partial += 1

        matches.append(
            {
                "settlement_id": st["settlement_id"],
                "document": st.get("document", ""),
                "buyer": buyer,
                "revenue_rub": rev,
                "period_month": st.get("period_month", ""),
                "bank_payment_ids": "|".join(ids),
                "n_payments": len(ids),
                "linked_amount_rub": linked_amt,
                "link_coverage": cov,
                "match_kind": kind,
                "name_overlap_max": round(max(c["ov"] for c in chosen), 3),
                "month_lag_min": min(c["lag"] for c in chosen),
                "doc_hit": "Y" if any(c["doc_hit"] for c in chosen) else "N",
                "counterparty_sample": (primary["p"].get("counterparty_raw") or "")[:80],
            }
        )

    fields = list(settles[0].keys())
    write_csv(W4 / "settlements.csv", settles, fields)

    linked_b2b = sum(1 for s in settles if s.get("channel") == "B2B" and s.get("bank_payment_id"))
    open_b2b_n = sum(1 for s in settles if s.get("channel") == "B2B" and not s.get("bank_payment_id"))
    stats = {
        "newly_full": newly_full,
        "newly_partial": newly_partial,
        "newly_total": newly_full + newly_partial,
        "linked_b2b": linked_b2b,
        "open_b2b": open_b2b_n,
        "by_kind": dict(Counter(m["match_kind"] for m in matches)),
        "new_linked_rev": round(sum(m["revenue_rub"] for m in matches), 2),
        "linked_rev_b2b": round(
            sum(fnum(s.get("revenue_rub")) or 0 for s in settles if s.get("channel") == "B2B" and s.get("bank_payment_id")),
            2,
        ),
        "open_rev_b2b": round(
            sum(fnum(s.get("revenue_rub")) or 0 for s in settles if s.get("channel") == "B2B" and not s.get("bank_payment_id")),
            2,
        ),
    }
    return matches, stats


def build_im_best_lag() -> tuple[list[dict], dict]:
    bank = list(csv.DictReader(open(H10 / "bank_in_classified.csv", encoding="utf-8")))
    acq_ip = defaultdict(float)
    acq_all = defaultdict(float)
    for r in bank:
        if r.get("in_class") != "ACQ_IM":
            continue
        amt = fnum(r.get("amount")) or 0
        pm = r.get("period_month") or ""
        acq_all[pm] += amt
        if r.get("legal_entity_id") == "LE-IP-YANINA":
            acq_ip[pm] += amt

    sales = defaultdict(float)
    for s in csv.DictReader(open(W4 / "sales_lines.csv", encoding="utf-8")):
        if s.get("dq_exclude_from_margin") == "Y":
            continue
        if s.get("channel") != "IM":
            continue
        sales[s.get("period_month") or ""] += fnum(s.get("revenue_rub")) or 0.0

    rows = []
    for m in sorted(pm for pm in sales if pm):
        s = sales[m]
        mi = month_idx(m)
        opts = []
        for lag, label in ((-1, "lag-1"), (0, "lag0"), (1, "lag+1")):
            cash = 0.0
            for pm, amt in acq_ip.items():
                if month_idx(pm) == (mi or 0) + lag:
                    cash += amt
            opts.append((label, cash, abs(cash - s) if cash else 1e18))
        best = min(opts, key=lambda x: x[2])
        label, cash, _ = best
        st = status_gap(s, cash)
        # also lag0 for reference
        lag0 = next(c for lab, c, _ in opts if lab == "lag0")
        rows.append(
            {
                "channel": "IM",
                "period_month": m,
                "sales_revenue_rub": round(s, 2),
                "bank_in_lag0_rub": round(lag0, 2),
                "bank_in_best_rub": round(cash, 2),
                "best_lag": label,
                "gap_rub": round(cash - s, 2),
                "gap_pct": round((cash - s) / s * 100, 1) if s else "",
                "status": st,
                "status_lag0": status_gap(s, lag0),
                "bank_in_all_le_lag0_rub": round(acq_all.get(m, 0), 2),
                "bank_class": "ACQ_IM_IP",
                "note": "best of lag-1/0/+1 on IP acquiring",
            }
        )

    stats = {
        "months": len(rows),
        "status_best": dict(Counter(r["status"] for r in rows)),
        "status_lag0": dict(Counter(r["status_lag0"] for r in rows)),
        "close_soft_best": sum(1 for r in rows if r["status"] in ("CLOSE", "SOFT")),
        "close_soft_lag0": sum(1 for r in rows if r["status_lag0"] in ("CLOSE", "SOFT")),
        "lag_pref": dict(Counter(r["best_lag"] for r in rows)),
        "coverage_best": round(
            sum(r["bank_in_best_rub"] for r in rows) / sum(r["sales_revenue_rub"] for r in rows), 3
        )
        if rows
        else None,
    }
    return rows, stats


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)
    MART.mkdir(parents=True, exist_ok=True)

    bank_rows = list(csv.DictReader(open(H10 / "bank_in_classified.csv", encoding="utf-8")))

    matches, settle_stats = buyer_pool_match(bank_rows)
    write_csv(
        MART / "settle_bank_b2b_h12.csv",
        matches,
        list(matches[0].keys()) if matches else ["settlement_id"],
    )
    write_csv(OUT / "settle_bank_b2b_h12.csv", matches, list(matches[0].keys()) if matches else ["settlement_id"])
    write_csv(
        W4 / "soft_matches_settle_bank_h12.csv",
        matches,
        list(matches[0].keys()) if matches else ["settlement_id"],
    )

    im_rows, im_stats = build_im_best_lag()
    write_csv(MART / "recon_im_best_lag.csv", im_rows, list(im_rows[0].keys()) if im_rows else ["period_month"])
    write_csv(OUT / "recon_im_best_lag.csv", im_rows, list(im_rows[0].keys()) if im_rows else ["period_month"])

    # merge IM best-lag into channel cash mart (update IM rows, keep TSUM)
    old = list(csv.DictReader(open(MART / "recon_channel_cash_month.csv", encoding="utf-8")))
    tsum = [r for r in old if r.get("channel") == "TSUM"]
    # unified fields
    merged = []
    for r in im_rows:
        merged.append(
            {
                "channel": "IM",
                "period_month": r["period_month"],
                "sales_revenue_rub": r["sales_revenue_rub"],
                "bank_in_rub": r["bank_in_best_rub"],
                "bank_in_all_le_rub": r["bank_in_all_le_lag0_rub"],
                "bank_in_next_month_rub": "",
                "best_bank_in_rub": r["bank_in_best_rub"],
                "best_gap_rub": r["gap_rub"],
                "best_status": r["status"],
                "best_lag_months": r["best_lag"],
                "gap_rub": r["gap_rub"],
                "gap_pct": r["gap_pct"],
                "status": r["status"],
                "bank_class": "ACQ_IM_IP",
                "note": r["note"],
            }
        )
    for r in tsum:
        merged.append(r)
    # write with union fields
    fields = [
        "channel",
        "period_month",
        "sales_revenue_rub",
        "bank_in_rub",
        "bank_in_all_le_rub",
        "bank_in_next_month_rub",
        "best_bank_in_rub",
        "best_gap_rub",
        "best_status",
        "best_lag_months",
        "gap_rub",
        "gap_pct",
        "status",
        "bank_class",
        "note",
    ]
    write_csv(MART / "recon_channel_cash_month.csv", merged, fields)

    summary = {
        "wave": "H12",
        "generated_at": NOW,
        "finding": (
            f"H12: B2B pool +{settle_stats['newly_total']} "
            f"(full {settle_stats['newly_full']}, partial {settle_stats['newly_partial']}) → "
            f"linked {settle_stats['linked_b2b']}/open {settle_stats['open_b2b']}; "
            f"IM best-lag CLOSE/SOFT {im_stats['close_soft_best']}/{im_stats['months']} "
            f"(was lag0 {im_stats['close_soft_lag0']})."
        ),
        "settle_b2b": settle_stats,
        "im_best_lag": im_stats,
        "not_sot": True,
    }
    (OUT / "h12_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    (EV / "h12_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    for name in ("settle_bank_b2b_h12.csv", "recon_im_best_lag.csv", "h12_summary.json"):
        src = OUT / name
        if src.exists():
            shutil.copy2(src, EV / name)

    wb = Workbook()
    ws = wb.active
    ws.title = "H12_Summary"
    ws["A1"] = "H12 B2B pool + IM lag"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A2"] = NOW
    ws["A4"] = summary["finding"]
    ws["A6"] = "B2B newly"
    ws["B6"] = settle_stats["newly_total"]
    ws["A7"] = "B2B linked/open"
    ws["B7"] = f"{settle_stats['linked_b2b']}/{settle_stats['open_b2b']}"
    ws["A8"] = "IM CLOSE/SOFT best"
    ws["B8"] = f"{im_stats['close_soft_best']}/{im_stats['months']}"
    if matches:
        ws2 = wb.create_sheet("B2B")
        ws2.append(list(matches[0].keys()))
        for r in matches:
            ws2.append(list(r.values()))
    wb.save(OUT / "H12_B2B_IM.xlsx")
    wb.save(EV / "H12_B2B_IM.xlsx")

    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
