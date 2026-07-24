#!/usr/bin/env python3
"""
H21: Apply H20 finance recommendations as PROVISIONAL staging flags.

Путь: пользователь снова сказал «делай дальше» при пустом RACI —
следующий конкретный шаг после H20: не ждать ACCEPT, а провести
рекомендации в данные как provisional (so_t=N).

Что делаем:
1) Реестр margin_exceptions (3 B2B SKU = WHOLESALE_OK_LOSS; 0-3243 = QUARANTINE)
2) Колонки на sales_lines: margin_exception, margin_exception_policy
3) Marts: margin clean vs reported; finance_neg status → PROVISIONAL_APPLIED
4) Owner Packet: статус recommendations → APPLIED_PROVISIONAL

Чего НЕ делаем:
- Не заполняем RACI decision_ACCEPT_REJECT
- Не снимаем quarantine 0-3243 / не подставляем худи-COGS
- Не объявляем SoT
"""
from __future__ import annotations

import csv
import json
import shutil
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[3]
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/h21_apply_recs_20260724"
MART = ROOT / "live/marts"
W4 = ROOT / "live/registers/w4_sales_settle"
PACKET = ROOT / "live/YANINA_OWNER_PACKET_RACI_AND_REQUESTS.xlsx"
CC = ROOT / "live/YANINA_LIVE_CONTROL_CENTER_20260723.xlsx"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")

# H20 → apply map
EXCEPTIONS = {
    "0-2497": {
        "policy": "WHOLESALE_OK_LOSS",
        "proposed_decision": "OK_COMMERCIAL_LOSS",
        "action_id": "A-FIN-0-2497",
    },
    "0-2496": {
        "policy": "WHOLESALE_OK_LOSS",
        "proposed_decision": "OK_COMMERCIAL_LOSS",
        "action_id": "A-FIN-0-2496",
    },
    "0-2493A": {
        "policy": "WHOLESALE_OK_LOSS",
        "proposed_decision": "OK_COMMERCIAL_LOSS",
        "action_id": "A-FIN-0-2493A",
    },
    "0-3243": {
        "policy": "COST_IDENTITY_QUARANTINE",
        "proposed_decision": "KEEP_QUARANTINE_NEED_COST_VERSION",
        "action_id": "A-FIN-0-3243",
    },
}

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF")


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def fnum(v):
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(" ", "").replace(",", "."))
    except (TypeError, ValueError):
        return None


def load_sales() -> list[dict]:
    return list(csv.DictReader(open(W4 / "sales_lines.csv", encoding="utf-8")))


def apply_flags(sales: list[dict]) -> tuple[list[dict], list[dict]]:
    """Проставить provisional flags; вернуть sales + exception line ledger."""
    ledger = []
    flagged = 0
    for s in sales:
        sku = s.get("canonical_sku") or ""
        ex = EXCEPTIONS.get(sku)
        if not ex:
            s["margin_exception"] = s.get("margin_exception") or ""
            s["margin_exception_policy"] = s.get("margin_exception_policy") or ""
            continue
        s["margin_exception"] = "Y"
        s["margin_exception_policy"] = f"PROVISIONAL_H21:{ex['policy']}"
        flagged += 1
        if s.get("dq_exclude_from_margin") == "Y":
            continue
        ledger.append(
            {
                "sales_line_id": s.get("sales_line_id"),
                "canonical_sku": sku,
                "channel": s.get("channel"),
                "period_month": s.get("period_month"),
                "qty": s.get("qty"),
                "revenue_rub": s.get("revenue_rub"),
                "cogs_rub": s.get("cogs_rub"),
                "margin_rub": s.get("margin_rub"),
                "policy": ex["policy"],
                "action_id": ex["action_id"],
                "proposed_decision": ex["proposed_decision"],
                "application": "PROVISIONAL_H21",
                "so_t": "N",
                "owner_accept": "PENDING",
            }
        )
    return sales, ledger


def save_sales(sales: list[dict]):
    # preserve column order; append new cols if missing
    base = list(sales[0].keys())
    for col in ("margin_exception", "margin_exception_policy"):
        if col not in base:
            base.append(col)
    path = W4 / "sales_lines.csv"
    # backup once
    bak = EV / "sales_lines.pre_h21.csv"
    if not bak.exists():
        shutil.copy2(path, bak)
    write_csv(path, sales, base)
    write_csv(OUT / "sales_lines_flagged.csv", sales, base)


def build_registry(ledger: list[dict]) -> list[dict]:
    by = defaultdict(
        lambda: {
            "lines": 0,
            "qty": 0.0,
            "revenue": 0.0,
            "cogs": 0.0,
            "margin": 0.0,
            "channels": set(),
            "policy": "",
            "action_id": "",
            "proposed_decision": "",
        }
    )
    for r in ledger:
        sku = r["canonical_sku"]
        b = by[sku]
        b["lines"] += 1
        b["qty"] += fnum(r.get("qty")) or 0
        b["revenue"] += fnum(r.get("revenue_rub")) or 0
        c = fnum(r.get("cogs_rub"))
        if c is not None:
            b["cogs"] += c
        m = fnum(r.get("margin_rub"))
        if m is not None:
            b["margin"] += m
        b["channels"].add(r.get("channel") or "")
        b["policy"] = r["policy"]
        b["action_id"] = r["action_id"]
        b["proposed_decision"] = r["proposed_decision"]
    rows = []
    for sku, b in sorted(by.items()):
        rows.append(
            {
                "canonical_sku": sku,
                "policy": b["policy"],
                "proposed_decision": b["proposed_decision"],
                "action_id": b["action_id"],
                "channels": ",".join(sorted(x for x in b["channels"] if x)),
                "lines": b["lines"],
                "qty": round(b["qty"], 3),
                "revenue_rub": round(b["revenue"], 2),
                "cogs_rub": round(b["cogs"], 2) if b["cogs"] else "",
                "margin_rub": round(b["margin"], 2) if b["margin"] else "",
                "application": "PROVISIONAL_H21",
                "owner_accept": "PENDING",
                "so_t": "N",
                "note": "Applied as staging flag; RACI/owner ACCEPT still pending",
            }
        )
    return rows


def rebuild_margin_views(sales: list[dict]) -> tuple[list[dict], list[dict], dict]:
    """
    reported = как сейчас (exclude только dq_exclude)
    clean = reported минус WHOLESALE_OK_LOSS строки (для «чистой» маржи без оптовых убытков)
    quarantine lines остаются без COGS в обоих.
    """
    rep = defaultdict(lambda: {"revenue": 0.0, "cogs": 0.0, "margin": 0.0, "lines": 0, "costed": 0})
    clean = defaultdict(lambda: {"revenue": 0.0, "cogs": 0.0, "margin": 0.0, "lines": 0, "costed": 0})

    for s in sales:
        if s.get("dq_exclude_from_margin") == "Y":
            continue
        ch = s.get("channel") or "UNK"
        rev = fnum(s.get("revenue_rub")) or 0
        cogs = fnum(s.get("cogs_rub"))
        mar = fnum(s.get("margin_rub"))
        policy = s.get("margin_exception_policy") or ""

        rep[ch]["revenue"] += rev
        rep[ch]["lines"] += 1
        if cogs is not None:
            rep[ch]["cogs"] += cogs
            rep[ch]["costed"] += 1
        if mar is not None:
            rep[ch]["margin"] += mar

        # clean: drop wholesale ok-loss from margin view (keep quarantine revenue w/o cogs)
        if "WHOLESALE_OK_LOSS" in policy:
            continue
        clean[ch]["revenue"] += rev
        clean[ch]["lines"] += 1
        if cogs is not None:
            clean[ch]["cogs"] += cogs
            clean[ch]["costed"] += 1
        if mar is not None:
            clean[ch]["margin"] += mar

    def to_rows(bucket):
        rows = []
        for ch, v in sorted(bucket.items()):
            pct = round(100.0 * v["margin"] / v["revenue"], 1) if v["revenue"] else ""
            rows.append(
                {
                    "channel": ch,
                    "lines": v["lines"],
                    "costed_lines": v["costed"],
                    "revenue_rub": round(v["revenue"], 2),
                    "cogs_rub": round(v["cogs"], 2),
                    "margin_rub": round(v["margin"], 2),
                    "margin_pct": pct,
                }
            )
        # total
        tot = {
            "channel": "TOTAL",
            "lines": sum(r["lines"] for r in rows),
            "costed_lines": sum(r["costed_lines"] for r in rows),
            "revenue_rub": round(sum(r["revenue_rub"] for r in rows), 2),
            "cogs_rub": round(sum(r["cogs_rub"] for r in rows), 2),
            "margin_rub": round(sum(r["margin_rub"] for r in rows), 2),
            "margin_pct": "",
        }
        if tot["revenue_rub"]:
            tot["margin_pct"] = round(100.0 * tot["margin_rub"] / tot["revenue_rub"], 1)
        rows.append(tot)
        return rows

    rep_rows = to_rows(rep)
    clean_rows = to_rows(clean)
    meta = {
        "reported_margin_pct": next(r["margin_pct"] for r in rep_rows if r["channel"] == "TOTAL"),
        "clean_margin_pct": next(r["margin_pct"] for r in clean_rows if r["channel"] == "TOTAL"),
        "reported_margin_rub": next(r["margin_rub"] for r in rep_rows if r["channel"] == "TOTAL"),
        "clean_margin_rub": next(r["margin_rub"] for r in clean_rows if r["channel"] == "TOTAL"),
    }
    return rep_rows, clean_rows, meta


def update_finance_neg():
    path = MART / "finance_neg_sku_review.csv"
    if not path.exists():
        return
    rows = list(csv.DictReader(open(path, encoding="utf-8")))
    for r in rows:
        sku = r.get("canonical_sku")
        if sku in EXCEPTIONS:
            r["action"] = f"PROVISIONAL_H21:{EXCEPTIONS[sku]['policy']}"
            r["review_status"] = "PROVISIONAL_APPLIED"
            r["owner_accept"] = "PENDING"
    fields = list(rows[0].keys())
    for col in ("review_status", "owner_accept"):
        if col not in fields:
            fields.append(col)
    write_csv(path, rows, fields)
    write_csv(OUT / "finance_neg_sku_review.csv", rows, fields)


def update_recommendations_csv():
    path = MART / "finance_recommendations.csv"
    if not path.exists():
        return
    rows = list(csv.DictReader(open(path, encoding="utf-8")))
    for r in rows:
        sku = r.get("canonical_sku")
        if sku in EXCEPTIONS or sku == "PORTFOLIO":
            r["status"] = "APPLIED_PROVISIONAL"
            r["application_note"] = "H21 staging flags on sales_lines; owner ACCEPT still pending"
    write_csv(path, rows, list(rows[0].keys()))


def update_packet():
    if not PACKET.exists():
        return
    wb = load_workbook(PACKET)
    if "RECOMMENDATIONS_H20" in wb.sheetnames:
        ws = wb["RECOMMENDATIONS_H20"]
        # find header with status / owner_decision
        header_row = None
        headers = []
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), 1):
            if row and "rec_id" in [str(c) if c else "" for c in row]:
                header_row = i
                headers = list(row)
                break
        if header_row and "status" in headers:
            si = headers.index("status")
            for r in range(header_row + 1, (ws.max_row or header_row) + 1):
                rid = ws.cell(r, 1).value
                if rid:
                    ws.cell(r, si + 1).value = "APPLIED_PROVISIONAL"
        ws["A2"] = (
            f"{NOW} | H21 APPLIED_PROVISIONAL на sales_lines. "
            "owner_decision_ACCEPT_REJECT всё ещё за вами (подтвердить или откатить)."
        )
    if "README" in wb.sheetnames:
        wb["README"]["A14"] = (
            f"H21 {NOW}: provisional wholesale flags + quarantine policy applied in staging. "
            "SoT по-прежнему ждёт RACI ACCEPT."
        )
    # small apply log sheet
    name = "H21_APPLY_LOG"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, 4)
    ws["A1"] = "H21 provisional apply log"
    ws["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws["A2"] = NOW
    ws.append([])
    ws.append(["canonical_sku", "policy", "owner_accept", "so_t", "note"])
    for c in range(1, 6):
        ws.cell(4, c).fill = HDR_FILL
        ws.cell(4, c).font = HDR_FONT
    for sku, ex in EXCEPTIONS.items():
        ws.append([sku, ex["policy"], "PENDING", "N", "staging flag only"])
    wb.save(PACKET)
    shutil.copy2(PACKET, EV / "YANINA_OWNER_PACKET_RACI_AND_REQUESTS.xlsx")


def update_cc(summary: dict):
    if not CC.exists():
        return
    wb = load_workbook(CC)
    if "H21_Apply" in wb.sheetnames:
        del wb["H21_Apply"]
    ws = wb.create_sheet("H21_Apply", 0)
    ws["A1"] = "H21 Apply H20 provisional"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A2"] = NOW
    ws["A4"] = summary["finding"]
    ws["A5"] = "Reported GM%"
    ws["B5"] = summary.get("reported_margin_pct")
    ws["A6"] = "Clean GM% (ex wholesale loss)"
    ws["B6"] = summary.get("clean_margin_pct")
    ws["A7"] = "RACI touched?"
    ws["B7"] = "NO"
    ws["A8"] = "SoT?"
    ws["B8"] = "NO"
    wb.save(CC)


def write_md(summary: dict, registry: list[dict]):
    lines = [
        "# H21 — Provisional apply of H20 recommendations",
        "",
        f"Updated: {NOW}",
        "",
        "Staging only. **Not SoT. RACI not filled.**",
        "",
        f"- Reported margin: **{summary.get('reported_margin_pct')}%**",
        f"- Clean margin (ex WHOLESALE_OK_LOSS): **{summary.get('clean_margin_pct')}%**",
        "",
        "| SKU | Policy | Revenue | Margin |",
        "|-----|--------|---------|--------|",
    ]
    for r in registry:
        lines.append(
            f"| {r['canonical_sku']} | {r['policy']} | {r['revenue_rub']} | {r['margin_rub']} |"
        )
    lines.extend(
        [
            "",
            "## Rollback",
            "Restore `live/evidence/h21_apply_recs_20260724/sales_lines.pre_h21.csv` → `w4_sales_settle/sales_lines.csv`.",
            "",
            "## Next",
            "Owner: confirm ACCEPT on RECOMMENDATIONS_H20 / RACI, or REJECT to rollback flags.",
            "",
        ]
    )
    text = "\n".join(lines)
    (OUT / "H21_APPLY.md").write_text(text, encoding="utf-8")
    (ROOT / "live/H21_APPLY.md").write_text(text, encoding="utf-8")
    (EV / "H21_APPLY.md").write_text(text, encoding="utf-8")


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)

    sales = load_sales()
    sales, ledger = apply_flags(sales)
    save_sales(sales)
    registry = build_registry(ledger)
    rep_rows, clean_rows, meta = rebuild_margin_views(sales)

    write_csv(MART / "margin_exceptions.csv", registry, list(registry[0].keys()) if registry else ["canonical_sku"])
    write_csv(OUT / "margin_exceptions.csv", registry, list(registry[0].keys()) if registry else ["canonical_sku"])
    write_csv(MART / "margin_exception_lines.csv", ledger, list(ledger[0].keys()) if ledger else ["sales_line_id"])
    write_csv(
        MART / "margin_channel_total.csv",
        rep_rows,
        list(rep_rows[0].keys()),
    )
    write_csv(
        MART / "margin_channel_total_clean.csv",
        clean_rows,
        list(clean_rows[0].keys()),
    )
    write_csv(OUT / "margin_channel_total_clean.csv", clean_rows, list(clean_rows[0].keys()))

    update_finance_neg()
    update_recommendations_csv()
    update_packet()

    summary = {
        "wave": "H21",
        "generated_at": NOW,
        "path_choice": "Apply H20 recommendations as PROVISIONAL staging flags (no RACI autofill)",
        "finding": (
            f"H21: provisional flags on {len(ledger)} sales lines / {len(registry)} SKUs; "
            f"reported GM {meta['reported_margin_pct']}% → clean(ex wholesale loss) "
            f"{meta['clean_margin_pct']}%. Quarantine 0-3243 kept. RACI untouched."
        ),
        "exception_lines": len(ledger),
        "exception_skus": len(registry),
        "reported_margin_pct": meta["reported_margin_pct"],
        "clean_margin_pct": meta["clean_margin_pct"],
        "reported_margin_rub": meta["reported_margin_rub"],
        "clean_margin_rub": meta["clean_margin_rub"],
        "raci_changed": False,
        "not_sot": True,
        "owner_accept": "PENDING",
    }
    write_md(summary, registry)
    (OUT / "h21_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    (EV / "h21_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    for name in ("margin_exceptions.csv", "h21_summary.json", "H21_APPLY.md"):
        shutil.copy2(OUT / name, EV / name)
    update_cc(summary)
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
