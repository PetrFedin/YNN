#!/usr/bin/env python3
"""
W3: REG-SKU + REG-COST + TSUM sales staging → пересечение и маржа.

Зачем:
- закрыть product/margin spine (SALES→SKU←COST);
- provisional canonical_sku / cost_version_id;
- измерить покрытие себестоимостью продаж ЦУМ.

Не SoT: разные авторы cost-файлов, нет единого cost_version approval.
"""
from __future__ import annotations

import csv
import hashlib
import json
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parents[3]
DOCS = ROOT / "documents"
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/w3_sku_cost_20260724"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")


def nfc(s: str) -> str:
    return unicodedata.normalize("NFC", s or "")


def sha16(*parts) -> str:
    raw = "|".join("" if p is None else str(p) for p in parts)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]


def to_float(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def parse_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if v is None:
        return None
    s = str(v).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


def normalize_article(raw) -> str:
    """Канонизация артикула для join (без размера/цвета)."""
    if raw is None:
        return ""
    s = nfc(str(raw)).strip().upper()
    s = s.replace("Ё", "Е")
    # Cyrillic lookalikes
    s = s.replace("Т-", "T-").replace("Т–", "T-").replace("Т—", "T-")
    s = s.replace("А", "A").replace("В", "B").replace("Е", "E").replace("К", "K").replace("М", "M").replace("Н", "H").replace("О", "O").replace("Р", "P").replace("С", "C").replace("Т", "T").replace("Х", "X")
    s = s.replace(" ", "").replace("\xa0", "")
    s = s.replace("–", "-").replace("—", "-")
    # drop trailing size-only junk if any
    s = re.sub(r"[/\\].*$", "", s)
    return s


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def load_catalog():
    return {
        nfc(r["file_name"]): r
        for r in csv.DictReader(open(ROOT / "live/registers/00_SOURCE_CATALOG_93.csv", encoding="utf-8-sig"))
    }


def resolve(name: str) -> Path | None:
    t = nfc(name)
    for p in DOCS.iterdir():
        if nfc(p.name) == t:
            return p
    return None


# ── TSUM sales ────────────────────────────────────────────────────
def parse_tsum(path: Path, source_file_id: str, year: int) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr_i = None
    header = None
    for i, r in enumerate(rows[:20]):
        if r and any(str(c).strip() == "Артикул" for c in r if c is not None):
            hdr_i = i
            header = [str(c).strip() if c is not None else "" for c in r]
            break
    if hdr_i is None:
        return []
    idx = {h: i for i, h in enumerate(header) if h}

    def col(*names):
        for n in names:
            if n in idx:
                return idx[n]
        return None

    i_art = col("Артикул")
    i_item = col("ItemID")
    i_color = col("ColorID")
    i_name = col("ЦветРазмер")
    i_pcs = col("Sales PCS")
    i_amt = col("Sales Cost RUR")
    i_store = col("Store")
    out = []
    for rnum, r in enumerate(rows[hdr_i + 1 :], start=hdr_i + 2):
        if not r or i_art is None or len(r) <= i_art or r[i_art] is None:
            continue
        art_raw = str(r[i_art]).strip()
        canon = normalize_article(art_raw)
        if not canon:
            continue
        pcs = to_float(r[i_pcs]) if i_pcs is not None else None
        amt = to_float(r[i_amt]) if i_amt is not None else None
        sales_line_id = "SL-" + sha16(source_file_id, rnum, canon, r[i_item] if i_item else "")
        out.append(
            {
                "sales_line_id": sales_line_id,
                "channel": "TSUM",
                "period_year": year,
                "period_month": "",  # grain in file = year aggregate lines
                "article_raw": art_raw,
                "canonical_sku": canon,
                "item_id": str(r[i_item]).strip() if i_item is not None and r[i_item] else "",
                "color": str(r[i_color]).strip() if i_color is not None and r[i_color] else "",
                "name_raw": str(r[i_name]).strip() if i_name is not None and r[i_name] else "",
                "qty": pcs if pcs is not None else "",
                "sales_amount_rub": amt if amt is not None else "",
                "amount_meaning": "TSUM Sales Cost RUR (transfer/wholesale-like, not retail check)",
                "store": str(r[i_store]).strip() if i_store is not None and r[i_store] else "",
                "source_file_id": source_file_id,
                "source_row_id": f"r{rnum}",
                "cost_version_id": "",
                "unit_cost_rub": "",
                "margin_rub": "",
                "margin_pct": "",
                "match_status": "UNMATCHED",
            }
        )
    return out


# ── COST: Zhukova ─────────────────────────────────────────────────
def parse_zhukova(path: Path, catalog: dict) -> list[dict]:
    meta = catalog.get(nfc(path.name), {})
    fid = meta.get("master_file_id") or "FILE-081"
    wb = load_workbook(path, read_only=True, data_only=True)
    out = []
    for sheet in ("2024", "2025", "2026"):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        hdr_i = None
        header = None
        for i, r in enumerate(rows[:10]):
            if r and any(str(c).strip() == "Артикул" for c in r if c is not None):
                hdr_i = i
                header = [str(c).strip() if c is not None else "" for c in r]
                break
        if hdr_i is None:
            continue
        idx = {h: i for i, h in enumerate(header) if h}

        def col(*names):
            for n in names:
                if n in idx:
                    return idx[n]
            return None

        i_art = col("Артикул")
        i_name = col("Наименование")
        i_dir = col("Направление/Проект")
        i_color = col("Цвет")
        i_size = col("Размер")
        i_qty = col("Кол-во")
        i_mat = col("Материалы для отшива")
        i_lab = col("Трудовые затраты")
        i_unit = col("Фактическая себестоимость")
        i_order = col("Стоимость заказа")
        i_date = col("Дата запуска") or col("Дата готовности")
        for rnum, r in enumerate(rows[hdr_i + 1 :], start=hdr_i + 2):
            if not r or i_art is None or len(r) <= i_art or not r[i_art]:
                continue
            art_raw = str(r[i_art]).strip()
            canon = normalize_article(art_raw)
            if not canon or canon in ("X", "Х", "-"):
                continue
            mat = to_float(r[i_mat]) if i_mat is not None else None
            lab = to_float(r[i_lab]) if i_lab is not None else None
            unit = to_float(r[i_unit]) if i_unit is not None else None
            qty = to_float(r[i_qty]) if i_qty is not None else None
            order = to_float(r[i_order]) if i_order is not None else None
            if unit is None and mat is not None and lab is not None:
                unit = mat + lab
            if unit is None or unit == 0:
                # keep zero-cost as record but flag
                pass
            d = parse_date(r[i_date]) if i_date is not None else None
            cvid = "CV-" + sha16("ZHUKOVA", sheet, fid, rnum, canon, unit)
            out.append(
                {
                    "cost_version_id": cvid,
                    "canonical_sku": canon,
                    "article_raw": art_raw,
                    "name": str(r[i_name]).strip() if i_name is not None and r[i_name] else "",
                    "direction": str(r[i_dir]).strip() if i_dir is not None and r[i_dir] else "",
                    "color": str(r[i_color]).strip() if i_color is not None and r[i_color] else "",
                    "size": str(r[i_size]).strip() if i_size is not None and r[i_size] else "",
                    "qty": qty if qty is not None else "",
                    "material_cost": mat if mat is not None else "",
                    "labor_cost": lab if lab is not None else "",
                    "unit_cost_rub": unit if unit is not None else "",
                    "order_cost_rub": order if order is not None else "",
                    "cost_date": d.isoformat() if d else "",
                    "cost_year": sheet,
                    "owner_file": "ZHUKOVA",
                    "channel_hint": "",
                    "source_file_id": fid,
                    "source_sheet": sheet,
                    "source_row_id": f"r{rnum}",
                }
            )
    wb.close()
    return out


# ── COST: Merkushina ──────────────────────────────────────────────
def parse_merkushina(path: Path, catalog: dict) -> list[dict]:
    meta = catalog.get(nfc(path.name), {})
    fid = meta.get("master_file_id") or "FILE-082"
    wb = load_workbook(path, read_only=True, data_only=True)
    out = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        hdr_i = None
        header = None
        for i, r in enumerate(rows[:8]):
            if r and any(str(c).strip() == "Артикул" for c in r if c is not None):
                hdr_i = i
                header = [str(c).strip() if c is not None else "" for c in r]
                break
        # some sheets have empty Артикул header cell but data in col1
        if hdr_i is None:
            for i, r in enumerate(rows[:8]):
                joined = " ".join(str(c) for c in r if c is not None)
                if "мат.расх на 1ед" in joined or "Фактическая себестоимость" in joined:
                    hdr_i = i
                    header = [str(c).strip() if c is not None else "" for c in r]
                    break
        if hdr_i is None:
            continue
        idx = {h: i for i, h in enumerate(header) if h}

        def col(*names):
            for n in names:
                if n in idx:
                    return idx[n]
            return None

        i_art = col("Артикул")
        # капсулы: артикул может быть во 2-й колонке без заголовка
        if i_art is None:
            i_art = 1
        i_name = col("Наименование")
        i_plat = col("Площадка")
        i_coll = col("КОЛЛЕКЦИЯ")
        i_color = col("Цвет")
        i_qty = col("Кол-во")
        i_mat_u = col("мат.расх на 1ед")
        i_lab_u = col("Трудовые затраты на 1ед")
        i_unit = col("Фактическая себестоимость 1ед")
        i_order = col("Стоимость заказа")
        for rnum, r in enumerate(rows[hdr_i + 1 :], start=hdr_i + 2):
            if not r:
                continue
            art_cell = r[i_art] if len(r) > i_art else None
            if art_cell is None or str(art_cell).strip() == "":
                continue
            art_raw = str(art_cell).strip()
            # skip section titles
            if len(art_raw) > 40 or " " in art_raw and not re.search(r"\d", art_raw):
                continue
            canon = normalize_article(art_raw)
            if not canon:
                continue
            mat_u = to_float(r[i_mat_u]) if i_mat_u is not None and len(r) > i_mat_u else None
            lab_u = to_float(r[i_lab_u]) if i_lab_u is not None and len(r) > i_lab_u else None
            unit = to_float(r[i_unit]) if i_unit is not None and len(r) > i_unit else None
            if unit is None and mat_u is not None and lab_u is not None:
                unit = mat_u + lab_u
            qty = to_float(r[i_qty]) if i_qty is not None and len(r) > i_qty else None
            order = to_float(r[i_order]) if i_order is not None and len(r) > i_order else None
            cvid = "CV-" + sha16("MERKUSHINA", sheet, fid, rnum, canon, unit)
            out.append(
                {
                    "cost_version_id": cvid,
                    "canonical_sku": canon,
                    "article_raw": art_raw,
                    "name": str(r[i_name]).strip() if i_name is not None and len(r) > i_name and r[i_name] else "",
                    "direction": str(r[i_plat]).strip() if i_plat is not None and len(r) > i_plat and r[i_plat] else sheet,
                    "color": str(r[i_color]).strip() if i_color is not None and len(r) > i_color and r[i_color] else "",
                    "size": "",
                    "qty": qty if qty is not None else "",
                    "material_cost": mat_u if mat_u is not None else "",
                    "labor_cost": lab_u if lab_u is not None else "",
                    "unit_cost_rub": unit if unit is not None else "",
                    "order_cost_rub": order if order is not None else "",
                    "cost_date": "",
                    "cost_year": "",
                    "owner_file": "MERKUSHINA",
                    "channel_hint": str(r[i_plat]).strip() if i_plat is not None and len(r) > i_plat and r[i_plat] else "",
                    "collection": str(r[i_coll]).strip() if i_coll is not None and len(r) > i_coll and r[i_coll] else "",
                    "source_file_id": fid,
                    "source_sheet": sheet,
                    "source_row_id": f"r{rnum}",
                }
            )
    wb.close()
    return out


# ── COST partial: Mokeeva (model + material fragment) ─────────────
def parse_mokeeva(path: Path, catalog: dict, year: str) -> list[dict]:
    meta = catalog.get(nfc(path.name), {})
    fid = meta.get("master_file_id") or f"FILE-MOK-{year}"
    try:
        # read_only ломается на битых merge в файле Мокеевой
        wb = load_workbook(path, read_only=False, data_only=True)
    except Exception as e:
        print("WARN mokeeva skip", path.name, e)
        return []
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr_i = None
    header = None
    for i, r in enumerate(rows[:5]):
        if r and any(str(c).strip() == "№ модели" for c in r if c is not None):
            hdr_i = i
            header = [str(c).strip() if c is not None else "" for c in r]
            break
    if hdr_i is None:
        return []
    idx = {h: i for i, h in enumerate(header) if h}
    i_model = idx.get("№ модели")
    i_name = idx.get("наименование изделия")
    i_mat_cost = idx.get("стоимость")  # first стоимость = main fabric
    out = []
    for rnum, r in enumerate(rows[hdr_i + 1 :], start=hdr_i + 2):
        if not r or i_model is None or len(r) <= i_model or not r[i_model]:
            continue
        art_raw = str(r[i_model]).strip()
        canon = normalize_article(art_raw)
        if not canon:
            continue
        mat = to_float(r[i_mat_cost]) if i_mat_cost is not None and len(r) > i_mat_cost else None
        # Mokeeva = material fragment only → not full unit cost
        cvid = "CV-" + sha16("MOKEEVA", year, fid, rnum, canon)
        out.append(
            {
                "cost_version_id": cvid,
                "canonical_sku": canon,
                "article_raw": art_raw,
                "name": str(r[i_name]).strip() if i_name is not None and r[i_name] else "",
                "direction": "MD",
                "color": "",
                "size": "",
                "qty": 1,
                "material_cost": mat if mat is not None else "",
                "labor_cost": "",
                "unit_cost_rub": "",  # incomplete
                "order_cost_rub": "",
                "cost_date": "",
                "cost_year": year,
                "owner_file": "MOKEEVA",
                "channel_hint": "partial_material_only",
                "source_file_id": fid,
                "source_sheet": year,
                "source_row_id": f"r{rnum}",
                "completeness": "PARTIAL_MATERIAL",
            }
        )
    return out


def pick_best_cost(versions: list[dict]) -> dict | None:
    """Выбрать лучшую версию себестоимости для SKU: полный unit_cost, свежий год, не Mokeeva."""
    full = [v for v in versions if v.get("unit_cost_rub") not in ("", None) and float(v["unit_cost_rub"] or 0) > 0]
    if not full:
        return None
    def score(v):
        owner = 2 if v["owner_file"] == "MERKUSHINA" else (1 if v["owner_file"] == "ZHUKOVA" else 0)
        year = int(v["cost_year"]) if str(v.get("cost_year") or "").isdigit() else 0
        return (owner, year, float(v["unit_cost_rub"]))
    return sorted(full, key=score, reverse=True)[0]


def build_sku_master(tsum_lines, cost_rows) -> list[dict]:
    by = {}
    for s in tsum_lines:
        k = s["canonical_sku"]
        by.setdefault(
            k,
            {
                "canonical_sku": k,
                "article_aliases": set(),
                "names": set(),
                "seen_in_tsum": False,
                "seen_in_cost": False,
                "tsum_qty": 0.0,
                "tsum_amount": 0.0,
            },
        )
        by[k]["article_aliases"].add(s["article_raw"])
        if s["name_raw"]:
            by[k]["names"].add(s["name_raw"][:80])
        by[k]["seen_in_tsum"] = True
        by[k]["tsum_qty"] += float(s["qty"] or 0)
        by[k]["tsum_amount"] += float(s["sales_amount_rub"] or 0)
    for c in cost_rows:
        k = c["canonical_sku"]
        by.setdefault(
            k,
            {
                "canonical_sku": k,
                "article_aliases": set(),
                "names": set(),
                "seen_in_tsum": False,
                "seen_in_cost": False,
                "tsum_qty": 0.0,
                "tsum_amount": 0.0,
            },
        )
        by[k]["article_aliases"].add(c["article_raw"])
        if c.get("name"):
            by[k]["names"].add(c["name"][:80])
        by[k]["seen_in_cost"] = True

    out = []
    for k, v in sorted(by.items()):
        out.append(
            {
                "canonical_sku": k,
                "article_aliases": " | ".join(sorted(v["article_aliases"])[:5]),
                "name_samples": " | ".join(sorted(v["names"])[:2]),
                "seen_in_tsum": "Y" if v["seen_in_tsum"] else "N",
                "seen_in_cost": "Y" if v["seen_in_cost"] else "N",
                "in_intersection": "Y" if v["seen_in_tsum"] and v["seen_in_cost"] else "N",
                "tsum_qty_total": round(v["tsum_qty"], 2),
                "tsum_amount_total": round(v["tsum_amount"], 2),
            }
        )
    return out


def enrich_tsum_with_cost(tsum_lines, cost_rows):
    by_sku = defaultdict(list)
    for c in cost_rows:
        by_sku[c["canonical_sku"]].append(c)
    best = {k: pick_best_cost(vs) for k, vs in by_sku.items()}
    matched = 0
    margin_rows = []
    for s in tsum_lines:
        b = best.get(s["canonical_sku"])
        if not b:
            continue
        unit = float(b["unit_cost_rub"])
        qty = float(s["qty"] or 0) or 1.0
        sales = float(s["sales_amount_rub"] or 0)
        cogs = unit * qty
        margin = sales - cogs
        pct = (margin / sales * 100) if sales else None
        s["cost_version_id"] = b["cost_version_id"]
        s["unit_cost_rub"] = round(unit, 2)
        s["margin_rub"] = round(margin, 2)
        s["margin_pct"] = round(pct, 2) if pct is not None else ""
        s["match_status"] = "MATCHED"
        s["cost_owner"] = b["owner_file"]
        matched += 1
        margin_rows.append(
            {
                "canonical_sku": s["canonical_sku"],
                "period_year": s["period_year"],
                "qty": qty,
                "sales_amount_rub": sales,
                "unit_cost_rub": unit,
                "cogs_rub": round(cogs, 2),
                "margin_rub": round(margin, 2),
                "margin_pct": round(pct, 2) if pct is not None else "",
                "cost_version_id": b["cost_version_id"],
                "cost_owner": b["owner_file"],
                "sales_line_id": s["sales_line_id"],
            }
        )
    return matched, margin_rows


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)
    catalog = load_catalog()

    # TSUM
    tsum_lines = []
    for year, fname, fid in [
        (2024, "Продажи ЦУМ 2024.xlsx", "FILE-057"),
        (2025, "Продажи ЦУМ 2025.xlsx", "FILE-058"),
        (2026, "Продажи ЦУМ 2026.xlsx", "FILE-059"),
    ]:
        p = resolve(fname)
        if not p:
            continue
        meta = catalog.get(nfc(fname), {})
        tsum_lines.extend(parse_tsum(p, meta.get("master_file_id") or fid, year))

    # COST
    cost_rows = []
    p = resolve("Себестоимость Жукова.xlsx")
    if p:
        cost_rows.extend(parse_zhukova(p, catalog))
    p = resolve("Себестоимость Меркушина.xlsx")
    if p:
        cost_rows.extend(parse_merkushina(p, catalog))
    # Мокеева: partial material only + битые merge в xlsx → опционально, не блокируем W3
    for year, fname in [("2025", "себестоимость Мокеева 2025.xlsx"), ("2026", "себестоимость Мокеева 2026.xlsx")]:
        p = resolve(fname)
        if not p:
            continue
        try:
            cost_rows.extend(parse_mokeeva(p, catalog, year))
        except Exception as e:
            print("WARN mokeeva failed", fname, e)

    # ensure collection key exists for all
    for c in cost_rows:
        c.setdefault("collection", "")
        c.setdefault("completeness", "FULL" if c.get("unit_cost_rub") not in ("", None) else "PARTIAL")

    sku = build_sku_master(tsum_lines, cost_rows)
    matched, margin_rows = enrich_tsum_with_cost(tsum_lines, cost_rows)

    # coverage stats
    tsum_skus = {s["canonical_sku"] for s in tsum_lines}
    cost_full_skus = {
        c["canonical_sku"]
        for c in cost_rows
        if c.get("unit_cost_rub") not in ("", None) and float(c.get("unit_cost_rub") or 0) > 0
    }
    inter = tsum_skus & cost_full_skus
    # by year coverage
    by_year = {}
    for y in (2024, 2025, 2026):
        lines_y = [s for s in tsum_lines if s["period_year"] == y]
        skus_y = {s["canonical_sku"] for s in lines_y}
        hit = skus_y & cost_full_skus
        amt_all = sum(float(s["sales_amount_rub"] or 0) for s in lines_y)
        amt_hit = sum(float(s["sales_amount_rub"] or 0) for s in lines_y if s["canonical_sku"] in hit)
        by_year[y] = {
            "tsum_lines": len(lines_y),
            "tsum_skus": len(skus_y),
            "skus_with_cost": len(hit),
            "sku_coverage_pct": round(100 * len(hit) / len(skus_y), 1) if skus_y else 0,
            "amount_all": round(amt_all, 2),
            "amount_with_cost": round(amt_hit, 2),
            "amount_coverage_pct": round(100 * amt_hit / amt_all, 1) if amt_all else 0,
        }

    # margin summary
    margins = [m for m in margin_rows if m["sales_amount_rub"]]
    total_sales = sum(m["sales_amount_rub"] for m in margins)
    total_cogs = sum(m["cogs_rub"] for m in margins)
    total_margin = total_sales - total_cogs

    sku_fields = [
        "canonical_sku", "article_aliases", "name_samples", "seen_in_tsum", "seen_in_cost",
        "in_intersection", "tsum_qty_total", "tsum_amount_total",
    ]
    cost_fields = [
        "cost_version_id", "canonical_sku", "article_raw", "name", "direction", "color", "size",
        "qty", "material_cost", "labor_cost", "unit_cost_rub", "order_cost_rub", "cost_date",
        "cost_year", "owner_file", "channel_hint", "collection", "completeness",
        "source_file_id", "source_sheet", "source_row_id",
    ]
    sales_fields = [
        "sales_line_id", "channel", "period_year", "period_month", "article_raw", "canonical_sku",
        "item_id", "color", "name_raw", "qty", "sales_amount_rub", "amount_meaning", "store",
        "source_file_id", "source_row_id", "cost_version_id", "unit_cost_rub", "margin_rub",
        "margin_pct", "match_status", "cost_owner",
    ]
    for s in tsum_lines:
        s.setdefault("cost_owner", "")

    write_csv(OUT / "sku_master.csv", sku, sku_fields)
    write_csv(OUT / "cost_versions.csv", cost_rows, cost_fields)
    write_csv(OUT / "tsum_sales_lines.csv", tsum_lines, sales_fields)
    write_csv(OUT / "margin_matched.csv", margin_rows, list(margin_rows[0].keys()) if margin_rows else ["canonical_sku"])

    # coverage table
    cov_rows = [{"period_year": y, **by_year[y]} for y in sorted(by_year)]
    write_csv(OUT / "coverage_tsum_vs_cost.csv", cov_rows, list(cov_rows[0].keys()) if cov_rows else ["period_year"])

    summary = {
        "generated_at": NOW,
        "wave": "W3",
        "tsum_sales_lines": len(tsum_lines),
        "tsum_unique_skus": len(tsum_skus),
        "cost_versions": len(cost_rows),
        "cost_full_skus": len(cost_full_skus),
        "sku_master": len(sku),
        "intersection_skus": len(inter),
        "matched_sales_lines": matched,
        "coverage_by_year": by_year,
        "margin_on_matched": {
            "sales_rub": round(total_sales, 2),
            "cogs_rub": round(total_cogs, 2),
            "margin_rub": round(total_margin, 2),
            "margin_pct": round(100 * total_margin / total_sales, 2) if total_sales else None,
            "n_lines": len(margins),
        },
        "cost_owners": dict(Counter(c["owner_file"] for c in cost_rows)),
        "finding": (
            f"W3: TSUM {len(tsum_lines)} lines / {len(tsum_skus)} SKU; "
            f"cost full SKUs {len(cost_full_skus)}; intersection {len(inter)}; "
            f"matched lines {matched}. "
            f"Matched margin {round(100*total_margin/total_sales,1) if total_sales else 'n/a'}% "
            f"on {round(total_sales/1e6,2)}M RUB sales with cost."
        ),
        "next": "W4 SALES/SETTLE or harden cost_version approval + article alias dictionary",
    }
    json.dump(summary, open(OUT / "w3_summary.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    json.dump(summary, open(EV / "w3_summary.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)

    # evidence xlsx
    wb = Workbook()
    ws = wb.active
    ws.title = "00_Summary"
    for i, (k, v) in enumerate(summary.items(), 1):
        ws.cell(i, 1, k)
        ws.cell(i, 2, json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v)
    fill = PatternFill("solid", fgColor="1F4E79")

    def add(name, rows):
        w = wb.create_sheet(name)
        if not rows:
            return
        headers = list(rows[0].keys())
        for c, h in enumerate(headers, 1):
            cell = w.cell(1, c, h)
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
        for ri, row in enumerate(rows[:5000], 2):
            for ci, h in enumerate(headers, 1):
                w.cell(ri, ci, row.get(h, ""))

    add("01_Coverage", cov_rows)
    add("02_Margin_Matched", margin_rows[:3000])
    add("03_SKU_Intersection", [s for s in sku if s["in_intersection"] == "Y"][:2000])
    wb.save(EV / "YANINA_W3_SKU_COST_EVIDENCE.xlsx")

    (OUT / "README.md").write_text(
        f"""# W3 SKU / COST / TSUM margin

Generated: {NOW}

- `sku_master.csv` — {len(sku)} SKU
- `cost_versions.csv` — {len(cost_rows)}
- `tsum_sales_lines.csv` — {len(tsum_lines)}
- `margin_matched.csv` — {len(margin_rows)}
- intersection SKUs: **{len(inter)}**
- matched sales lines: **{matched}**

Evidence: `../../evidence/w3_sku_cost_20260724/`

Controlled Staging — не SoT.
""",
        encoding="utf-8",
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
