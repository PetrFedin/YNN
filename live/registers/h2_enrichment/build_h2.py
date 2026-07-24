#!/usr/bin/env python3
"""
H2: доводка связей + извлечение сумм/кандидатов владельцев из документов.

1) Backfill HIGH/MED settle↔bank → settlements_enriched + patch W4 settlements
2) Строгий fuzzy SUP name ↔ bank (без ложных «си»)
3) Извлечение сумм УСН (и аналогичных line-code полей) из PDF
4) Кандидаты ФИО из подписей налоговых деклараций → Owner Packet (не auto-ACCEPT)

Не SoT. Не подставляет RACI decision.
"""
from __future__ import annotations

import csv
import hashlib
import json
import re
import shutil
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from pypdf import PdfReader

ROOT = Path(__file__).resolve().parents[3]
OUT = Path(__file__).resolve().parent
EV = ROOT / "live/evidence/h2_enrichment_20260724"
W1 = ROOT / "live/registers/w1_bank_cash"
W4 = ROOT / "live/registers/w4_sales_settle"
W5 = ROOT / "live/registers/w5_sup_exp_mat"
W6 = ROOT / "live/registers/w6_tax_bud"
H1 = ROOT / "live/registers/h1_spine_links"
DOCS = ROOT / "documents"
OWNER = ROOT / "live/YANINA_OWNER_PACKET_RACI_AND_REQUESTS.xlsx"
NOW = datetime.now().strftime("%Y-%m-%d %H:%M")


def nfc(s: str) -> str:
    return unicodedata.normalize("NFC", s or "")


def sha16(*parts) -> str:
    raw = "|".join("" if p is None else str(p) for p in parts)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]


def write_csv(path: Path, rows: list[dict], fields: list[str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def norm_name(s: str) -> str:
    s = nfc(s).lower()
    s = re.sub(r"[\"«»'`]", "", s)
    s = re.sub(r"\b(ооо|ип|зао|оао|ао|пао|им)\b", " ", s)
    s = re.sub(r"[^a-zа-я0-9]+", " ", s, flags=re.I)
    return re.sub(r"\s+", " ", s).strip()


def significant_tokens(s: str) -> set[str]:
    # отсекаем короткие токены (ложные «си», «им»)
    return {t for t in norm_name(s).split() if len(t) >= 4}


def resolve(name: str) -> Path | None:
    t = nfc(name)
    for p in DOCS.iterdir():
        if nfc(p.name) == t:
            return p
    return None


# ── 1) Settle backfill ────────────────────────────────────────────
def backfill_settlements() -> dict:
    matches = list(csv.DictReader(open(H1 / "settle_bank_by_doc.csv", encoding="utf-8")))
    settles = list(csv.DictReader(open(W4 / "settlements.csv", encoding="utf-8")))
    by_id = {r["settlement_id"]: r for r in settles}

    # backup once
    bak = W4 / "settlements_pre_h2.csv"
    if not bak.exists():
        shutil.copy2(W4 / "settlements.csv", bak)

    applied = []
    for m in matches:
        if m.get("confidence") not in ("HIGH", "MED"):
            continue
        st = by_id.get(m["settlement_id"])
        if not st:
            continue
        if st.get("bank_payment_id"):
            continue
        st["bank_payment_id"] = m["bank_payment_id"]
        st["status"] = f"LINKED_H2_{m['confidence']}"
        applied.append(
            {
                "settlement_id": m["settlement_id"],
                "bank_payment_id": m["bank_payment_id"],
                "confidence": m["confidence"],
                "match_method": m["match_method"],
                "revenue_rub": m["revenue_rub"],
                "bank_amount": m["bank_amount"],
            }
        )

    fields = list(settles[0].keys())
    write_csv(W4 / "settlements.csv", settles, fields)
    write_csv(OUT / "settle_backfill_applied.csv", applied, list(applied[0].keys()) if applied else ["settlement_id"])
    write_csv(W4 / "settlements_enriched.csv", settles, fields)
    return {"applied": len(applied), "by_conf": dict(Counter(a["confidence"] for a in applied))}


# ── 2) Fuzzy SUP ↔ bank ───────────────────────────────────────────
# (реализация ниже, после extract_signer)


# ── 3) Tax PDF amounts (USN line codes) ───────────────────────────
LINE_AMT_RE = re.compile(
    r"(?<!\d)(\d{3})\s+((?:\d\s+)+\d)\s*(?:-\s*){2,}",
)


def spaced_to_int(s: str) -> int | None:
    digits = re.sub(r"\s+", "", s)
    if not digits.isdigit():
        return None
    # слишком длинные — скорее ИНН/ОКТМО
    if len(digits) > 12:
        return None
    return int(digits)


def extract_usn_amounts(path: Path) -> dict:
    try:
        reader = PdfReader(str(path))
        text = "\n".join((p.extract_text() or "") for p in reader.pages)
    except Exception as e:
        return {"ok": False, "error": str(e)[:80]}
    lines = {}
    for m in LINE_AMT_RE.finditer(text):
        code, raw = m.group(1), m.group(2)
        val = spaced_to_int(raw)
        if val is None:
            continue
        # keep last occurrence per code (forms repeat headers)
        lines[code] = val
    # USN key fields
    mapping = {
        "020": "advance_q1_payable",
        "040": "advance_q2_payable",
        "070": "advance_q3_payable",
        "100": "tax_year_payable",
        "110": "income_q1",
        "111": "income_h1",
        "112": "income_9m",
        "113": "income_year",
        "130": "tax_computed_q1",
        "131": "tax_computed_h1",
        "132": "tax_computed_9m",
        "133": "tax_computed_year",
        "140": "contrib_deduct_q1",
        "141": "contrib_deduct_h1",
        "142": "contrib_deduct_9m",
        "143": "contrib_deduct_year",
    }
    out = {"ok": True, "pages": len(reader.pages), "raw_line_codes": lines}
    for code, name in mapping.items():
        if code in lines:
            out[name] = lines[code]
    return out


SIGNER_RE = re.compile(
    r"Достоверность[\s\S]{0,200}?1\s*\n\s*([А-ЯЁ][А-ЯЁ\s]{2,})\s*\n\s*([А-ЯЁ][А-ЯЁ\s]{2,})\s*\n\s*([А-ЯЁ][А-ЯЁ\s]{2,})",
    re.I,
)


def extract_signer(path: Path) -> str:
    """ФИО подписанта с титула (буквы через пробел), без названия ООО."""
    try:
        text = PdfReader(str(path)).pages[0].extract_text() or ""
    except Exception:
        return ""
    # зона подписи
    low = text.lower()
    start = low.find("достоверность")
    end = low.find("подпись")
    region = text[start:end] if start >= 0 and end > start else text
    STOP = {
        "ОБЩЕСТВО", "ОГРАНИЧЕННОЙ", "ОТВЕТСТВЕННОСТЬЮ", "ДЕКОР", "ООО", "ИП",
        "НАЛОГОПЛАТЕЛЬЩИК", "ПРЕДСТАВИТЕЛЬ", "СТРАНИЦАХ", "ЛИСТАХ", "КОПИЙ",
        "ДОКУМЕНТОВ", "ПОДТВЕРЖДАЮЩИХ", "ПРИЛОЖЕНИЕМ", "ФЕДЕРАЦИИ", "РОССИИ",
    }
    parts = []
    for m in re.finditer(r"(?:[А-ЯЁ] ){2,}[А-ЯЁ]", region):
        w = re.sub(r"\s+", "", m.group(0))
        if w in STOP or not (3 <= len(w) <= 15):
            continue
        if not w.isalpha():
            continue
        parts.append(w)
    # убрать дубли подряд
    dedup = []
    for p in parts:
        if not dedup or dedup[-1] != p:
            dedup.append(p)
    # ФИО = 3 слова; предпочитаем блок, где среднее похоже на имя (окончания)
    for i in range(len(dedup) - 2):
        a, b, c = dedup[i : i + 3]
        if a in STOP or b in STOP or c in STOP:
            continue
        # отчество часто на -ВИЧ/-ВНА/-ИЧНА
        if c.endswith(("ВИЧ", "ВНА", "ИЧНА", "ЬЕВИЧ", "ЬЕВНА")) or len(c) >= 6:
            return _normalize_signer_fio(f"{a} {b} {c}")
    return ""


def _normalize_signer_fio(fio: str) -> str:
    """Чинит артефакты spaced-OCR: ЯНИНАД→ЯНИНА, ЯНИНН→ЯНИН."""
    parts = fio.split()
    if not parts:
        return fio
    s = parts[0]
    if s.startswith("ЯНИНА"):
        parts[0] = "ЯНИНА"
    elif s.startswith("ЯНИН"):
        parts[0] = "ЯНИН"
    return " ".join(parts)


def fuzzy_sup_bank() -> tuple[list[dict], dict]:
    """Только HIGH: точное имя или бренд (1 токен) ⊂ bank name."""
    suppliers = list(csv.DictReader(open(W5 / "suppliers.csv", encoding="utf-8")))
    bank = list(csv.DictReader(open(W1 / "bank_payments.csv", encoding="utf-8")))
    inn_links = list(csv.DictReader(open(H1 / "sup_bank_by_inn.csv", encoding="utf-8")))
    matched_ids = {r["counterparty_id"] for r in inn_links if r["status"] == "MATCHED"}

    cp_groups: dict[str, list] = defaultdict(list)
    for p in bank:
        if p.get("is_internal") == "Y":
            continue
        key = norm_name(p.get("counterparty_raw") or "")
        if key:
            cp_groups[key].append(p)

    rows = []
    for s in suppliers:
        if s["counterparty_id"] in matched_ids:
            continue
        sn = norm_name(s["name"])
        stoks = significant_tokens(s["name"])
        if not sn:
            continue
        best = None
        for bn, pays in cp_groups.items():
            exact = sn == bn
            # бренд: единственный значимый токен длины>=6 входит в bank name
            brand = False
            if len(stoks) == 1:
                tok = next(iter(stoks))
                if len(tok) >= 6 and tok in bn.split():
                    brand = True
            # полное ФИО ИП: все значимые токены поставщика ⊆ bank
            fio = False
            if len(stoks) >= 3 and stoks <= set(bn.split()):
                fio = True
            if not (exact or brand or fio):
                continue
            conf = "HIGH"
            score = 1.0 if exact or fio else 0.9
            cand = (score, conf, bn, pays)
            if best is None or cand[0] > best[0]:
                best = cand
        if not best:
            continue
        score, conf, bn, pays = best
        out_pays = [p for p in pays if p.get("direction") == "out"]
        out_amt = sum(float(p["amount"] or 0) for p in out_pays)
        rows.append(
            {
                "link_id": "SF-" + sha16(s["counterparty_id"], bn),
                "counterparty_id": s["counterparty_id"],
                "supplier_name": s["name"],
                "inn": s.get("inn", ""),
                "bank_name_norm": bn,
                "name_score": round(score, 3),
                "confidence": conf,
                "bank_out_count": len(out_pays),
                "bank_out_rub": round(out_amt, 2),
                "sample_bank_payment_id": (out_pays[0]["bank_payment_id"] if out_pays else ""),
                "match_method": "fuzzy_name_strict",
            }
        )
    stats = {
        "candidates": len(rows),
        "by_conf": dict(Counter(r["confidence"] for r in rows)),
        "out_rub": round(sum(r["bank_out_rub"] for r in rows), 2),
    }
    return rows, stats


def build_tax_enrichment(catalog: list[dict]) -> tuple[list[dict], list[dict], dict]:
    amount_rows = []
    signer_rows = []
    for row in catalog:
        cat = row.get("category") or ""
        if not cat.startswith("tax"):
            continue
        fname = row["file_name"]
        path = resolve(fname)
        if not path or not fname.lower().endswith(".pdf"):
            continue
        signer = extract_signer(path)
        if signer:
            signer_rows.append(
                {
                    "source_file_id": row["master_file_id"],
                    "source_file_name": fname,
                    "tax_type": cat,
                    "signer_fio_candidate": signer,
                    "role_hint": "Tax/Finance (подпись декларации)",
                    "status": "CANDIDATE_NEEDS_CONFIRM",
                }
            )
        if cat == "tax_usn":
            ext = extract_usn_amounts(path)
            amount_rows.append(
                {
                    "source_file_id": row["master_file_id"],
                    "source_file_name": fname,
                    "tax_type": "USN",
                    "extract_ok": "Y" if ext.get("ok") else "N",
                    "income_year": ext.get("income_year", ""),
                    "tax_computed_year": ext.get("tax_computed_year", ""),
                    "tax_year_payable": ext.get("tax_year_payable", ""),
                    "advance_q1_payable": ext.get("advance_q1_payable", ""),
                    "advance_q2_payable": ext.get("advance_q2_payable", ""),
                    "contrib_deduct_year": ext.get("contrib_deduct_year", ""),
                    "raw_codes_json": json.dumps(ext.get("raw_line_codes") or {}, ensure_ascii=False),
                    "amount_status": "EXTRACTED_USN_LINES" if ext.get("ok") else "FAIL",
                }
            )
    # dedupe signers
    uniq = {}
    for r in signer_rows:
        uniq[r["signer_fio_candidate"]] = r
    signer_rows = list(uniq.values())
    stats = {
        "usn_files": len(amount_rows),
        "usn_with_income_year": sum(1 for r in amount_rows if r.get("income_year") != ""),
        "signer_candidates": len(signer_rows),
        "signers": [r["signer_fio_candidate"] for r in signer_rows],
    }
    return amount_rows, signer_rows, stats


def patch_w6_obligations(amount_rows: list[dict]):
    """Проставить amount в tax_obligations для USN, если извлечено."""
    p = W6 / "tax_obligations.csv"
    if not p.exists():
        return 0
    rows = list(csv.DictReader(open(p, encoding="utf-8")))
    by_file = {r["source_file_id"]: r for r in amount_rows if r.get("income_year") != "" or r.get("tax_year_payable") != ""}
    n = 0
    for o in rows:
        if o["tax_type"] != "USN":
            continue
        ar = by_file.get(o["source_file_id"])
        if not ar:
            continue
        # приоритет: к уплате за год, иначе исчисленный, иначе доходы как reference
        amt = ar.get("tax_year_payable") or ar.get("tax_computed_year") or ""
        if amt == "":
            continue
        o["amount"] = amt
        o["amount_status"] = ar["amount_status"]
        n += 1
    write_csv(p, rows, list(rows[0].keys()))
    return n


def update_owner_packet(signers: list[dict]):
    if not OWNER.exists() or not signers:
        return
    wb = load_workbook(OWNER)
    name = "CANDIDATES_FROM_DOCS"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, 1)
    fill = PatternFill("solid", fgColor="1F4E79")
    hdr = Font(color="FFFFFF", bold=True)
    warn = PatternFill("solid", fgColor="FFF2CC")
    ws["A1"] = "Кандидаты ФИО из документов — НЕ auto-ACCEPT в RACI"
    ws["A1"].font = Font(bold=True, size=12, color="1F4E79")
    ws["A2"] = f"Извлечено: {NOW}. Подтвердите личность и перенесите в лист RACI вручную."
    headers = ["signer_fio_candidate", "role_hint", "source_file_name", "status", "action"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(4, c, h)
        cell.fill = fill
        cell.font = hdr
    for ri, r in enumerate(signers, 5):
        ws.cell(ri, 1, r["signer_fio_candidate"]).fill = warn
        ws.cell(ri, 2, r["role_hint"])
        ws.cell(ri, 3, r["source_file_name"])
        ws.cell(ri, 4, r["status"])
        ws.cell(ri, 5, "Если это ваш сотрудник/представитель — внесите в RACI Tax/Finance")
    # README pointer
    if "README" in wb.sheetnames:
        wb["README"]["A12"] = f"Кандидат(ы) из tax PDF: {', '.join(s['signer_fio_candidate'] for s in signers)} — лист CANDIDATES_FROM_DOCS"
    wb.save(OWNER)


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    EV.mkdir(parents=True, exist_ok=True)
    catalog = list(csv.DictReader(open(ROOT / "live/registers/00_SOURCE_CATALOG_93.csv", encoding="utf-8-sig")))

    settle_stats = backfill_settlements()
    fuzzy_rows, fuzzy_stats = fuzzy_sup_bank()
    amount_rows, signer_rows, tax_stats = build_tax_enrichment(catalog)
    usn_patched = patch_w6_obligations(amount_rows)
    update_owner_packet(signer_rows)

    write_csv(
        OUT / "sup_bank_fuzzy.csv",
        fuzzy_rows,
        [
            "link_id", "counterparty_id", "supplier_name", "inn", "bank_name_norm",
            "name_score", "confidence", "bank_out_count", "bank_out_rub",
            "sample_bank_payment_id", "match_method",
        ],
    )
    write_csv(
        OUT / "tax_usn_amounts.csv",
        amount_rows,
        [
            "source_file_id", "source_file_name", "tax_type", "extract_ok",
            "income_year", "tax_computed_year", "tax_year_payable",
            "advance_q1_payable", "advance_q2_payable", "contrib_deduct_year",
            "raw_codes_json", "amount_status",
        ],
    )
    write_csv(
        OUT / "signer_candidates.csv",
        signer_rows,
        ["source_file_id", "source_file_name", "tax_type", "signer_fio_candidate", "role_hint", "status"],
    )

    summary = {
        "generated_at": NOW,
        "wave": "H2",
        "settle_backfill": settle_stats,
        "sup_fuzzy": fuzzy_stats,
        "tax_extract": tax_stats,
        "usn_obligations_patched": usn_patched,
        "finding": (
            f"H2: settle backfill {settle_stats}; "
            f"SUP fuzzy {fuzzy_stats}; "
            f"USN amounts {tax_stats['usn_with_income_year']}/{tax_stats['usn_files']}; "
            f"signer candidates {tax_stats['signers']}."
        ),
        "next": "Confirm signer candidate(s) in Owner Packet → fill RACI → SoT gate",
        "not_sot": True,
    }
    json.dump(summary, open(OUT / "h2_summary.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    json.dump(summary, open(EV / "h2_summary.json", "w", encoding="utf-8"), ensure_ascii=False, indent=2)

    wb = Workbook()
    ws = wb.active
    ws.title = "00_Summary"
    for i, (k, v) in enumerate(summary.items(), 1):
        ws.cell(i, 1, k)
        ws.cell(i, 2, json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v)
    fill = PatternFill("solid", fgColor="1F4E79")

    def add(name, rows):
        w = wb.create_sheet(name)
        if not rows:
            return
        headers = list(rows[0].keys())
        for c, h in enumerate(headers, 1):
            cell = w.cell(1, c, h)
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
        for ri, row in enumerate(rows[:4000], 2):
            for ci, h in enumerate(headers, 1):
                w.cell(ri, ci, row.get(h, ""))

    add("01_Settle_Backfill", list(csv.DictReader(open(OUT / "settle_backfill_applied.csv"))) if (OUT / "settle_backfill_applied.csv").exists() else [])
    add("02_SUP_Fuzzy", fuzzy_rows)
    add("03_USN_Amounts", amount_rows)
    add("04_Signers", signer_rows)
    wb.save(EV / "YANINA_H2_ENRICHMENT_EVIDENCE.xlsx")

    (OUT / "README.md").write_text(
        f"""# H2 Enrichment

Generated: {NOW}

- settle backfill → W4 settlements: {settle_stats}
- SUP fuzzy: {fuzzy_stats}
- USN PDF amounts: {tax_stats}
- Owner Packet ← CANDIDATES_FROM_DOCS (confirm manually)

Evidence: `../../evidence/h2_enrichment_20260724/`
""",
        encoding="utf-8",
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
